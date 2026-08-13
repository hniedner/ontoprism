from __future__ import annotations

import hashlib
import zipfile
from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook

from ontolib.repositories.icdo import ingest

if TYPE_CHECKING:
    from pathlib import Path

pytestmark = pytest.mark.unit


def _workbook(path: Path) -> None:
    workbook = Workbook()
    morphology = workbook.active
    morphology.title = "Morphology"
    morphology.append(["ICD-O-4 Morphology"])
    morphology.append(
        [
            "ICDO4",
            "Level",
            "Term",
            "Code reference",
            "obs",
            "See also",
            "See note",
            "Includes",
            "Excludes",
            "Other text",
        ]
    )
    morphology.append(["80000/0", "Preferred", "Neoplasm, benign"])
    morphology.append(["80000/0", "Synonym", "Benign tumour"])
    morphology.append(["85032/0", "Synonym", "Publisher exception"])
    topography = workbook.create_sheet("Topography")
    topography.append(["ICD-O-4 Topography"])
    topography.append(
        [
            "ICDO4",
            "Level",
            "Term",
            "Note",
            "code_reference",
            "obs",
            "See also",
            "See note",
            "Includes",
            "Excludes",
            "Other text",
        ]
    )
    topography.append([None, "3", "OTHER AND ILL-DEFINED SITES"])
    topography.append(["C76.0", "Preferred", "Head, face or neck, NOS"])
    topography.append(["C76.0", "Related", "Head NOS"])
    workbook.save(path)


def test_synthetic_xlsx_preserves_optional_state_and_category_hierarchy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "ICD-O-4.xlsx"
    _workbook(source)
    monkeypatch.setattr(
        ingest, "ICDO4_SOURCE_SHA256", hashlib.sha256(source.read_bytes()).hexdigest()
    )
    result = ingest.ingest_icdo4(source)
    assert [row.code for row in result.morphology.records] == ["80000/0", "85032/0"]
    assert result.morphology.records[1].preferred is None
    assert result.topography.records[0].code == "C76"
    assert result.topography.records[1].parent_code == "C76"


def test_zip_and_annex_digest_contracts(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "ICD-O-4.xlsx"
    morphology_annex = tmp_path / "Morphology_annexes.xlsx"
    topography_annex = tmp_path / "Topography_annexes.xlsx"
    archive = tmp_path / "ICD-O-4.zip"
    _workbook(source)
    morphology_annex.write_bytes(b"morphology annex")
    topography_annex.write_bytes(b"topography annex")
    with zipfile.ZipFile(archive, "w") as output:
        output.write(source, "ICD-O-4.xlsx")
        output.write(morphology_annex, "Morphology_annexes.xlsx")
        output.write(topography_annex, "Topography_annexes.xlsx")
    monkeypatch.setattr(
        ingest, "ICDO4_ARCHIVE_SHA256", hashlib.sha256(archive.read_bytes()).hexdigest()
    )
    monkeypatch.setattr(
        ingest, "ICDO4_SOURCE_SHA256", hashlib.sha256(source.read_bytes()).hexdigest()
    )
    monkeypatch.setattr(
        ingest,
        "ICDO4_MORPHOLOGY_ANNEX_SHA256",
        hashlib.sha256(morphology_annex.read_bytes()).hexdigest(),
    )
    monkeypatch.setattr(
        ingest,
        "ICDO4_TOPOGRAPHY_ANNEX_SHA256",
        hashlib.sha256(topography_annex.read_bytes()).hexdigest(),
    )
    result = ingest.ingest_icdo4(
        archive,
        morphology_annex_path=morphology_annex,
        topography_annex_path=topography_annex,
    )
    assert result.morphology.archive_sha256 == ingest.ICDO4_ARCHIVE_SHA256
    topography_annex.write_bytes(b"changed")
    with pytest.raises(ingest.SourceFormatError, match="annex SHA-256"):
        ingest.ingest_icdo4(
            archive,
            morphology_annex_path=morphology_annex,
            topography_annex_path=topography_annex,
        )


def test_source_shape_and_required_value_gates_fail_closed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "ICD-O-4.xlsx"
    _workbook(source)
    monkeypatch.setattr(ingest, "ICDO4_SOURCE_SHA256", "0" * 64)
    with pytest.raises(ingest.SourceFormatError, match="workbook SHA-256"):
        ingest.ingest_icdo4(source)
    workbook = Workbook()
    workbook.save(source)
    with pytest.raises(ingest.SourceFormatError, match="sheet names"):
        ingest.ingest_icdo4(source, verify_identity=False)


def test_term_and_cell_gates_fail_closed() -> None:
    duplicate = [("80000/0", "Preferred", "First"), ("80000/0", "Preferred", "Second")]
    with pytest.raises(ingest.SourceFormatError, match="multiple preferred"):
        ingest._records(duplicate, edition="4.0", axis="morphology")
    with pytest.raises(ingest.SourceFormatError, match="preferred term is absent"):
        ingest._records(
            [("80000/0", "Synonym", "Only synonym")], edition="4.0", axis="morphology"
        )
    with pytest.raises(ingest.SourceFormatError, match="unsupported source cell type"):
        ingest._text(object())
