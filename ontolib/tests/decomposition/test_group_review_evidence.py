from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from scripts.adjudication import _parser
from scripts.research import group_review_packet as group_review

pytestmark = pytest.mark.unit

_ROOT = Path(__file__).parents[3]
_GOLDEN = Path(__file__).with_name("golden")
_PACKET = _ROOT / "tmp/m1-6-group-review-packet.json"
_MARKDOWN = _ROOT / "evidence/group-review-rationale-26.07d.md"
_SIDECAR = _ROOT / "evidence/group-review-rationale-26.07d.json"
_SOURCE_MARKDOWN = _ROOT / "tmp/m1-6-group-review-rationale-template.md"


def _loaded():
    packet = group_review.load_group_review_packet(_PACKET)
    return packet, group_review.load_group_review_rationale_evidence(
        markdown_path=_MARKDOWN,
        sidecar_path=_SIDECAR,
        packet=packet,
    )


def _mutated_markdown(tmp_path: Path, mutate) -> tuple[Path, Path]:
    markdown = tmp_path / _MARKDOWN.name
    markdown.write_bytes(mutate(_MARKDOWN.read_bytes()))
    sidecar = tmp_path / _SIDECAR.name
    sidecar.write_bytes(_SIDECAR.read_bytes())
    return markdown, sidecar


def test_tracked_markdown_is_frozen_verbatim_and_strictly_bound() -> None:
    packet, evidence = _loaded()

    assert packet.ncit_version == "26.07d"
    assert _MARKDOWN.read_bytes() == _SOURCE_MARKDOWN.read_bytes()
    assert evidence.reviewer == "R. Hannes Niedner, M.D."
    assert evidence.review_date == "2026-08-28"
    assert len(evidence.rows) == 18
    assert tuple(row.concept_code for row in evidence.rows) == tuple(
        row.concept_code for row in packet.review_rows
    )
    assert tuple(row.review_row_identity for row in evidence.rows) == tuple(
        row.row_identity for row in packet.review_rows
    )
    assert all(row.rationale and row.rationale.strip() for row in evidence.rows)
    assert Counter(row.decision for row in evidence.rows) == {
        "Approve intentional normalization": 3,
        "Require source-reproducible correction": 11,
        "Abstain / escalate": 4,
    }
    assert all(
        row.pair_decision == row.decision
        for row in evidence.rows
        if row.review_type == "pair-only"
    )
    assert all(
        row.pair_decision is None
        for row in evidence.rows
        if row.review_type == "grouping"
    )


def test_admission_writes_frozen_markdown_before_computing_sidecar(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    packet = group_review.load_group_review_packet(_PACKET)
    source = tmp_path / "completed.md"
    source.write_bytes(_SOURCE_MARKDOWN.read_bytes())
    (tmp_path / "evidence").mkdir()
    monkeypatch.chdir(tmp_path)

    sidecar = group_review.admit_group_review_rationale_evidence(
        packet=packet,
        source_markdown=source,
        markdown_output=Path("evidence/group-review-rationale-26.07d.md"),
        sidecar_output=Path("evidence/group-review-rationale-26.07d.json"),
    )

    admitted = Path(sidecar.markdown_path)
    assert admitted.read_bytes() == source.read_bytes()
    assert (
        sidecar.markdown_sha256
        == group_review.hashlib.sha256(admitted.read_bytes()).hexdigest()
    )
    assert (
        group_review.load_group_review_rationale_evidence(
            markdown_path=admitted,
            sidecar_path=Path("evidence/group-review-rationale-26.07d.json"),
            packet=packet,
        ).sidecar
        == sidecar
    )


def test_parser_treats_only_human_rationale_block_as_opaque_text() -> None:
    _packet_value, evidence = _loaded()
    first = next(row for row in evidence.rows if row.concept_code == "C181564")

    assert "Questions for the human rationale" not in first.rationale
    assert "Why is separating" not in first.rationale
    assert "HL7 FHIR mCODE interoperability standards" in first.rationale
    assert "\nfrom the staging result" in first.rationale


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda value: value.replace(b"HL7", b"hL7", 1), "markdown digest"),
        (
            lambda value: (
                value[: value.index(b"## 2.")] + value[value.index(b"## 3.") :]
            ),
            "review sections",
        ),
        (
            lambda value: value.replace(b"## 2. C186620", b"## 2. C181564", 1),
            "review sections",
        ),
        (
            lambda value: value.replace(b"## 1.", b"## 19.", 1),
            "review sections",
        ),
        (
            lambda value: value.replace(b"Human rationale:", b"Human rationale", 1),
            "Human rationale marker",
        ),
    ],
)
def test_loader_rejects_changed_deleted_duplicate_reordered_or_malformed_markdown(
    tmp_path: Path, mutation, message: str
) -> None:
    packet = group_review.load_group_review_packet(_PACKET)
    markdown, sidecar = _mutated_markdown(tmp_path, mutation)

    with pytest.raises(ValueError, match=message):
        group_review.load_group_review_rationale_evidence(
            markdown_path=markdown, sidecar_path=sidecar, packet=packet
        )


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda value: value | {"unexpected": True}, "extra"),
        (lambda value: value | {"markdown_sha256": "0" * 64}, "markdown digest"),
        (
            lambda value: (
                value
                | {
                    "rows": [
                        (
                            {**row, "rationale_chars": row["rationale_chars"] + 1}
                            if i == 0
                            else row
                        )
                        for i, row in enumerate(value["rows"])
                    ]
                }
            ),
            "rationale character count",
        ),
        (lambda value: value | {"sidecar_identity": "0" * 64}, "sidecar identity"),
    ],
)
def test_loader_rejects_malformed_digest_row_and_unknown_sidecar_fields(
    tmp_path: Path, mutation, message: str
) -> None:
    packet = group_review.load_group_review_packet(_PACKET)
    payload = mutation(json.loads(_SIDECAR.read_bytes()))
    if "unexpected" not in payload and payload["sidecar_identity"] != "0" * 64:
        payload["sidecar_identity"] = group_review._identity(
            {key: value for key, value in payload.items() if key != "sidecar_identity"}
        )
    sidecar = tmp_path / _SIDECAR.name
    sidecar.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises((ValueError, ValidationError), match=message):
        group_review.load_group_review_rationale_evidence(
            markdown_path=_MARKDOWN, sidecar_path=sidecar, packet=packet
        )


def test_governed_transcription_round_trips_real_workbook_and_dry_run(
    tmp_path: Path,
) -> None:
    packet, evidence = _loaded()
    workbook = tmp_path / "m1-6-group-review-workbook-reviewed.xlsx"
    registry_path = tmp_path / "m1-6-group-review-decisions.json"
    dry_run_path = tmp_path / "m1-6-group-review-dry-run.json"

    registry, dry_run = group_review.transcribe_group_review_rationale_evidence(
        packet=packet,
        evidence=evidence,
        workbook=workbook,
        registry_output=registry_path,
        dry_run_output=dry_run_path,
    )

    assert tuple(
        (
            row.review_row_identity,
            row.concept_code,
            row.review_type,
            row.pair_decision,
            row.decision,
            group_review.rationale_sha256(row.rationale),
            len(row.rationale),
            row.reviewer,
            row.review_date,
        )
        for row in registry.decisions
    ) == tuple(
        (
            row.review_row_identity,
            row.concept_code,
            row.review_type,
            row.pair_decision,
            row.decision,
            row.rationale_sha256,
            row.rationale_chars,
            evidence.reviewer,
            evidence.review_date,
        )
        for row in evidence.rows
    )
    assert registry == group_review.load_group_decision_registry(registry_path)
    assert json.loads(dry_run_path.read_bytes()) == dry_run.model_dump(mode="json")
    assert dry_run.writes_performed is False
    assert dry_run.unresolved_count == 11
    assert dry_run.deferred_count == 4

    book = load_workbook(workbook, data_only=False)
    assert book["Bindings"].sheet_state == "veryHidden"
    assert all(sheet.protection.sheet for sheet in book.worksheets)
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in book.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_first_evidence_admission_inventory_is_exact_and_rejects_binary() -> None:
    observed = tuple(
        sorted(
            path.relative_to(_ROOT).as_posix()
            for path in (_ROOT / "evidence").iterdir()
        )
    )
    expected = (
        "evidence/README.md",
        "evidence/group-review-rationale-26.07d.json",
        "evidence/group-review-rationale-26.07d.md",
    )

    assert observed == expected
    group_review.validate_first_evidence_inventory(observed, ncit_version="26.07d")
    with pytest.raises(ValueError, match="evidence inventory"):
        group_review.validate_first_evidence_inventory(
            (*observed, "evidence/review.xlsx"), ncit_version="26.07d"
        )


def test_evidence_docs_record_exact_governed_commands_and_open_status() -> None:
    docs = (_ROOT / "docs/evidence/README.md").read_text(encoding="utf-8")
    golden = (_GOLDEN / "README.md").read_text(encoding="utf-8")
    command = (
        "pdm run adjudication transcribe-group-review-evidence --packet "
        "tmp/m1-6-group-review-packet.json --markdown "
        "evidence/group-review-rationale-26.07d.md --sidecar "
        "evidence/group-review-rationale-26.07d.json --reviewed-xlsx "
        "tmp/m1-6-group-review-workbook-reviewed.xlsx --registry-output "
        "tmp/m1-6-group-review-decisions.json --dry-run-output "
        "tmp/m1-6-group-review-dry-run.json"
    )

    for text in (docs, golden):
        assert command in text
        assert "11 corrections" in text
        assert "4 escalations" in text
        assert "#274" in text
        assert "#127" in text
        assert "publication" in text.casefold()


def test_group_review_evidence_cli_names_every_boundary_path() -> None:
    args = _parser().parse_args(
        [
            "transcribe-group-review-evidence",
            "--packet",
            "packet.json",
            "--markdown",
            "rationale.md",
            "--sidecar",
            "rationale.json",
            "--reviewed-xlsx",
            "reviewed.xlsx",
            "--registry-output",
            "registry.json",
            "--dry-run-output",
            "dry-run.json",
        ]
    )

    assert (
        args.packet,
        args.markdown,
        args.sidecar,
        args.reviewed_xlsx,
        args.registry_output,
        args.dry_run_output,
    ) == tuple(
        Path(value)
        for value in (
            "packet.json",
            "rationale.md",
            "rationale.json",
            "reviewed.xlsx",
            "registry.json",
            "dry-run.json",
        )
    )
