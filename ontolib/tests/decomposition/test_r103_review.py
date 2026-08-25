from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import cast
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from defusedxml import ElementTree as DefusedET
from openpyxl import load_workbook
from scripts.adjudication import _parser

from ontolib.decomposition import r103_review
from ontolib.decomposition.proposal_registry import load_proposal_registry
from ontolib.decomposition.r103_review import (
    R103ReviewValidationError,
    build_r103_review_packet,
    dry_run_r103_review,
    import_r103_review_decisions,
    load_r103_review_packet,
    write_r103_review_packet,
    write_r103_review_workbook,
)

EXPECTED = (
    ("C2860", "R103", "C12950"),
    ("C3264", "R103", "C12950"),
    ("C3716", "R103", "C34228"),
)
OUTCOMES = (
    "source-supported",
    "correction-proposal",
    "concept-scoped-accuracy-exclusion",
    "review-required",
)
HUMAN_FIELDS = ("Outcome", "Rationale", "Reviewer", "Date")
NCIT = "http://ncicb.nci.nih.gov/xml/owl/EVS/Thesaurus.owl#"
RDF = "http://www.w3.org/1999/02/22-rdf-syntax-ns#"
RDFS = "http://www.w3.org/2000/01/rdf-schema#"
OWL = "http://www.w3.org/2002/07/owl#"


def _class(
    code: str,
    label: str,
    definition: str,
    members: tuple[tuple[str, ...], ...] = (),
) -> str:
    rendered = []
    for member in members:
        if member[0] == "genus":
            rendered.append(f'<rdf:Description rdf:about="{NCIT}{member[1]}"/>')
        else:
            rendered.append(
                "<owl:Restriction>"
                f'<owl:onProperty rdf:resource="{NCIT}{member[1]}"/>'
                f'<owl:someValuesFrom rdf:resource="{NCIT}{member[2]}"/>'
                "</owl:Restriction>"
            )
    equivalent = (
        "<owl:equivalentClass><owl:Class>"
        '<owl:intersectionOf rdf:parseType="Collection">'
        + "".join(rendered)
        + "</owl:intersectionOf></owl:Class></owl:equivalentClass>"
        if rendered
        else ""
    )
    return (
        f'<owl:Class rdf:about="{NCIT}{code}">{equivalent}'
        f"<P97>{definition}</P97><rdfs:label>{label}</rdfs:label></owl:Class>"
    )


@pytest.mark.unit
def test_definition_parser_ignores_subclass_intersection_without_equivalent_class() -> (
    None
):
    element = DefusedET.fromstring(
        f'<owl:Class xmlns:rdf="{RDF}" xmlns:rdfs="{RDFS}" xmlns:owl="{OWL}" '
        f'rdf:about="{NCIT}C2860"><rdfs:subClassOf><owl:Class>'
        '<owl:intersectionOf rdf:parseType="Collection">'
        f'<rdf:Description rdf:about="{NCIT}C7617"/>'
        "</owl:intersectionOf></owl:Class></rdfs:subClassOf></owl:Class>"
    )

    assert r103_review._definition_from_class("C2860", element) is None


@pytest.fixture
def source_boundary(tmp_path: Path) -> tuple[Path, Path, Path, Path]:
    classes = (
        _class(
            "C2860",
            "Adrenal Rest Tumor",
            "Adrenal-rest definition",
            (
                ("genus", "C7617"),
                ("restriction", "R101", "C12841"),
                ("restriction", "R103", "C12950"),
            ),
        ),
        _class(
            "C3264",
            "Embryonal Neoplasm",
            "Embryonal definition",
            (
                ("genus", "C7062"),
                ("restriction", "R103", "C12950"),
                ("restriction", "R104", "C13054"),
            ),
        ),
        _class(
            "C3716",
            "Primitive Neuroectodermal Tumor",
            "PNET definition",
            (
                ("genus", "C3264"),
                ("genus", "C9305"),
                ("restriction", "R103", "C34228"),
                ("restriction", "R104", "C42050"),
            ),
        ),
        _class(
            "C3708",
            "Germ Cell Tumor",
            "Germ-cell definition",
            (
                ("genus", "C4741"),
                ("restriction", "R103", "C54105"),
                ("restriction", "R104", "C12597"),
            ),
        ),
        _class("R103", "Disease_Has_Normal_Tissue_Origin", "R103 definition"),
        _class("R104", "Disease_Has_Normal_Cell_Origin", "R104 definition"),
        *(
            _class(code, f"Label {code}", f"Definition {code}")
            for code in (
                "C12950",
                "C34228",
                "C54105",
                "C7617",
                "C7062",
                "C9305",
                "C12841",
                "C13054",
                "C42050",
                "C4741",
                "C12597",
            )
        ),
    )
    owl = tmp_path / "Thesaurus-stated.owl"
    owl.write_text(
        f'<rdf:RDF xmlns:rdf="{RDF}" xmlns:rdfs="{RDFS}" xmlns:owl="{OWL}" '
        f'xmlns="{NCIT}"><owl:Ontology rdf:about="{NCIT[:-1]}">'
        "<owl:versionInfo>26.07d</owl:versionInfo></owl:Ontology>"
        + "".join(classes)
        + "</rdf:RDF>",
        encoding="utf-8",
    )
    owl_sha = hashlib.sha256(owl.read_bytes()).hexdigest()
    manifest = tmp_path / "candidate.json"
    proposals = Path("ontolib/tests/decomposition/golden/proposal-registry.json")
    manifest_payload = {
        "schema_version": 3,
        "ontology_version": "26.07d",
        "source_identity": "b" * 64,
        "stated_artifact": {
            "path": str(owl),
            "sha256": owl_sha,
            "size_bytes": owl.stat().st_size,
            "variant": "stated",
            "artifact_identity": "a" * 64,
        },
    }
    manifest.write_text(json.dumps(manifest_payload, sort_keys=True), encoding="utf-8")
    oracle = tmp_path / "oracle.json"
    oracle.write_text('{"unchanged":true}\n', encoding="utf-8")
    return owl, manifest, proposals, oracle


@pytest.fixture
def packet(source_boundary):
    owl, manifest, proposals, _oracle = source_boundary
    return build_r103_review_packet(owl, manifest, proposals)


@pytest.mark.unit
def test_packet_has_exact_source_derived_inventory_and_complete_evidence(
    packet,
) -> None:
    assert (
        tuple((row.subject_code, row.role_code, row.filler_code) for row in packet.rows)
        == EXPECTED
    )
    assert packet.source_release == "26.07d"
    assert packet.source_identity == "b" * 64
    assert packet.source_artifact_sha256
    assert packet.candidate_manifest_identity
    assert packet.query_contract_identity
    assert packet.tool_identity
    assert packet.packet_identity
    assert (
        packet.inventory_scope == "issue-declared assertions, source-presence certified"
    )
    assert packet.method_reference.subject_code == "C3708"
    assert packet.method_reference.filler_code == "C54105"
    assert packet.method_reference.is_decision_row is False
    assert all(row.role_is_non_defining for row in packet.rows)
    for row in packet.rows:
        assert row.subject_label
        assert row.subject_p97_definition
        assert row.role_label
        assert row.role_p97_definition
        assert row.filler_label
        assert row.filler_p97_definition
        assert row.complete_definition_identity
        assert row.source_fact_identity
        assert row.source_group_identity
        assert row.source_occurrence_identity
        assert row.anchor_code == row.subject_code
        assert row.depth == 0
        assert row.structural_path[-1] == row.member_position
        assert row.genus_facts
        assert all(fact.role_code in {"R103", "R104"} for fact in row.coasserted_facts)
        assert row.current_state in {"projected", "suppressed", "review-required"}
        assert row.machine_evidence.startswith("Mechanical evidence only;")
        assert "recommend" not in row.machine_evidence.casefold()
        assert row.contrast_to_method.startswith("Comparison only;")
    assert [fact.role_code for fact in packet.rows[1].coasserted_facts] == [
        "R103",
        "R104",
    ]
    assert packet.rows[2].genus_facts[0].member_position == 0
    assert packet.rows[2].genus_facts[1].member_position == 1


@pytest.mark.unit
def test_packet_regeneration_is_deterministic_and_source_bound(
    source_boundary, tmp_path: Path
) -> None:
    owl, manifest, proposals, _oracle = source_boundary
    first = build_r103_review_packet(owl, manifest, proposals)
    second = build_r103_review_packet(owl, manifest, proposals)
    assert first == second
    first_path = tmp_path / "first.json"
    second_path = tmp_path / "second.json"
    write_r103_review_packet(first_path, first)
    write_r103_review_packet(second_path, second)
    assert first_path.read_bytes() == second_path.read_bytes()
    first_workbook = tmp_path / "first.xlsx"
    second_workbook = tmp_path / "second.xlsx"
    write_r103_review_workbook(first_workbook, first)
    write_r103_review_workbook(second_workbook, second)
    assert first_workbook.read_bytes() == second_workbook.read_bytes()
    assert load_r103_review_packet(first_path) == first
    owl.write_text(
        owl.read_text().replace("PNET definition", "changed"), encoding="utf-8"
    )
    with pytest.raises(R103ReviewValidationError, match="stale source artifact"):
        build_r103_review_packet(owl, manifest, proposals)


@pytest.mark.unit
def test_workbook_has_only_three_rows_four_blank_human_fields_and_no_decision_language(
    packet, tmp_path: Path
) -> None:
    path = tmp_path / "review.xlsx"
    write_r103_review_workbook(path, packet)
    book = load_workbook(path)
    assert book.sheetnames == [
        "Instructions",
        "R103 Review",
        "Method Reference",
        "Bindings",
    ]
    assert book["Bindings"].sheet_state == "veryHidden"
    sheet = book["R103 Review"]
    headers = tuple(cell.value for cell in sheet[1])
    assert headers[-4:] == HUMAN_FIELDS
    assert sheet.max_row == 4
    assert tuple(sheet.cell(row, 1).value for row in range(2, 5)) == tuple(
        item[0] for item in EXPECTED
    )
    assert all(
        sheet.cell(row, column).value is None
        for row in range(2, 5)
        for column in range(sheet.max_column - 3, sheet.max_column + 1)
    )
    visible = "\n".join(str(cell.value or "") for row in sheet for cell in row)
    assert "content approval" not in visible.casefold()
    assert "recommended outcome" not in visible.casefold()
    method_links = tuple(
        sheet.cell(row, headers.index("Method Reference") + 1).value
        for row in range(2, 5)
    )
    assert method_links == ("Method Reference row 2",) * 3

    method = book["Method Reference"]
    method_headers = tuple(cell.value for cell in method[1])
    assert {
        "Subject Code",
        "Subject Label",
        "Subject P97 Definition",
        "Role Code",
        "Filler Code",
        "Filler Label",
        "Filler P97 Definition",
        "Source Restriction",
        "Coasserted R103/R104 Facts",
        "Comparison Scope",
        "Decision Row",
    } <= set(method_headers)
    assert method.cell(2, method_headers.index("Subject Code") + 1).value == "C3708"
    assert method.cell(2, method_headers.index("Role Code") + 1).value == "R103"
    assert method.cell(2, method_headers.index("Filler Code") + 1).value == "C54105"
    assert method.cell(2, method_headers.index("Decision Row") + 1).value == "No"


def _reviewed(
    packet, path: Path, outcomes: tuple[str, str, str] = OUTCOMES[:3]
) -> None:
    write_r103_review_workbook(path, packet)
    book = load_workbook(path)
    sheet = book["R103 Review"]
    headers = {cell.value: cast("int", cell.column) for cell in sheet[1]}
    for row, outcome in enumerate(outcomes, start=2):
        rationale: str = "Human review rationale"
        if outcome == "correction-proposal":
            rationale = json.dumps(
                {
                    "human_rationale": "Source needs correction",
                    "proposed_correction": "Replace the asserted filler with C54105",
                    "duplicate_search_evidence": (
                        "NCIt 26.07d query: no equivalent correction"
                    ),
                    "provenance": "SME review of source-bound packet",
                    "lifecycle": "proposed",
                },
                sort_keys=True,
            )
        sheet.cell(row, headers["Outcome"], outcome)
        sheet.cell(row, headers["Rationale"], rationale)
        sheet.cell(row, headers["Reviewer"], "TEST SME")
        sheet.cell(row, headers["Date"], "2099-01-02")
    book.save(path)


@pytest.mark.unit
def test_import_binds_exact_decisions_and_preserves_narrow_exclusion(
    packet, tmp_path: Path
) -> None:
    workbook = tmp_path / "reviewed.xlsx"
    _reviewed(packet, workbook)
    registry = import_r103_review_decisions(
        packet, workbook, tmp_path / "registry.json"
    )
    assert len(registry.decisions) == 3
    assert {row.outcome for row in registry.decisions} == set(OUTCOMES[:3])
    assert all(
        row.packet_identity == packet.packet_identity for row in registry.decisions
    )
    assert all(
        row.workbook_identity == registry.workbook_identity
        for row in registry.decisions
    )
    assert all(
        row.source_identity == packet.source_identity for row in registry.decisions
    )
    exclusion = registry.exclusion_preview[0]
    assert exclusion.model_dump() == {
        "subject_code": "C3716",
        "role_code": "R103",
        "filler_code": "C34228",
        "source_identity": packet.source_identity,
        "source_release": "26.07d",
    }
    assert registry.proposal_preview[0].lifecycle == "proposed"


@pytest.mark.unit
def test_import_normalizes_excel_datetime_to_iso_date(packet, tmp_path: Path) -> None:
    workbook = tmp_path / "reviewed.xlsx"
    _reviewed(packet, workbook)
    book = load_workbook(workbook)
    headers = {cell.value: cast("int", cell.column) for cell in book["R103 Review"][1]}
    book["R103 Review"].cell(2, headers["Date"], datetime(2099, 1, 2, 13, 45))
    book.save(workbook)

    registry = import_r103_review_decisions(
        packet, workbook, tmp_path / "registry.json"
    )

    assert registry.decisions[0].review_date == "2099-01-02"


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("formula", "formula"),
        ("macro", "macro"),
        ("external", "external link"),
        ("hidden", "hidden row"),
        ("edit", "immutable row"),
        ("reorder", "immutable row"),
        ("duplicate", "exactly three"),
        ("missing", "exactly three"),
        ("extra", "exactly three"),
        ("unknown", "closed outcome"),
        ("blank", "required human field"),
        ("date", "ISO date"),
    ],
)
def test_import_refuses_workbook_tampering_and_invalid_human_fields(
    packet, tmp_path: Path, mutation: str, message: str
) -> None:
    path = tmp_path / f"{mutation}.xlsx"
    _reviewed(packet, path)
    if mutation in {"macro", "external"}:
        member = (
            "xl/vbaProject.bin"
            if mutation == "macro"
            else "xl/externalLinks/externalLink1.xml"
        )
        with ZipFile(path, "a", ZIP_DEFLATED) as archive:
            archive.writestr(member, b"x")
    else:
        _mutate_workbook(path, mutation)
    with pytest.raises(R103ReviewValidationError, match=message):
        import_r103_review_decisions(packet, path, tmp_path / "must-not-exist.json")
    assert not (tmp_path / "must-not-exist.json").exists()


def _mutate_workbook(path: Path, mutation: str) -> None:  # noqa: C901
    book = load_workbook(path)
    sheet = book["R103 Review"]
    headers = {cell.value: cast("int", cell.column) for cell in sheet[1]}
    if mutation == "formula":
        sheet.cell(2, headers["Outcome"], "=1+1")
    elif mutation == "hidden":
        sheet.row_dimensions[2].hidden = True
    elif mutation == "edit":
        sheet.cell(2, 2, "R104")
    elif mutation == "reorder":
        sheet.move_range(f"A2:{sheet.cell(3, sheet.max_column).coordinate}", rows=1)
    elif mutation == "duplicate":
        sheet.append([cell.value for cell in sheet[2]])
    elif mutation == "missing":
        sheet.delete_rows(4)
    elif mutation == "extra":
        sheet.append(["C999"])
    elif mutation == "unknown":
        sheet.cell(2, headers["Outcome"], "approved")
    elif mutation == "blank":
        sheet.cell(2, headers["Reviewer"]).value = None
    elif mutation == "date":
        sheet.cell(2, headers["Date"], "02/01/2099")
    book.save(path)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("payload_mutation", "message"),
    [
        (lambda value: value.pop("proposed_correction"), "proposed_correction"),
        (
            lambda value: value.pop("duplicate_search_evidence"),
            "duplicate_search_evidence",
        ),
        (lambda value: value.pop("provenance"), "provenance"),
        (lambda value: value.update(lifecycle="locally-approved"), "lifecycle"),
    ],
)
def test_correction_proposal_requires_explicit_proposed_evidence(
    packet, tmp_path: Path, payload_mutation, message: str
) -> None:
    path = tmp_path / "proposal.xlsx"
    _reviewed(packet, path)
    book = load_workbook(path)
    sheet = book["R103 Review"]
    headers = {cell.value: cast("int", cell.column) for cell in sheet[1]}
    payload = json.loads(cast("str", sheet.cell(3, headers["Rationale"]).value))
    payload_mutation(payload)
    sheet.cell(3, headers["Rationale"], json.dumps(payload))
    book.save(path)
    with pytest.raises(R103ReviewValidationError, match=message):
        import_r103_review_decisions(packet, path, tmp_path / "none.json")


@pytest.mark.unit
def test_dry_run_reports_no_writes_and_preserves_oracle_and_proposal_registry(
    packet, source_boundary, tmp_path: Path
) -> None:
    _owl, _manifest, proposal_path, oracle_path = source_boundary
    workbook = tmp_path / "reviewed.xlsx"
    registry_path = tmp_path / "decisions.json"
    _reviewed(packet, workbook)
    registry = import_r103_review_decisions(packet, workbook, registry_path)
    proposal_before = proposal_path.read_bytes()
    oracle_before = oracle_path.read_bytes()
    proposal_registry = load_proposal_registry(proposal_path)
    result = dry_run_r103_review(
        packet, registry, oracle_path=oracle_path, proposal_registry_path=proposal_path
    )
    assert result.writes_performed is False
    assert result.outcome_counts == dict.fromkeys(OUTCOMES[:3], 1)
    assert len(result.proposal_previews) == 1
    assert len(result.exclusion_previews) == 1
    assert result.unresolved == 0
    assert result.readiness == "ready-for-separate-application"
    assert proposal_path.read_bytes() == proposal_before
    assert oracle_path.read_bytes() == oracle_before
    assert load_proposal_registry(proposal_path) == proposal_registry


@pytest.mark.unit
def test_import_refuses_stale_packet_source(packet, tmp_path: Path) -> None:
    workbook = tmp_path / "reviewed.xlsx"
    _reviewed(packet, workbook)
    stale = packet.model_copy(update={"source_identity": "0" * 64})
    with pytest.raises(R103ReviewValidationError, match="packet identity"):
        import_r103_review_decisions(stale, workbook, tmp_path / "none.json")


@pytest.mark.unit
def test_cli_exposes_offline_prepare_import_and_dry_run() -> None:
    parser = _parser()
    prepare = parser.parse_args(
        [
            "prepare-r103-review-packet",
            "--stated-owl",
            "source.owl",
            "--source-manifest",
            "source.json",
            "--proposal-registry",
            "proposals.json",
            "--output-packet",
            "packet.json",
            "--output-xlsx",
            "review.xlsx",
        ]
    )
    assert prepare.command == "prepare-r103-review-packet"
    importer = parser.parse_args(
        [
            "import-r103-review-decisions",
            "--packet",
            "packet.json",
            "--reviewed-xlsx",
            "reviewed.xlsx",
            "--output",
            "decisions.json",
        ]
    )
    assert importer.command == "import-r103-review-decisions"
    dry_run = parser.parse_args(
        [
            "dry-run-r103-review",
            "--packet",
            "packet.json",
            "--registry",
            "decisions.json",
            "--oracle",
            "oracle.json",
            "--proposal-registry",
            "proposals.json",
            "--output",
            "dry-run.json",
        ]
    )
    assert dry_run.command == "dry-run-r103-review"
