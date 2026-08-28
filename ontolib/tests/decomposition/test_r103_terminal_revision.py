from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import load_workbook
from scripts.adjudication import _parser

from ontolib.decomposition import r103_review_promotion
from ontolib.decomposition.axes import (
    UNSUPPORTED_FILLERS_BY_CONCEPT_ROLE,
    is_unsupported_filler,
)
from ontolib.decomposition.r103_review import R103ReviewValidationError
from ontolib.decomposition.r103_review_promotion import load_r103_corroboration

GOLDEN = Path(__file__).with_name("golden")
REV1 = GOLDEN / "r103-review-state-26.07d.json"
REV2 = GOLDEN / "r103-review-state-26.07d-rev2.json"
CORROBORATION = GOLDEN / "r103-c3264-corroboration-26.07d.json"
ASSERTION = ("C3264", "R103", "C12950")
REVIEWER = "R. Hannes Niedner, M.D."
REVIEW_DATE = "2026-08-28"
OUTCOME = "concept-scoped-accuracy-exclusion"
RATIONALE = (
    "My Recommendation: Concept-scoped exclusion\n\n"
    "Rationale from Scientific Literature: The modern understanding of CNS and "
    "peripheral embryonal tumors (such as medulloblastoma, atypical "
    "teratoid/rhabdoid tumor, and ETMR) has shifted significantly with molecular "
    "profiling. Literature confirms that these tumors arise from the transformation "
    "of very specific local progenitor populations or stem cells (e.g., transitional "
    "cerebellar progenitors in the rhombic lip for certain medulloblastomas) whose "
    'developmental program stalls—not from a generalized pool of "embryonic tissue."'
    '\n\nThe term "embryonal" in oncology refers primarily to the primitive, '
    "undifferentiated morphologic appearance of the tumor cells (small round blue "
    "cells resembling those in a developing embryo) rather than a literal derivation "
    "from generic embryonic tissue. Because R103 represents a strict causal origin "
    '(Disease_Has_Normal_Tissue_Origin), applying a broad "Embryonic Tissue" filler '
    "across this entire umbrella misrepresents the biology.\n\n"
    "Therefore, selecting Concept-scoped exclusion is the correct semantic action. "
    "It safely prevents this morphologic resemblance from being falsely projected as "
    "a strict anatomic origin in downstream reasoning, while fully preserving the "
    "original NCIt source assertion and its provenance in the graph."
)
QUALIFICATION = (
    "R103 is non-defining; this exclusion applies exactly to the C3264/R103/C12950 "
    "source assertion, and individual descendants may have specific embryonic or "
    "fetal origins."
)


def _revision_api() -> tuple[Any, Any, Any, Any]:
    prepare = cast(
        "Any", getattr(r103_review_promotion, "prepare_r103_review_revision", None)
    )
    transcribe = cast(
        "Any", getattr(r103_review_promotion, "transcribe_r103_review_revision", None)
    )
    promote = cast(
        "Any", getattr(r103_review_promotion, "promote_r103_review_revision", None)
    )
    loader = cast(
        "Any",
        getattr(r103_review_promotion, "load_r103_promoted_review_revision", None),
    )
    assert callable(prepare), "governed R103 revision preparation is missing"
    assert callable(transcribe), "governed explicit-user transcription is missing"
    assert callable(promote), "append-only R103 revision promotion is missing"
    assert callable(loader), "strict R103 revision loader is missing"
    return prepare, transcribe, promote, loader


def _build_revision(tmp_path: Path):
    prepare, transcribe, promote, loader = _revision_api()
    blank = tmp_path / "r103-revision-blank.xlsx"
    reviewed = tmp_path / "r103-revision-transcribed.xlsx"
    registry = tmp_path / "r103-revision-decisions.json"
    dry_run = tmp_path / "r103-revision-dry-run.json"
    output = tmp_path / "r103-review-state-26.07d-rev2.json"
    prepare(predecessor_path=REV1, output_workbook_path=blank)
    transcribe(
        predecessor_path=REV1,
        blank_workbook_path=blank,
        output_workbook_path=reviewed,
        assertion=ASSERTION,
        outcome=OUTCOME,
        rationale=RATIONALE,
        reviewer=REVIEWER,
        review_date=REVIEW_DATE,
    )
    revision = promote(
        predecessor_path=REV1,
        reviewed_workbook_path=reviewed,
        oracle_path=GOLDEN / "neoplasm-adjudicated.json",
        proposal_registry_path=GOLDEN / "proposal-registry.json",
        qualification=QUALIFICATION,
        output_registry_path=registry,
        output_dry_run_path=dry_run,
        output_path=output,
    )
    return revision, output, blank, reviewed, registry, dry_run, loader


@pytest.mark.unit
def test_governed_revision_transcribes_exact_human_decision_and_binds_predecessor(
    tmp_path: Path,
) -> None:
    rev1_before = REV1.read_bytes()
    revision, _output, blank, reviewed, _registry, _dry_run, _loader = _build_revision(
        tmp_path
    )

    assert REV1.read_bytes() == rev1_before
    assert blank.read_bytes() != reviewed.read_bytes()
    assert revision.predecessor_artifact_identity == (
        "90ea507e93cebaf6399b3aa5bea92081e6d3dba50b7631783666d9382d267d1a"
    )
    decision = revision.registry.decisions[1]
    assert (
        decision.subject_code,
        decision.role_code,
        decision.filler_code,
    ) == ASSERTION
    assert (decision.outcome, decision.reviewer, decision.review_date) == (
        OUTCOME,
        REVIEWER,
        REVIEW_DATE,
    )
    assert decision.rationale == RATIONALE
    assert revision.machine_qualification == QUALIFICATION
    assert revision.machine_qualification not in decision.rationale
    assert revision.transcription.actor == "software-transcriber"
    assert revision.transcription.authorship_claimed is False
    assert revision.transcription.authority == "explicit-human-instruction"
    assert (
        revision.transcription.workbook_identity == revision.registry.workbook_identity
    )
    assert revision.packet == revision.predecessor.packet

    book = load_workbook(reviewed, data_only=False, keep_links=False)
    bindings = {
        row[0].value: row[1].value
        for row in book["Bindings"].iter_rows(min_row=2, max_col=2)
    }
    assert bindings["packet_identity"] == revision.packet.packet_identity
    assert bindings["source_identity"] == revision.packet.source_identity


@pytest.mark.unit
def test_revision_is_exact_write_free_effective_decision_state(tmp_path: Path) -> None:
    oracle = GOLDEN / "neoplasm-adjudicated.json"
    proposals = GOLDEN / "proposal-registry.json"
    before = (oracle.read_bytes(), proposals.read_bytes())
    revision, _output, _blank, _reviewed, _registry, _dry_run, _loader = (
        _build_revision(tmp_path)
    )

    assert tuple(
        (row.subject_code, row.outcome) for row in revision.registry.decisions
    ) == (
        ("C2860", "source-supported"),
        ("C3264", OUTCOME),
        ("C3716", "source-supported"),
    )
    assert revision.registry.proposal_preview == ()
    assert tuple(item.model_dump() for item in revision.registry.exclusion_preview) == (
        {
            "subject_code": "C3264",
            "role_code": "R103",
            "filler_code": "C12950",
            "source_identity": revision.packet.source_identity,
            "source_release": "26.07d",
        },
    )
    assert revision.dry_run.writes_performed is False
    assert revision.dry_run.unresolved == 0
    assert revision.dry_run.readiness == "ready-for-separate-application"
    assert revision.dry_run.proposal_previews == ()
    assert len(revision.dry_run.exclusion_previews) == 1
    assert (
        revision.dry_run.oracle_identity_before
        == revision.dry_run.oracle_identity_after
    )
    assert (
        revision.dry_run.proposal_registry_identity_before
        == revision.dry_run.proposal_registry_identity_after
    )
    assert (oracle.read_bytes(), proposals.read_bytes()) == before

    source = revision.packet.rows[1]
    predecessor_source = revision.predecessor.packet.rows[1]
    assert (
        source.complete_definition_identity,
        source.source_fact_identity,
        source.source_group_identity,
        source.source_occurrence_identity,
    ) == (
        predecessor_source.complete_definition_identity,
        predecessor_source.source_fact_identity,
        predecessor_source.source_group_identity,
        predecessor_source.source_occurrence_identity,
    )


@pytest.mark.unit
@pytest.mark.parametrize(
    "mutation", ["predecessor", "qualification", "transcription", "nested"]
)
def test_revision_loader_rejects_tampering_even_with_recomputed_outer_identity(
    tmp_path: Path, mutation: str
) -> None:
    _revision, output, _blank, _reviewed, _registry, _dry_run, loader = _build_revision(
        tmp_path
    )
    payload = json.loads(output.read_text(encoding="ascii"))
    if mutation == "predecessor":
        payload["predecessor_artifact_identity"] = "0" * 64
    elif mutation == "qualification":
        payload["machine_qualification"] = "R103 is defining."
    elif mutation == "transcription":
        payload["transcription"]["authorship_claimed"] = True
    else:
        payload["registry"]["decisions"][1]["rationale"] += " rewritten"
    payload["artifact_identity"] = hashlib.sha256(
        json.dumps(
            {
                key: value
                for key, value in payload.items()
                if key != "artifact_identity"
            },
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=True,
        ).encode("ascii")
    ).hexdigest()
    output.write_text(json.dumps(payload), encoding="ascii")

    with pytest.raises(R103ReviewValidationError):
        loader(output)


@pytest.mark.unit
def test_revision_workflow_reject_branches_are_live(tmp_path: Path) -> None:
    prepare, transcribe, promote, _loader = _revision_api()
    blank = tmp_path / "blank.xlsx"
    reviewed = tmp_path / "reviewed.xlsx"
    prepare(predecessor_path=REV1, output_workbook_path=blank)

    common = {
        "predecessor_path": REV1,
        "blank_workbook_path": blank,
        "output_workbook_path": reviewed,
        "assertion": ASSERTION,
        "outcome": OUTCOME,
        "rationale": RATIONALE,
        "reviewer": REVIEWER,
        "review_date": REVIEW_DATE,
    }
    with pytest.raises(R103ReviewValidationError, match="explicit revision values"):
        transcribe(**{**common, "outcome": "source-supported"})
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_text("not an XLSX", encoding="utf-8")
    with pytest.raises(R103ReviewValidationError, match="invalid review workbook"):
        transcribe(**{**common, "blank_workbook_path": invalid})

    transcribe(**common)
    with pytest.raises(R103ReviewValidationError, match="not blank"):
        transcribe(**{**common, "blank_workbook_path": reviewed})

    promotion = {
        "predecessor_path": REV1,
        "reviewed_workbook_path": reviewed,
        "oracle_path": GOLDEN / "neoplasm-adjudicated.json",
        "proposal_registry_path": GOLDEN / "proposal-registry.json",
        "qualification": QUALIFICATION,
        "output_registry_path": tmp_path / "registry.json",
        "output_dry_run_path": tmp_path / "dry-run.json",
        "output_path": tmp_path / "rev2.json",
    }
    with pytest.raises(R103ReviewValidationError, match="qualification"):
        promote(**{**promotion, "qualification": "R103 is defining."})
    first = promote(**promotion)
    assert promote(**promotion) == first
    Path(promotion["output_path"]).write_text("conflict\n", encoding="utf-8")
    with pytest.raises(R103ReviewValidationError, match="output conflict"):
        promote(**promotion)


@pytest.mark.unit
@pytest.mark.parametrize(
    "mutation", ["citation", "qualification", "identity", "binding"]
)
def test_corroboration_loader_rejects_tampering(tmp_path: Path, mutation: str) -> None:
    _prepare, _transcribe, _promote, loader = _revision_api()
    revision = loader(REV2)
    payload = json.loads(CORROBORATION.read_text(encoding="ascii"))
    if mutation == "citation":
        payload["citations"][0]["doi"] = "10.1000/wrong"
    elif mutation == "qualification":
        payload["scope_qualification"] = "descendants are excluded"
    elif mutation == "identity":
        payload["corroboration_identity"] = "0" * 64
    else:
        payload["effective_decision_identity"] = "0" * 64
    if mutation in {"citation", "qualification", "binding"}:
        payload["corroboration_identity"] = hashlib.sha256(
            json.dumps(
                {
                    key: value
                    for key, value in payload.items()
                    if key != "corroboration_identity"
                },
                sort_keys=True,
                separators=(",", ":"),
                ensure_ascii=True,
            ).encode("ascii")
        ).hexdigest()
    changed = tmp_path / f"corroboration-{mutation}.json"
    changed.write_text(json.dumps(payload), encoding="ascii")

    with pytest.raises(R103ReviewValidationError):
        load_r103_corroboration(changed, revision=revision)


@pytest.mark.unit
def test_tracked_rev2_and_corroboration_are_strict_source_bound_contracts() -> None:
    _prepare, _transcribe, _promote, loader = _revision_api()
    revision = loader(REV2)
    corroboration_loader = cast(
        "Any", getattr(r103_review_promotion, "load_r103_corroboration", None)
    )
    assert callable(corroboration_loader), "strict corroboration consumer is missing"
    corroboration = cast("Any", corroboration_loader(CORROBORATION, revision=revision))

    assert revision.registry.decisions[1].decision_identity == (
        corroboration.effective_decision_identity
    )
    assert corroboration.relationship == "corroboration-not-proof"
    assert corroboration.scope_qualification == QUALIFICATION
    assert tuple((item.doi, item.pmid) for item in corroboration.citations) == (
        ("10.1038/nature09587", "21150899"),
        ("10.1038/s41586-019-1158-7", "31043743"),
        ("10.1111/bpa.13059", "35266242"),
        ("10.1016/j.neuron.2022.07.012", "35985323"),
        ("10.3390/genes12020318", "33672414"),
    )
    decided_subset = {
        (item.subject_code, item.role_code): frozenset({item.filler_code})
        for item in revision.registry.exclusion_preview
    }
    assert decided_subset.items() <= UNSUPPORTED_FILLERS_BY_CONCEPT_ROLE.items()
    assert is_unsupported_filler("C3264", "R103", "C12950")
    assert not is_unsupported_filler("C3716", "R103", "C12950")
    assert not is_unsupported_filler("C3716", "R103", "C34228")


@pytest.mark.unit
def test_cli_exposes_governed_prepare_transcribe_and_promote_revision_commands() -> (
    None
):
    parser = _parser()
    prepare = parser.parse_args(
        [
            "prepare-r103-review-revision",
            "--predecessor",
            "rev1.json",
            "--output-xlsx",
            "blank.xlsx",
        ]
    )
    transcribe = parser.parse_args(
        [
            "transcribe-r103-review-revision",
            "--predecessor",
            "rev1.json",
            "--blank-xlsx",
            "blank.xlsx",
            "--output-xlsx",
            "reviewed.xlsx",
            "--subject",
            "C3264",
            "--role",
            "R103",
            "--filler",
            "C12950",
            "--outcome",
            OUTCOME,
            "--rationale-file",
            "rationale.txt",
            "--reviewer",
            REVIEWER,
            "--review-date",
            REVIEW_DATE,
        ]
    )
    promote = parser.parse_args(
        [
            "promote-r103-review-revision",
            "--predecessor",
            "rev1.json",
            "--reviewed-xlsx",
            "reviewed.xlsx",
            "--oracle",
            "oracle.json",
            "--proposal-registry",
            "proposals.json",
            "--output-registry",
            "registry.json",
            "--output-dry-run",
            "dry-run.json",
            "--output",
            "rev2.json",
            "--output-corroboration",
            "corroboration.json",
        ]
    )

    assert prepare.command == "prepare-r103-review-revision"
    assert transcribe.command == "transcribe-r103-review-revision"
    assert promote.command == "promote-r103-review-revision"
