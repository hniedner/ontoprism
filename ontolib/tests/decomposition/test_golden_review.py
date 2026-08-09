from __future__ import annotations

import hashlib
import json
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from scripts.adjudication import main as adjudication_main
from scripts.research import golden_review
from scripts.research.golden_review import (
    _ROW_DECISION_ADAPTER,
    CandidateOutcomes,
    EngineAcceptance,
    ExpectedTriple,
    GoldenSetValidationError,
    KeptRow,
    RowDecisionCrossTab,
    UnusedCandidateRow,
    evaluate_adjudication,
    export_row_decisions,
    import_adjudication_workbook,
    load_adjudication,
    load_row_decisions,
    load_scorable_golden,
    read_json_without_duplicates,
    write_evaluation_report,
)

from ontolib.decomposition.minting import MintedConcept
from ontolib.decomposition.proposal_registry import (
    ConceptProposal,
    CrossOntologyMapping,
    DuplicateCheck,
    ProposalRegistry,
    ProposalStatus,
    RelationProposal,
    relation_proposal_id,
)

if TYPE_CHECKING:
    from collections.abc import Callable

    from openpyxl.worksheet.worksheet import Worksheet

_EXCEL_MAX_COLUMN = 16384
_DIGEST_A = "a" * 64
_DIGEST_B = "b" * 64
_DIGEST_C = "c" * 64
_DETECTOR_IDENTITY = "6" * 64
_REQUIRED_SEEDS = ("C4791", "C35756", "C89995")


def _constituent(
    axis: str = "op:StageValue",
    filler: str = "C27970",
    *,
    group: str | None = None,
    needs_review: bool = False,
    provenance_status: str | None = None,
    proposal_id: str | None = None,
) -> dict[str, object]:
    status = provenance_status or (
        "locally-approved" if filler.startswith("MINT-") else "ncit-26.07d"
    )
    return {
        "axis": axis,
        "filler": filler,
        "relationship_group": group,
        "needs_review": needs_review,
        "provenance_status": status,
        "proposal_id": proposal_id or (filler if filler.startswith("MINT-") else None),
    }


def _proposal_registry(
    *, status: ProposalStatus = "locally-approved"
) -> ProposalRegistry:
    name = "associated prior disease"
    relation = RelationProposal(
        id=relation_proposal_id(name),
        axis="op:AssociatedPriorDisease",
        preferred_name=name,
        definition="Relates a disease to a distinct disease present earlier.",
        domain="C7057",
        range="C7057",
        source_roles=("R126",),
        source_examples=("C100051->C27262",),
        rationale="The source role conflates several disease relationships.",
        duplicate_checks=(
            DuplicateCheck(
                resource="RO",
                version="2026-08-05",
                query=name,
                result="no-equivalent",
                evidence_url="https://example.test/ro-review",
            ),
        ),
        submission_target="RO",
        status=status,
    )
    return ProposalRegistry(
        source_identity=_DIGEST_A,
        ontology_version="26.07d",
        proposals=(relation,),
    )


def _metric_proposal_registry() -> ProposalRegistry:
    approved = _proposal_registry().proposals[0]
    name = "caused by associated disease"
    proposed = RelationProposal(
        id=relation_proposal_id(name),
        axis="op:CausedByAssociatedDisease",
        preferred_name=name,
        definition="Relates a disease to another disease stated to cause it.",
        domain="C7057",
        range="C7057",
        source_roles=("R126",),
        source_examples=("C0->C999999",),
        rationale="Direct causation requires a univocal relation.",
        duplicate_checks=(
            DuplicateCheck(
                resource="RO",
                version="2026-08-05",
                query=name,
                result="no-equivalent",
                evidence_url="https://example.test/ro-review",
            ),
        ),
        submission_target="RO",
    )
    return ProposalRegistry(
        source_identity=_DIGEST_A,
        ontology_version="26.07d",
        proposals=(approved, proposed),
    )


_EMPTY_PROPOSAL_REGISTRY = ProposalRegistry(
    source_identity=_DIGEST_A,
    ontology_version="26.07d",
    proposals=(),
)


def _engine_constituent(
    axis: str = "op:StageValue",
    filler: str = "C27970",
    *,
    group: str | None = None,
    needs_review: bool = False,
) -> dict[str, object]:
    return {
        "axis": axis,
        "filler": filler,
        "relationship_group": group,
        "needs_review": needs_review,
    }


def _accepted(
    code: str,
    *,
    outcome: str = "decomposed",
    constituents: list[dict[str, object]] | None = None,
    semantic_types: list[str] | None = None,
) -> dict[str, object]:
    return {
        "code": code,
        "label": f"Reviewed {code}",
        "adjudication": {
            "status": "accepted",
            "rationale": "Reviewed against the stated NCIt definition.",
        },
        "expected": {
            "outcome": outcome,
            "semantic_types": semantic_types or ["Neoplastic Process"],
            "constituents": (
                constituents
                if constituents is not None
                else ([] if outcome != "decomposed" else [_constituent()])
            ),
        },
    }


def _artifact(
    concepts: list[dict[str, object]],
    *,
    engine_evidence_identity: str | None = None,
    corpus_evidence_identity: str | None = None,
    proposal_registry: ProposalRegistry = _EMPTY_PROPOSAL_REGISTRY,
) -> dict[str, object]:
    payload = {
        "_meta": {
            "schema_version": 3,
            "status": "SME-ADJUDICATED",
            "ncit_version": "26.07d",
            "source_identity": _DIGEST_A,
            "sample_manifest_identity": _DIGEST_B,
            "run_id": "neoplasm-run-1",
            "run_fingerprint_identity": _DIGEST_C,
            "engine_artifact_identity": "d" * 64,
            "engine_evidence_identity": (
                engine_evidence_identity
                or cast("str", _engine_evidence()["evidence_identity"])
            ),
            "corpus_evidence_identity": (
                corpus_evidence_identity
                or cast("str", _corpus_evidence()["evidence_identity"])
            ),
            "detector_identity": _DETECTOR_IDENTITY,
            "proposal_registry_identity": proposal_registry.registry_identity,
            "workbook_identity": "e" * 64,
            "reviewer": {
                "name": "Example Reviewer",
                "qualification_or_role": "NCIt ontology curator",
                "reviewed_at": "2026-07-30",
            },
        },
        "concepts": concepts,
    }
    payload["artifact_identity"] = _identity(payload)
    return payload


def _identity(value: object) -> str:
    return hashlib.sha256(
        json.dumps(
            value,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=True,
        ).encode()
    ).hexdigest()


def _sign(value: dict[str, object]) -> dict[str, object]:
    value.pop("evidence_identity", None)
    value["evidence_identity"] = _identity(value)
    return value


def _resign_artifact(value: dict[str, object]) -> dict[str, object]:
    value.pop("artifact_identity", None)
    value["artifact_identity"] = _identity(value)
    return value


def _bind_artifact_to_engine(
    concepts: list[dict[str, object]],
    engine: dict[str, object],
    corpus: dict[str, object] | None = None,
    proposal_registry: ProposalRegistry = _EMPTY_PROPOSAL_REGISTRY,
) -> dict[str, object]:
    bound_corpus = corpus or _corpus_evidence()
    return _artifact(
        concepts,
        engine_evidence_identity=cast("str", engine["evidence_identity"]),
        corpus_evidence_identity=cast("str", bound_corpus["evidence_identity"]),
        proposal_registry=proposal_registry,
    )


def _corpus_evidence(
    *, denominator: list[str] | None = None, residual: list[str] | None = None
) -> dict[str, object]:
    return _sign(
        {
            "name": "issue-154-corpus-sample",
            "source_identity": _DIGEST_A,
            "sample_manifest_identity": "9" * 64,
            "run_id": "sample",
            "run_fingerprint_identity": "8" * 64,
            "engine_artifact_identity": "7" * 64,
            "detector_identity": _DETECTOR_IDENTITY,
            "denominator_codes": denominator or ["C1"],
            "residual_codes": residual or [],
        }
    )


def _m1_concepts() -> list[dict[str, object]]:
    codes = [f"C{index}" for index in range(17)] + list(_REQUIRED_SEEDS)
    return [_accepted(code) for code in codes]


def _write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value), encoding="utf-8")


@pytest.mark.unit
def test_adjudication_rejects_draft_duplicate_keys_and_duplicate_concepts(
    tmp_path: Path,
) -> None:
    draft = tmp_path / "draft.json"
    _write_json(draft, {"_meta": {"status": "AUTO-DRAFT"}, "concepts": []})
    with pytest.raises(GoldenSetValidationError, match="not SME-adjudicated"):
        load_adjudication(draft)

    duplicate_key = tmp_path / "duplicate-key.json"
    duplicate_key.write_text('{"_meta": {}, "_meta": {}, "concepts": []}')
    with pytest.raises(GoldenSetValidationError, match="duplicate JSON key: _meta"):
        load_adjudication(duplicate_key)

    duplicate_concept = tmp_path / "duplicate-concept.json"
    concepts = _m1_concepts()
    concepts[-1] = concepts[0]
    _write_json(duplicate_concept, _artifact(concepts))
    with pytest.raises(GoldenSetValidationError, match="concept codes must be unique"):
        load_adjudication(duplicate_concept)


@pytest.mark.unit
def test_adjudication_rejects_tampered_payload_and_empty_accepted_cohort(
    tmp_path: Path,
) -> None:
    tampered = _artifact(_m1_concepts())
    tampered["concepts"][0]["label"] = "Tampered after signature"
    path = tmp_path / "tampered.json"
    _write_json(path, tampered)
    with pytest.raises(GoldenSetValidationError, match="identity does not match"):
        load_adjudication(path)

    concepts = [
        {
            "code": concept["code"],
            "label": concept["label"],
            "adjudication": {
                "status": "rejected",
                "rationale": "Unsuitable for this oracle.",
            },
            "expected": None,
        }
        for concept in _m1_concepts()
    ]
    path = tmp_path / "no-accepted.json"
    _write_json(path, _artifact(concepts))
    with pytest.raises(GoldenSetValidationError, match="at least one accepted"):
        load_adjudication(path)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutate", "message"),
    [
        (lambda value: value["_meta"].update({"unknown": "x"}), "extra"),
        (
            lambda value: value["_meta"]["reviewer"].update(
                {"qualification_or_role": ""}
            ),
            "qualification",
        ),
        (
            lambda value: value["concepts"][0].update({"unexpected": "x"}),
            "extra",
        ),
        (
            lambda value: value["concepts"][0]["expected"]["constituents"].append(
                _constituent(group="another-group")
            ),
            "axis/filler pairs must be unique",
        ),
        (
            lambda value: value["concepts"][0]["expected"].update(
                {"outcome": "residual"}
            ),
            "non-decomposed expectation",
        ),
        (
            lambda value: value["concepts"][0]["expected"]["constituents"][0].update(
                {"filler": "not-an-ncit-code"}
            ),
            "NCIt constituent filler",
        ),
    ],
)
def test_adjudication_schema_rejects_untrusted_shapes(
    tmp_path: Path,
    mutate: object,
    message: str,
) -> None:
    payload = _artifact(_m1_concepts())
    mutate(payload)  # type: ignore[operator]
    _resign_artifact(payload)
    path = tmp_path / "invalid.json"
    _write_json(path, payload)

    with pytest.raises(GoldenSetValidationError, match=message):
        load_adjudication(path)


@pytest.mark.unit
def test_nonaccepted_decisions_require_rationale_and_cannot_claim_expectations(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = {
        "code": "C0",
        "label": "Unsuitable",
        "adjudication": {"status": "rejected", "rationale": ""},
        "expected": None,
    }
    path = tmp_path / "missing-rationale.json"
    _write_json(path, _artifact(concepts))
    with pytest.raises(GoldenSetValidationError, match="rationale"):
        load_adjudication(path)

    concepts[0]["adjudication"]["rationale"] = "Unsuitable source case."
    concepts[0]["expected"] = _accepted("C0")["expected"]
    _write_json(path, _artifact(concepts))
    with pytest.raises(GoldenSetValidationError, match="must not define expected"):
        load_adjudication(path)


@pytest.mark.unit
def test_m1_cohort_validation_is_mandatory(tmp_path: Path) -> None:
    too_small = tmp_path / "small.json"
    _write_json(too_small, _artifact(_m1_concepts()[:19]))
    with pytest.raises(GoldenSetValidationError, match="20 to 50"):
        load_adjudication(too_small)

    missing_seed = tmp_path / "missing-seed.json"
    _write_json(
        missing_seed,
        _artifact([_accepted(f"C{index}") for index in range(20)]),
    )
    with pytest.raises(GoldenSetValidationError, match="C35756, C4791, C89995"):
        load_adjudication(missing_seed)


@pytest.mark.unit
def test_scorable_view_uses_accepted_decisions_and_retains_review_exclusions(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent(),
            _constituent(
                "op:AssociatedRegion",
                "C12418",
                needs_review=True,
            ),
        ],
    )
    concepts[1] = {
        "code": "C1",
        "label": "Rejected",
        "adjudication": {
            "status": "rejected",
            "rationale": "Unsuitable for this oracle.",
        },
        "expected": None,
    }
    path = tmp_path / "adjudicated.json"
    _write_json(path, _artifact(concepts))

    golden = load_scorable_golden(path)

    assert "C1" not in golden.expected
    assert golden.expected["C0"] == frozenset({("op:StageValue", "C27970")})
    assert golden.review_exclusions["C0"] == frozenset(
        {("op:AssociatedRegion", "C12418")}
    )
    assert golden.reviewer_qualification == "NCIt ontology curator"
    assert golden.expectations["C0"].constituents[0].provenance_status == (
        "ncit-26.07d"
    )


@pytest.mark.unit
def test_augmented_expectation_is_bound_to_matching_registry_record(
    tmp_path: Path,
) -> None:
    registry = _proposal_registry()
    proposal_id = registry.proposals[0].id
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent(
                "op:AssociatedPriorDisease",
                "C27262",
                provenance_status="locally-approved",
                proposal_id=proposal_id,
            )
        ],
    )
    path = tmp_path / "adjudicated.json"
    _write_json(path, _artifact(concepts, proposal_registry=registry))

    artifact = load_adjudication(path, registry)

    assert artifact.meta.proposal_registry_identity == registry.registry_identity
    assert artifact.concepts[0].expected is not None
    assert artifact.concepts[0].expected.constituents[0].proposal_id == proposal_id

    concepts[0]["expected"]["constituents"][0]["proposal_id"] = "RELPROP-unknown"
    _write_json(path, _artifact(concepts, proposal_registry=registry))
    with pytest.raises(GoldenSetValidationError, match="unknown proposal"):
        load_adjudication(path, registry)


def _accepted_concept_registry() -> ProposalRegistry:
    name = "Malignant Non-Seminomatous Germ Cell"
    concept = ConceptProposal(
        id=MintedConcept(axis="op:CellType", label=name).id,
        axis="op:CellType",
        preferred_name=name,
        definition="A malignant germ cell with non-seminomatous differentiation.",
        parent_concepts=("C12917",),
        semantic_types=("Cell",),
        source_concepts=("C27787",),
        source_roles=("R105",),
        rationale="No existing NCIt cell concept expresses the required intersection.",
        duplicate_checks=(
            DuplicateCheck(
                resource="NCIt",
                version="26.07d",
                query=name,
                result="no-equivalent",
                evidence_url="https://example.test/ncit-review",
            ),
        ),
        mappings=(
            CrossOntologyMapping(
                system="SNOMED CT US",
                version="2025-09-01",
                concept_id="128766005",
                label="Germ cell tumor, nonseminomatous",
                predicate="relatedMatch",
                evidence_url="https://example.test/snomed/128766005",
            ),
        ),
        submission_target="NCIt",
        status="accepted",
        replacement_ncit_code="C999999",
    )
    return ProposalRegistry(
        source_identity=_DIGEST_A,
        ontology_version="26.07d",
        proposals=(concept,),
    )


@pytest.mark.unit
def test_accepted_proposal_expectation_must_cite_the_replacement_code(
    tmp_path: Path,
) -> None:
    """`accepted-in-ncit` is the terminal lifecycle step (D60).

    Once NCI assigns a real code, the augmented expectation must cite that code,
    not the MINT-* placeholder that stood in for it.
    """
    registry = _accepted_concept_registry()
    proposal = registry.proposals[0]
    path = tmp_path / "adjudicated.json"

    def _write(filler: str) -> None:
        concepts = _m1_concepts()
        concepts[0] = _accepted(
            "C0",
            constituents=[
                _constituent(
                    "op:CellType",
                    filler,
                    provenance_status="accepted-in-ncit",
                    proposal_id=proposal.id,
                )
            ],
        )
        _write_json(path, _artifact(concepts, proposal_registry=registry))

    _write("C999999")
    artifact = load_adjudication(path, registry)
    assert artifact.concepts[0].expected is not None
    assert artifact.concepts[0].expected.constituents[0].filler == "C999999"

    _write(proposal.id)
    with pytest.raises(
        GoldenSetValidationError, match="proposal filler does not match"
    ):
        load_adjudication(path, registry)


@pytest.mark.unit
def test_augmented_expectation_rejects_missing_id_status_or_axis_binding(
    tmp_path: Path,
) -> None:
    local_registry = _proposal_registry()
    proposal_id = local_registry.proposals[0].id
    base = _constituent(
        "op:AssociatedPriorDisease",
        "C27262",
        provenance_status="locally-approved",
        proposal_id=proposal_id,
    )
    path = tmp_path / "invalid-adjudicated.json"

    for mutation, registry, message in (
        ({"proposal_id": None}, local_registry, "proposal ID"),
        ({}, _proposal_registry(status="submitted"), "status"),
        ({"axis": "op:CausedByAssociatedDisease"}, local_registry, "axis"),
        ({"filler": "free text"}, local_registry, "NCIt code"),
    ):
        constituent = base | mutation
        concepts = _m1_concepts()
        concepts[0] = _accepted("C0", constituents=[constituent])
        _write_json(path, _artifact(concepts, proposal_registry=registry))
        with pytest.raises(GoldenSetValidationError, match=message):
            load_adjudication(path, registry)


def _create_workbook(
    path: Path, *, pending: bool = False, formula: bool = False
) -> None:
    wb = Workbook()
    start = wb.active
    start.title = "START HERE"
    reviewer = wb.create_sheet("Reviewer & Attestation")
    reviewer["B5"] = "Example Reviewer"
    reviewer["B6"] = "NCIt ontology curator"
    reviewer["B7"] = date(2026, 7, 30)
    reviewer["B8"] = "NCIt 26.07d"
    reviewer["B9"] = "ATTESTED"
    concepts = wb.create_sheet("Concept Decisions")
    concept_headers = [
        "Order",
        "Concept Code",
        "Source Label",
        "Source Semantic Types",
        "Expected Semantic Types",
        "Engine Suggested Outcome",
        "SME Decision Status",
        "Expected Outcome",
        "Rationale / Required Follow-up",
        "Source Reviewed?",
        "Concept Complete?",
    ]
    for column, header in enumerate(concept_headers, start=1):
        concepts.cell(4, column, header)
    codes = [f"C{index}" for index in range(17)] + list(_REQUIRED_SEEDS)
    for order, code in enumerate(codes, start=1):
        row = order + 4
        values = [
            order,
            code,
            f"Reviewed {code}",
            "Neoplastic Process",
            "Neoplastic Process",
            "decomposed",
            "PENDING" if pending and order == 1 else "accepted",
            "decomposed",
            "Reviewed against the stated source.",
            "YES",
            "YES",
        ]
        for column, value in enumerate(values, start=1):
            concepts.cell(row, column, value)
    constituents = wb.create_sheet("Constituent Decisions")
    constituent_headers = [
        "Concept Order",
        "Concept Code",
        "Source Label",
        "Row Type",
        "Engine Axis",
        "Engine Filler",
        "Engine Filler Label",
        "Engine Group",
        "Engine needs_review",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "Expected Proposal ID",
        "SME Notes",
        "Row Complete?",
    ]
    for column, header in enumerate(constituent_headers, start=1):
        constituents.cell(4, column, header)
    for order, code in enumerate(codes, start=1):
        row = order + 4
        values = [
            order,
            code,
            f"Reviewed {code}",
            "ENGINE SUGGESTION",
            "op:StageValue",
            "C27970",
            "Stage III",
            None,
            "FALSE",
            "include",
            "op:StageValue",
            "C27970",
            None,
            "FALSE",
            "ncit-26.07d",
            None,
            "",
            "YES",
        ]
        for column, value in enumerate(values, start=1):
            constituents.cell(row, column, value)
    wb.create_sheet("Validation Summary")
    wb.create_sheet("Worked Examples")
    wb.create_sheet("Prior SME Evidence")
    evidence = wb.create_sheet("Source & Run Evidence")
    rows = [
        ("NCIt release", "26.07d"),
        ("Source identity", _DIGEST_A),
        ("Sample identity", _DIGEST_B),
        ("Engine run", "neoplasm-run-1"),
        ("Run fingerprint identity", _DIGEST_C),
        ("Artifact SHA-256", "d" * 64),
        ("Engine evidence identity", _engine_evidence()["evidence_identity"]),
        ("Corpus evidence identity", _corpus_evidence()["evidence_identity"]),
        ("Detector identity", _DETECTOR_IDENTITY),
        ("Proposal registry identity", _EMPTY_PROPOSAL_REGISTRY.registry_identity),
    ]
    for row, (key, value) in enumerate(rows, start=5):
        evidence.cell(row, 1, key)
        evidence.cell(row, 2, value)
    if formula:
        concepts["I5"] = '=CONCAT("not", " authored")'
    wb.save(path)


@pytest.mark.unit
def test_openpyxl_contract_preserves_formula_and_boolean_cell_types(
    tmp_path: Path,
) -> None:
    path = tmp_path / "contract.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=1+1"
    ws["A2"] = True
    wb.save(path)

    loaded = load_workbook(path, data_only=False)

    assert loaded.active["A1"].data_type == "f"
    assert loaded.active["A2"].data_type == "b"


@pytest.mark.unit
def test_workbook_import_preserves_reviewer_values_and_provenance(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)

    artifact = import_adjudication_workbook(workbook)

    assert artifact.meta.reviewer.name == "Example Reviewer"
    assert artifact.meta.reviewer.qualification_or_role == "NCIt ontology curator"
    assert artifact.meta.workbook_identity != _DIGEST_A
    assert artifact.meta.run_fingerprint_identity == _DIGEST_C
    assert artifact.concepts[0].expected is not None
    assert artifact.concepts[0].expected.constituents[0].needs_review is False


@pytest.mark.unit
def test_workbook_import_rejects_hidden_reviewer_schema_column(tmp_path: Path) -> None:
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    hidden = load_workbook(workbook)
    hidden["Constituent Decisions"].column_dimensions["J"].hidden = True
    hidden.save(workbook)

    with pytest.raises(GoldenSetValidationError, match="hidden reviewer columns"):
        import_adjudication_workbook(workbook)


@pytest.mark.unit
def test_workbook_import_rejects_a_grouped_hidden_reviewer_column(
    tmp_path: Path,
) -> None:
    """openpyxl stores a grouped hide under the range's first letter only.

    `column_dimensions["T"]` on an `S:T` group auto-creates a fresh default with
    `hidden=False`, so a per-letter probe lets a reviewer conceal a data column by
    hiding a range whose first column is blank -- and the SME then attests to
    a sheet they could not fully see.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    hidden = load_workbook(workbook)
    sheet = hidden["Constituent Decisions"]
    # "T" carries concealed reviewer content; "S" is the blank leading column of the
    # hidden range, which is the only key openpyxl stores the group under.
    sheet["T5"] = "concealed reviewer note"
    sheet.column_dimensions.group("S", "T", hidden=True)
    assert sheet.column_dimensions["T"].hidden is False
    hidden.save(workbook)

    with pytest.raises(GoldenSetValidationError, match="hidden reviewer columns"):
        import_adjudication_workbook(workbook)


@pytest.mark.unit
def test_trailing_hidden_columns_neither_reject_nor_materialize_the_grid(
    tmp_path: Path,
) -> None:
    """Excel writes `<col min="8" max="16384" hidden="1"/>` for a trailing hide.

    That is an ordinary, harmless reviewer action. An uncapped expansion would
    report ~16k indexes and, once the caller reads cells, materialise the whole
    grid -- turning a legitimate import into a hang rather than a verdict.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    hidden = load_workbook(workbook)
    sheet = hidden["Concept Decisions"]
    used_columns = sheet.max_column
    sheet.column_dimensions.group(
        get_column_letter(used_columns + 1), "XFD", hidden=True
    )
    hidden.save(workbook)

    reloaded = load_workbook(workbook)["Concept Decisions"]
    assert (
        reloaded.column_dimensions[get_column_letter(used_columns + 1)].max
        == _EXCEL_MAX_COLUMN
    )
    # Bounded by the used range, so the caller never materialises the full grid.
    assert (
        max(
            golden_review.hidden_column_indexes(reloaded, limit=reloaded.max_column),
            default=0,
        )
        <= used_columns
    )
    assert reloaded.max_column == used_columns

    assert import_adjudication_workbook(workbook).concepts


@pytest.mark.unit
def test_a_visible_column_dimension_is_not_reported_as_hidden(tmp_path: Path) -> None:
    """Only the reject direction was pinned; a width change must still import.

    openpyxl populates `min`/`max` on every `<col>` element it reads, so a gate
    that ignored `hidden` would reject every workbook a reviewer had ever resized.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    widened = load_workbook(workbook)
    widened["Concept Decisions"].column_dimensions["A"].width = 40
    widened.save(workbook)

    assert import_adjudication_workbook(workbook).concepts


@pytest.mark.unit
def test_hidden_column_without_a_stored_range_is_still_reported() -> None:
    """An in-memory hide leaves `min`/`max` unset until `reindex()` at save time.

    Skipping those dimensions would let the anti-tamper helper degrade to
    "nothing is hidden". Not reachable through ``load_workbook`` -- a file's
    ``<col>`` always carries ``min`` -- but the helper must not depend on that.
    """
    sheet = Workbook().active
    assert sheet is not None
    sheet["D1"] = "concealed"
    sheet.column_dimensions["D"].hidden = True
    assert sheet.column_dimensions["D"].min is None

    assert golden_review.hidden_column_indexes(sheet, limit=sheet.max_column) == {4}


@pytest.mark.unit
def test_metric_modality_rejects_unknown_normalized_axis() -> None:
    with pytest.raises(GoldenSetValidationError, match="unknown normalized axis"):
        golden_review._pair_modality(("op:NormalTissueOrgin", "C49276"))


@pytest.mark.unit
def test_workbook_import_hashes_the_parsed_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    snapshot = workbook.read_bytes()
    real_load_workbook = golden_review.load_workbook

    def replace_after_load(source: object, **kwargs: object) -> Workbook:
        loaded = real_load_workbook(source, **kwargs)
        workbook.write_bytes(b"changed after the review snapshot was opened")
        return loaded

    monkeypatch.setattr(golden_review, "load_workbook", replace_after_load)

    artifact = import_adjudication_workbook(workbook)

    assert artifact.meta.workbook_identity == hashlib.sha256(snapshot).hexdigest()


@pytest.mark.unit
def test_workbook_import_allows_semantic_bundle_decision_sheet(tmp_path: Path) -> None:
    workbook = tmp_path / "semantic-review.xlsx"
    _create_workbook(workbook)
    loaded = load_workbook(workbook)
    loaded.create_sheet("Semantic Bundle Decisions")
    loaded.save(workbook)

    artifact = import_adjudication_workbook(workbook)

    assert artifact.meta.reviewer.name == "Example Reviewer"


@pytest.mark.unit
@pytest.mark.parametrize(
    ("pending", "formula", "message"),
    [
        (True, False, "pending adjudication"),
        (False, True, "formula cells are not permitted"),
    ],
)
def test_workbook_import_fails_closed_on_unresolved_or_computed_input(
    tmp_path: Path,
    pending: bool,
    formula: bool,
    message: str,
) -> None:
    workbook = tmp_path / "invalid.xlsx"
    _create_workbook(workbook, pending=pending, formula=formula)

    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (
            lambda wb: setattr(
                wb["Concept Decisions"].row_dimensions[5], "hidden", True
            ),
            "hidden concept rows",
        ),
        (
            lambda wb: setattr(wb["Worked Examples"], "sheet_state", "hidden"),
            "reviewer input sheet must be visible",
        ),
        (
            lambda wb: (
                setattr(wb["Reviewer & Attestation"]["A16"], "value", "Hidden note"),
                setattr(
                    wb["Reviewer & Attestation"].row_dimensions[16],
                    "hidden",
                    True,
                ),
            ),
            "hidden reviewer rows",
        ),
        (
            lambda wb: setattr(
                wb["Constituent Decisions"].row_dimensions[5], "hidden", True
            ),
            "hidden constituent rows",
        ),
        (
            lambda wb: setattr(
                wb["Constituent Decisions"]["D5"], "value", "UNRECOGNIZED"
            ),
            "invalid row type",
        ),
        (
            lambda wb: setattr(wb["Constituent Decisions"]["B5"], "value", "C999999"),
            "unknown concepts",
        ),
        (
            lambda wb: setattr(wb["Concept Decisions"]["B5"], "value", None),
            "populated concept row has blank concept code",
        ),
        (
            lambda wb: setattr(wb["Constituent Decisions"]["B5"], "value", None),
            "populated constituent row has blank concept code",
        ),
        (
            lambda wb: (
                setattr(wb["Source & Run Evidence"]["A11"], "value", "Source identity"),
                setattr(wb["Source & Run Evidence"]["B11"], "value", _DIGEST_A),
            ),
            "duplicate Source & Run Evidence key",
        ),
    ],
)
def test_workbook_import_rejects_hidden_duplicate_or_orphaned_truth(
    tmp_path: Path,
    mutation: object,
    message: str,
) -> None:
    workbook_path = tmp_path / "invalid-structure.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    mutation(workbook)  # type: ignore[operator]
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook_path)


@pytest.mark.unit
def test_workbook_import_requires_reviewer_authored_expected_semantic_types(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "missing-expected-types.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Concept Decisions"].delete_cols(5)
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match="Expected Semantic Types"):
        import_adjudication_workbook(workbook_path)


def _engine_evidence(
    *, wrong_group: bool = False, wrong_identity: bool = False
) -> dict:
    codes = [f"C{index}" for index in range(17)] + list(_REQUIRED_SEEDS)
    return _sign(
        {
            "schema_version": 1,
            "ncit_version": "26.07d",
            "source_identity": _DIGEST_A,
            "sample_manifest_identity": _DIGEST_B,
            "run_id": "neoplasm-run-1",
            "run_fingerprint_identity": ("f" * 64 if wrong_identity else _DIGEST_C),
            "engine_artifact_identity": "d" * 64,
            "detector_identity": _DETECTOR_IDENTITY,
            "concepts": [
                {
                    "code": code,
                    "outcome": "decomposed",
                    "semantic_types": ["Neoplastic Process"],
                    "constituents": [
                        {
                            "axis": "op:StageValue",
                            "filler": "C27970",
                            "relationship_group": (
                                "actual" if wrong_group and code == "C0" else None
                            ),
                            "needs_review": False,
                        }
                    ],
                }
                for code in codes
            ],
            "residual_precoordinated_codes": codes,
        }
    )


@pytest.mark.unit
def test_evaluation_reports_outcomes_groups_and_d21_exclusions(tmp_path: Path) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent(group="expected"),
            _constituent(
                "op:AssociatedRegion",
                "C12418",
                needs_review=True,
            ),
        ],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"].append(
        _engine_constituent("op:AssociatedRegion", "C12418")
    )
    _sign(engine)
    corpus = _corpus_evidence(denominator=["C10", "C11"], residual=["C10"])
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine, corpus))

    report = evaluate_adjudication(load_adjudication(artifact_path), engine, corpus)

    first = report["concepts"][0]
    assert first["pair_score"]["ncit_bound"]["extra"] == []
    assert first["pair_score"]["augmented"]["extra"] == []
    assert first["expected_review_exclusions"] == [["op:AssociatedRegion", "C12418"]]
    assert first["group_match"] is False
    assert first["outcome_match"] is True
    assert report["residual_comparison"]["adjudication"]["count"] == 20
    assert report["residual_comparison"]["corpus_sample"]["count"] == 1
    assert report["residual_comparison"]["rates_averaged"] is False


@pytest.mark.unit
def test_evaluation_derives_ncit_bound_and_augmented_views_from_provenance(
    tmp_path: Path,
) -> None:
    registry = _metric_proposal_registry()
    approved, proposed = registry.proposals
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent("op:CellType", "C36903"),
            _constituent(
                "op:AssociatedPriorDisease",
                "C27262",
                provenance_status="locally-approved",
                proposal_id=approved.id,
            ),
            _constituent(
                "op:CausedByAssociatedDisease",
                "C999999",
                provenance_status="proposed",
                proposal_id=proposed.id,
            ),
        ],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"] = [
        {
            "axis": "op:CellType",
            "filler": "C36903",
            "relationship_group": None,
            "needs_review": False,
        }
    ]
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(
        artifact_path,
        _bind_artifact_to_engine(
            concepts,
            engine,
            proposal_registry=registry,
        ),
    )

    report = evaluate_adjudication(
        load_adjudication(artifact_path, registry), engine, _corpus_evidence()
    )

    assert report["pair_micro"]["ncit_bound"] == {
        "expected": 20,
        "actual": 20,
        "true_positive": 20,
        "precision": 1.0,
        "recall": 1.0,
    }
    assert report["pair_micro"]["augmented"] == {
        "expected": 21,
        "actual": 20,
        "true_positive": 20,
        "precision": 1.0,
        "recall": 20 / 21,
    }
    assert report["pair_micro"]["defining_only"] == {
        "expected": 20,
        "actual": 20,
        "true_positive": 20,
        "precision": 1.0,
        "recall": 1.0,
    }
    assert report["pair_micro"]["non_defining"] == {
        "expected": 0,
        "actual": 0,
        "true_positive": 0,
        "precision": None,
        "recall": None,
    }
    assert report["expected_pair_provenance"] == {
        "locally-approved": 1,
        "ncit-26.07d": 20,
        "proposed": 1,
    }
    assert report["pair_by_axis"]["ncit_bound"]["op:CellType"] == {
        "expected": 1,
        "actual": 1,
        "true_positive": 1,
        "precision": 1.0,
        "recall": 1.0,
    }
    assert report["pair_by_axis"]["augmented"]["op:CellType"] == {
        "expected": 1,
        "actual": 1,
        "true_positive": 1,
        "precision": 1.0,
        "recall": 1.0,
    }
    assert report["pair_by_axis"]["augmented"]["op:AssociatedPriorDisease"] == {
        "expected": 1,
        "actual": 0,
        "true_positive": 0,
        "precision": None,
        "recall": 0.0,
    }
    for view in ("ncit_bound", "augmented", "defining_only", "non_defining"):
        assert {
            field: sum(axis[field] for axis in report["pair_by_axis"][view].values())
            for field in ("expected", "actual", "true_positive")
        } == {
            field: report["pair_micro"][view][field]
            for field in ("expected", "actual", "true_positive")
        }
    assert report["expected_pair_deferrals"] == {
        "augmented": {"deferred": 0, "engine_matches": 0, "expected": 21},
        "defining_only": {"deferred": 0, "engine_matches": 0, "expected": 20},
        "ncit_bound": {"deferred": 0, "engine_matches": 0, "expected": 20},
        "non_defining": {"deferred": 0, "engine_matches": 0, "expected": 0},
    }
    assert report["proposal_governance"] == {
        "augmented_expected": 1,
        "distinct_engine_proposals": 0,
        "engine_emissions": 0,
        "expected_by_status": {"locally-approved": 1, "proposed": 1},
    }
    first = report["concepts"][0]["pair_score"]
    assert first["ncit_bound"]["missing"] == []
    assert first["augmented"]["missing"] == [["op:AssociatedPriorDisease", "C27262"]]


@pytest.mark.unit
def test_evaluation_derives_non_defining_stratum_from_axis_contract(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent("op:CellType", "C36903"),
            _constituent("op:NormalTissueOrigin", "C12345"),
        ],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"] = [
        _engine_constituent("op:CellType", "C36903"),
        _engine_constituent("op:NormalTissueOrigin", "C12345"),
    ]
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    assert report["pair_micro"]["defining_only"] == {
        "expected": 20,
        "actual": 20,
        "true_positive": 20,
        "precision": 1.0,
        "recall": 1.0,
    }
    assert report["pair_micro"]["non_defining"] == {
        "expected": 1,
        "actual": 1,
        "true_positive": 1,
        "precision": 1.0,
        "recall": 1.0,
    }
    first = report["concepts"][0]["pair_score"]
    assert first["defining_only"]["expected"] == 1
    assert first["non_defining"]["expected"] == 1
    assert report["concepts"][0]["expected_pair_modality"] == [
        {"axis": "op:CellType", "filler": "C36903", "modality": "asserted"},
        {
            "axis": "op:NormalTissueOrigin",
            "filler": "C12345",
            "modality": "non-defining",
        },
    ]


@pytest.mark.unit
def test_local_relation_with_ncit_filler_remains_in_ncit_bound_view(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[_constituent("op:AssociatedPriorDisease", "C3270")],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"] = [
        _engine_constituent("op:AssociatedPriorDisease", "C3270")
    ]
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    assert report["concepts"][0]["pair_score"]["ncit_bound"]["true_positive"] == 1


@pytest.mark.unit
def test_reviewer_resolution_scores_an_engine_flagged_pair(tmp_path: Path) -> None:
    concepts = _m1_concepts()
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"][0]["needs_review"] = True
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    first = report["concepts"][0]
    assert first["engine_review_flags"] == [["op:StageValue", "C27970"]]
    assert first["pair_score"]["ncit_bound"]["true_positive"] == 1
    assert report["expected_pair_deferrals"]["ncit_bound"] == {
        "deferred": 0,
        "engine_matches": 0,
        "expected": 20,
    }


@pytest.mark.unit
def test_engine_flagged_sme_rejected_pair_is_a_false_positive(tmp_path: Path) -> None:
    concepts = _m1_concepts()
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"][0]["filler"] = "C27971"
    engine["concepts"][0]["constituents"][0]["needs_review"] = True
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    first = report["concepts"][0]
    assert first["engine_review_flags"] == [["op:StageValue", "C27971"]]
    assert first["pair_score"]["ncit_bound"]["extra"] == [["op:StageValue", "C27971"]]
    assert first["pair_score"]["ncit_bound"]["precision"] == 0.0


@pytest.mark.unit
def test_group_comparison_ignores_group_names_but_not_membership(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent(group="expected-name"),
            _constituent("op:StageValue", "C27971", group="expected-name"),
        ],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["constituents"] = [
        _engine_constituent(group="different-name"),
        _engine_constituent("op:StageValue", "C27971", group="different-name"),
    ]
    _sign(engine)
    corpus = _corpus_evidence()
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine, corpus))

    report = evaluate_adjudication(load_adjudication(artifact_path), engine, corpus)

    assert report["concepts"][0]["group_match"] is True
    assert report["group_partition_agreement"] == {
        "concepts_agree": 20,
        "concepts_disagree": 0,
    }

    engine = _engine_evidence()
    engine["concepts"][0]["constituents"] = [
        _engine_constituent(group="first"),
        _engine_constituent("op:StageValue", "C27971", group="second"),
    ]
    _sign(engine)
    artifact_path = tmp_path / "different-partition.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine, corpus))

    report = evaluate_adjudication(load_adjudication(artifact_path), engine, corpus)

    assert report["concepts"][0]["group_match"] is False
    assert report["group_partition_agreement"] == {
        "concepts_agree": 19,
        "concepts_disagree": 1,
    }


@pytest.mark.unit
def test_evaluation_rejects_identity_drift_and_invalid_residual_subset(
    tmp_path: Path,
) -> None:
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _artifact(_m1_concepts()))
    artifact = load_adjudication(artifact_path)
    corpus = _corpus_evidence(denominator=["C1"], residual=["C2"])

    with pytest.raises(GoldenSetValidationError, match="run fingerprint identity"):
        evaluate_adjudication(artifact, _engine_evidence(wrong_identity=True), corpus)
    with pytest.raises(
        GoldenSetValidationError, match="residual codes must be a subset"
    ):
        evaluate_adjudication(artifact, _engine_evidence(), corpus)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("ncit_version", "26.05d", "NCIt version"),
        ("source_identity", "f" * 64, "source identity"),
        ("sample_manifest_identity", "f" * 64, "sample manifest identity"),
        ("run_id", "neoplasm-run-9", "run id"),
        ("run_fingerprint_identity", "f" * 64, "run fingerprint identity"),
        ("engine_artifact_identity", "f" * 64, "engine artifact identity"),
        ("detector_identity", "f" * 64, "detector identity"),
    ],
)
def test_every_engine_identity_binding_is_enforced(
    tmp_path: Path,
    field: str,
    value: str,
    message: str,
) -> None:
    """Each of the eight identity bindings must discriminate on its own.

    A binding that is pinned but never compared is the "identity declared, never
    checked" defect: the artifact would claim provenance the engine run does not
    have. `engine evidence identity` is covered separately below, because drifting
    it requires leaving the artifact bound to a different payload.
    """
    engine = _engine_evidence()
    engine[field] = value
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(_m1_concepts(), engine))

    with pytest.raises(GoldenSetValidationError, match=message):
        evaluate_adjudication(
            load_adjudication(artifact_path), engine, _corpus_evidence()
        )


@pytest.mark.unit
def test_engine_evidence_identity_binding_is_enforced(tmp_path: Path) -> None:
    engine = _engine_evidence()
    artifact_path = tmp_path / "artifact.json"
    _write_json(
        artifact_path,
        _artifact(_m1_concepts(), engine_evidence_identity="f" * 64),
    )

    with pytest.raises(
        GoldenSetValidationError, match="engine evidence identity does not match"
    ):
        evaluate_adjudication(
            load_adjudication(artifact_path), engine, _corpus_evidence()
        )


@pytest.mark.unit
def test_engine_worklist_order_must_match_adjudication_order(tmp_path: Path) -> None:
    """Per-concept scoring pairs adjudication[i] with engine[i].

    A permuted engine worklist with the same code set would score every concept
    against another concept's expectation while validation stayed green.
    """
    engine = _engine_evidence()
    concepts = cast("list[dict[str, object]]", engine["concepts"])
    concepts[0], concepts[1] = concepts[1], concepts[0]
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(_m1_concepts(), engine))

    with pytest.raises(GoldenSetValidationError, match="adjudication order"):
        evaluate_adjudication(
            load_adjudication(artifact_path), engine, _corpus_evidence()
        )


@pytest.mark.unit
def test_deferral_counters_report_pairs_the_engine_also_emitted(
    tmp_path: Path,
) -> None:
    """`engine_matches` is the signal that a run cannot quietly bury a deferral.

    Every other assertion on `expected_pair_deferrals` uses a fixture where both
    accumulators sit at zero, so neither ever has to accumulate anything.
    """
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        constituents=[
            _constituent(),
            _constituent("op:AssociatedRegion", "C12418", needs_review=True),
            _constituent("op:CellType", "C36903", needs_review=True),
        ],
    )
    engine = _engine_evidence()
    # The engine emits one of the two deferred pairs and not the other.
    cast("list[dict[str, object]]", engine["concepts"][0]["constituents"]).append(
        _engine_constituent("op:AssociatedRegion", "C12418")
    )
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    deferrals = report["expected_pair_deferrals"]
    assert deferrals["ncit_bound"] == {
        "deferred": 2,
        "engine_matches": 1,
        "expected": 22,
    }
    assert deferrals["augmented"]["deferred"] == 2
    assert deferrals["augmented"]["engine_matches"] == 1
    assert deferrals["non_defining"] == {
        "deferred": 0,
        "engine_matches": 0,
        "expected": 0,
    }


@pytest.mark.unit
def test_evaluation_rejects_tampered_or_non_decomposed_residual_evidence(
    tmp_path: Path,
) -> None:
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _artifact(_m1_concepts()))
    artifact = load_adjudication(artifact_path)
    engine = _engine_evidence()
    engine["concepts"][0]["semantic_types"] = ["Finding"]
    with pytest.raises(GoldenSetValidationError, match="identity does not match"):
        evaluate_adjudication(artifact, engine, _corpus_evidence())

    engine = _engine_evidence()
    engine["concepts"][0]["outcome"] = "atomic-no-op"
    engine["concepts"][0]["constituents"] = []
    _sign(engine)
    with pytest.raises(
        GoldenSetValidationError, match="residual codes require decomposed outcomes"
    ):
        evaluate_adjudication(artifact, engine, _corpus_evidence())


@pytest.mark.unit
def test_residual_denominator_uses_actual_decomposed_and_types_are_sets(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted(
        "C0",
        semantic_types=["Neoplastic Process", "Disease or Syndrome"],
    )
    engine = _engine_evidence()
    engine["concepts"][0]["semantic_types"] = [
        "Disease or Syndrome",
        "Neoplastic Process",
    ]
    engine["concepts"][1]["outcome"] = "atomic-no-op"
    engine["concepts"][1]["constituents"] = []
    engine["residual_precoordinated_codes"].remove("C1")
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    assert report["concepts"][0]["semantic_types_match"] is True
    assert report["residual_comparison"]["adjudication"]["denominator"] == 19


@pytest.mark.unit
def test_residual_numerator_excludes_decomposed_concepts_the_detector_cleared(
    tmp_path: Path,
) -> None:
    """Gate liveness for the residual filter, which the tracked baseline cannot give.

    On the tracked M1 evidence every decomposed concept is also residual, so
    `count == denominator == 18` there and `rate == 1.0`. That saturation makes the
    tracked assertions satisfiable by the identity `count == denominator` — exactly
    what deleting the `code in engine.residual_precoordinated_codes` filter
    produces. This fixture is deliberately unsaturated: 20 decomposed concepts of
    which the detector flagged 18, so the numerator can only be 18 if the filter
    runs. Deleting the filter yields 20/20; inverting it yields 2/20.
    """
    engine = _engine_evidence()
    cleared = {"C0", "C1"}
    engine["residual_precoordinated_codes"] = [
        code for code in engine["residual_precoordinated_codes"] if code not in cleared
    ]
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(_m1_concepts(), engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    adjudication = report["residual_comparison"]["adjudication"]
    assert adjudication["denominator"] == 20
    assert adjudication["count"] == 18
    assert adjudication["rate"] == 0.9
    assert set(adjudication["denominator_codes"]) >= cleared
    assert set(adjudication["residual_codes"]).isdisjoint(cleared)
    assert (
        set(adjudication["residual_codes"])
        == set(adjudication["denominator_codes"]) - cleared
    )
    # The report subtracts the two rates; it never averages them. An average of
    # 0.9 and the corpus sample's 0.0 would be 0.45.
    assert report["residual_comparison"]["corpus_sample"]["rate"] == 0.0
    assert report["residual_comparison"]["absolute_rate_delta"] == 0.9


@pytest.mark.unit
def test_the_residual_delta_is_a_distance_not_a_signed_difference(
    tmp_path: Path,
) -> None:
    """The corpus may be *more* residual than the adjudicated sample.

    Every other residual fixture in this module, and the tracked M1 evidence,
    puts the adjudication rate at or above the corpus rate, so `a - b` is already
    non-negative and dropping `abs()` changes nothing. Here the adjudication is
    0/20 and the corpus 1/1, so the unsigned difference is -1.0 and only a
    distance reports 1.0.
    """
    engine = _engine_evidence()
    engine["residual_precoordinated_codes"] = []
    _sign(engine)
    corpus = _corpus_evidence(denominator=["C1"], residual=["C1"])
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(_m1_concepts(), engine, corpus))

    report = evaluate_adjudication(load_adjudication(artifact_path), engine, corpus)

    residual = report["residual_comparison"]
    assert residual["adjudication"]["rate"] == 0.0
    assert residual["corpus_sample"]["rate"] == 1.0
    assert residual["absolute_rate_delta"] == 1.0


@pytest.mark.unit
def test_zero_pair_metrics_are_undefined_and_detector_drift_is_rejected(
    tmp_path: Path,
) -> None:
    concepts = _m1_concepts()
    concepts[0] = _accepted("C0", outcome="atomic-no-op", constituents=[])
    for index in range(1, len(concepts)):
        code = concepts[index]["code"]
        concepts[index] = {
            "code": code,
            "label": f"Rejected {code}",
            "adjudication": {
                "status": "rejected",
                "rationale": "Unsuitable for this oracle.",
            },
            "expected": None,
        }
    engine = _engine_evidence()
    engine["concepts"][0]["outcome"] = "atomic-no-op"
    engine["concepts"][0]["constituents"] = []
    engine["residual_precoordinated_codes"].remove("C0")
    _sign(engine)
    artifact_path = tmp_path / "artifact.json"
    _write_json(artifact_path, _bind_artifact_to_engine(concepts, engine))

    report = evaluate_adjudication(
        load_adjudication(artifact_path), engine, _corpus_evidence()
    )

    assert report["pair_micro"]["ncit_bound"]["precision"] is None
    assert report["pair_micro"]["ncit_bound"]["recall"] is None
    assert report["pair_micro"]["augmented"]["precision"] is None
    assert report["pair_micro"]["augmented"]["recall"] is None
    drifted_corpus = _corpus_evidence()
    drifted_corpus["detector_identity"] = "5" * 64
    _sign(drifted_corpus)
    with pytest.raises(GoldenSetValidationError, match="detector identity"):
        evaluate_adjudication(load_adjudication(artifact_path), engine, drifted_corpus)


@pytest.mark.unit
def test_evaluation_report_is_byte_reproducible(tmp_path: Path) -> None:
    corpus = _corpus_evidence(residual=["C1"])
    artifact_path = tmp_path / "artifact.json"
    _write_json(
        artifact_path,
        _artifact(
            _m1_concepts(),
            corpus_evidence_identity=cast("str", corpus["evidence_identity"]),
        ),
    )
    artifact = load_adjudication(artifact_path)
    report = evaluate_adjudication(artifact, _engine_evidence(), corpus)
    first = tmp_path / "first.json"
    second = tmp_path / "second.json"

    write_evaluation_report(report, first)
    write_evaluation_report(report, second)

    assert first.read_bytes() == second.read_bytes()


@pytest.mark.unit
def test_evidence_json_reader_rejects_duplicate_provenance_keys(tmp_path: Path) -> None:
    path = tmp_path / "duplicate.json"
    path.write_text('{"run_id":"first","run_id":"second"}', encoding="utf-8")

    with pytest.raises(GoldenSetValidationError, match="duplicate JSON key: run_id"):
        read_json_without_duplicates(path)


@pytest.mark.unit
def test_a_failed_cli_write_leaves_the_previous_artifact_intact(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A half-written tracked artifact is worse than no new one.

    `Path.write_text` truncates the destination the moment it opens it, and the
    CLI pointed it straight at a tracked golden file. The double emulates exactly
    that: truncate, then fail. Staging in a sibling directory and `os.replace`-ing
    into place means only the staged copy is lost, and it leaves nothing behind.
    """
    workbook = tmp_path / "review.xlsx"
    output = tmp_path / "rows.json"
    _create_workbook(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(output)])
    previous = output.read_bytes()
    real_write_text = Path.write_text

    def _truncate_then_fail(self: Path, *args: object, **kwargs: object) -> int:
        self.write_bytes(b"")
        raise OSError(28, "No space left on device")

    monkeypatch.setattr(Path, "write_text", _truncate_then_fail)
    with pytest.raises(OSError, match="No space left on device"):
        adjudication_main(["export-row-decisions", str(workbook), str(output)])
    monkeypatch.setattr(Path, "write_text", real_write_text)

    assert output.read_bytes() == previous
    assert sorted(item.name for item in tmp_path.iterdir()) == [
        "review.xlsx",
        "rows.json",
    ]


@pytest.mark.unit
def test_adjudication_cli_imports_workbook_and_evaluates_report(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "review.xlsx"
    artifact_path = tmp_path / "artifact.json"
    engine_path = tmp_path / "engine.json"
    corpus_path = tmp_path / "corpus.json"
    registry_path = tmp_path / "registry.json"
    report_path = tmp_path / "report.json"
    engine = _engine_evidence()
    corpus = _corpus_evidence(residual=["C1"])
    _create_workbook(workbook)
    review = load_workbook(workbook)
    evidence = review["Source & Run Evidence"]
    for row in range(5, evidence.max_row + 1):
        if evidence.cell(row, 1).value == "Corpus evidence identity":
            evidence.cell(row, 2, corpus["evidence_identity"])
    review.save(workbook)
    _write_json(engine_path, engine)
    registry_path.write_text(
        _EMPTY_PROPOSAL_REGISTRY.model_dump_json(),
        encoding="utf-8",
    )
    _write_json(
        corpus_path,
        corpus,
    )

    adjudication_main(
        [
            "import-workbook",
            str(workbook),
            str(registry_path),
            str(artifact_path),
        ]
    )
    adjudication_main(
        [
            "evaluate",
            str(artifact_path),
            str(engine_path),
            str(corpus_path),
            str(registry_path),
            str(report_path),
        ]
    )

    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert artifact["_meta"]["reviewer"]["name"] == "Example Reviewer"
    assert report["accepted_concepts"] == 20


_EXTRA_DECISION_ROWS = (
    # code, row type, engine axis, engine filler, action, expected axis,
    # expected filler, complete
    ("C0", "ADD IF MISSING", None, None, "include", "op:PrimarySite", "C12345", "YES"),
    ("C1", "ADD IF MISSING", None, None, "exclude", None, None, "YES"),
    ("C2", "ADD IF MISSING", None, None, "not-needed", None, None, "NO"),
    # A revision that replaced the engine's pair: the reviewer kept the row but not
    # what the engine offered, so it counts in `revise` and not in `pair_preserved`.
    (
        "C3",
        "ENGINE SUGGESTION",
        "op:StageValue",
        "C27974",
        "revise",
        "op:StageValue",
        "C27971",
        "YES",
    ),
    # An excluded row that still carries a stale expectation, as three rows of the
    # attested #57 workbook do. Its pair differs from C4's kept pair, so a reader
    # that filtered on "has an expected pair" instead of the SME action would
    # produce a triple the oracle does not contain.
    (
        "C4",
        "ENGINE SUGGESTION",
        "op:StageValue",
        "C27972",
        "exclude",
        "op:StageValue",
        "C27972",
        "YES",
    ),
    # A revision that kept the engine's pair — 32 of the attested workbook's 42
    # revised suggestions have this shape, which is why `pair_preserved` and
    # `include` are different questions.
    (
        "C5",
        "ENGINE SUGGESTION",
        "op:StageValue",
        "C27973",
        "revise",
        "op:StageValue",
        "C27973",
        "YES",
    ),
)


def _append_decision_rows(path: Path) -> None:
    """Append reviewer decision rows covering every SME action and row type."""
    workbook = load_workbook(path)
    sheet = workbook["Constituent Decisions"]
    for offset, (
        code,
        row_type,
        engine_axis,
        engine_filler,
        action,
        axis,
        filler,
        complete,
    ) in enumerate(_EXTRA_DECISION_ROWS):
        row = sheet.max_row + 1 + offset
        sheet.cell(row, 2, code)
        sheet.cell(row, 4, row_type)
        sheet.cell(row, 5, engine_axis)
        sheet.cell(row, 6, engine_filler)
        sheet.cell(row, 10, action)
        sheet.cell(row, 11, axis)
        sheet.cell(row, 12, filler)
        sheet.cell(row, 14, "FALSE")
        sheet.cell(row, 15, "ncit-26.07d")
        sheet.cell(row, 18, complete)
    workbook.save(path)


@pytest.mark.unit
def test_row_decision_export_keeps_every_reviewer_decision(tmp_path: Path) -> None:
    """Rejected and not-needed rows survive the export; the importer discards them."""
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    _append_decision_rows(workbook)

    export = export_row_decisions(workbook)

    assert len(export.rows) == 26
    assert export.meta.ncit_version == "26.07d"
    assert export.meta.source_workbook == "review.xlsx"
    assert export.meta.reviewer.name == "Example Reviewer"
    assert (
        export.meta.workbook_identity
        == hashlib.sha256(workbook.read_bytes()).hexdigest()
    )
    assert export.cross_tab() == RowDecisionCrossTab(
        engine_suggestion=EngineAcceptance(
            include=20, revise=2, exclude=1, pair_preserved=21
        ),
        add_if_missing=CandidateOutcomes(include=1, revise=0, exclude=1, not_needed=1),
    )


@pytest.mark.unit
def test_row_decision_export_reproduces_the_imported_expected_set(
    tmp_path: Path,
) -> None:
    """`include` + `revise` rows are exactly the constituents the importer keeps.

    This is the correspondence the tracked artifacts rely on. Keying on the SME
    action matters: the excluded C4 row carries a stale expectation the importer
    drops, so a pair-presence filter would over-collect.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    _append_decision_rows(workbook)

    export = export_row_decisions(workbook)
    artifact = import_adjudication_workbook(workbook)

    assert export.expected_pairs() == {
        ExpectedTriple(code=concept.code, axis=item.axis, filler=item.filler)
        for concept in artifact.concepts
        if concept.expected is not None
        for item in concept.expected.constituents
    }
    assert (
        ExpectedTriple(code="C4", axis="op:StageValue", filler="C27972")
        not in export.expected_pairs()
    )


def _hide_a_constituent_row(workbook: Workbook) -> None:
    workbook["Constituent Decisions"].row_dimensions[5].hidden = True


def _hide_a_reviewer_column(workbook: Workbook) -> None:
    workbook["Constituent Decisions"].column_dimensions["J"].hidden = True


def _plant_a_reviewer_formula(workbook: Workbook) -> None:
    workbook["Concept Decisions"]["I5"] = '=CONCAT("not", " authored")'


def _withdraw_the_attestation(workbook: Workbook) -> None:
    workbook["Reviewer & Attestation"]["B9"] = "PENDING"


def _remove_a_required_evidence_key(workbook: Workbook) -> None:
    sheet = workbook["Source & Run Evidence"]
    for row in range(5, sheet.max_row + 1):
        if sheet.cell(row, 1).value == "Detector identity":
            # `sheet.cell(row, column, None)` is a *read*: openpyxl treats a None
            # `value` argument as "no value given" and returns the cell untouched.
            sheet.cell(row, 1).value = None
            sheet.cell(row, 2).value = None
            return
    raise AssertionError("fixture no longer carries a detector identity")


def _corrupt_a_row_type(workbook: Workbook) -> None:
    workbook["Constituent Decisions"]["D5"] = "ENGINE HINT"


def _rename_a_sheet(workbook: Workbook) -> None:
    workbook["Worked Examples"].title = "Worked Example"


@pytest.mark.unit
@pytest.mark.parametrize(
    ("tamper", "message"),
    [
        (_rename_a_sheet, "workbook sheets do not match the review contract"),
        (_hide_a_constituent_row, "hidden constituent rows"),
        (_hide_a_reviewer_column, "hidden reviewer columns"),
        (_plant_a_reviewer_formula, "formula cells are not permitted"),
        (_withdraw_the_attestation, "reviewer attestation is pending"),
        (_remove_a_required_evidence_key, "Source & Run Evidence is missing"),
        (_corrupt_a_row_type, "invalid row type"),
    ],
    ids=[
        "sheet-contract",
        "hidden-row",
        "hidden-column",
        "formula",
        "attestation",
        "evidence-key",
        "row-identity",
    ],
)
def test_row_decision_export_fails_closed_on_a_tampered_workbook(
    tmp_path: Path, tamper: Callable[[Workbook], None], message: str
) -> None:
    """The export runs the workbook-level tamper gates the oracle import runs.

    One case per gate family named in `export_row_decisions_bytes`: the sheet
    contract, hidden rows, hidden columns, formula cells in reviewer input, the
    attestation, the required evidence keys and the constituent row identity. The
    gates the export does *not* run are pinned by
    `test_row_decision_export_accepts_kept_constituent_defects_the_import_rejects`.
    """
    workbook_path = tmp_path / "tampered.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    tamper(workbook)
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match=message):
        export_row_decisions(workbook_path)


@pytest.mark.unit
def test_row_decision_export_accepts_kept_constituent_defects_the_import_rejects(
    tmp_path: Path,
) -> None:
    """The export stops before `_kept_constituent`; it never builds the expectation.

    It does read the row identity *and* `Expected Axis` / `Expected Filler`, so
    "stops at the row identity" was wrong in both directions. What it does not run
    is `_kept_constituent`, which is where the `Expected Provenance Status` and
    `Expected needs_review` gates live, so those are the import's alone. A workbook
    corrupt in either column is a valid row-decision export and an invalid oracle.
    The export is therefore not a substitute for `import-workbook`, and this pins
    that boundary so the claim cannot silently widen.
    """
    for column, value, message in (
        (15, "not-a-status", "provenance_status"),
        (14, "MAYBE", "expected needs_review must be TRUE or FALSE"),
    ):
        workbook_path = tmp_path / f"kept-gate-{column}.xlsx"
        _create_workbook(workbook_path)
        workbook = load_workbook(workbook_path)
        workbook["Constituent Decisions"].cell(5, column, value)
        workbook.save(workbook_path)

        export = export_row_decisions(workbook_path)

        first = export.rows[0]
        assert isinstance(first, KeptRow)
        assert first.code == "C0"
        assert first.sme_action == "include"
        assert first.expected.axis == "op:StageValue"
        assert first.expected.filler == "C27970"
        with pytest.raises(GoldenSetValidationError, match=message):
            import_adjudication_workbook(workbook_path)


@pytest.mark.unit
def test_row_decision_reader_rejects_a_padded_axis_on_an_excluded_row(
    tmp_path: Path,
) -> None:
    """Non-kept rows are canonically validated too — a deliberate tightening.

    The pre-export reader skipped `exclude` and `not-needed` rows before reading
    `Expected Axis`/`Expected Filler`, so a padded value there was accepted. The
    export writes those cells out verbatim, so it must read them, and it holds them
    to the same canonical form as a kept row's on both paths.
    """
    workbook_path = tmp_path / "padded-excluded-axis.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Constituent Decisions"]
    sheet["J5"] = "exclude"
    sheet["K5"] = " op:StageValue"
    workbook.save(workbook_path)

    message = "C0 expected axis must be non-empty without outer whitespace"
    with pytest.raises(GoldenSetValidationError, match=message):
        export_row_decisions(workbook_path)
    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook_path)


@pytest.mark.unit
def test_row_decision_export_rejects_an_unrecognized_sme_action(
    tmp_path: Path,
) -> None:
    """An action outside the reviewed vocabulary cannot be silently recorded."""
    workbook_path = tmp_path / "invalid-action.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Constituent Decisions"]["J5"] = "maybe"
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match="invalid SME action: maybe"):
        export_row_decisions(workbook_path)


@pytest.mark.unit
def test_row_decision_loader_rejects_a_kept_row_without_an_expected_pair(
    tmp_path: Path,
) -> None:
    """A hand-edited export that drops a kept row's pair cannot be loaded.

    `KeptRow` requires both halves, so the refusal comes from the field rather
    than from a validator that a `model_construct` could step around.
    """
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    assert load_row_decisions(export_path).rows[0].sme_action == "include"
    payload["rows"][0]["expected"]["filler"] = None
    _write_json(export_path, payload)

    with pytest.raises(GoldenSetValidationError, match=r"expected\.filler"):
        load_row_decisions(export_path)


@pytest.mark.unit
def test_row_decision_loader_rejects_duplicate_kept_pairs(tmp_path: Path) -> None:
    """Two kept rows cannot claim the same concept, axis and filler."""
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    payload["rows"][1]["code"] = payload["rows"][0]["code"]
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(GoldenSetValidationError, match="must be unique on"):
        load_row_decisions(export_path)


def _resign_row_decisions(payload: dict[str, Any]) -> dict[str, Any]:
    """Recompute `payload_identity` so a test isolates the structural rules.

    Without this a structural test would depend on `_validate_rows` checking
    structure before identity, and would silently start asserting tamper detection
    instead if that order ever changed.
    """
    payload.pop("payload_identity", None)
    payload["payload_identity"] = _identity(payload)
    return payload


@pytest.mark.unit
def test_row_decision_loader_rejects_duplicate_withdrawn_pairs(
    tmp_path: Path,
) -> None:
    """Uniqueness covers every row carrying an expected pair, not only kept ones.

    A withdrawn expectation duplicated across two `exclude` rows is one reviewer
    decision counted twice, and it enlarges the denominator of the acceptance rate
    exactly as a duplicated kept row would.
    """
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    _append_decision_rows(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    withdrawn = next(
        row
        for row in payload["rows"]
        if row["sme_action"] == "exclude" and row["expected"] is not None
    )
    payload["rows"].append(dict(withdrawn))
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(GoldenSetValidationError, match="must be unique on"):
        load_row_decisions(export_path)


@pytest.mark.unit
def test_row_decision_loader_rejects_a_pair_both_kept_and_withdrawn(
    tmp_path: Path,
) -> None:
    """One triple cannot be simultaneously in and out of the oracle.

    `expected_pairs()` keys on the row variant, so a triple present on both a kept
    and an excluded row was silently counted as kept while also being reported as
    a rejection — the numerator and the denominator disagreeing about the same
    decision.
    """
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    _append_decision_rows(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    kept = next(row for row in payload["rows"] if row["sme_action"] == "include")
    withdrawn = next(
        row
        for row in payload["rows"]
        if row["sme_action"] == "exclude" and row["expected"] is not None
    )
    withdrawn.update(code=kept["code"], expected=dict(kept["expected"]))
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(GoldenSetValidationError, match="both kept and withdrawn"):
        load_row_decisions(export_path)


@pytest.mark.unit
def test_rows_carrying_no_expected_pair_may_repeat(tmp_path: Path) -> None:
    """The stated limit of the uniqueness rule, pinned so it is not mistaken.

    A row with no expected pair has no identity in this payload: the export does
    not record the engine's suggested axis and filler, which is the only thing that
    distinguishes two excluded suggestions on one concept. The attested #57
    workbook contains six such legitimately identical tuples — three excluded
    suggestions and three not-needed candidates — so uniqueness cannot extend to
    them without rejecting the real review. Duplication of *these* rows is caught
    by `payload_identity` at load and by the per-concept engine-run binding in
    `test_m1_baseline.py`, not here.
    """
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    _append_decision_rows(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    blank = next(row for row in payload["rows"] if row["sme_action"] == "not-needed")
    payload["rows"].append(dict(blank))
    _write_json(export_path, _resign_row_decisions(payload))

    export = load_row_decisions(export_path)

    assert export.cross_tab().add_if_missing.not_needed == 2


@pytest.mark.unit
def test_row_decision_export_binds_the_rows_to_the_engine_run_they_measure(
    tmp_path: Path,
) -> None:
    """An acceptance rate is meaningless without the run whose output it grades.

    `_required_evidence` already demanded these three keys and then discarded all
    but `NCIt release`, so the exported rows named no engine run, no source and no
    engine evidence. Nothing tied "48 of 106 suggestions accepted" to the
    suggestions of a particular run.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)

    export = export_row_decisions(workbook)
    artifact = import_adjudication_workbook(workbook)

    assert export.meta.source_identity == artifact.meta.source_identity
    assert export.meta.run_id == artifact.meta.run_id
    assert (
        export.meta.engine_evidence_identity == artifact.meta.engine_evidence_identity
    )
    assert (
        export.meta.engine_evidence_identity == _engine_evidence()["evidence_identity"]
    )


@pytest.mark.unit
def test_row_decision_export_signs_the_rows_it_exports(tmp_path: Path) -> None:
    """The payload identity covers the rows, not merely the workbook they came from.

    `workbook_identity` hashes the `.xlsx`; it says nothing about the row set
    derived from it. Every sibling artifact in this module carries a
    payload-covering identity and validates it at load; this one did not.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)

    export = export_row_decisions(workbook)

    assert export.meta.schema_version == 4
    assert export.payload_identity == _identity(
        export.model_dump(mode="json", by_alias=True, exclude={"payload_identity"})
    )
    assert export.payload_identity != export.meta.workbook_identity


def _row_decision_payload(tmp_path: Path) -> tuple[Path, dict[str, Any]]:
    workbook = tmp_path / "review.xlsx"
    export_path = tmp_path / "rows.json"
    _create_workbook(workbook)
    _append_decision_rows(workbook)
    adjudication_main(["export-row-decisions", str(workbook), str(export_path)])
    return export_path, json.loads(export_path.read_text(encoding="utf-8"))


def _drop_excluded_rows(payload: dict[str, Any]) -> None:
    payload["rows"] = [row for row in payload["rows"] if row["sme_action"] != "exclude"]


def _relabel_a_candidate_row(payload: dict[str, Any]) -> None:
    """Promote an SME-added row to a suggestion, engine pair and all.

    Relabelling alone is now refused by `_AdjudicatedRow`, so the forgery has to
    invent the suggestion too. That leaves a structurally valid row set, and the
    payload identity is what refuses it.
    """
    for row in payload["rows"]:
        if row["row_type"] == "ADD IF MISSING" and row["sme_action"] == "include":
            row["row_type"] = "ENGINE SUGGESTION"
            row["engine"] = dict(row["expected"])
            return
    raise AssertionError("fixture no longer contains an included candidate row")


def _duplicate_an_excluded_row(payload: dict[str, Any]) -> None:
    excluded = next(row for row in payload["rows"] if row["sme_action"] == "exclude")
    payload["rows"].extend([dict(excluded)] * 20)


def _blank_a_withdrawn_expectation(payload: dict[str, Any]) -> None:
    for row in payload["rows"]:
        if row["sme_action"] == "exclude" and row["expected"] is not None:
            row["expected"] = None
            return
    raise AssertionError("fixture no longer contains a withdrawn expectation")


@pytest.mark.unit
@pytest.mark.parametrize(
    "edit",
    [
        _drop_excluded_rows,
        _relabel_a_candidate_row,
        _duplicate_an_excluded_row,
        _blank_a_withdrawn_expectation,
    ],
    ids=["deleted", "relabelled", "duplicated", "blanked"],
)
def test_row_decision_loader_rejects_a_hand_edited_row_set(
    tmp_path: Path, edit: Callable[[dict[str, Any]], None]
) -> None:
    """Every edit that moves the acceptance rate is now a load failure.

    Each of these loaded clean before the payload identity existed, and each moves
    the published denominator: deleting the excluded rows, relabelling candidate
    rows as suggestions, duplicating one excluded row, or blanking a withdrawn
    expectation — which destroys exactly the property
    `test_row_decisions_and_the_oracle_agree_on_the_expected_set` relies on.
    """
    export_path, payload = _row_decision_payload(tmp_path)
    before = load_row_decisions(export_path)
    edit(payload)
    _write_json(export_path, payload)

    with pytest.raises(GoldenSetValidationError, match="payload identity"):
        load_row_decisions(export_path)
    assert payload["rows"] != [row.model_dump(mode="json") for row in before.rows], (
        "the edit must actually change the row set"
    )


_UNDECLARED_CODE = "C999999"


def _append_constituent_row(workbook: Workbook, code: str) -> None:
    """Append one complete `ENGINE SUGGESTION` / `include` row for `code`."""
    sheet = workbook["Constituent Decisions"]
    row = sheet.max_row + 1
    sheet.cell(row, 2, code)
    sheet.cell(row, 4, "ENGINE SUGGESTION")
    sheet.cell(row, 5, "op:StageValue")
    sheet.cell(row, 6, "C27971")
    sheet.cell(row, 10, "include")
    sheet.cell(row, 11, "op:StageValue")
    sheet.cell(row, 12, "C27971")
    sheet.cell(row, 14, "FALSE")
    sheet.cell(row, 15, "ncit-26.07d")
    sheet.cell(row, 18, "YES")


@pytest.mark.unit
def test_row_decision_export_rejects_a_constituent_row_with_no_concept(
    tmp_path: Path,
) -> None:
    """An orphan constituent row inflates the denominator with an unreviewed code.

    `import-workbook` has always rejected a constituent row whose concept is absent
    from `Concept Decisions` — the SME never adjudicated that concept, so nothing
    licenses its rows. The export accepted it and wrote it into the tracked
    artifact, so the same workbook produced a clean acceptance denominator and an
    unloadable oracle. The export now applies the same check.
    """
    workbook_path = tmp_path / "orphan.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    _append_constituent_row(workbook, _UNDECLARED_CODE)
    workbook.save(workbook_path)

    message = f"constituent rows reference unknown concepts: {_UNDECLARED_CODE}"
    with pytest.raises(GoldenSetValidationError, match=message):
        export_row_decisions(workbook_path)
    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook_path)


def _conceal_a_concept_declaration(workbook: Workbook) -> None:
    """Declare a concept on a hidden row and give it a constituent row.

    A human reading the workbook sees neither the concept nor a reason for the
    constituent row that references it, but `_declared_concept_codes` reads the
    cell regardless, so the orphan gate is satisfied by an invisible declaration.
    """
    concepts = workbook["Concept Decisions"]
    row = concepts.max_row + 1
    for column, value in enumerate(
        [
            row - 4,
            _UNDECLARED_CODE,
            f"Reviewed {_UNDECLARED_CODE}",
            "Neoplastic Process",
            "Neoplastic Process",
            "decomposed",
            "accepted",
            "decomposed",
            "Reviewed against the stated source.",
            "YES",
            "YES",
        ],
        start=1,
    ):
        concepts.cell(row, column, value)
    concepts.row_dimensions[row].hidden = True
    _append_constituent_row(workbook, _UNDECLARED_CODE)


def _add_a_concept_row_with_no_code(workbook: Workbook) -> None:
    """Populate a concept row while leaving its `Concept Code` cell empty."""
    concepts = workbook["Concept Decisions"]
    row = concepts.max_row + 1
    concepts.cell(row, 3, "Reviewed but unnamed")
    concepts.cell(row, 7, "accepted")


@pytest.mark.unit
@pytest.mark.parametrize(
    "read",
    [export_row_decisions, import_adjudication_workbook],
    ids=["export", "import"],
)
@pytest.mark.parametrize(
    ("tamper", "message"),
    [
        (_conceal_a_concept_declaration, "hidden concept rows are not permitted"),
        (
            _add_a_concept_row_with_no_code,
            "populated concept row has blank concept code",
        ),
    ],
    ids=["hidden", "blank-code"],
)
def test_both_entry_points_apply_the_concept_sheet_preconditions(
    tmp_path: Path,
    read: Callable[[Path], object],
    tamper: Callable[[Workbook], None],
    message: str,
) -> None:
    """The orphan gate is only as strong as the sheet it reads its codes from.

    `_declared_concept_codes` accepts any textual code, including one on a hidden
    row, and ignores a populated row whose code cell is blank. The import reached
    the orphan check only after `_workbook_concepts` had refused both; the export
    called `_concept_headers` and `_declared_concept_codes` directly and so
    accepted a workbook the import then rejected — a clean acceptance denominator
    beside an unloadable oracle, from one file. The concealed case is the worse
    one: the concept licensing the constituent row is invisible to a reviewer.

    Both paths now share `_concept_decision_sheet`, so the two guards cannot
    diverge again without moving the code both of them call.
    """
    workbook_path = tmp_path / "concept-sheet.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    tamper(workbook)
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match=message):
        read(workbook_path)


@pytest.mark.unit
@pytest.mark.parametrize(
    "read",
    [export_row_decisions, import_adjudication_workbook],
    ids=["export", "import"],
)
def test_a_constituent_code_that_is_not_an_ncit_code_is_refused(
    tmp_path: Path, read: Callable[[Path], object]
) -> None:
    """The row model is the only thing checking the shape of a constituent code.

    `_constituent_row_identity` asks whether the cell holds text and nothing more,
    so `^C[0-9]+$` is enforced solely by `_ROW_DECISION_ADAPTER.validate_python`,
    whose `except` wrap turns the pydantic failure into a `GoldenSetValidationError`.
    Every other way that call can fail is pre-empted by an explicit check upstream,
    which left the pattern as the one live reject on that line — and it was
    uncovered on both entry points.

    The same value goes into `Concept Decisions!B5`, so the concept is declared and
    the orphan gate is not what fires. The assertion is on the message rather than
    the exception type, because a `GoldenSetValidationError` from any of the
    upstream checks would satisfy the type and prove nothing about the pattern.
    """
    workbook_path = tmp_path / "not-an-ncit-code.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Constituent Decisions"]["B5"] = "banana"
    workbook["Concept Decisions"]["B5"] = "banana"
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match="should match pattern"):
        read(workbook_path)


@pytest.mark.unit
def test_engine_suggestion_cannot_be_left_not_needed() -> None:
    """`not-needed` is a candidate-row action; on a suggestion it is a non-decision.

    `PENDING` already stops an unadjudicated engine suggestion reaching the
    measurement. `not-needed` reached the same state through a second door: it
    parses, it is retained by the export, it is dropped by the import, and it lands
    in the acceptance denominator as a suggestion the SME never accepted or
    rejected. It is now unrepresentable rather than rejected by a runtime check:
    `UnusedCandidateRow` pins `row_type`, so the combination has no inhabitant and
    a later variant cannot reintroduce it without changing that literal.
    """
    with pytest.raises(ValidationError, match="ADD IF MISSING"):
        UnusedCandidateRow(
            code="C0",
            row_type=cast("Any", "ENGINE SUGGESTION"),
            sme_action="not-needed",
            engine=None,
        )
    with pytest.raises(ValidationError, match="ADD IF MISSING"):
        _ROW_DECISION_ADAPTER.validate_python(
            {
                "code": "C0",
                "row_type": "ENGINE SUGGESTION",
                "sme_action": "not-needed",
                "engine": None,
            }
        )

    unused = UnusedCandidateRow(
        code="C0", row_type="ADD IF MISSING", sme_action="not-needed", engine=None
    )
    assert not isinstance(unused, KeptRow)
    assert not hasattr(unused, "expected")


@pytest.mark.unit
@pytest.mark.parametrize("blank", ["axis", "filler"])
@pytest.mark.parametrize(
    "read",
    [export_row_decisions, import_adjudication_workbook],
    ids=["export", "import"],
)
def test_a_half_named_withdrawal_is_refused_on_both_paths(
    tmp_path: Path, read: Callable[[Path], object], blank: str
) -> None:
    """An excluded row names the pair it withdrew, or it names nothing.

    `expected_axis` and `expected_filler` were independently `str | None`, so
    `(axis="op:StageValue", filler=None)` was a third state nothing rejected:
    `withdrawn_triple` returned `None` for it, which silently exempted the row from
    both the `(code, axis, filler)` uniqueness rule and the kept/withdrawn
    disjointness rule. Both entry points accepted such a workbook end to end.
    """
    workbook_path = tmp_path / f"half-withdrawal-{blank}.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Constituent Decisions"]
    sheet["J5"] = "exclude"
    sheet["K5" if blank == "axis" else "L5"] = None
    workbook.save(workbook_path)

    with pytest.raises(
        GoldenSetValidationError,
        match="C0 expected pair must name both an axis and a filler, or neither",
    ):
        read(workbook_path)


@pytest.mark.unit
def test_row_decision_export_records_the_pair_each_suggestion_offered(
    tmp_path: Path,
) -> None:
    """`Engine Axis` and `Engine Filler` are real columns nothing was reading.

    Without them the export recorded only what the reviewer wrote, so "48 accepted
    unchanged" was a claim about a label — relabelling `revise` rows whose pair the
    engine had in fact emitted moved the published rate from 0.4528 to 0.7547 with
    every test green. Recording the suggested pair makes "unchanged" a comparison
    the export can perform rather than a word in a docstring.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    _append_decision_rows(workbook)

    export = export_row_decisions(workbook)
    suggestion = next(row for row in export.rows if row.row_type == "ENGINE SUGGESTION")
    candidate = next(row for row in export.rows if row.row_type == "ADD IF MISSING")

    assert suggestion.engine is not None
    assert (suggestion.engine.axis, suggestion.engine.filler) == (
        "op:StageValue",
        "C27970",
    )
    assert candidate.engine is None


def _blank_the_engine_pair_of_a_suggestion(sheet: Worksheet) -> None:
    sheet["E5"] = None
    sheet["F5"] = None


def _relabel_a_suggestion_as_a_candidate(sheet: Worksheet) -> None:
    sheet["D5"] = "ADD IF MISSING"


@pytest.mark.unit
@pytest.mark.parametrize(
    ("tamper", "message"),
    [
        (_blank_the_engine_pair_of_a_suggestion, "C0 engine axis must be text"),
        (
            _relabel_a_suggestion_as_a_candidate,
            "C0 candidate row must not record an engine suggestion",
        ),
    ],
    ids=["suggestion-without-a-pair", "candidate-with-a-pair"],
)
@pytest.mark.parametrize(
    "read",
    [export_row_decisions, import_adjudication_workbook],
    ids=["export", "import"],
)
def test_the_engine_pair_is_present_exactly_on_the_engine_s_own_rows(
    tmp_path: Path,
    read: Callable[[Path], object],
    tamper: Callable[[Worksheet], None],
    message: str,
) -> None:
    """`ENGINE SUGGESTION` means the engine suggested something; the pair proves it.

    Row type was a free-text label the reviewer's sheet carried and nothing
    corroborated, so a candidate row could be relabelled a suggestion — the edit
    that moved the denominator from 106 to 121. The engine pair is now required on
    a suggestion and forbidden on a candidate, on both entry points, so the label
    cannot move without the evidence for it moving too.
    """
    workbook_path = tmp_path / "engine-pair.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    tamper(workbook["Constituent Decisions"])
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match=message):
        read(workbook_path)


@pytest.mark.unit
def test_a_revised_row_that_kept_the_engine_pair_is_counted_as_preserved(
    tmp_path: Path,
) -> None:
    """`include` counts a label; `pair_preserved` counts an equality.

    On the attested workbook the two differ: 32 of the 42 `revise` suggestions
    carry the engine's exact pair, so `include` is not "the rows whose pair
    survived". The fixture reproduces that shape — one `revise` row whose expected
    pair equals the engine's — so `pair_preserved` exceeds `include` and neither
    number can be derived from the other.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    _append_decision_rows(workbook)

    acceptance = export_row_decisions(workbook).cross_tab().engine_suggestion

    assert acceptance.include == 20
    assert acceptance.revise == 2
    assert acceptance.pair_preserved == 21
    assert acceptance.included_rate is not None
    assert round(acceptance.included_rate, 4) == round(20 / 23, 4)


@pytest.mark.unit
def test_more_pairs_cannot_be_preserved_than_the_reviewer_kept() -> None:
    """`pair_preserved` counts kept suggestions, so it cannot exceed them.

    A hand-edited tally claiming more preserved pairs than `include + revise` rows
    describes no row set, and nothing may accept it as evidence.
    """
    with pytest.raises(
        ValidationError, match="pair_preserved cannot exceed the kept suggestions"
    ):
        EngineAcceptance(include=1, revise=1, exclude=2, pair_preserved=3)

    assert (
        EngineAcceptance(
            include=1, revise=1, exclude=2, pair_preserved=2
        ).pair_preserved
        == 2
    )


@pytest.mark.unit
def test_a_kept_row_cannot_be_built_without_its_expected_pair() -> None:
    """`KeptRow` requires both halves, so no caller has to assert they are there.

    The pair used to be `str | None` on every row with a runtime validator saying
    "kept implies both present", which forced four unchecked `cast("str", ...)`
    sites. A `model_construct` or a new variant would have skipped the validator
    and left those casts asserting something no longer true.
    """
    for missing in ("axis", "filler"):
        expected = {"axis": "op:StageValue", "filler": "C27970"}
        del expected[missing]
        with pytest.raises(ValidationError, match=missing):
            _ROW_DECISION_ADAPTER.validate_python(
                {
                    "code": "C0",
                    "row_type": "ENGINE SUGGESTION",
                    "sme_action": "include",
                    "engine": {"axis": "op:StageValue", "filler": "C27970"},
                    "expected": expected,
                }
            )

    with pytest.raises(ValidationError, match="expected"):
        _ROW_DECISION_ADAPTER.validate_python(
            {
                "code": "C0",
                "row_type": "ENGINE SUGGESTION",
                "sme_action": "include",
                "engine": {"axis": "op:StageValue", "filler": "C27970"},
            }
        )

    kept = _ROW_DECISION_ADAPTER.validate_python(
        {
            "code": "C0",
            "row_type": "ENGINE SUGGESTION",
            "sme_action": "include",
            "engine": {"axis": "op:StageValue", "filler": "C27970"},
            "expected": {"axis": "op:StageValue", "filler": "C27970"},
        }
    )
    assert isinstance(kept, KeptRow)
    assert kept.expected_triple == ExpectedTriple("C0", "op:StageValue", "C27970")
    assert kept.pair_preserved


@pytest.mark.unit
def test_a_not_needed_candidate_row_cannot_record_an_expectation(
    tmp_path: Path,
) -> None:
    """`not-needed` says the reviewer never had to fill the row in.

    A pair in those cells contradicts the action, and `UnusedCandidateRow` has
    nowhere to put it. Before, both were recorded and then silently ignored by
    every consumer, so the workbook and the export disagreed with no error.
    """
    workbook_path = tmp_path / "not-needed-with-pair.xlsx"
    _create_workbook(workbook_path)
    _append_decision_rows(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Constituent Decisions"]
    row = next(
        index
        for index in range(5, sheet.max_row + 1)
        if sheet.cell(index, 10).value == "not-needed"
    )
    sheet.cell(row, 11, "op:PrimarySite")
    sheet.cell(row, 12, "C12345")
    workbook.save(workbook_path)

    message = "C2 not-needed candidate row must not record an expected pair"
    with pytest.raises(GoldenSetValidationError, match=message):
        export_row_decisions(workbook_path)
    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook_path)


@pytest.mark.unit
@pytest.mark.parametrize("complete", ["YES", "NO"])
def test_workbook_rejects_a_not_needed_engine_suggestion(
    tmp_path: Path, complete: str
) -> None:
    """`Row Complete?` cannot buy an engine suggestion out of being adjudicated.

    The waiver used to read "not an engine suggestion *and* `not-needed`", so an
    `ENGINE SUGGESTION` / `not-needed` row exempted itself from `Row Complete?` as
    well, passed both the export and the import, and shifted the acceptance
    denominator with no error. The combination is now refused before completeness
    is consulted, so the value in that cell cannot change the outcome.
    """
    workbook_path = tmp_path / f"not-needed-{complete}.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Constituent Decisions"]
    sheet["J5"] = "not-needed"
    sheet["R5"] = complete
    workbook.save(workbook_path)

    message = "C0 engine suggestion cannot be left not-needed"
    with pytest.raises(GoldenSetValidationError, match=message):
        export_row_decisions(workbook_path)
    with pytest.raises(GoldenSetValidationError, match=message):
        import_adjudication_workbook(workbook_path)


@pytest.mark.unit
def test_an_incomplete_engine_suggestion_is_still_rejected(tmp_path: Path) -> None:
    """The completeness waiver reaches only `not-needed` rows, nothing else.

    Pinned beside the case above so the reordering cannot quietly widen the
    waiver: an `ENGINE SUGGESTION` carrying a real action must still be complete.
    """
    workbook_path = tmp_path / "incomplete.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Constituent Decisions"]["R5"] = "NO"
    workbook.save(workbook_path)

    with pytest.raises(
        GoldenSetValidationError, match="C0 has incomplete constituent row"
    ):
        export_row_decisions(workbook_path)


@pytest.mark.unit
def test_the_cell_no_engine_suggestion_row_can_occupy_has_no_field(
    tmp_path: Path,
) -> None:
    """An impossible cell is absent from the *type*, not reported as a zero.

    Reporting `("ENGINE SUGGESTION", "not-needed"): 0` invited the reading that the
    combination is merely unused in this workbook. It cannot occur, so
    `EngineAcceptance` has nowhere to put it, and a reader cannot ask.
    """
    workbook = tmp_path / "review.xlsx"
    _create_workbook(workbook)
    _append_decision_rows(workbook)

    cross_tab = export_row_decisions(workbook).cross_tab()

    assert set(EngineAcceptance.model_fields) == {
        "include",
        "revise",
        "exclude",
        "pair_preserved",
    }
    assert set(CandidateOutcomes.model_fields) == {
        "include",
        "revise",
        "exclude",
        "not_needed",
    }
    assert not hasattr(cross_tab.engine_suggestion, "not_needed")
    assert cross_tab.add_if_missing.not_needed == 1


@pytest.mark.unit
def test_acceptance_is_undefined_when_the_run_suggested_nothing() -> None:
    """A review of a run that emitted no constituent has no rate, not a zero.

    `include / adjudicated` is a division the tracked evidence never performs with
    an empty denominator, so nothing else proves this branch is reachable — and a
    bare `float` return would have made `cross_tab()` raise `ZeroDivisionError` on
    a legitimate export whose rows are all SME-added candidates.
    """
    empty = EngineAcceptance(include=0, revise=0, exclude=0, pair_preserved=0)
    ruled_on = EngineAcceptance(include=1, revise=1, exclude=2, pair_preserved=1)

    assert empty.included_rate is None
    assert ruled_on.adjudicated == 4
    assert ruled_on.included_rate == 0.25


@pytest.mark.unit
def test_an_unrecognized_action_is_reported_before_the_completeness_gate(
    tmp_path: Path,
) -> None:
    """A typo in `SME Action` names the action, not the completeness column.

    The vocabulary check ran after the completeness check, so an unrecognised
    action on an incomplete row was reported as an incomplete row and the reviewer
    was sent to the wrong cell.
    """
    workbook_path = tmp_path / "invalid-action-incomplete.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Constituent Decisions"]
    sheet["J5"] = "maybe"
    sheet["R5"] = "NO"
    workbook.save(workbook_path)

    with pytest.raises(GoldenSetValidationError, match="invalid SME action: maybe"):
        export_row_decisions(workbook_path)


@pytest.mark.unit
def test_row_decision_loader_rejects_an_empty_row_set(tmp_path: Path) -> None:
    """An export with no rows is not a review, and the refusal names only the rows.

    The rejection is reachable only through the JSON loader: a workbook with no
    populated constituent row is a real possibility, but `export_row_decisions`
    reaches this gate with whatever the sheet held, so the empty case has to be
    provoked here. The message must not drag `_meta` into the failure text — a
    model-level check reported the whole validated input, and these land verbatim
    in a CLI error and a CI log.
    """
    export_path, payload = _row_decision_payload(tmp_path)
    payload["rows"] = []
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(GoldenSetValidationError) as error:
        load_row_decisions(export_path)

    assert "row decisions must not be empty" in str(error.value)
    assert "_meta" not in str(error.value)


@pytest.mark.unit
@pytest.mark.parametrize("field", ["axis", "filler"])
@pytest.mark.parametrize("action", ["include", "exclude"])
def test_row_decision_loader_rejects_a_padded_expectation(
    tmp_path: Path, action: str, field: str
) -> None:
    """The row model's canonical-text gate is live, and only on this path.

    `_cell_text`/`_optional_text` canonicalize before the row model is built, so
    from the workbook the model's own check can never fire. The signed JSON export
    is a real second entry point — `test_m1_baseline.py` reads the tracked file
    through it — and there the padded value reaches the model directly.
    """
    export_path, payload = _row_decision_payload(tmp_path)
    row = next(
        item
        for item in payload["rows"]
        if item["sme_action"] == action and item["expected"] is not None
    )
    row["expected"][field] = " " + row["expected"][field]
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(
        GoldenSetValidationError,
        match="must be non-empty without outer whitespace",
    ):
        load_row_decisions(export_path)


@pytest.mark.unit
@pytest.mark.parametrize("field", ["ncit_version", "source_workbook", "run_id"])
def test_row_decision_loader_rejects_padded_metadata_text(
    tmp_path: Path, field: str
) -> None:
    """A hand-edited `_meta` cannot smuggle a padded release, label or run id."""
    export_path, payload = _row_decision_payload(tmp_path)
    payload["_meta"][field] = " " + payload["_meta"][field]
    _write_json(export_path, _resign_row_decisions(payload))

    with pytest.raises(
        GoldenSetValidationError,
        match=f"{field} must be non-empty without outer whitespace",
    ):
        load_row_decisions(export_path)


@pytest.mark.unit
def test_row_decision_export_rejects_a_padded_evidence_value(
    tmp_path: Path,
) -> None:
    """`Source & Run Evidence` values are never canonicalized on the way in.

    `_evidence_values` reads both columns raw, so `RowDecisionMetadata`'s text gate
    is the only thing standing between a padded `NCIt release` cell and the tracked
    export — and reaching it is what exercises the `_model_error` wrap in
    `export_row_decisions_bytes`, which turns a pydantic failure into the
    `GoldenSetValidationError` every caller of this module handles.
    """
    workbook_path = tmp_path / "padded-release.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Source & Run Evidence"]
    sheet.cell(5, 2, " 26.07d")
    workbook.save(workbook_path)

    with pytest.raises(
        GoldenSetValidationError,
        match="ncit_version must be non-empty without outer whitespace",
    ):
        export_row_decisions(workbook_path)


@pytest.mark.unit
def test_row_decision_export_rejects_a_padded_source_workbook_label(
    tmp_path: Path,
) -> None:
    """The display label is caller-supplied, so it is held to the same form.

    `export_row_decisions` passes `Path.name`, which preserves a leading space in
    the file name; `export_row_decisions_bytes` accepts any label at all.
    """
    workbook_path = tmp_path / "review.xlsx"
    _create_workbook(workbook_path)

    with pytest.raises(
        GoldenSetValidationError,
        match="source_workbook must be non-empty without outer whitespace",
    ):
        golden_review.export_row_decisions_bytes(
            workbook_path.read_bytes(), " review.xlsx"
        )


@pytest.mark.unit
@pytest.mark.parametrize(
    "read",
    [export_row_decisions, import_adjudication_workbook],
    ids=["export", "import"],
)
def test_an_unreadable_workbook_path_is_reported_as_such(
    tmp_path: Path, read: Callable[[Path], object]
) -> None:
    """A missing or unreadable `.xlsx` fails as a workbook, not as a stack trace."""
    with pytest.raises(
        GoldenSetValidationError, match="cannot read adjudication workbook"
    ):
        read(tmp_path / "absent.xlsx")


@pytest.mark.unit
def test_row_decision_loader_rejects_a_json_document_that_is_not_an_object(
    tmp_path: Path,
) -> None:
    """A JSON array parses cleanly and is not a row-decision export."""
    export_path = tmp_path / "rows.json"
    _write_json(export_path, [])

    with pytest.raises(
        GoldenSetValidationError, match="row decisions must be a JSON object"
    ):
        load_row_decisions(export_path)
