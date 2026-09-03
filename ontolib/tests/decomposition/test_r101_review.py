from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path
from types import SimpleNamespace
from typing import TYPE_CHECKING, Any, cast
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
import pytest_asyncio
from defusedxml import ElementTree as DefusedET
from openpyxl import load_workbook
from scripts import adjudication
from scripts.adjudication import _parser

import ontolib.decomposition.r101_review as r101_review_module
from ontolib.decomposition.provenance_models import NcitSourceSnapshot
from ontolib.decomposition.r101_conservation import (
    R101ConservationValidationError,
    load_r101_conservation_report,
    validate_r101_publication,
)
from ontolib.decomposition.r101_review import (
    AtomicDecision,
    DiseaseProposition,
    QLeverReviewLabels,
    R101DecisionRegistry,
    R101ReviewPacket,
    R101ReviewValidationError,
    ReviewMembership,
    ReviewPath,
    ReviewPattern,
    build_r101_review_packet,
    dry_run_r101_decision_expansion,
    import_r101_review_decisions,
    load_r101_decision_registry,
    load_r101_review_packet,
    write_r101_review_workbook,
)

if TYPE_CHECKING:
    from collections.abc import Iterator

GOLDEN = Path(__file__).parent / "golden"
REPORT_PATH = GOLDEN / "neoplasm-r101-v4-conservation.json.gz"

APPROVE = "Approve non-exclusive coverage except marked exceptions"
REJECT = "Reject; retain broader site in projection"
INDIVIDUAL = "Require individual disease review"
ABSTAIN = "Abstain / escalate"
DECISIONS = (APPROVE, REJECT, INDIVIDUAL, ABSTAIN)
DENIALS = {
    "scope_non_exclusive": True,
    "source_preserved": True,
    "not_equivalent": True,
    "not_universal": True,
    "not_exclusive": True,
}
PATTERN_HEADERS = (
    "Pattern Number",
    "Review Proposition",
    "Broader Site",
    "Retained More-Specific Site",
    "Human-readable R82 path(s)",
    "Affected Disease Count",
    "Affected Diseases",
    "Source Occurrence Count",
    "One-step Count",
    "Transitive Count",
    "Min Path Length",
    "Max Path Length",
    "Context/Risk Summary",
    "Fixed Scope",
    "Decision",
    "Rationale",
    "Reviewer Identity",
    "Review Date",
)
DISEASE_HEADERS = (
    "Pattern Number",
    "Disease",
    "Specific Proposition",
    "Broader Site",
    "Retained Site",
    "Readable R82 path",
    "Source Occurrence Count",
    "Context Summary",
    "Review Priority / Risk Flags",
    "Exception?",
    "Exception Rationale",
)


class _Labels:
    def __init__(self, overrides: dict[str, list[str]] | None = None) -> None:
        self.query_count = 0
        self.requested_codes: tuple[str, ...] = ()
        self.overrides = overrides or {}

    async def labels_for_review(
        self, codes: tuple[str, ...]
    ) -> dict[str, tuple[str, ...]]:
        self.query_count += 1
        self.requested_codes = codes
        return {
            code: tuple(self.overrides.get(code, [f"TEST label {code}"]))
            for code in codes
        }


@pytest.fixture(scope="module")
def source_manifest(tmp_path_factory: pytest.TempPathFactory) -> Iterator[Path]:
    report = load_r101_conservation_report(REPORT_PATH)
    path = tmp_path_factory.mktemp("r101-source") / "manifest.json"
    path.write_text("{}", encoding="utf-8")
    patch = pytest.MonkeyPatch()
    patch.setattr(
        r101_review_module,
        "validate_ncit_sibling_manifest",
        lambda _path: NcitSourceSnapshot(
            source_identity=report.source_identity,
            ontology_version=report.source_release_id,
        ),
    )
    yield path
    patch.undo()


@pytest_asyncio.fixture(scope="module", loop_scope="module")
async def packet_and_labels(source_manifest: Path):
    report = load_r101_conservation_report(REPORT_PATH)
    labels = _Labels()
    packet = await build_r101_review_packet(report, source_manifest, labels)
    return report, packet, labels


def _headers(sheet: Any) -> dict[str, int]:
    return {str(cell.value): cast("int", cell.column) for cell in sheet[1]}


def _test_identity(payload: object) -> str:
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def _fill_pattern_decisions(path: Path, decision: str = APPROVE) -> None:
    book = load_workbook(path)
    sheet = book["Pattern Review"]
    headers = _headers(sheet)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, headers["Decision"], decision)
        sheet.cell(row, headers["Rationale"], "TEST-ONLY reviewed rationale")
        sheet.cell(row, headers["Reviewer Identity"], "TEST-ONLY reviewer")
        sheet.cell(row, headers["Review Date"], "2099-01-01")
    book.save(path)


def _workbook(packet: Any, path: Path, decision: str = APPROVE) -> None:
    write_r101_review_workbook(path, packet)
    _fill_pattern_decisions(path, decision)


@pytest.mark.unit
async def test_v3_packet_freezes_exact_human_and_atomic_inventories(
    packet_and_labels,
) -> None:
    report, packet, labels = packet_and_labels
    assert packet.schema_version == 3
    assert len(packet.patterns) == 162
    assert len(packet.disease_propositions) == 2800
    assert len(packet.occurrences) == 3291
    assert len(packet.membership) == 2800
    assert len({row.disease_code for row in packet.disease_propositions}) == 1916
    assert sum(row.occurrence_count for row in packet.patterns) == 3291
    assert (
        sorted(row.occurrence_count for row in packet.disease_propositions).count(1)
        == 2322
    )
    assert (
        sorted(row.occurrence_count for row in packet.disease_propositions).count(2)
        == 465
    )
    assert (
        sorted(row.occurrence_count for row in packet.disease_propositions).count(3)
        == 13
    )
    assert max(row.occurrence_count for row in packet.disease_propositions) == 3
    assert max(row.occurrence_count for row in packet.patterns) == 245
    assert {row.occurrence_id for row in packet.occurrences} == {
        row.occurrence_id
        for row in report.occurrences
        if row.disposition == "covered-by-retained-r82"
    }
    assert labels.query_count == 1
    assert len(labels.requested_codes) > 1900
    assert packet.guidance_identity
    assert packet.visible_rows_identity
    assert packet.membership_identity
    assert packet.packet_identity
    assert all(
        row.model_dump(include=set(DENIALS)) == DENIALS
        for row in (*packet.patterns, *packet.disease_propositions)
    )
    assert all(row.review_proposition for row in packet.patterns)
    assert all(row.specific_proposition for row in packet.disease_propositions)


@pytest.mark.unit
async def test_propositions_state_nonexclusive_projection_scope_without_overclaim(
    packet_and_labels,
) -> None:
    _, packet, _ = packet_and_labels
    text = "\n".join(
        [row.review_proposition for row in packet.patterns]
        + [row.specific_proposition for row in packet.disease_propositions]
    )
    required = (
        "valid, more-precise primary-site coverage",
        "in the curated projection",
        "broader source assertions remain preserved",
        "does not mean every case occurs only",
        "only valid site",
    )
    assert all(value in text for value in required)
    assert "equivalent" not in text.casefold()
    for row in packet.disease_propositions:
        assert f"{row.broader_label} ({row.broader_code})" in row.specific_proposition
        assert f"{row.retained_label} ({row.retained_code})" in row.specific_proposition


@pytest.mark.unit
async def test_workbook_is_human_centered_and_contains_no_internal_ids_or_json(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    path = tmp_path / "v3.xlsx"
    write_r101_review_workbook(path, packet)
    book = load_workbook(path)
    assert book.sheetnames == [
        "Instructions and Semantics",
        "Pattern Review",
        "Disease Propositions",
        "Column Definitions",
        "Review Examples",
        "Bindings",
    ]
    assert "Occurrence Evidence" not in book.sheetnames
    assert book["Bindings"].sheet_state == "veryHidden"
    assert tuple(cell.value for cell in book["Pattern Review"][1]) == PATTERN_HEADERS
    assert (
        tuple(cell.value for cell in book["Disease Propositions"][1]) == DISEASE_HEADERS
    )
    assert book["Pattern Review"].max_row == 163
    assert book["Disease Propositions"].max_row == 2801
    pattern_headers = _headers(book["Pattern Review"])
    disease_headers = _headers(book["Disease Propositions"])
    assert all(
        book["Pattern Review"].cell(row, pattern_headers["Decision"]).value is None
        for row in range(2, 164)
    )
    assert [
        book["Disease Propositions"].cell(row, disease_headers["Exception?"]).value
        for row in range(2, 2802)
    ] == ["No"] * 2800
    assert all(
        book["Disease Propositions"]
        .cell(row, disease_headers["Exception Rationale"])
        .value
        is None
        for row in range(2, 2802)
    )
    forbidden_headers = re.compile(
        r"(?:Pattern ID|Row Identity|Occurrence ID|Source Fact ID|"
        r"Source Group ID|Path Identity|Hash|JSON)",
        re.I,
    )
    allowed_binding_names = {
        "packet_identity",
        "guidance_identity",
        "visible_rows_identity",
        "membership_identity",
        "schema_version",
        "source_release_id",
    }
    for sheet in book.worksheets:
        if sheet.title == "Bindings":
            assert {
                str(row[0].value) for row in sheet.iter_rows(min_row=2)
            } == allowed_binding_names
            continue
        assert not any(forbidden_headers.search(str(cell.value)) for cell in sheet[1])
        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                assert "r101-" not in value
                assert not re.search(r"\b[0-9a-f]{64}\b", value)
                assert not value.startswith("{")
                assert not value.startswith("[")


@pytest.mark.unit
async def test_workbook_guidance_editability_and_formula_free_container(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    path = tmp_path / "v3.xlsx"
    write_r101_review_workbook(path, packet)
    book = load_workbook(path)
    guidance = "\n".join(
        str(cell.value or "")
        for row in book["Instructions and Semantics"].iter_rows()
        for cell in row
    )
    guidance = guidance.casefold()
    assert all(
        text.casefold() in guidance
        for text in (
            "projection coverage, not disease exclusivity",
            "non-exclusive projection coverage",
            "source assertions remain preserved",
            "zero strict rule-eligible cases",
            "no safe workload reduction",
            "Hiddenness is not security",
            "packet",
            "error signs",
            "generated no",
            "change only true exceptions to yes",
        )
    )
    assert "SEER" not in tuple(str(cell.value) for cell in book["Pattern Review"][1])
    unlocked = [
        (sheet.title, str(sheet.cell(1, cast("int", cell.column)).value))
        for sheet in book.worksheets
        for row in sheet.iter_rows(min_row=2)
        for cell in row
        if not cell.protection.locked
    ]
    assert len(unlocked) == 162 * 4 + 2800 * 2
    assert {header for _, header in unlocked} == {
        "Decision",
        "Rationale",
        "Reviewer Identity",
        "Review Date",
        "Exception?",
        "Exception Rationale",
    }
    assert all(sheet.protection.sheet for sheet in book.worksheets)
    assert book.calculation.calcMode == "auto"
    assert book.calculation.fullCalcOnLoad is False
    with ZipFile(path) as archive:
        roots = (
            DefusedET.fromstring(archive.read(name))
            for name in archive.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
        formula = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f"
        assert not any(any(True for _ in root.iter(formula)) for root in roots)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("decision", "atomic_kind"),
    [
        (APPROVE, "approved-non-exclusive-coverage"),
        (REJECT, "rejected-retain-broader"),
        (INDIVIDUAL, "individual-review-required"),
        (ABSTAIN, "escalated"),
    ],
)
async def test_every_pattern_decision_expands_to_atomic_frozen_membership(
    packet_and_labels, tmp_path: Path, decision: str, atomic_kind: str
) -> None:
    report, packet, _ = packet_and_labels
    workbook = tmp_path / f"{atomic_kind}.xlsx"
    registry_path = tmp_path / f"{atomic_kind}.json"
    _workbook(packet, workbook, decision)
    registry = import_r101_review_decisions(
        packet, workbook, registry_path, provenance="test-only"
    )
    assert load_r101_decision_registry(registry_path) == registry
    assert len(registry.pattern_decisions) == 162
    assert len(registry.atomic_decisions) == 3291
    assert {row.outcome for row in registry.atomic_decisions} == {atomic_kind}
    assert all(
        row.source_packet_identity == packet.packet_identity
        for row in registry.atomic_decisions
    )
    assert all(row.proposition_text for row in registry.atomic_decisions)
    assert all(
        row.model_dump(include=set(DENIALS)) == DENIALS
        for row in registry.atomic_decisions
    )
    result = dry_run_r101_decision_expansion(report, packet, registry)
    assert result.writes_performed is False
    assert result.atomic_decisions == 3291
    assert report.content_authorization.status == "pending"
    with pytest.raises(R101ConservationValidationError, match="authorization-missing"):
        validate_r101_publication(report)


@pytest.mark.unit
async def test_approved_exceptions_exclude_all_disease_occurrences(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    workbook = tmp_path / "exceptions.xlsx"
    _workbook(packet, workbook)
    book = load_workbook(workbook)
    disease = book["Disease Propositions"]
    headers = _headers(disease)
    target = packet.disease_propositions[0]
    disease.cell(2, headers["Exception?"], "Yes")
    disease.cell(2, headers["Exception Rationale"], "Disease-specific context differs")
    book.save(workbook)
    registry = import_r101_review_decisions(
        packet, workbook, tmp_path / "registry.json", provenance="test-only"
    )
    affected = [
        row
        for row in registry.atomic_decisions
        if row.pattern_number == target.pattern_number
        and row.disease_code == target.disease_code
    ]
    assert len(affected) == target.occurrence_count
    assert {row.outcome for row in affected} == {"disease-exception"}
    assert (
        sum(
            row.outcome == "approved-non-exclusive-coverage"
            for row in registry.atomic_decisions
        )
        == 3291 - target.occurrence_count
    )

    missing = tmp_path / "missing-rationale.xlsx"
    _workbook(packet, missing)
    book = load_workbook(missing)
    disease = book["Disease Propositions"]
    headers = _headers(disease)
    disease.cell(2, headers["Exception?"], "Yes")
    book.save(missing)
    with pytest.raises(R101ReviewValidationError, match="exception rationale"):
        import_r101_review_decisions(
            packet, missing, tmp_path / "none.json", provenance="test-only"
        )


@pytest.mark.unit
async def test_exceptions_are_invalid_for_nonapprove_patterns(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    path = tmp_path / "invalid-exception.xlsx"
    _workbook(packet, path, REJECT)
    book = load_workbook(path)
    disease = book["Disease Propositions"]
    headers = _headers(disease)
    disease.cell(2, headers["Exception?"], "Yes")
    disease.cell(2, headers["Exception Rationale"], "not allowed")
    book.save(path)
    with pytest.raises(R101ReviewValidationError, match="non-approve"):
        import_r101_review_decisions(
            packet, path, tmp_path / "none.json", provenance="test-only"
        )


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("pattern-count-minus", "162 pattern rows"),
        ("disease-count-plus", "2800 disease rows"),
        ("substitute-disease", "disease proposition differs"),
        ("edit-label", "pattern visible row differs"),
        ("edit-proposition", "pattern visible row differs"),
        ("reorder", "pattern visible row differs"),
        ("binding", "binding cells differ"),
    ],
)
async def test_named_visible_inventory_and_digest_refusal_gates_are_live(
    packet_and_labels, tmp_path: Path, mutation: str, message: str
) -> None:
    _, packet, _ = packet_and_labels
    path = tmp_path / f"{mutation}.xlsx"
    _workbook(packet, path)
    book = load_workbook(path)
    patterns = book["Pattern Review"]
    diseases = book["Disease Propositions"]
    pattern_headers = _headers(patterns)
    disease_headers = _headers(diseases)
    if mutation == "pattern-count-minus":
        patterns.delete_rows(patterns.max_row)
    elif mutation == "disease-count-plus":
        diseases.append([2801])
    elif mutation == "substitute-disease":
        diseases.cell(2, disease_headers["Disease"], "Substituted Disease (C999999)")
    elif mutation == "edit-label":
        patterns.cell(2, pattern_headers["Broader Site"], "Edited (C1)")
    elif mutation == "edit-proposition":
        patterns.cell(2, pattern_headers["Review Proposition"], "Universal claim")
    elif mutation == "reorder":
        first = [cell.value for cell in patterns[2]]
        second = [cell.value for cell in patterns[3]]
        for column, value in enumerate(second, 1):
            patterns.cell(2, column, value)
        for column, value in enumerate(first, 1):
            patterns.cell(3, column, value)
    elif mutation == "binding":
        book["Bindings"]["B2"] = "0" * 64
    book.save(path)
    with pytest.raises(R101ReviewValidationError, match=message):
        import_r101_review_decisions(
            packet, path, tmp_path / "must-not-exist.json", provenance="test-only"
        )
    assert not (tmp_path / "must-not-exist.json").exists()


@pytest.mark.unit
async def test_saved_reopened_and_openpyxl_resaved_workbooks_import_by_cells(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    original = tmp_path / "original.xlsx"
    _workbook(packet, original)
    original_bytes = original.read_bytes()
    book = load_workbook(original)
    resaved = tmp_path / "resaved.xlsx"
    book.save(resaved)
    # Container bytes are deliberately not an acceptance identity; either equality
    # or inequality after a third-party resave is benign.
    assert original_bytes
    assert resaved.read_bytes()
    registry = import_r101_review_decisions(
        packet, resaved, tmp_path / "registry.json", provenance="test-only"
    )
    assert len(registry.atomic_decisions) == 3291


@pytest.mark.unit
async def test_stale_v2_packet_and_workbook_refuse(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    stale_packet = tmp_path / "v2.json"
    payload = packet.model_dump(mode="json")
    payload["schema_version"] = 2
    stale_packet.write_text(json.dumps(payload))
    with pytest.raises(R101ReviewValidationError, match="schema_version"):
        load_r101_review_packet(stale_packet)

    stale_workbook = tmp_path / "v2.xlsx"
    _workbook(packet, stale_workbook)
    book = load_workbook(stale_workbook)
    book["Bindings"]["B6"] = 2
    book.save(stale_workbook)
    with pytest.raises(R101ReviewValidationError, match="binding cells differ"):
        import_r101_review_decisions(
            packet, stale_workbook, tmp_path / "none.json", provenance="test-only"
        )


@pytest.mark.unit
@pytest.mark.parametrize(
    ("field", "delta", "message"),
    [
        ("patterns", -1, "162 patterns"),
        ("disease_propositions", -1, "2800 disease propositions"),
        ("disease_propositions", 1, "2800 disease propositions"),
        ("occurrences", -1, "3291 occurrences"),
        ("occurrences", 1, "3291 occurrences"),
    ],
)
async def test_packet_count_refusal_gates_are_live(
    packet_and_labels, field: str, delta: int, message: str
) -> None:
    _, packet, _ = packet_and_labels
    payload = packet.model_dump(mode="python")
    rows = list(payload[field])
    if delta < 0:
        rows.pop()
    else:
        rows.append(rows[-1])
    payload[field] = tuple(rows)
    with pytest.raises(ValueError, match=message):
        R101ReviewPacket.model_validate(payload)


class _SelectRows:
    def __init__(self, labels: dict[str, tuple[str, ...]]) -> None:
        self.labels = labels
        self.batch_sizes: list[int] = []

    async def select(self, query: str, *, required_variables=()):
        codes = sorted(set(re.findall(r"#(C[0-9]+)>", query)))
        self.batch_sizes.append(len(codes))
        assert tuple(required_variables) == ("c", "label")
        return [
            {
                "c": f"http://ncicb.nci.nih.gov/xml/owl/EVS/Thesaurus.owl#{code}",
                "label": label,
            }
            for code in codes
            for label in self.labels.get(code, (f"Label {code}",))
        ]


@pytest.mark.unit
async def test_qlever_labels_are_bounded_and_refuse_ambiguous_rows(
    source_manifest: Path,
) -> None:
    codes = tuple(f"C{index}" for index in range(1201))
    client = _SelectRows({})
    reader = QLeverReviewLabels(client, max_batch_size=500)
    labels = await reader.labels_for_review(codes)
    assert len(labels) == 1201
    assert reader.query_count == math.ceil(1201 / 500)
    assert client.batch_sizes == [500, 500, 201]

    report = load_r101_conservation_report(REPORT_PATH)
    for override, message in (
        ((), "missing label"),
        (("one", "two"), "multiple labels"),
    ):
        with pytest.raises(R101ReviewValidationError, match=message):
            await build_r101_review_packet(
                report, source_manifest, _Labels({"C12727": list(override)})
            )


@pytest.mark.unit
async def test_prepare_certifies_source_before_batched_label_reads(
    monkeypatch, tmp_path: Path, capsys, source_manifest: Path
) -> None:
    report = load_r101_conservation_report(REPORT_PATH)
    events: list[str] = []

    async def source_snapshot(manifest: Path, endpoint: str) -> NcitSourceSnapshot:
        events.append("source-certified")
        return NcitSourceSnapshot(
            source_identity=report.source_identity,
            ontology_version=report.source_release_id,
        )

    class ClientContext:
        async def __aenter__(self):
            events.append("label-client-opened")
            return object()

        async def __aexit__(self, *_args):
            return None

    class Labels:
        query_count = 5

        def __init__(self, _client) -> None:
            events.append("label-reader-created")

    packet = SimpleNamespace(
        packet_identity="a" * 64,
        patterns=(1,) * 162,
        disease_propositions=(1,) * 2800,
        occurrences=(1,) * 3291,
    )

    async def build(*_args):
        events.append("labels-read")
        return packet

    monkeypatch.setattr(adjudication, "_source_snapshot", source_snapshot)
    monkeypatch.setattr(
        adjudication, "ncit_sparql_client", lambda _endpoint: ClientContext()
    )
    monkeypatch.setattr(adjudication, "QLeverReviewLabels", Labels)
    monkeypatch.setattr(adjudication, "build_r101_review_packet", build)
    monkeypatch.setattr(adjudication, "write_r101_review_packet", lambda *_args: None)
    monkeypatch.setattr(adjudication, "write_r101_review_workbook", lambda *_args: None)
    await adjudication._prepare_r101_review(
        cast(
            "Any",
            SimpleNamespace(
                report=REPORT_PATH,
                source_manifest=source_manifest,
                endpoint="http://qlever.test",
                output_packet=tmp_path / "packet.json",
                output_xlsx=tmp_path / "packet.xlsx",
            ),
        )
    )
    assert events == [
        "source-certified",
        "label-client-opened",
        "label-reader-created",
        "labels-read",
    ]
    output = capsys.readouterr().err
    assert "patterns=162 diseases=2800 occurrences=3291" in output
    assert "source_checks=9 label_reads=5 qlever_reads=14" in output


@pytest.mark.unit
def test_cli_exposes_decision_expansion_not_authorization() -> None:
    parser = _parser()
    common = [
        "dry-run-r101-decision-expansion",
        "--report",
        "report.gz",
        "--packet",
        "packet.json",
        "--registry",
        "registry.json",
        "--output",
        "result.json",
    ]
    assert parser.parse_args(common).command == "dry-run-r101-decision-expansion"
    with pytest.raises(SystemExit):
        parser.parse_args(["dry-run-r101-authorization"])


@pytest.mark.unit
async def test_archive_macros_links_formulas_and_duplicate_json_refuse(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    duplicate = tmp_path / "duplicate.json"
    duplicate.write_text('{"schema_version":3,"schema_version":3}')
    with pytest.raises(R101ReviewValidationError, match="duplicate JSON key"):
        load_r101_review_packet(duplicate)
    path = tmp_path / "review.xlsx"
    _workbook(packet, path)
    for member, message in (
        ("xl/vbaProject.bin", "macro"),
        ("xl/externalLinks/externalLink1.xml", "external link"),
    ):
        changed = tmp_path / Path(member).name
        changed.write_bytes(path.read_bytes())
        with ZipFile(changed, "a", ZIP_DEFLATED) as archive:
            archive.writestr(member, b"TEST")
        with pytest.raises(R101ReviewValidationError, match=message):
            import_r101_review_decisions(
                packet, changed, tmp_path / "none.json", provenance="test-only"
            )
    book = load_workbook(path)
    book["Pattern Review"]["P2"] = "=1+1"
    book.save(path)
    with pytest.raises(R101ReviewValidationError, match="formula"):
        import_r101_review_decisions(
            packet, path, tmp_path / "none.json", provenance="test-only"
        )


@pytest.mark.unit
def test_packet_row_identity_and_shape_refusal_gates_are_live(
    packet_and_labels,
) -> None:
    _, packet, _ = packet_and_labels
    path_payload = packet.patterns[0].paths[0].model_dump(mode="python")
    for field, value, message in (
        ("labels", (*path_payload["labels"], "extra"), "path labels"),
        ("fact_identities", (), "at least 1 item"),
        ("path_identity", "0" * 64, "path identity"),
    ):
        changed = {**path_payload, field: value}
        with pytest.raises(ValueError, match=message):
            ReviewPath.model_validate(changed)

    for mutation, message in (
        (
            lambda value: value["evidence_kind_counts"].update(
                {"one_step": value["evidence_kind_counts"]["one_step"] + 1}
            ),
            "evidence counts",
        ),
        (
            lambda value: value.update(
                {"disease_codes": (*value["disease_codes"], value["disease_codes"][0])}
            ),
            "membership is not unique",
        ),
        (lambda value: value.update({"row_identity": "0" * 64}), "row identity"),
    ):
        changed = packet.patterns[0].model_dump(mode="python")
        mutation(changed)
        with pytest.raises(ValueError, match=message):
            ReviewPattern.model_validate(changed)

    proposition = packet.disease_propositions[0].model_dump(mode="python")
    proposition["proposition_identity"] = "0" * 64
    with pytest.raises(ValueError, match="proposition identity"):
        DiseaseProposition.model_validate(proposition)

    member = packet.membership[0].model_dump(mode="python")
    member["membership_identity"] = "0" * 64
    with pytest.raises(ValueError, match="membership row identity"):
        ReviewMembership.model_validate(member)


@pytest.mark.unit
async def test_packet_aggregate_identity_and_membership_refusals(
    packet_and_labels,
) -> None:
    _, packet, _ = packet_and_labels
    for mutation, message in (
        (
            lambda value: value.update({"guidance_identity": "0" * 64}),
            "guidance identity",
        ),
        (
            lambda value: value.update({"visible_rows_identity": "0" * 64}),
            "visible rows identity",
        ),
        (
            lambda value: value.update({"membership_identity": "0" * 64}),
            "membership identity",
        ),
        (
            lambda value: value.update({"packet_identity": "0" * 64}),
            "packet identity",
        ),
    ):
        changed = packet.model_dump(mode="python")
        mutation(changed)
        with pytest.raises(ValueError, match=message):
            R101ReviewPacket.model_validate(changed)

    for field, value, message in (
        (
            "disease_code",
            packet.membership[1].disease_code,
            "proposition membership",
        ),
        (
            "occurrence_ids",
            packet.membership[1].occurrence_ids,
            "occurrence audit records",
        ),
    ):
        changed = packet.model_dump(mode="python")
        member = dict(changed["membership"][0])
        member[field] = value
        identity_payload = {
            key: item for key, item in member.items() if key != "membership_identity"
        }
        member["membership_identity"] = _test_identity(identity_payload)
        changed["membership"] = (member, *changed["membership"][1:])
        with pytest.raises(ValueError, match=message):
            R101ReviewPacket.model_validate(changed)


@pytest.mark.unit
async def test_builder_source_and_covered_row_refusal_gates_are_live(
    packet_and_labels,
    source_manifest: Path,
) -> None:
    report, _, _ = packet_and_labels
    for changed, message in (
        (report.model_copy(update={"source_identity": "0" * 64}), "source manifest"),
        (report.model_copy(update={"mechanical_status": "incomplete"}), "incomplete"),
        (report.model_copy(update={"grouping_presentation": ()}), "does not exhaust"),
    ):
        with pytest.raises(R101ReviewValidationError, match=message):
            await build_r101_review_packet(changed, source_manifest, _Labels())

    index = next(
        index
        for index, row in enumerate(report.occurrences)
        if row.disposition == "covered-by-retained-r82"
    )
    covered = report.occurrences[index]
    for changed_row, message in (
        (covered.model_copy(update={"old_links": ()}), "review-groupable"),
        (
            covered.model_copy(
                update={
                    "retained_r82_target": covered.retained_r82_target.model_copy(
                        update={"axis": "op:Other"}
                    )
                }
            ),
            "cross-axis",
        ),
    ):
        rows = list(report.occurrences)
        rows[index] = changed_row
        with pytest.raises(R101ReviewValidationError, match=message):
            await build_r101_review_packet(
                report.model_copy(update={"occurrences": tuple(rows)}),
                source_manifest,
                _Labels(),
            )


@pytest.mark.unit
async def test_workbook_guidance_headers_decisions_and_exception_refusals(
    packet_and_labels, tmp_path: Path
) -> None:
    _, packet, _ = packet_and_labels
    mutations = (
        ("visible-bindings", "visibility"),
        ("guidance", "guidance differs"),
        ("pattern-header", "pattern headers"),
        ("disease-header", "disease proposition headers"),
        ("invalid-decision", "closed values"),
        ("missing-exception", "generated No or a justified Yes"),
        ("invalid-exception", "generated No or a justified Yes"),
        ("no-with-rationale", "allowed only"),
    )
    for name, message in mutations:
        path = tmp_path / f"{name}.xlsx"
        _workbook(packet, path)
        book = load_workbook(path)
        if name == "visible-bindings":
            book["Bindings"].sheet_state = "visible"
        elif name == "guidance":
            book["Instructions and Semantics"]["B2"] = "changed"
        elif name == "pattern-header":
            book["Pattern Review"]["A1"] = "Wrong"
        elif name == "disease-header":
            book["Disease Propositions"]["A1"] = "Wrong"
        elif name == "invalid-decision":
            book["Pattern Review"]["O2"] = "maybe"
        elif name == "missing-exception":
            book["Disease Propositions"]["J2"] = None
        elif name == "invalid-exception":
            book["Disease Propositions"]["J2"] = "Maybe"
        else:
            book["Disease Propositions"]["K2"] = "not allowed for No"
        book.save(path)
        with pytest.raises(R101ReviewValidationError, match=message):
            import_r101_review_decisions(
                packet, path, tmp_path / "none.json", provenance="test-only"
            )


@pytest.mark.unit
async def test_registry_identity_atomic_identity_and_stale_bindings_refuse(
    packet_and_labels, tmp_path: Path
) -> None:
    report, packet, _ = packet_and_labels
    workbook = tmp_path / "review.xlsx"
    _workbook(packet, workbook)
    registry = import_r101_review_decisions(
        packet, workbook, tmp_path / "registry.json", provenance="test-only"
    )
    atomic_payload = registry.atomic_decisions[0].model_dump(mode="python")
    atomic_payload["atomic_decision_identity"] = "0" * 64
    with pytest.raises(ValueError, match="atomic decision identity"):
        AtomicDecision.model_validate(atomic_payload)

    registry_payload = registry.model_dump(mode="python")
    registry_payload["registry_identity"] = "0" * 64
    with pytest.raises(ValueError, match="registry identity"):
        R101DecisionRegistry.model_validate(registry_payload)

    for stale, message in (
        (registry.model_copy(update={"packet_identity": "0" * 64}), "packet identity"),
        (registry.model_copy(update={"report_identity": "0" * 64}), "report identity"),
        (registry.model_copy(update={"source_identity": "0" * 64}), "source identity"),
    ):
        with pytest.raises(R101ReviewValidationError, match=message):
            dry_run_r101_decision_expansion(report, packet, stale)
