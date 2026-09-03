from __future__ import annotations

import json
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import cast
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from scripts.adjudication import _parser
from scripts.research import group_review_packet as group_review
from scripts.research.current_evidence import CurrentComparison, CurrentEngineEvidence
from scripts.research.group_review_packet import (
    GroupReviewPacket,
    HumanReviewPendingDisposition,
    build_group_review_packet,
    diagnose_grouping,
    generate_group_review_packet,
    load_group_review_packet,
)

from ontolib.decomposition.r101_conservation import load_r101_conservation_report

pytestmark = pytest.mark.unit

_ROOT = Path(__file__).parents[3]
_GOLDEN = Path(__file__).with_name("golden")
_EVIDENCE = _GOLDEN / "neoplasm-current-engine-evidence.json"
_COMPARISON = _GOLDEN / "neoplasm-current-comparison.json"
_R101 = _GOLDEN / "neoplasm-r101-v4-conservation.json.gz"

APPROVE = "Approve intentional normalization"
CORRECT = "Require source-reproducible correction"
REJECT = "Reject proposed regrouping"
ABSTAIN = "Abstain / escalate"


def _sheet_headers(sheet) -> dict[str, int]:
    return {str(cell.value): cast("int", cell.column) for cell in sheet[1]}


def _inputs() -> tuple[CurrentEngineEvidence, CurrentComparison]:
    return (
        CurrentEngineEvidence.model_validate_json(_EVIDENCE.read_bytes()),
        CurrentComparison.model_validate_json(_COMPARISON.read_bytes()),
    )


def _packet() -> GroupReviewPacket:
    evidence, comparison = _inputs()
    return build_group_review_packet(
        evidence=evidence,
        comparison=comparison,
        r101_report=load_r101_conservation_report(_R101),
    )


def test_packet_derives_current_cohort_metrics_and_controls() -> None:
    evidence, comparison = _inputs()
    packet = _packet()

    assert packet.historical_full_partition_agreement.model_dump() == {
        "numerator": 2,
        "denominator": 20,
        "provenance": "historical-57",
    }
    assert packet.current_metrics == comparison.metrics
    assert packet.cohort.outcome_counts == dict(
        sorted(Counter(item.outcome for item in evidence.concepts).items())
    )
    assert packet.cohort.decomposed_codes == tuple(
        item.code for item in evidence.concepts if item.outcome == "decomposed"
    )
    assert {item.outcome for item in packet.cohort.controls} == {
        "semantic-excluded",
        "atomic-no-op",
    }
    assert all(
        item.interpretation == "empty-partition-control-not-grouping-success"
        for item in packet.cohort.controls
    )
    highest = max(
        evidence.concepts,
        key=lambda item: (len(item.all_source_occurrences), item.code),
    )
    observed_highest = (
        packet.cohort.highest_fanout_code,
        packet.cohort.highest_fanout_occurrences,
    )
    assert observed_highest == (
        highest.code,
        len(highest.all_source_occurrences),
    )


def test_every_disagreement_has_total_pair_group_and_disposition_diagnosis() -> None:
    packet = _packet()

    assert len(packet.concepts) == 18
    assert packet.schema_version == 4
    assert any(item.pair_relations.expected_not_emitted for item in packet.concepts)
    assert any(item.pair_relations.current_only_scoreable for item in packet.concepts)
    assert {item.grouping_diagnosis.kind for item in packet.concepts} >= {
        "agrees-on-common-pairs",
        "over-merge",
        "over-split",
    }
    a, b, c = ("op:A", "C1"), ("op:B", "C2"), ("op:C", "C3")
    assert diagnose_grouping(((a, b), (c,)), ((a, c), (b,))).kind == "misassignment"
    assert all(item.disposition.status != "accepted" for item in packet.concepts)
    assert any(
        (
            item.pair_relations.expected_not_emitted
            or item.pair_relations.current_only_scoreable
        )
        and item.grouping_diagnosis.kind
        in {"over-merge", "over-split", "misassignment"}
        for item in packet.concepts
    )
    assert {item.disposition.status for item in packet.concepts} == {
        "human-review-pending"
    }
    for item in packet.concepts:
        assert isinstance(item.disposition, HumanReviewPendingDisposition)
        assert item.disposition.acceptance == "not-recorded"


def test_packet_copies_pair_relations_and_exposes_non_scoreable_occurrences() -> None:
    _evidence, comparison = _inputs()
    packet = _packet()
    comparison_by_code = {item.code: item for item in comparison.concepts}

    for concept, row in zip(packet.concepts, packet.review_rows, strict=True):
        expected = comparison_by_code[concept.code].pair_relations
        assert concept.pair_relations == expected
        assert row.pair_relations == expected
        assert row.current_scoreable_group_ids == tuple(
            group.normalized_group_id for group in concept.actual_groups
        )
        for context in concept.non_scoreable_emitted_pairs:
            assert context.normalized_group_membership == "excluded-non-scoreable"
            assert context.relation in {
                "expected-emitted-review-bearing",
                "current-only-review-bearing",
                "current-only-proposed",
            }
            if context.relation != "current-only-proposed":
                assert context.source_occurrences
                assert context.source_occurrence_ids == tuple(
                    item.occurrence_id for item in context.source_occurrences
                )

    expected_review = {
        ("C101539", "C47806"),
        ("C101539", "C47817"),
        ("C132677", "C40557"),
        ("C132677", "C40989"),
        ("C132677", "C48322"),
        ("C100054", "C36027"),
        ("C100054", "C8326"),
    }
    observed_review = {
        (concept.code, context.pair[1])
        for concept in packet.concepts
        for context in concept.non_scoreable_emitted_pairs
        if context.relation == "expected-emitted-review-bearing"
    }
    assert expected_review <= observed_review
    c132677 = next(item for item in packet.concepts if item.code == "C132677")
    assert ("op:ClinicalFinding", "C41444") in (
        c132677.pair_relations.expected_not_emitted
    )
    c100054 = next(item for item in packet.concepts if item.code == "C100054")
    assert not c100054.pair_relations.expected_not_emitted


def test_every_actual_group_cites_exact_pair_and_source_occurrences() -> None:
    evidence, _comparison = _inputs()
    packet = _packet()
    evidence_by_code = {item.code: item for item in evidence.concepts}

    for concept in packet.concepts:
        expected_pairs = {
            pair for group in concept.expected_groups for pair in group.pairs
        }
        actual_pairs = {
            pair.pair for group in concept.actual_groups for pair in group.pairs
        }
        assert expected_pairs == {
            pair for block in concept.expected_partition for pair in block
        }
        assert actual_pairs == {
            pair for block in concept.actual_partition for pair in block
        }
        assert all(
            group.source_evidence.availability == "unavailable-historical-oracle"
            for group in concept.expected_groups
        )
        source_occurrences = {
            item.occurrence_id: item
            for item in evidence_by_code[concept.code].all_source_occurrences
        }
        for group in concept.actual_groups:
            assert not group.normalized_group_id.startswith("source:")
            for pair in group.pairs:
                if pair.availability == "unavailable-upstream":
                    assert pair.reason == "source-occurrence-unavailable-upstream"
                    continue
                assert pair.occurrences
                assert pair.occurrence_ids == tuple(
                    item.occurrence_id for item in pair.occurrences
                )
                assert all(
                    source_occurrences[item.occurrence_id].model_dump()
                    == item.model_dump()
                    for item in pair.occurrences
                )
                assert all(
                    item.filler_code == pair.pair[1] for item in pair.occurrences
                )


def test_rule_boundary_names_complete_catalog_without_fabricating_evidence() -> None:
    packet = _packet()

    assert tuple(item.kind for item in packet.transformation_rule_catalog) == (
        "co-assertion-preservation",
        "routing",
        "specificity-collapse",
        "repeated-pairs",
        "reviewed-regrouping",
    )
    assert packet.review_boundary.status == "human-review-pending"
    assert (
        packet.review_boundary.reason
        == "machine-evidence-complete-human-decision-blank"
    )
    assert packet.review_boundary.sme_adjudication == "not-recorded"


def test_wrong_highest_fanout_normalized_partition_is_rejected() -> None:
    evidence, comparison = _inputs()
    highest = max(
        evidence.concepts,
        key=lambda item: (len(item.all_source_occurrences), item.code),
    )
    target = next(item for item in comparison.concepts if item.code == highest.code)
    wrong_full = target.full_partition.model_copy(
        update={"actual_partition": target.full_partition.expected_partition}
    )
    wrong_target = target.model_copy(update={"full_partition": wrong_full})
    wrong = comparison.model_copy(
        update={
            "concepts": tuple(
                wrong_target if item.code == highest.code else item
                for item in comparison.concepts
            )
        }
    )

    with pytest.raises(ValueError, match="actual normalized partition"):
        build_group_review_packet(
            evidence=evidence,
            comparison=wrong,
            r101_report=load_r101_conservation_report(_R101),
        )


def test_packet_rejects_rebound_and_aliased_group_identity() -> None:
    evidence, comparison = _inputs()
    rebound = comparison.model_copy(update={"source_identity": "f" * 64})
    with pytest.raises(ValueError, match="source identity"):
        build_group_review_packet(
            evidence=evidence,
            comparison=rebound,
            r101_report=load_r101_conservation_report(_R101),
        )

    packet = _packet()
    payload = packet.model_dump(mode="json")
    actual = payload["concepts"][0]["actual_groups"][0]
    actual["normalized_group_id"] = actual["source_group_ids"][0]
    with pytest.raises(ValidationError, match="normalized group identity"):
        GroupReviewPacket.model_validate_json(json.dumps(payload))


def test_generator_writes_canonical_identity_checked_json(tmp_path: Path) -> None:
    output = tmp_path / "group-review.json"
    generated = generate_group_review_packet(
        evidence_path=_EVIDENCE,
        comparison_path=_COMPARISON,
        r101_report_path=_R101,
        output=output,
    )

    assert load_group_review_packet(output) == generated
    assert output.read_bytes().endswith(b"\n")
    persisted_identity = json.loads(output.read_bytes())["packet_identity"]
    assert persisted_identity == generated.packet_identity

    payload = json.loads(output.read_bytes())
    duplicate = payload["concepts"][0]["pair_relations"]["expected_matched_scoreable"][
        0
    ]
    payload["concepts"][0]["pair_relations"]["expected_not_emitted"].append(duplicate)
    output.write_text(json.dumps(payload), encoding="utf-8")
    with pytest.raises(ValueError, match=r"packet identity|pair relations"):
        load_group_review_packet(output)


def test_current_loader_accepts_schema_4_and_rejects_historical_schema_3(
    tmp_path: Path,
) -> None:
    current = tmp_path / "current.json"
    current.write_text(_packet().model_dump_json(), encoding="utf-8")

    assert load_group_review_packet(current).schema_version == 4
    with pytest.raises(ValidationError, match="schema_version"):
        load_group_review_packet(
            _ROOT / "evidence/group-review-packet-26.07d-schema3.json"
        )


def test_strict_boundary_rejects_coercion_and_unknown_fields() -> None:
    payload = _packet().model_dump(mode="json")
    payload["cohort"]["highest_fanout_occurrences"] = "99"
    payload["unexpected"] = True

    with pytest.raises(ValidationError):
        GroupReviewPacket.model_validate(payload)


def test_group_review_cli_requires_both_bound_inputs_and_output() -> None:
    args = _parser().parse_args(
        [
            "generate-group-review-packet",
            "--current-evidence",
            "evidence.json",
            "--current-comparison",
            "comparison.json",
            "--r101-report",
            "r101.json.gz",
            "--output",
            "packet.json",
            "--workbook",
            "review.xlsx",
            "--correction-audit",
            "audit.xlsx",
            "--blank-validation",
            "blank-validation.json",
        ]
    )

    assert args.current_evidence == Path("evidence.json")
    assert args.current_comparison == Path("comparison.json")
    assert args.r101_report == Path("r101.json.gz")
    assert args.output == Path("packet.json")
    assert args.workbook == Path("review.xlsx")
    assert args.correction_audit == Path("audit.xlsx")
    assert args.blank_validation == Path("blank-validation.json")


def _review_boundary():
    evidence, comparison = _inputs()
    return group_review.build_machine_group_review_packet(
        evidence=evidence, comparison=comparison, r101_report_path=_R101
    )


@pytest.mark.unit
def test_machine_rule_evidence_joins_exact_source_and_output_witnesses() -> None:
    packet = _review_boundary()
    evidence, _ = _inputs()
    occurrences = {
        row.occurrence_id: row
        for concept in evidence.concepts
        for row in concept.all_source_occurrences
    }
    by_kind = {
        kind: [row for row in packet.rule_evidence if row.kind == kind]
        for kind in packet.rule_kinds
    }

    assert all(by_kind.values())
    for row in packet.rule_evidence:
        assert row.source_occurrence_ids
        assert row.source_fact_ids
        assert row.source_group_ids
        assert all(value in occurrences for value in row.source_occurrence_ids)
        assert {
            occurrences[value].source_fact_id for value in row.source_occurrence_ids
        } == set(row.source_fact_ids)
        assert {
            occurrences[value].source_group_id for value in row.source_occurrence_ids
        } == set(row.source_group_ids)
    assert any(row.output_group_ids for row in by_kind["co-assertion-preservation"])
    assert all(row.output_pairs for row in by_kind["routing"])
    assert any(row.r82_path for row in by_kind["specificity-collapse"])
    assert all(len(row.source_occurrence_ids) > 1 for row in by_kind["repeated-pairs"])
    assert all(
        row.proposed_partition and row.affected_co_membership
        for row in by_kind["reviewed-regrouping"]
    )
    assert all(
        row.machine_evidence_limitation
        == "historical expected partition has no source citations"
        for row in by_kind["reviewed-regrouping"]
    )


@pytest.mark.unit
def test_blank_workbook_has_no_machine_generated_human_text(tmp_path: Path) -> None:
    packet = _review_boundary()
    workbook = tmp_path / "group-review.xlsx"
    group_review.write_group_review_workbook(workbook, packet)
    book = load_workbook(workbook)
    sheet = book["Group Review"]
    headers = _sheet_headers(sheet)
    assert sheet.max_row - 1 == len(packet.review_rows)
    assert len(packet.review_rows) == sum(
        not row.full_partition.agrees for row in _inputs()[1].concepts
    )
    for name in ("Decision", "Rationale", "Reviewer", "Date"):
        assert all(
            sheet.cell(row, headers[name]).value is None
            for row in range(2, sheet.max_row + 1)
        )


@pytest.mark.unit
def test_workbook_exposes_linked_human_readable_evidence_and_expected_warning(
    tmp_path: Path,
) -> None:
    packet = _review_boundary()
    workbook = tmp_path / "group-review.xlsx"
    group_review.write_group_review_workbook(workbook, packet)
    book = load_workbook(workbook)

    assert book.sheetnames == [
        "Instructions",
        "Group Review",
        "Group Evidence",
        "Pair Relation Evidence",
        "Source Evidence",
        "Rule Evidence",
        "Bindings",
    ]
    instructions = "\n".join(
        str(cell.value or "") for row in book["Instructions"] for cell in row
    )
    assert "expected-side source evidence is unavailable" in instructions.casefold()
    assert "oracle proposal" in instructions.casefold()
    review = book["Group Review"]
    headers = _sheet_headers(review)
    assert all(
        "Rule Evidence row" in str(review.cell(row, headers["Evidence Links"]).value)
        and "Group Evidence row"
        in str(review.cell(row, headers["Evidence Links"]).value)
        for row in range(2, review.max_row + 1)
    )
    source = book["Source Evidence"]
    source_headers = _sheet_headers(source)
    assert {
        "Occurrence ID",
        "Source Fact ID",
        "Source Group ID",
        "Anchor Code",
        "Depth",
        "Structural Path",
        "Member Position",
        "Role Code",
        "Filler Code",
        "Label Availability",
        "Definition Availability",
    } <= set(source_headers)
    rules = book["Rule Evidence"]
    rule_headers = _sheet_headers(rules)
    assert {
        "Kind",
        "Source Occurrence IDs",
        "Source Fact IDs",
        "Source Group IDs",
        "Output Group IDs",
        "Output Pairs",
        "R82 Path",
        "Proposed Partition",
        "Affected Co-membership",
        "Evidence Limitation",
    } <= set(rule_headers)
    relation_evidence = book["Pair Relation Evidence"]
    relation_headers = _sheet_headers(relation_evidence)
    assert {
        "Concept",
        "Pair Relation",
        "Pair",
        "Occurrence Context",
        "Normalized Group Membership",
    } <= set(relation_headers)
    review_bearing_rows = [
        row
        for row in relation_evidence.iter_rows(min_row=2, values_only=True)
        if "review-bearing" in str(row[relation_headers["Pair Relation"] - 1])
    ]
    assert review_bearing_rows
    forbidden = {"missing", "omitted", "historical-only", "restore", "add"}
    assert not any(
        word in " ".join(str(value or "").casefold() for value in row)
        for row in review_bearing_rows
        for word in forbidden
    )
    assert all(
        row[relation_headers["Normalized Group Membership"] - 1]
        == "Excluded from scoreable grouping"
        for row in review_bearing_rows
    )
    groups = book["Group Evidence"]
    assert any(
        "EXPECTED SOURCE UNAVAILABLE" in str(cell.value)
        for row in groups.iter_rows()
        for cell in row
    )


@pytest.mark.unit
def test_workbook_names_pair_relations_and_scoreable_groups_accurately(
    tmp_path: Path,
) -> None:
    packet = _review_boundary()
    workbook = tmp_path / "group-review.xlsx"
    group_review.write_group_review_workbook(workbook, packet)
    book = load_workbook(workbook)
    review = book["Group Review"]
    headers = _sheet_headers(review)

    assert "Pair Delta" not in headers
    assert "Actual Group IDs" not in headers
    assert "Pair Relations" in headers
    assert "Current Scoreable Group IDs" in headers
    c100054_row = next(
        row
        for row in range(2, review.max_row + 1)
        if review.cell(row, headers["Concept"]).value == "C100054"
    )
    relations = json.loads(
        str(review.cell(c100054_row, headers["Pair Relations"]).value)
    )
    assert relations["expected_not_emitted"] == []
    assert all(
        review.cell(row, headers[name]).protection.locked
        for row in range(2, review.max_row + 1)
        for name in (
            "Pair Relations",
            "Current Scoreable Group IDs",
            "Machine Evidence / Suggestion",
        )
    )


@pytest.mark.unit
def test_blank_correction_audit_and_validation_cannot_authorize_readiness(
    tmp_path: Path,
) -> None:
    packet = _review_boundary()
    review = tmp_path / "group-review-rev2.xlsx"
    audit = tmp_path / "group-correction-audit-rev2.xlsx"
    validation_path = tmp_path / "blank-validation-rev2.json"
    registry = tmp_path / "must-not-exist-decision-registry.json"
    group_review.write_group_review_workbook(review, packet)
    group_review.write_group_correction_audit(audit, packet)

    with pytest.raises(ValueError, match="all human fields are required"):
        group_review.import_group_review_decisions(packet, review, registry)
    assert not registry.exists()
    result = group_review.validate_blank_group_review_outputs(
        packet=packet,
        review_workbook=review,
        correction_audit=audit,
        output=validation_path,
    )

    assert result.model_dump(mode="json") == {
        "schema_version": 1,
        "packet_identity": packet.packet_identity,
        "review_workbook_identity": group_review.hashlib.sha256(
            review.read_bytes()
        ).hexdigest(),
        "correction_audit_identity": group_review.hashlib.sha256(
            audit.read_bytes()
        ).hexdigest(),
        "writes_performed": False,
        "human_decisions_present": False,
        "authorization": False,
        "readiness_eligible": False,
    }
    assert json.loads(validation_path.read_bytes()) == result.model_dump(mode="json")
    book = load_workbook(audit, data_only=False)
    assert book.sheetnames == ["Instructions", "Correction Audit", "Bindings"]
    sheet = book["Correction Audit"]
    headers = _sheet_headers(sheet)
    human = ("Correction Action", "Correction Rationale", "Reviewer", "Date")
    assert all(
        sheet.cell(row, headers[name]).value is None
        and sheet.cell(row, headers[name]).protection.locked is False
        for row in range(2, sheet.max_row + 1)
        for name in human
    )
    relations = {
        str(sheet.cell(row, headers["Pair Relation"]).value)
        for row in range(2, sheet.max_row + 1)
    }
    assert {
        "expected-not-emitted",
        "expected-emitted-review-bearing",
        "current-only-scoreable",
        "current-only-proposed",
        "scoreable-grouping",
    } <= relations
    assert not any(
        "decision registry" in str(cell.value or "").casefold()
        for worksheet in book.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


@pytest.mark.unit
def test_group_workbook_bytes_are_deterministic(tmp_path: Path) -> None:
    packet = _review_boundary()
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    group_review.write_group_review_workbook(first, packet)
    time.sleep(2.1)
    group_review.write_group_review_workbook(second, packet)
    assert first.read_bytes() == second.read_bytes()


@pytest.mark.unit
def test_pair_decision_is_editable_and_required_for_pair_only_rows(
    tmp_path: Path,
) -> None:
    packet = _review_boundary()
    path = tmp_path / "review.xlsx"
    group_review.write_group_review_workbook(path, packet)
    book = load_workbook(path)
    sheet = book["Group Review"]
    headers = _sheet_headers(sheet)
    pair_only_row = next(
        index
        for index, row in enumerate(packet.review_rows, start=2)
        if row.review_type == "pair-only"
    )
    assert (
        sheet.cell(pair_only_row, headers["Pair Decision"]).protection.locked is False
    )

    for index, row in enumerate(packet.review_rows, start=2):
        sheet.cell(index, headers["Decision"], APPROVE)
        if row.review_type == "pair-only" and index != pair_only_row:
            sheet.cell(index, headers["Pair Decision"], APPROVE)
        sheet.cell(index, headers["Rationale"], "TEST-ONLY SME rationale")
        sheet.cell(index, headers["Reviewer"], "TEST-ONLY reviewer")
        sheet.cell(index, headers["Date"], "2099-01-01")
    book.save(path)

    with pytest.raises(ValueError, match="pair decision is required for pair-only"):
        group_review.import_group_review_decisions(
            packet, path, tmp_path / "must-not-exist.json"
        )
    assert not (tmp_path / "must-not-exist.json").exists()


def _filled_workbook(tmp_path: Path):
    packet = _review_boundary()
    path = tmp_path / "review.xlsx"
    group_review.write_group_review_workbook(path, packet)
    book = load_workbook(path)
    sheet = book["Group Review"]
    headers = _sheet_headers(sheet)
    for row, review_row in enumerate(packet.review_rows, start=2):
        sheet.cell(row, headers["Decision"], APPROVE)
        if review_row.review_type == "pair-only":
            sheet.cell(row, headers["Pair Decision"], APPROVE)
        sheet.cell(row, headers["Rationale"], "TEST-ONLY SME rationale")
        sheet.cell(row, headers["Reviewer"], "TEST-ONLY reviewer")
        sheet.cell(row, headers["Date"], "2099-01-01")
    book.save(path)
    return packet, path, headers


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("immutable", "immutable"),
        ("missing", "review rows"),
        ("duplicate", "review rows"),
        ("blank", "human fields"),
        ("machine-rationale", "machine-generated text"),
        ("stale", "binding"),
        ("formula", "formula"),
        ("macro", "macro"),
        ("external", "external link"),
        ("contradiction", "contradictory"),
    ],
)
def test_import_refuses_unsafe_incomplete_or_contradictory_review(
    tmp_path: Path, mutation: str, message: str
) -> None:
    packet, path, headers = _filled_workbook(tmp_path)
    if mutation in {"macro", "external"}:
        member = (
            "xl/vbaProject.bin"
            if mutation == "macro"
            else "xl/externalLinks/externalLink1.xml"
        )
        with ZipFile(path, "a", ZIP_DEFLATED) as archive:
            archive.writestr(member, b"TEST")
    else:
        book = load_workbook(path)
        sheet = book["Group Review"]
        if mutation == "immutable":
            sheet.cell(2, headers["Concept"], "C999999")
        elif mutation == "missing":
            sheet.delete_rows(sheet.max_row)
        elif mutation == "duplicate":
            sheet.append([cell.value for cell in sheet[2]])
        elif mutation == "blank":
            sheet.cell(2, headers["Rationale"]).value = None
        elif mutation == "machine-rationale":
            sheet.cell(
                2,
                headers["Rationale"],
                packet.review_rows[0].machine_suggestion,
            )
        elif mutation == "stale":
            book["Bindings"]["B2"] = "0" * 64
        elif mutation == "formula":
            sheet.cell(2, headers["Rationale"], "=1+1")
        else:
            sheet.cell(2, headers["Decision"], APPROVE)
            sheet.cell(2, headers["Pair Decision"], CORRECT)
        book.save(path)
    with pytest.raises(ValueError, match=message):
        group_review.import_group_review_decisions(
            packet, path, tmp_path / "must-not-exist.json"
        )
    assert not (tmp_path / "must-not-exist.json").exists()


@pytest.mark.unit
def test_import_closes_decisions_and_dry_run_reports_deferred_without_writes(
    tmp_path: Path,
) -> None:
    packet, path, headers = _filled_workbook(tmp_path)
    book = load_workbook(path)
    sheet = book["Group Review"]
    sheet.cell(2, headers["Decision"], ABSTAIN)
    sheet.cell(3, headers["Decision"], CORRECT)
    book.save(path)
    registry = group_review.import_group_review_decisions(
        packet, path, tmp_path / "decisions.json"
    )
    assert all(
        decision.pair_decision == decision.decision
        for decision in registry.decisions
        if decision.review_type == "pair-only"
    )
    result = group_review.dry_run_group_review_decisions(packet, registry)
    assert result.writes_performed is False
    assert result.unresolved_count == 1
    assert result.deferred_count == 1
    assert result.affected_concepts
    assert result.affected_groups
    assert result.affected_traces


@pytest.mark.unit
def test_import_normalizes_excel_datetime_to_iso_date(tmp_path: Path) -> None:
    packet, path, headers = _filled_workbook(tmp_path)
    book = load_workbook(path)
    book["Group Review"].cell(2, headers["Date"], datetime(2099, 1, 2, 13, 45))
    book.save(path)

    registry = group_review.import_group_review_decisions(
        packet, path, tmp_path / "decisions.json"
    )

    assert registry.decisions[0].review_date == "2099-01-02"


@pytest.mark.unit
def test_macro_archive_check_is_case_insensitive(tmp_path: Path) -> None:
    packet, path, _headers = _filled_workbook(tmp_path)
    with ZipFile(path, "a", ZIP_DEFLATED) as archive:
        archive.writestr("XL/VBAPROJECT.BIN", b"TEST")

    with pytest.raises(ValueError, match="macro"):
        group_review.import_group_review_decisions(
            packet, path, tmp_path / "must-not-exist.json"
        )


@pytest.mark.unit
def test_group_concept_rejects_stored_common_pair_flags_that_drift() -> None:
    concept = _review_boundary().concepts[0]
    payload = concept.model_dump(mode="python")
    payload["common_pair_eligible"] = not concept.common_pair_eligible

    with pytest.raises(ValidationError, match="common-pair"):
        group_review.GroupReviewConcept.model_validate(payload)


@pytest.mark.unit
def test_rule_evidence_rejects_kind_specific_evidence_drift() -> None:
    row = next(item for item in _review_boundary().rule_evidence if item.r82_path)
    payload = row.model_dump(mode="python")
    payload["kind"] = "routing"
    payload["row_identity"] = group_review._identity(
        {key: value for key, value in payload.items() if key != "row_identity"}
    )

    with pytest.raises(ValidationError, match="R82 path"):
        group_review.RuleEvidenceRow.model_validate(payload)


@pytest.mark.unit
def test_highest_fanout_generation_loads_each_authoritative_source_once(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    reads: Counter[Path] = Counter()
    original = Path.read_bytes

    def counted(path: Path) -> bytes:
        reads[path.resolve()] += 1
        return original(path)

    monkeypatch.setattr(Path, "read_bytes", counted)
    output = tmp_path / "packet.json"
    workbook = tmp_path / "packet.xlsx"
    group_review.generate_group_review_boundary(
        evidence_path=_EVIDENCE,
        comparison_path=_COMPARISON,
        r101_report_path=_R101,
        output=output,
        workbook=workbook,
        correction_audit=tmp_path / "audit.xlsx",
        blank_validation=tmp_path / "blank-validation.json",
    )
    assert reads[_EVIDENCE.resolve()] == 1
    assert reads[_COMPARISON.resolve()] == 1
    assert reads[_R101.resolve()] == 1
    assert output.is_file()
    assert workbook.is_file()
    assert (tmp_path / "audit.xlsx").is_file()
    assert (tmp_path / "blank-validation.json").is_file()


@pytest.mark.unit
def test_readme_gives_exact_group_post_sme_import_and_dry_run_commands() -> None:
    readme = (_GOLDEN / "README.md").read_text(encoding="utf-8")
    assert (
        "pdm run adjudication import-group-review --packet "
        "tmp/m1-6-group-review-packet-rev2.json --reviewed-xlsx "
        "tmp/m1-6-group-review-workbook-rev2-reviewed.xlsx --output "
        "tmp/m1-6-group-review-decisions-rev2.json"
    ) in readme
    assert (
        "pdm run adjudication dry-run-group-review --packet "
        "tmp/m1-6-group-review-packet-rev2.json --registry "
        "tmp/m1-6-group-review-decisions-rev2.json --output "
        "tmp/m1-6-group-review-dry-run-rev2.json"
    ) in readme
