from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, cast

import pytest
from openpyxl import Workbook, load_workbook
from scripts.research import stage_bundle_pilot
from scripts.research.stage_bundle_pilot import (
    EVIDENCE_REGISTRY,
    STAGE_BUNDLE_CANDIDATES,
    ConstituentCorrection,
    _candidate_rows,
    _engine_pairs,
    _parser,
    _payload_identity,
    _rule,
    apply_constituent_corrections,
    build_provenance_ledger,
    build_stage_bundle_report,
    build_verification_manifest,
    import_review_decisions,
    load_constituent_corrections,
    validate_source_audit,
    write_review_workbook,
)

from ontolib.decomposition.proposal_registry import load_proposal_registry
from ontolib.decomposition.semantic_bundles import (
    BundleAxis,
    MemberRole,
    ProjectedConstituentEvidence,
    canonical_restriction_fact_id,
    validate_candidate_evidence,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

_SOURCE_IDENTITY = "f54dd2910a31245a30cea094dc72ce6a5c8d7b5a9c4e484007a35a1c343624c8"
_PROPOSAL_REGISTRY = load_proposal_registry(
    Path(__file__).with_name("golden") / "proposal-registry.json"
)


def _candidate_artifact() -> dict[str, object]:
    report = build_stage_bundle_report({})
    report["artifact_identity"] = _payload_identity(report)
    return report


def _engine_evidence_for_candidates() -> dict[
    str, tuple[ProjectedConstituentEvidence, ...]
]:
    by_code: defaultdict[
        str, dict[tuple[BundleAxis, str], ProjectedConstituentEvidence]
    ] = defaultdict(dict)
    for candidate in STAGE_BUNDLE_CANDIDATES:
        for member in candidate.members:
            by_code[candidate.subject_code][member.pair] = ProjectedConstituentEvidence(
                axis=member.axis,
                filler_code=member.filler_code,
                needs_review=False,
                relationship_group="stage",
                source_role="R88" if member.source_occurrences else None,
                axis_source="role" if member.source_occurrences else None,
                source_fact_ids=tuple(
                    occurrence.fact_id for occurrence in member.source_occurrences
                ),
            )
    return {code: tuple(items.values()) for code, items in by_code.items()}


@pytest.mark.unit
def test_registry_contains_typed_review_candidates_not_accepted_rules() -> None:
    counts = Counter(candidate.subject_code for candidate in STAGE_BUNDLE_CANDIDATES)

    assert len(STAGE_BUNDLE_CANDIDATES) == 15
    assert counts["C115057"] == 2
    assert counts["C27787"] == 2
    assert counts["C35756"] == 2
    for candidate in STAGE_BUNDLE_CANDIDATES:
        validate_candidate_evidence(candidate, EVIDENCE_REGISTRY)
        assert len(candidate.members) == 2

    artifact = _candidate_artifact()
    candidates = {
        item["candidate_id"]: item
        for item in cast(
            "list[dict[str, object]]", artifact["semantic_bundle_candidates"]
        )
    }
    ajcc_sources = cast(
        "list[str]",
        candidates["stage-c115057-ajcc-v6"]["evidence_source_ids"],
    )
    figo_sources = cast(
        "list[str]",
        candidates["stage-c162226-figo-2018"]["evidence_source_ids"],
    )
    assert "ajcc-official-staging-system" in ajcc_sources
    assert "figo-cervix-2018" in figo_sources


@pytest.mark.unit
def test_candidate_evidence_sources_are_registered_unique_and_nonempty() -> None:
    report = _candidate_artifact()
    sources = cast("list[dict[str, str]]", report["evidence_sources"])
    registered = [source["evidence_id"] for source in sources]
    referenced = {
        evidence_id
        for candidate in STAGE_BUNDLE_CANDIDATES
        for evidence_id in candidate.evidence_source_ids
    }

    assert len(registered) == len(set(registered))
    assert referenced <= set(registered)
    assert all(all(value.strip() for value in source.values()) for source in sources)

    candidate = STAGE_BUNDLE_CANDIDATES[0]
    with pytest.raises(ValueError, match="unknown candidate reference"):
        _rule(
            "unregistered-evidence-candidate",
            candidate.subject_code,
            "Unregistered evidence candidate",
            candidate.members[0],
            candidate.members[1],
            authority=candidate.classification.authority,
            version=candidate.classification.version,
            evidence_ids=("unregistered-source",),
        )


@pytest.mark.unit
@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("semantic-identity", "semantic identity"),
        ("member-axis", "stage-type role"),
        ("unknown-source", "evidence source|unknown candidate reference"),
        ("extra-field", "fields do not match"),
    ],
)
def test_candidate_artifact_rows_require_typed_semantic_contract(
    mutation: str,
    message: str,
) -> None:
    artifact = _candidate_artifact()
    candidates = cast("list[dict[str, object]]", artifact["semantic_bundle_candidates"])
    candidate = candidates[0]
    if mutation == "semantic-identity":
        candidate["semantic_identity"] = "0" * 64
    elif mutation == "member-axis":
        members = cast("list[dict[str, object]]", candidate["members"])
        members[0]["axis"] = "op:StageValue"
    elif mutation == "unknown-source":
        candidate["evidence_source_ids"] = ["unregistered-source"]
    else:
        candidate["unexpected"] = "value"
    artifact["artifact_identity"] = _payload_identity(
        {key: value for key, value in artifact.items() if key != "artifact_identity"}
    )

    with pytest.raises(ValueError, match=message):
        _candidate_rows(artifact)


@pytest.mark.unit
def test_valg_method_claim_does_not_fabricate_an_ncit_occurrence() -> None:
    valg = next(
        candidate
        for candidate in STAGE_BUNDLE_CANDIDATES
        if candidate.candidate_id == "stage-c35756-valg-extensive"
    )
    method = next(
        member for member in valg.members if member.role is MemberRole.STAGING_METHOD
    )

    assert method.filler_code == "C141685"
    assert method.source_occurrences == ()
    assert method.evidence_claim_ids == ("mcode-4.0.0-valg-method",)


@pytest.mark.unit
def test_source_occurrences_are_canonical_and_must_exist_in_audit() -> None:
    occurrences = {
        occurrence
        for candidate in STAGE_BUNDLE_CANDIDATES
        for member in candidate.members
        for occurrence in member.source_occurrences
    }
    facts = [
        {
            "root_code": item.root_code,
            "role_code": item.source_role,
            "filler_code": item.filler_code,
            "anchor_code": item.anchor_code,
            "depth": item.depth,
            "group_id": item.source_group_id,
        }
        for item in occurrences
    ]
    facts.append(
        {
            "root_code": "C198031",
            "role_code": "R88",
            "filler_code": "C198023",
            "anchor_code": "C198031",
            "depth": 0,
            "group_id": (
                "0d414f8ad31ecc05baa4617d99f8aa622c9c1a684f55f49120ffe79e78b594cf"
            ),
        }
    )
    audit = {"ncit_release": "26.07d", "facts": facts}

    validate_source_audit(audit)
    occurrence = next(iter(occurrences))
    assert occurrence.fact_id == canonical_restriction_fact_id(
        occurrence.anchor_code,
        occurrence.source_group_id,
        occurrence.source_role,
        occurrence.filler_code,
    )
    facts.pop()
    with pytest.raises(ValueError, match="missing semantic-bundle source fact"):
        validate_source_audit(audit)


@pytest.mark.unit
def test_report_separates_available_deferred_missing_and_observed_scores() -> None:
    evidence = _engine_evidence_for_candidates()
    incomplete_code = "C132677"
    deferred_code = "C181564"
    evidence[incomplete_code] = tuple(
        item for item in evidence[incomplete_code] if item.filler_code != "C132248"
    )
    evidence[deferred_code] = tuple(
        ProjectedConstituentEvidence(
            axis=item.axis,
            filler_code=item.filler_code,
            needs_review=item.filler_code == "C27966",
            relationship_group=item.relationship_group,
            source_role=item.source_role,
            axis_source=item.axis_source,
            source_fact_ids=item.source_fact_ids,
        )
        for item in evidence[deferred_code]
    )

    report = build_stage_bundle_report(evidence)
    availability = cast("dict[str, object]", report["engine_pair_availability"])

    assert report["status"] == "FINAL-REVIEW-PENDING"
    assert availability["candidate_counts"] == {
        "expected": 15,
        "available": 13,
        "deferred": 1,
        "incomplete": 1,
    }
    semantic_scores = cast("dict[str, dict[str, str]]", availability["semantic_scores"])
    assert semantic_scores["exact_bundle"]["status"] == "not-evaluable"
    assert semantic_scores["association"]["status"] == "not-evaluable"


@pytest.mark.unit
def test_engine_parser_preserves_review_and_source_metadata() -> None:
    raw = {
        "schema_version": 1,
        "ncit_version": "26.07d",
        "concepts": [
            {
                "code": "C1",
                "outcome": "decomposed",
                "constituents": [
                    {
                        "axis": "op:StageValue",
                        "filler": "C2",
                        "needs_review": True,
                        "relationship_group": "stage",
                        "source_role": "R88",
                        "axis_source": "role",
                        "source_fact_ids": ["a" * 64],
                    },
                    {
                        "axis": "op:Morphology",
                        "filler": "C3",
                        "needs_review": False,
                        "relationship_group": None,
                    },
                ],
            }
        ],
    }

    parsed = _engine_pairs(raw)["C1"]

    assert len(parsed) == 1
    assert parsed[0].needs_review is True
    assert parsed[0].relationship_group == "stage"
    assert parsed[0].source_role == "R88"
    assert parsed[0].axis_source == "role"
    assert parsed[0].source_fact_ids == ("a" * 64,)


@pytest.mark.unit
def test_provenance_ledger_dispositions_exactly_304_occurrences() -> None:
    facts = [
        {
            "root_code": f"C{index + 1000}",
            "role_code": "R101",
            "filler_code": f"C{index + 2000}",
            "anchor_code": f"C{index + 3000}",
            "depth": 0,
            "group_id": f"group-{index}",
        }
        for index in range(303)
    ]
    contracted_fact = {
        "root_code": "C9000",
        "role_code": "R105",
        "filler_code": "C9001",
        "anchor_code": "C9002",
        "depth": 1,
        "group_id": "group-contracted",
    }
    facts.append(contracted_fact)
    audit = {"fact_count": 304, "facts": facts}
    disposition = {
        "source_identity": _SOURCE_IDENTITY,
        "ontology_version": "26.07d",
        "rows": [
            {
                "root_code": "C9000",
                "role_code": "R105",
                "filler_code": "C9001",
                "disposition": "expected-current",
                "anchors": [["C9002", 1]],
            }
        ],
    }
    engine = {
        "concepts": [
            {"code": fact["root_code"], "outcome": "decomposed"} for fact in facts
        ]
    }

    ledger = build_provenance_ledger(audit, disposition, engine)

    assert ledger["occurrence_count"] == 304
    assert ledger["counts"] == {
        "constituent-workbook-review": 303,
        "contracted-role-disposition": 1,
    }
    rows = cast("list[dict[str, object]]", ledger["occurrences"])
    assert all(len(cast("str", row["fact_id"])) == 64 for row in rows)
    with pytest.raises(ValueError, match="fact count does not match"):
        build_provenance_ledger(audit | {"facts": facts[:-1]}, disposition, engine)


def _blank_workbook(path: Path) -> None:
    workbook = Workbook()
    cast("Worksheet", workbook.active).title = "START HERE"
    reviewer = workbook.create_sheet("Reviewer & Attestation")
    reviewer["B5"] = "Example Reviewer"
    reviewer["B6"] = "NCIt ontology curator"
    reviewer["B7"] = date(2026, 8, 4)
    reviewer["B8"] = "NCIt 26.07d"
    reviewer["B9"] = "ATTESTED"
    concepts = workbook.create_sheet("Concept Decisions")
    concept_headers = (
        "Order",
        "Concept Code",
        "Source Label",
        "Source Semantic Types",
        "Expected Semantic Types",
        "Engine Suggested Outcome",
        "SME Decision Status",
        "Expected Outcome",
        "Rationale / Required Follow-up",
        "Source Reviewed?",
        "Concept Complete?",
    )
    for column, header in enumerate(concept_headers, start=1):
        concepts.cell(4, column, header)
    codes = [f"C{index}" for index in range(17)] + ["C4791", "C35756", "C89995"]
    for order, code in enumerate(codes, start=1):
        values = (
            order,
            code,
            f"Reviewed {code}",
            "Neoplastic Process",
            "Neoplastic Process",
            "decomposed",
            "accepted",
            "decomposed",
            "Reviewed against the stated source.",
            "YES",
            "YES",
        )
        for column, value in enumerate(values, start=1):
            concepts.cell(order + 4, column, value)
    constituents = workbook.create_sheet("Constituent Decisions")
    constituent_headers = (
        "Concept Order",
        "Concept Code",
        "Source Label",
        "Row Type",
        "Engine Axis",
        "Engine Filler",
        "Engine Filler Label",
        "Engine Group",
        "Engine needs_review",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "Expected Proposal ID",
        "SME Notes",
        "Row Complete?",
    )
    for column, header in enumerate(constituent_headers, start=1):
        constituents.cell(4, column, header)
    for order, code in enumerate(codes, start=1):
        values = (
            order,
            code,
            f"Reviewed {code}",
            "ENGINE SUGGESTION",
            "op:StageValue",
            "C27970",
            "Stage III",
            None,
            "FALSE",
            "include",
            "op:StageValue",
            "C27970",
            None,
            "FALSE",
            "ncit-26.07d",
            None,
            "",
            "YES",
        )
        for column, value in enumerate(values, start=1):
            constituents.cell(order + 4, column, value)
    workbook.create_sheet("Validation Summary")
    workbook.create_sheet("Worked Examples")
    workbook.create_sheet("Prior SME Evidence")
    evidence = workbook.create_sheet("Source & Run Evidence")
    evidence_rows = (
        ("NCIt release", "26.07d"),
        ("Source identity", _SOURCE_IDENTITY),
        ("Sample identity", "b" * 64),
        ("Engine run", "neoplasm-run-1"),
        ("Run fingerprint identity", "c" * 64),
        ("Artifact SHA-256", "d" * 64),
        ("Engine evidence identity", "e" * 64),
        ("Corpus evidence identity", "f" * 64),
        ("Detector identity", "0" * 64),
        ("Proposal registry identity", _PROPOSAL_REGISTRY.registry_identity),
    )
    for row, (key, value) in enumerate(evidence_rows, start=5):
        evidence.cell(row, 1, key)
        evidence.cell(row, 2, value)
    workbook.save(path)


def _attest_review_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    workbook["Reviewer & Attestation"]["B9"] = "ATTESTED"
    sheet = workbook["Semantic Bundle Decisions"]
    sheet["B5"] = "ATTESTED"
    for row in range(9, sheet.max_row + 1):
        sheet.cell(row, 8, "ACCEPT" if row == 9 else "REJECT")
        sheet.cell(row, 9, "Reviewer assessed the exact proposed association.")
        sheet.cell(row, 10, "Domain Reviewer")
        sheet.cell(row, 11, "2026-08-04")
    workbook.save(path)


@pytest.mark.unit
def test_constituent_corrections_remove_and_add_exact_pairs() -> None:
    workbook = Workbook()
    sheet = workbook.create_sheet("Constituent Decisions")
    headers = (
        "Concept Order",
        "Concept Code",
        "Source Label",
        "Row Type",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "Expected Role Modality",
        "SME Notes",
        "Row Complete?",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    values = (
        1,
        "C1",
        "Source concept",
        "ADD IF MISSING",
        "include",
        "op:AssociatedRegion",
        "C2",
        None,
        "FALSE",
        "ncit-26.07d",
        "asserted",
        "Original decision.",
        "YES",
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(5, column, value)

    apply_constituent_corrections(
        workbook,
        (
            ConstituentCorrection(
                action="remove",
                concept_code="C1",
                axis="op:AssociatedRegion",
                filler_code="C2",
                rationale="The source-backed correction removes this broader pair.",
            ),
            ConstituentCorrection(
                action="add",
                concept_code="C1",
                axis="op:AssociatedRegion",
                filler_code="C3",
                rationale="The complete source audit supports this missing pair.",
                provenance_status="ncit-26.07d",
                needs_review=False,
            ),
        ),
    )

    assert sheet["E5"].value == "exclude"
    assert sheet["E6"].value == "include"
    assert sheet["F6"].value == "op:AssociatedRegion"
    assert sheet["G6"].value == "C3"
    assert sheet["I6"].value == "FALSE"
    assert sheet["J6"].value == "ncit-26.07d"
    assert sheet["M6"].value == "YES"


@pytest.mark.unit
def test_add_correction_requires_explicit_provenance_and_review_state() -> None:
    with pytest.raises(ValueError, match="add correction requires"):
        ConstituentCorrection(
            action="add",
            concept_code="C1",
            axis="op:AssociatedRegion",
            filler_code="C3",
            rationale="The reviewer added this pair.",
        )


@pytest.mark.unit
def test_correction_loader_requires_typed_add_provenance(tmp_path: Path) -> None:
    path = tmp_path / "corrections.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "corrections": [
                    {
                        "action": "add",
                        "concept_code": "C1",
                        "axis": "op:AssociatedRegion",
                        "filler_code": "C3",
                        "rationale": "The source audit supports this pair.",
                        "provenance_status": "ncit-26.07d",
                        "needs_review": False,
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    loaded = load_constituent_corrections(path)

    assert loaded[0].provenance_status == "ncit-26.07d"
    assert loaded[0].needs_review is False

    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["corrections"][0]["provenance_status"] = "invented"
    path.write_text(json.dumps(payload), encoding="utf-8")
    with pytest.raises(ValueError, match="provenance"):
        load_constituent_corrections(path)


@pytest.mark.unit
def test_ncit_add_correction_requires_bound_source_audit_support(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)

    with pytest.raises(ValueError, match="source audit"):
        write_review_workbook(
            base,
            _candidate_artifact(),
            review,
            (
                ConstituentCorrection(
                    action="add",
                    concept_code="C1",
                    axis="op:AssociatedRegion",
                    filler_code="C999999",
                    rationale="This pair has no source occurrence.",
                    provenance_status="ncit-26.07d",
                    needs_review=False,
                ),
            ),
            proposal_registry=_PROPOSAL_REGISTRY,
        )


@pytest.mark.unit
def test_constituent_corrections_reject_duplicate_headers() -> None:
    workbook = Workbook()
    sheet = workbook.create_sheet("Constituent Decisions")
    headers = (
        "Concept Order",
        "Concept Code",
        "Source Label",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "SME Notes",
        "Row Complete?",
        "SME Action",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)

    with pytest.raises(ValueError, match="duplicate header"):
        apply_constituent_corrections(workbook, ())


@pytest.mark.unit
def test_semantic_review_accepts_excel_date_cells(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    workbook = load_workbook(review)
    assert workbook["Semantic Bundle Decisions"]["K9"].number_format == "@"
    workbook.save(review)
    _attest_review_workbook(review)
    workbook = load_workbook(review)
    workbook["Semantic Bundle Decisions"]["K9"] = datetime(2026, 8, 4)
    workbook.save(review)

    canonical = import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)

    decisions = cast("list[dict[str, str]]", canonical["decisions"])
    assert decisions[0]["reviewed_at"] == "2026-08-04"


@pytest.mark.unit
def test_review_workbook_binds_augmented_pair_to_proposal_registry(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    workbook = load_workbook(base)
    sheet = workbook["Constituent Decisions"]
    headers = {
        sheet.cell(4, column).value: column for column in range(1, sheet.max_column + 1)
    }
    sheet.cell(5, headers["Expected Axis"], "op:AssociatedPriorDisease")
    sheet.cell(5, headers["Expected Filler"], "C27262")
    sheet.cell(5, headers["Expected Provenance Status"], "locally-approved")
    workbook.save(base)
    artifact = _candidate_artifact()

    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)

    bound = load_workbook(review)
    bound_sheet = bound["Constituent Decisions"]
    bound_headers = {
        bound_sheet.cell(4, column).value: column
        for column in range(1, bound_sheet.max_column + 1)
    }
    assert bound_sheet.cell(5, bound_headers["Expected Proposal ID"]).value == (
        "RELPROP-8637e8500dff"
    )
    _attest_review_workbook(review)
    canonical = import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)
    assert canonical["proposal_registry_identity"] == (
        _PROPOSAL_REGISTRY.registry_identity
    )

    tampered = load_workbook(review)
    tampered_sheet = tampered["Constituent Decisions"]
    tampered_sheet.cell(5, bound_headers["Expected Proposal ID"], "RELPROP-unknown")
    tampered.save(review)
    with pytest.raises(ValueError, match="unknown proposal"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)


@pytest.mark.unit
def test_relation_proposal_workbook_pair_requires_an_ncit_filler(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    workbook = load_workbook(base)
    sheet = workbook["Constituent Decisions"]
    headers = {
        sheet.cell(4, column).value: column for column in range(1, sheet.max_column + 1)
    }
    sheet.cell(5, headers["Expected Axis"], "op:AssociatedPriorDisease")
    sheet.cell(5, headers["Expected Filler"], "free text")
    sheet.cell(5, headers["Expected Provenance Status"], "locally-approved")
    workbook.save(base)

    with pytest.raises(ValueError, match="NCIt code"):
        write_review_workbook(
            base,
            _candidate_artifact(),
            review,
            proposal_registry=_PROPOSAL_REGISTRY,
        )


@pytest.mark.unit
def test_canonical_rules_can_only_be_imported_from_attested_decisions(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)

    workbook = load_workbook(review)
    workbook["Reviewer & Attestation"]["B9"] = "ATTESTED"
    workbook.save(review)
    with pytest.raises(ValueError, match="metadata"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)

    workbook = load_workbook(review)
    workbook["Reviewer & Attestation"]["B9"] = "PENDING"
    sheet = workbook["Semantic Bundle Decisions"]
    sheet["B5"] = "ATTESTED"
    for row in range(9, sheet.max_row + 1):
        sheet.cell(row, 8, "ACCEPT" if row == 9 else "REJECT")
        sheet.cell(row, 9, "Reviewer assessed the exact proposed association.")
        sheet.cell(row, 10, "Domain Reviewer")
        sheet.cell(row, 11, "2026-08-04")
    workbook.save(review)

    with pytest.raises(ValueError, match="attestation is pending"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)

    workbook = load_workbook(review)
    workbook["Reviewer & Attestation"]["B9"] = "ATTESTED"
    workbook.save(review)
    canonical = import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)

    assert canonical["status"] == "ATTESTED"
    assert len(cast("str", canonical["adjudication_artifact_identity"])) == 64
    rules = cast("list[dict[str, object]]", canonical["semantic_bundle_rules"])
    assert [item["candidate_id"] for item in rules] == [
        STAGE_BUNDLE_CANDIDATES[0].candidate_id
    ]
    workbook = load_workbook(review)
    workbook["Semantic Bundle Decisions"]["B9"] = "0" * 64
    workbook.save(review)
    with pytest.raises(ValueError, match="candidate fields changed"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)


@pytest.mark.unit
def test_canonical_import_requires_complete_constituent_review(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    _attest_review_workbook(review)
    workbook = load_workbook(review)
    workbook["Constituent Decisions"]["J5"] = "PENDING"
    workbook.save(review)

    with pytest.raises(ValueError, match="pending constituent action"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)


@pytest.mark.unit
@pytest.mark.parametrize("hidden_target", ["sheet", "row", "column"])
def test_canonical_import_rejects_hidden_semantic_review_content(
    tmp_path: Path,
    hidden_target: str,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    _attest_review_workbook(review)
    workbook = load_workbook(review)
    sheet = workbook["Semantic Bundle Decisions"]
    if hidden_target == "sheet":
        sheet.sheet_state = "hidden"
    elif hidden_target == "row":
        sheet.row_dimensions[9].hidden = True
    else:
        sheet.column_dimensions["H"].hidden = True
    workbook.save(review)

    with pytest.raises(ValueError, match=r"visible|hidden"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)


@pytest.mark.unit
def test_canonical_import_validates_semantic_review_metadata(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    _attest_review_workbook(review)
    workbook = load_workbook(review)
    workbook["Semantic Bundle Decisions"]["B3"] = "stale-release"
    workbook.save(review)

    with pytest.raises(ValueError, match="metadata"):
        import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)


@pytest.mark.unit
def test_canonical_import_hashes_the_parsed_workbook_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    _attest_review_workbook(review)
    snapshot = review.read_bytes()
    real_load_workbook = stage_bundle_pilot.load_workbook

    def replace_after_load(source: object, **kwargs: object) -> Workbook:
        workbook = real_load_workbook(source, **kwargs)
        review.write_bytes(b"changed after the review snapshot was opened")
        return workbook

    monkeypatch.setattr(stage_bundle_pilot, "load_workbook", replace_after_load)

    canonical = import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)

    binding = cast("dict[str, str]", canonical["review_workbook"])
    assert binding["sha256"] == hashlib.sha256(snapshot).hexdigest()


@pytest.mark.unit
def test_external_manifest_hash_binds_pending_candidate_and_workbook(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    candidate_path = tmp_path / "candidate.json"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    candidate_path.write_text(
        json.dumps(artifact, sort_keys=True),
        encoding="utf-8",
    )
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)

    manifest = build_verification_manifest(
        candidate_path, review, proposal_registry=_PROPOSAL_REGISTRY
    )

    assert manifest["status"] == "FINAL-REVIEW-PENDING"
    files = cast("dict[str, dict[str, str]]", manifest["files"])
    assert len(files["candidate_artifact"]["sha256"]) == 64
    assert len(files["review_workbook"]["sha256"]) == 64

    workbook = load_workbook(review)
    workbook["Semantic Bundle Decisions"]["B2"] = "0" * 64
    workbook.save(review)
    with pytest.raises(ValueError, match="candidate artifact identity"):
        build_verification_manifest(
            candidate_path, review, proposal_registry=_PROPOSAL_REGISTRY
        )


@pytest.mark.unit
def test_attested_manifest_validates_canonical_payload_and_review_binding(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    review = tmp_path / "review.xlsx"
    other_review = tmp_path / "other-review.xlsx"
    candidate_path = tmp_path / "candidate.json"
    canonical_path = tmp_path / "canonical.json"
    _blank_workbook(base)
    artifact = _candidate_artifact()
    candidate_path.write_text(json.dumps(artifact, sort_keys=True), encoding="utf-8")
    write_review_workbook(base, artifact, review, proposal_registry=_PROPOSAL_REGISTRY)
    _attest_review_workbook(review)

    canonical_path.write_text(
        json.dumps({"status": "ATTESTED"}),
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="canonical semantic rules"):
        build_verification_manifest(
            candidate_path,
            review,
            canonical_path,
            proposal_registry=_PROPOSAL_REGISTRY,
        )

    canonical = import_review_decisions(review, artifact, _PROPOSAL_REGISTRY)
    canonical_path.write_text(json.dumps(canonical, sort_keys=True), encoding="utf-8")
    other_review.write_bytes(review.read_bytes())
    workbook = load_workbook(other_review)
    workbook["Semantic Bundle Decisions"]["B5"] = "FINAL-REVIEW-PENDING"
    workbook.save(other_review)

    with pytest.raises(ValueError, match="review workbook binding"):
        build_verification_manifest(
            candidate_path,
            other_review,
            canonical_path,
            proposal_registry=_PROPOSAL_REGISTRY,
        )

    manifest = build_verification_manifest(
        candidate_path,
        review,
        canonical_path,
        proposal_registry=_PROPOSAL_REGISTRY,
    )
    assert manifest["status"] == "ATTESTED"


@pytest.mark.unit
def test_cli_separates_packet_preparation_from_review_finalization() -> None:
    parser = _parser()
    prepared = parser.parse_args(
        [
            "prepare",
            "--workbook",
            "base.xlsx",
            "--source-audit",
            "audit.json",
            "--engine-evidence",
            "engine.json",
            "--contracted-disposition",
            "disposition.json",
            "--proposal-registry",
            "registry.json",
            "--output",
            "candidate.json",
            "--review-workbook-output",
            "review.xlsx",
        ]
    )
    finalized = parser.parse_args(
        [
            "finalize",
            "--candidate",
            "candidate.json",
            "--review-workbook",
            "reviewed.xlsx",
            "--proposal-registry",
            "registry.json",
            "--canonical-rules-output",
            "canonical.json",
            "--manifest-output",
            "manifest.json",
        ]
    )

    assert prepared.command == "prepare"
    assert finalized.command == "finalize"
    assert finalized.review_workbook.name == "reviewed.xlsx"
    with pytest.raises(SystemExit):
        parser.parse_args(
            [
                "finalize",
                "--workbook",
                "base.xlsx",
                "--candidate",
                "candidate.json",
                "--review-workbook",
                "reviewed.xlsx",
                "--proposal-registry",
                "registry.json",
                "--canonical-rules-output",
                "canonical.json",
                "--manifest-output",
                "manifest.json",
            ]
        )
