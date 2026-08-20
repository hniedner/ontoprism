from __future__ import annotations

import math
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import load_workbook
from scripts.decompose import _source_snapshot

from ontolib.decomposition.collapse_policy import (
    NO_COLLAPSE_VETO_POLICY,
    load_packaged_collapse_veto_policy,
)
from ontolib.decomposition.fanout_baseline import _CountingClient
from ontolib.decomposition.r101_conservation import load_r101_conservation_report
from ontolib.decomposition.r101_review import (
    QLeverReviewLabels,
    build_r101_review_packet,
    write_r101_review_workbook,
)
from ontolib.decomposition.run import _decompose_one, _qualify_collapse_policy
from ontolib.terminologies.ncit.client import ncit_sparql_client


class _RecordedLabels:
    def __init__(self, values: dict[str, str]) -> None:
        self.values = values
        self.requested_codes: tuple[str, ...] = ()

    async def labels_for_review(
        self, codes: tuple[str, ...]
    ) -> dict[str, tuple[str, ...]]:
        self.requested_codes = codes
        return {code: (self.values[code],) for code in codes}


@pytest.mark.integration
@pytest.mark.full_store
async def test_r101_review_labels_match_real_qlever_in_bounded_batches(
    tmp_path: Path,
) -> None:
    report = load_r101_conservation_report(
        Path("ontolib/tests/decomposition/golden/neoplasm-r101-v4-conservation.json.gz")
    )
    manifest_path = Path("data/qlever-ncit/.ontoprism-ncit-candidate.json")
    source = await _source_snapshot(manifest_path, "http://localhost:7888")
    assert source.source_identity == report.source_identity
    assert source.ontology_version == report.source_release_id
    async with ncit_sparql_client("http://localhost:7888") as client:
        labels = QLeverReviewLabels(client)
        packet = await build_r101_review_packet(report, manifest_path, labels)

    assert labels.query_count == math.ceil(len(labels.requested_codes) / 500)
    assert len(labels.requested_codes) > 1900
    assert len(packet.patterns) == 162
    assert len(packet.disease_propositions) == 2800
    assert len(packet.occurrences) == 3291
    assert max(row.occurrence_count for row in packet.patterns) == 245
    assert max(row.occurrence_count for row in packet.disease_propositions) == 3
    assert {row.disease_code for row in packet.disease_propositions} <= set(
        labels.requested_codes
    )
    by_code = (
        {row.broader_code: row.broader_label for row in packet.patterns}
        | {row.retained_code: row.retained_label for row in packet.patterns}
        | {row.disease_code: row.disease_label for row in packet.disease_propositions}
    )
    assert by_code["C12727"] == "Heart"
    assert by_code["C12869"] == "Left Atrium"

    delivered_workbook = tmp_path / "r101-review-workbook-v3.xlsx"
    write_r101_review_workbook(delivered_workbook, packet)
    assert delivered_workbook.stat().st_size > 0
    workbook = load_workbook(delivered_workbook, read_only=True)
    pattern_rows = workbook["Pattern Review"].iter_rows(min_row=2, values_only=True)
    disease_rows = workbook["Disease Propositions"].iter_rows(
        min_row=2, values_only=True
    )
    assert sum(row[14] is not None for row in pattern_rows) == 0
    assert sum(row[9] == "No" and row[10] is None for row in disease_rows) == 2800

    for pattern in packet.patterns:
        for path in pattern.paths:
            by_code.update(zip(path.code_path, path.labels, strict=True))
    replay_labels = _RecordedLabels(by_code)
    replay = await build_r101_review_packet(report, manifest_path, replay_labels)
    assert replay == packet
    assert replay_labels.requested_codes == labels.requested_codes


@pytest.mark.integration
@pytest.mark.full_store
async def test_c5292_policy_matches_source_and_retains_review_sites() -> None:
    report = load_r101_conservation_report(
        Path("ontolib/tests/decomposition/golden/neoplasm-r101-v4-conservation.json.gz")
    )
    policy = load_packaged_collapse_veto_policy()

    async def no_label_match(_surface: str) -> str | None:
        return None

    async with ncit_sparql_client("http://localhost:7888") as client:
        counted = _CountingClient(client)
        await _qualify_collapse_policy(
            policy,
            cast("Any", counted),
            source_identity=report.source_identity,
            walker_max_depth=7,
        )
        qualification_queries = counted.logical_select_count
        result = await _decompose_one(
            "C5292",
            cast("Any", counted),
            label=None,
            label_lookup=no_label_match,
            source_identity=report.source_identity,
            collapse_policy=policy,
            walker_max_depth=7,
        )
        policy_item_queries = counted.logical_select_count - qualification_queries
    async with ncit_sparql_client("http://localhost:7888") as client:
        baseline_client = _CountingClient(client)
        baseline = await _decompose_one(
            "C5292",
            cast("Any", baseline_client),
            label=None,
            label_lookup=no_label_match,
            source_identity=report.source_identity,
            collapse_policy=NO_COLLAPSE_VETO_POLICY,
            walker_max_depth=7,
        )

    assert qualification_queries == 17
    assert policy_item_queries == baseline_client.logical_select_count

    assert result.decomposition is not None
    assert result.decomposition.complete_definition is not None
    source_r101 = {
        row.filler_code
        for row in result.decomposition.complete_definition.occurrences
        if row.role_code == "R101"
    }
    assert source_r101 == {
        "C12438",
        "C12351",
        "C32292",
        "C32639",
        "C12512",
        "C12789",
        "C12439",
    }
    assert "C12349" not in source_r101
    primary_sites = {
        row.filler_code: row
        for row in result.decomposition.constituents
        if row.axis == "op:PrimarySite"
    }
    assert {"C12351", "C12439", "C12512", "C32639"} <= set(primary_sites)
    assert all(primary_sites[code].needs_review for code in primary_sites)
    assert {primary_sites[code].group for code in primary_sites} == {"op:PrimarySite"}
    prior = {
        link.filler_code
        for row in report.occurrences
        if row.concept_code == "C5292" and row.new_links
        for link in row.new_links
        if link.axis == "op:PrimarySite"
    }
    assert {"C12351", "C12439", "C12512"}.isdisjoint(prior)
    assert baseline.decomposition is not None
    baseline_pairs = {
        (row.axis, row.filler_code) for row in baseline.decomposition.constituents
    }
    policy_pairs = {
        (row.axis, row.filler_code) for row in result.decomposition.constituents
    }
    authorized_delta = {
        ("op:PrimarySite", "C12351"),
        ("op:PrimarySite", "C12439"),
        ("op:PrimarySite", "C12512"),
    }
    assert policy_pairs == baseline_pairs | authorized_delta
    assert {
        ("op:AssociatedRegion", "C32292"),
        ("op:AssociatedSite", "C12351"),
        ("op:Morphology", "C4959"),
        ("op:NormalTissueOrigin", "C12349"),
        ("op:PrimarySite", "C12789"),
        ("op:PrimarySite", "C32639"),
    } <= policy_pairs
    baseline_by_pair = {
        (row.axis, row.filler_code): row for row in baseline.decomposition.constituents
    }
    policy_by_pair = {
        (row.axis, row.filler_code): row for row in result.decomposition.constituents
    }
    assert all(
        policy_by_pair[pair] == row
        for pair, row in baseline_by_pair.items()
        if pair[0] != "op:PrimarySite"
    )
