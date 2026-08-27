from __future__ import annotations

import hashlib
import importlib.util
import json
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

import pytest
from openpyxl import load_workbook
from scripts.adjudication import _parser
from scripts.adjudication import main as adjudication_main

from ontolib.decomposition import r103_review, r103_review_promotion
from ontolib.decomposition.proposal_registry import load_proposal_registry
from ontolib.decomposition.r103_review import (
    R103Decision,
    R103DecisionRegistry,
    R103ReviewDryRun,
    R103ReviewPacket,
    R103ReviewValidationError,
    dry_run_r103_review,
    import_r103_review_decisions,
    write_r103_review_dry_run,
    write_r103_review_packet,
    write_r103_review_workbook,
)

if TYPE_CHECKING:
    from types import ModuleType

CURRENT_DECISIONS = (
    (
        "source-supported",
        "C12950 is supported by C2860’s stated derivation from adrenal embryonic "  # noqa: RUF001
        "rest cells. It is anatomically broad but is the most specific currently "
        "available NCIt tissue-origin filler; no range-compatible, more specific "
        "NCIt replacement has been identified.",
    ),
    (
        "review-required",
        "Fetal-tissue resemblance describes morphology, not necessarily normal "
        "tissue of origin, so the R103 assertion should not be projected without "
        "stronger evidence.",
    ),
    (
        "source-supported",
        "C3716 explicitly states origin in neuroectoderm, and C34228 defines that "
        "embryologic tissue; the R103 assertion directly represents the stated "
        "normal tissue of origin while R104 separately preserves cell-origin context.",
    ),
)


def _review_fixture_module() -> ModuleType:
    path = Path(__file__).with_name("test_r103_review.py")
    spec = importlib.util.spec_from_file_location("r103_review_fixture_module", path)
    if spec is None or spec.loader is None:
        raise AssertionError("R103 review fixture module cannot be loaded")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _current_review_inputs(tmp_path: Path):
    review_fixtures = cast("Any", _review_fixture_module())
    source_boundary_factory = review_fixtures.source_boundary.__wrapped__
    packet_factory = review_fixtures.packet.__wrapped__
    fixture_boundary = source_boundary_factory(tmp_path)
    owl_path, manifest_path, tracked_proposals, oracle_path = fixture_boundary
    proposal_registry_path = tmp_path / "proposal-registry.json"
    proposal_registry_path.write_bytes(tracked_proposals.read_bytes())
    source_boundary = (
        owl_path,
        manifest_path,
        proposal_registry_path,
        oracle_path,
    )
    packet = packet_factory(source_boundary)
    _owl, _manifest, proposal_registry_path, oracle_path = source_boundary
    packet_path = tmp_path / "packet.json"
    workbook_path = tmp_path / "reviewed.xlsx"
    registry_path = tmp_path / "registry.json"
    dry_run_path = tmp_path / "dry-run.json"
    write_r103_review_packet(packet_path, packet)
    write_r103_review_workbook(workbook_path, packet)
    workbook = load_workbook(workbook_path)
    sheet = workbook["R103 Review"]
    columns = {cell.value: cast("int", cell.column) for cell in sheet[1]}
    for row_number, (outcome, rationale) in enumerate(CURRENT_DECISIONS, start=2):
        sheet.cell(row_number, columns["Outcome"], outcome)
        sheet.cell(row_number, columns["Rationale"], rationale)
        sheet.cell(row_number, columns["Reviewer"], "R. Hannes Niedner, M.D.")
        sheet.cell(row_number, columns["Date"], "2026-08-26")
    workbook.save(workbook_path)
    registry = import_r103_review_decisions(packet, workbook_path, registry_path)
    dry_run = dry_run_r103_review(
        packet,
        registry,
        oracle_path=oracle_path,
        proposal_registry_path=proposal_registry_path,
    )
    write_r103_review_dry_run(dry_run_path, dry_run)
    return (
        packet,
        registry,
        dry_run,
        packet_path,
        registry_path,
        dry_run_path,
        oracle_path,
        proposal_registry_path,
    )


def _identity(value: object) -> str:
    return hashlib.sha256(
        json.dumps(
            value, sort_keys=True, separators=(",", ":"), ensure_ascii=True
        ).encode("ascii")
    ).hexdigest()


def _promotion_functions():
    operation = cast(
        "Any", getattr(r103_review_promotion, "promote_r103_review_state", None)
    )
    loader = cast(
        "Any",
        getattr(r103_review_promotion, "load_r103_promoted_review_state", None),
    )
    assert callable(operation), "R103 review-state promotion operation is missing"
    assert callable(loader), "R103 promoted review-state loader is missing"
    return operation, loader


def _promote_current(tmp_path: Path):
    operation, loader = _promotion_functions()
    inputs = _current_review_inputs(tmp_path)
    output_path = tmp_path / "promoted.json"
    promoted = cast(
        "Any",
        operation(
            packet_path=inputs[3],
            registry_path=inputs[4],
            dry_run_path=inputs[5],
            oracle_path=inputs[6],
            proposal_registry_path=inputs[7],
            output_path=output_path,
        ),
    )
    return promoted, output_path, inputs, loader


@pytest.mark.unit
def test_promotes_complete_current_review_state_without_mutating_inputs(
    tmp_path: Path,
) -> None:
    operation, loader = _promotion_functions()
    (
        packet,
        registry,
        dry_run,
        packet_path,
        registry_path,
        dry_run_path,
        oracle_path,
        proposal_registry_path,
    ) = _current_review_inputs(tmp_path)
    inputs = {
        path: path.read_bytes()
        for path in (
            packet_path,
            registry_path,
            dry_run_path,
            oracle_path,
            proposal_registry_path,
        )
    }
    output_path = tmp_path / "promoted.json"

    promoted = cast(
        "Any",
        operation(
            packet_path=packet_path,
            registry_path=registry_path,
            dry_run_path=dry_run_path,
            oracle_path=oracle_path,
            proposal_registry_path=proposal_registry_path,
            output_path=output_path,
        ),
    )

    assert promoted.schema_version == 1
    assert promoted.packet == packet
    assert promoted.registry == registry
    assert promoted.dry_run == dry_run
    assert (
        promoted.oracle_file_sha256
        == hashlib.sha256(oracle_path.read_bytes()).hexdigest()
    )
    assert (
        promoted.proposal_registry_file_sha256
        == hashlib.sha256(proposal_registry_path.read_bytes()).hexdigest()
    )
    identity_payload = promoted.model_dump(mode="json", exclude={"artifact_identity"})
    expected_identity = hashlib.sha256(
        json.dumps(
            identity_payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True
        ).encode("ascii")
    ).hexdigest()
    assert promoted.artifact_identity == expected_identity
    assert {path: path.read_bytes() for path in inputs} == inputs
    for path in inputs:
        path.unlink()
    assert loader(output_path) == promoted


@pytest.mark.unit
@pytest.mark.parametrize("mutation", ["identity", "unknown", "duplicate", "embedded"])
def test_promoted_loader_is_strict_and_revalidates_nested_state(
    tmp_path: Path, mutation: str
) -> None:
    _promoted, output_path, _inputs, loader = _promote_current(tmp_path)
    payload = json.loads(output_path.read_text(encoding="ascii"))
    if mutation == "identity":
        payload["artifact_identity"] = "0" * 64
    elif mutation == "unknown":
        payload["unknown"] = True
    elif mutation == "duplicate":
        rendered = output_path.read_text(encoding="ascii").replace(
            '  "schema_version": 1,',
            '  "schema_version": 1,\n  "schema_version": 1,',
            1,
        )
        output_path.write_text(rendered, encoding="ascii")
    else:
        payload["dry_run"]["readiness"] = "ready-for-separate-application"
        payload["artifact_identity"] = _identity(
            {key: value for key, value in payload.items() if key != "artifact_identity"}
        )
    if mutation != "duplicate":
        output_path.write_text(json.dumps(payload), encoding="ascii")

    with pytest.raises(R103ReviewValidationError):
        loader(output_path)


@pytest.mark.unit
def test_loader_refuses_meaningful_registry_outcome_flip_with_recomputed_identities(
    tmp_path: Path,
) -> None:
    _promoted, output_path, _inputs, loader = _promote_current(tmp_path)
    payload = json.loads(output_path.read_text(encoding="ascii"))
    decision = payload["registry"]["decisions"][1]
    decision["outcome"] = "source-supported"
    decision["decision_identity"] = _identity(
        {key: value for key, value in decision.items() if key != "decision_identity"}
    )
    registry = payload["registry"]
    registry["registry_identity"] = _identity(
        {key: value for key, value in registry.items() if key != "registry_identity"}
    )
    payload["artifact_identity"] = _identity(
        {key: value for key, value in payload.items() if key != "artifact_identity"}
    )
    output_path.write_text(json.dumps(payload), encoding="ascii")

    with pytest.raises(R103ReviewValidationError, match="human review values"):
        loader(output_path)


@pytest.mark.unit
@pytest.mark.parametrize("binding", ["row", "workbook", "source"])
def test_loader_refuses_packet_decision_workbook_and_source_join_drift(
    tmp_path: Path, binding: str
) -> None:
    _promoted, output_path, _inputs, loader = _promote_current(tmp_path)
    payload = json.loads(output_path.read_text(encoding="ascii"))
    registry = payload["registry"]
    if binding == "row":
        registry["decisions"][0]["row_identity"] = registry["decisions"][1][
            "row_identity"
        ]
    elif binding == "workbook":
        registry["workbook_identity"] = "0" * 64
        for decision in registry["decisions"]:
            decision["workbook_identity"] = "0" * 64
    else:
        registry["source_identity"] = "0" * 64
        for decision in registry["decisions"]:
            decision["source_identity"] = "0" * 64
    for decision in registry["decisions"]:
        decision["decision_identity"] = _identity(
            {
                key: value
                for key, value in decision.items()
                if key != "decision_identity"
            }
        )
    registry["registry_identity"] = _identity(
        {key: value for key, value in registry.items() if key != "registry_identity"}
    )
    payload["artifact_identity"] = _identity(
        {key: value for key, value in payload.items() if key != "artifact_identity"}
    )
    output_path.write_text(json.dumps(payload), encoding="ascii")

    with pytest.raises(R103ReviewValidationError):
        loader(output_path)


@pytest.mark.unit
@pytest.mark.parametrize("mutation", ["dry-run", "oracle"])
def test_promotion_recomputes_dry_run_and_refuses_stale_production_inputs(
    tmp_path: Path, mutation: str
) -> None:
    operation, _loader = _promotion_functions()
    inputs = _current_review_inputs(tmp_path)
    if mutation == "dry-run":
        payload = json.loads(inputs[5].read_text(encoding="ascii"))
        payload["unresolved"] = 0
        inputs[5].write_text(json.dumps(payload), encoding="ascii")
    else:
        inputs[6].write_bytes(inputs[6].read_bytes() + b"oracle drift")
    output = tmp_path / "must-not-exist.json"

    with pytest.raises(R103ReviewValidationError, match="dry-run"):
        operation(
            packet_path=inputs[3],
            registry_path=inputs[4],
            dry_run_path=inputs[5],
            oracle_path=inputs[6],
            proposal_registry_path=inputs[7],
            output_path=output,
        )
    assert not output.exists()


@pytest.mark.unit
def test_promotion_binds_proposal_semantics_and_bytes_but_oracle_bytes_only(
    tmp_path: Path,
) -> None:
    operation, _loader = _promotion_functions()
    inputs = _current_review_inputs(tmp_path)
    proposal = load_proposal_registry(inputs[7])
    proposal_payload = json.loads(inputs[7].read_text(encoding="utf-8"))
    inputs[7].write_text(
        json.dumps(proposal_payload, indent=4) + "\n", encoding="utf-8"
    )
    inputs[6].write_bytes(b"not a cohort document")
    refreshed = dry_run_r103_review(
        inputs[0],
        inputs[1],
        oracle_path=inputs[6],
        proposal_registry_path=inputs[7],
    )
    write_r103_review_dry_run(inputs[5], refreshed)

    promoted = cast(
        "Any",
        operation(
            packet_path=inputs[3],
            registry_path=inputs[4],
            dry_run_path=inputs[5],
            oracle_path=inputs[6],
            proposal_registry_path=inputs[7],
            output_path=tmp_path / "promoted.json",
        ),
    )

    assert promoted.proposal_registry_identity == proposal.registry_identity
    assert (
        promoted.proposal_registry_file_sha256
        == hashlib.sha256(inputs[7].read_bytes()).hexdigest()
    )
    assert (
        promoted.oracle_file_sha256
        == hashlib.sha256(b"not a cohort document").hexdigest()
    )

    changed_proposal = type(proposal).model_validate(
        {
            **proposal.model_dump(mode="python", exclude={"registry_identity"}),
            "source_identity": "0" * 64,
            "registry_identity": "",
        }
    )
    inputs[7].write_text(
        json.dumps(changed_proposal.model_dump(mode="json")), encoding="utf-8"
    )
    with pytest.raises(R103ReviewValidationError, match="semantic binding"):
        operation(
            packet_path=inputs[3],
            registry_path=inputs[4],
            dry_run_path=inputs[5],
            oracle_path=inputs[6],
            proposal_registry_path=inputs[7],
            output_path=tmp_path / "must-not-exist.json",
        )


@pytest.mark.unit
def test_promotion_is_idempotent_refuses_conflicts_and_cleans_atomic_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    operation, _loader = _promotion_functions()
    inputs = _current_review_inputs(tmp_path)
    output = tmp_path / "promoted.json"
    kwargs = {
        "packet_path": inputs[3],
        "registry_path": inputs[4],
        "dry_run_path": inputs[5],
        "oracle_path": inputs[6],
        "proposal_registry_path": inputs[7],
        "output_path": output,
    }
    operation(**kwargs)
    original = output.read_bytes()
    operation(**kwargs)
    assert output.read_bytes() == original
    output.write_bytes(b"conflict\n")
    with pytest.raises(R103ReviewValidationError, match="conflict"):
        operation(**kwargs)
    assert output.read_bytes() == b"conflict\n"

    output.unlink()
    monkeypatch.setattr(
        r103_review.os,
        "replace",
        lambda *_args: (_ for _ in ()).throw(OSError("boom")),
    )
    with pytest.raises(OSError, match="boom"):
        operation(**kwargs)
    assert not output.exists()
    assert not tuple(tmp_path.glob(f".{output.name}.*"))


@pytest.mark.unit
def test_cli_parses_required_promotion_paths_and_dispatches_synchronously(
    tmp_path: Path,
) -> None:
    _promoted, expected, inputs, loader = _promote_current(tmp_path)
    expected.unlink()
    arguments = [
        "promote-r103-review-state",
        "--packet",
        str(inputs[3]),
        "--registry",
        str(inputs[4]),
        "--dry-run",
        str(inputs[5]),
        "--oracle",
        str(inputs[6]),
        "--proposal-registry",
        str(inputs[7]),
        "--output",
        str(expected),
    ]
    parsed = _parser().parse_args(arguments)
    assert parsed.command == "promote-r103-review-state"
    adjudication_main(arguments)
    assert cast("Any", loader(expected)).artifact_identity


@pytest.mark.unit
def test_promotion_does_not_change_upstream_model_shapes() -> None:
    assert set(R103ReviewPacket.model_fields) == {
        "schema_version",
        "source_release",
        "source_identity",
        "source_artifact_sha256",
        "source_artifact_size",
        "candidate_manifest_identity",
        "proposal_registry_identity",
        "query_contract_identity",
        "tool_identity",
        "source_pass_count",
        "inventory_scope",
        "rows",
        "method_reference",
        "packet_identity",
    }
    assert set(R103DecisionRegistry.model_fields) == {
        "schema_version",
        "packet_identity",
        "workbook_identity",
        "source_identity",
        "source_release",
        "decisions",
        "proposal_preview",
        "exclusion_preview",
        "registry_identity",
    }
    assert set(R103ReviewDryRun.model_fields) == {
        "writes_performed",
        "outcome_counts",
        "proposal_previews",
        "exclusion_previews",
        "unresolved",
        "readiness",
        "oracle_identity_before",
        "oracle_identity_after",
        "proposal_registry_identity_before",
        "proposal_registry_identity_after",
    }
    assert "workbook_identity" in R103Decision.model_fields


@pytest.mark.unit
def test_tracked_promotion_loads_with_exact_review_and_noncohort_oracle_boundary() -> (
    None
):
    golden = Path(__file__).with_name("golden")
    promoted = r103_review_promotion.load_r103_promoted_review_state(
        golden / "r103-review-state-26.07d.json"
    )
    assert promoted.artifact_identity == (
        "90ea507e93cebaf6399b3aa5bea92081e6d3dba50b7631783666d9382d267d1a"
    )
    assert tuple(
        (decision.subject_code, decision.outcome)
        for decision in promoted.registry.decisions
    ) == (
        ("C2860", "source-supported"),
        ("C3264", "review-required"),
        ("C3716", "source-supported"),
    )
    assert tuple(
        (decision.rationale, decision.reviewer, decision.review_date)
        for decision in promoted.registry.decisions
    ) == tuple(
        (rationale, "R. Hannes Niedner, M.D.", "2026-08-26")
        for _, rationale in CURRENT_DECISIONS
    )
    assert all(row.source_occurrence_identity for row in promoted.packet.rows)
    assert promoted.dry_run.unresolved == 1
    assert promoted.dry_run.readiness == "review-incomplete"
    oracle = json.loads(
        (golden / "neoplasm-adjudicated.json").read_text(encoding="utf-8")
    )
    cohort = {concept["code"] for concept in oracle["concepts"]}
    assert {"C2860", "C3264", "C3716"}.isdisjoint(cohort)
