"""Deterministic readers for the certified ICD-O source workbooks."""

from __future__ import annotations

import hashlib
import io
import json
import zipfile
from collections import defaultdict
from typing import TYPE_CHECKING, Literal, cast

import xlrd
from openpyxl import load_workbook
from pydantic import ValidationError

if TYPE_CHECKING:
    from collections.abc import Iterable
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

from ontolib.repositories.icdo.models import (
    CanonicalDataset,
    Icdo4Datasets,
    IcdoRecord,
    MorphologyCode32,
    MorphologyCode40,
    SourceShape,
    TopographyCode40,
)

ICDO32_SHA256 = "7ca51dcb66107d6462b43212b26aa65d52f6b0e306c6295e8c751416b3278a21"
ICDO4_ARCHIVE_SHA256 = (
    "395cfef9d039bd2978efefa79f086ba20eb51a4f1ea14fa0b69a44a2943d25d2"
)
ICDO4_SOURCE_SHA256 = "280ae87dc8bfea873a2346e7a5bee380877da1c84f8339697155fa5e77f3deef"
ICDO4_MORPHOLOGY_ANNEX_SHA256 = (
    "754904e347e3749dc93484e8f7af09e7540dc1965a1f3eea1463a2834aa70900"
)
ICDO4_TOPOGRAPHY_ANNEX_SHA256 = (
    "694da4b4fb997cb213afcf1bd012eaaa83d989ee7d1e83d48221d31662ea33b6"
)
_HEADERS = (
    "ICDO3.2",
    "Level",
    "Term",
    "Code reference",
    "obs",
    "See also",
    "See note",
    "Includes",
    "Excludes",
    "Other text",
)
_ROW_WIDTH = {"morphology": 10, "topography": 11}


class SourceFormatError(ValueError):
    """Publisher source does not satisfy the certified workbook contract."""


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _text(value: object) -> str | None:
    if value is None or value == "":
        return None
    if not isinstance(value, (str, int, float)):
        raise SourceFormatError(f"unsupported source cell type: {type(value).__name__}")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _append(target: dict[str, list[str]], key: str, value: object) -> None:
    text = _text(value)
    if text is not None and text not in target[key]:
        target[key].append(text)


def _parse_code(code: str, *, edition: str, axis: str) -> str | None:
    try:
        if axis == "morphology":
            if edition == "3.2":
                MorphologyCode32(value=code)
            else:
                MorphologyCode40(value=code)
            return "morphology"
        return TopographyCode40(value=code).level
    except ValidationError:
        return None


def _topography_category_code(
    code: str | None, level: str | None, term: str | None, *, axis: str
) -> str | None:
    if (
        axis == "topography"
        and code is None
        and level == "3"
        and term == "OTHER AND ILL-DEFINED SITES"
    ):
        return "C76"
    return code


def _valid_row(code: str | None, level: str | None, *, axis: str) -> bool:
    record_levels = {"Preferred", "Synonym", "Related"}
    if axis == "topography":
        record_levels.add("3")
    return code is not None and level in record_levels


def _term_key(level: str) -> str:
    return {
        "Preferred": "preferred",
        "Synonym": "synonyms",
        "Related": "related",
        "3": "preferred",
    }[level]


def _required_term(term: str | None, row_number: int) -> str:
    if term is None:
        raise SourceFormatError(f"row {row_number}: required term is empty")
    return term


def _missing_preferred_allowed(edition: str, axis: str, code: str) -> bool:
    return edition == "4.0" and axis == "morphology" and code == "85032/0"


def _morphology_fields(code: str, edition: str) -> dict[str, str | None]:
    parsed = (
        MorphologyCode32(value=code)
        if edition == "3.2"
        else MorphologyCode40(value=code)
    )
    return {
        "base_morphology": parsed.base,
        "specificity": (
            parsed.specificity if isinstance(parsed, MorphologyCode40) else None
        ),
        "behaviour": parsed.behaviour,
    }


def _structural_fields(
    code: str, *, edition: str, axis: str, level: str
) -> dict[str, str | None]:
    if axis == "morphology":
        return _morphology_fields(code, edition)
    return {"parent_code": code[:3]} if level == "leaf" else {}


def _optional_fields(axis: str) -> tuple[tuple[str, int], ...]:
    if axis == "topography":
        return (
            ("notes", 3),
            ("code_references", 4),
            ("notes", 5),
            ("see_also", 6),
            ("see_notes", 7),
            ("includes", 8),
            ("excludes", 9),
            ("other_text", 10),
        )
    return (
        ("code_references", 3),
        ("notes", 4),
        ("see_also", 5),
        ("see_notes", 6),
        ("includes", 7),
        ("excludes", 8),
        ("other_text", 9),
    )


def _collect_rows(
    rows: Iterable[tuple[object, ...]], *, edition: str, axis: str
) -> tuple[
    dict[str, dict[str, list[str]]],
    dict[str, Literal["morphology", "category", "leaf"]],
]:
    values: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    levels: dict[str, Literal["morphology", "category", "leaf"]] = {}
    for row_number, row in enumerate(rows, start=3):
        width = _ROW_WIDTH[axis]
        padded = (*row, *(None for _ in range(max(0, width - len(row)))))
        code, level, term = (_text(padded[index]) for index in range(3))
        code = _topography_category_code(code, level, term, axis=axis)
        if not _valid_row(code, level, axis=axis):
            continue
        code = cast("str", code)
        level = cast("str", level)
        term = _required_term(term, row_number)
        record_level = cast(
            "Literal['morphology', 'category', 'leaf'] | None",
            _parse_code(code, edition=edition, axis=axis),
        )
        if record_level is None:
            raise SourceFormatError(
                f"row {row_number}: invalid ICD-O-{edition} {axis} code: {code}"
            )
        levels[code] = record_level
        _append(values[code], _term_key(level), term)
        for name, index in _optional_fields(axis):
            _append(values[code], name, padded[index])
    return values, levels


def _build_records(
    values: dict[str, dict[str, list[str]]],
    levels: dict[str, Literal["morphology", "category", "leaf"]],
    *,
    edition: str,
    axis: str,
) -> tuple[IcdoRecord, ...]:
    output: list[IcdoRecord] = []
    for code in sorted(values):
        data = values[code]
        preferred = data["preferred"]
        if len(preferred) > 1:
            raise SourceFormatError(f"{code}: multiple preferred terms")
        if not preferred and not _missing_preferred_allowed(edition, axis, code):
            raise SourceFormatError(f"{code}: required preferred term is absent")
        kwargs = _structural_fields(
            code, edition=edition, axis=axis, level=levels[code]
        )
        output.append(
            IcdoRecord(
                code=code,
                level=levels[code],
                preferred=preferred[0] if preferred else None,
                synonyms=tuple(data["synonyms"]),
                related=tuple(data["related"]),
                notes=tuple(data["notes"]),
                code_references=tuple(data["code_references"]),
                see_also=tuple(data["see_also"]),
                see_notes=tuple(data["see_notes"]),
                includes=tuple(data["includes"]),
                excludes=tuple(data["excludes"]),
                other_text=tuple(data["other_text"]),
                **kwargs,
            )
        )
    return tuple(output)


def _records(
    rows: Iterable[tuple[object, ...]], *, edition: str, axis: str
) -> tuple[IcdoRecord, ...]:
    values, levels = _collect_rows(rows, edition=edition, axis=axis)
    return _build_records(values, levels, edition=edition, axis=axis)


def canonical_bytes(dataset: CanonicalDataset) -> bytes:
    return json.dumps(
        dataset.model_dump(
            mode="json", exclude={"source_shape", "term_counts", "level_counts"}
        ),
        ensure_ascii=True,
        separators=(",", ":"),
        sort_keys=True,
    ).encode()


def _require_digest(payload: bytes, expected: str, label: str) -> None:
    if _sha256(payload) != expected:
        raise SourceFormatError(f"{label} SHA-256 does not match certification")


def _require_value(actual: object, expected: object, message: str) -> None:
    if actual != expected:
        raise SourceFormatError(message)


def ingest_icdo32_morphology(path: Path) -> CanonicalDataset:
    payload = path.read_bytes()
    _require_digest(payload, ICDO32_SHA256, "ICD-O-3.2 source")
    workbook = xlrd.open_workbook(file_contents=payload, formatting_info=True)
    _require_value(
        workbook.sheet_names(),
        ["ICD-O-3.2 Morphology"],
        "ICD-O-3.2 sheet names do not match certification",
    )
    sheet = workbook.sheet_by_index(0)
    headers = tuple(_text(sheet.cell_value(1, index)) or "" for index in range(10))
    _require_value(headers, _HEADERS, "ICD-O-3.2 headers do not match certification")
    rows = [
        tuple(sheet.cell_value(row, col) for col in range(10))
        for row in range(2, sheet.nrows)
    ]
    return CanonicalDataset(
        edition="3.2",
        axis="morphology",
        records=_records(rows, edition="3.2", axis="morphology"),
        source_shape=SourceShape(
            sheet_names=tuple(workbook.sheet_names()),
            headers=headers,
            merged_ranges=("A1:H1",),
            trailing_blank_rows=0,
        ),
        source_sha256=_sha256(payload),
    )


def _xlsx_payload(path: Path) -> tuple[bytes, str | None]:
    payload = path.read_bytes()
    if path.suffix.lower() != ".zip":
        return payload, None
    _require_digest(payload, ICDO4_ARCHIVE_SHA256, "ICD-O-4 archive")
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        names = archive.namelist()
        _require_value(
            names,
            ["ICD-O-4.xlsx", "Morphology_annexes.xlsx", "Topography_annexes.xlsx"],
            "ICD-O-4 archive members do not match certification",
        )
        member = archive.read("ICD-O-4.xlsx")
    return member, _sha256(payload)


def _require_annexes(morphology: Path | None, topography: Path | None) -> None:
    for annex, expected in (
        (morphology, ICDO4_MORPHOLOGY_ANNEX_SHA256),
        (topography, ICDO4_TOPOGRAPHY_ANNEX_SHA256),
    ):
        if annex is not None:
            _require_digest(annex.read_bytes(), expected, "ICD-O-4 annex")


def _icdo4_dataset(
    workbook: Workbook,
    *,
    axis: Literal["morphology", "topography"],
    merged: str,
    annex: Path | None,
    payload: bytes,
    archive_sha: str | None,
) -> CanonicalDataset:
    sheet = workbook[axis.title()]
    headers = tuple(_text(cell.value) or "" for cell in sheet[2])
    rows = (tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=3))
    return CanonicalDataset(
        edition="4.0",
        axis=axis,
        records=_records(rows, edition="4.0", axis=axis),
        source_shape=SourceShape(
            sheet_names=tuple(workbook.sheetnames),
            headers=headers,
            merged_ranges=(merged,),
            trailing_blank_rows=0,
        ),
        source_sha256=_sha256(payload),
        archive_sha256=archive_sha,
        annex_sha256=_sha256(annex.read_bytes()) if annex is not None else None,
    )


def ingest_icdo4(
    path: Path,
    *,
    morphology_annex_path: Path | None = None,
    topography_annex_path: Path | None = None,
    verify_identity: bool = True,
) -> Icdo4Datasets:
    payload, archive_sha = _xlsx_payload(path)
    if verify_identity:
        _require_digest(payload, ICDO4_SOURCE_SHA256, "ICD-O-4 workbook")
    _require_annexes(morphology_annex_path, topography_annex_path)
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    _require_value(
        workbook.sheetnames,
        ["Morphology", "Topography"],
        "ICD-O-4 sheet names do not match certification",
    )
    morphology = _icdo4_dataset(
        workbook,
        axis="morphology",
        merged="A1:H1",
        annex=morphology_annex_path,
        payload=payload,
        archive_sha=archive_sha,
    )
    topography = _icdo4_dataset(
        workbook,
        axis="topography",
        merged="A1:I1",
        annex=topography_annex_path,
        payload=payload,
        archive_sha=archive_sha,
    )
    return Icdo4Datasets(morphology=morphology, topography=topography)
