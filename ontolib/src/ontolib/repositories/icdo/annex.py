"""Identity-bound reconciliation of publisher morphology annexes."""

from __future__ import annotations

import hashlib
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from ontolib.repositories.icdo.store import canonical_sha256

if TYPE_CHECKING:
    from pathlib import Path

    from ontolib.repositories.icdo.models import CanonicalDataset


class _Model(BaseModel):
    model_config = ConfigDict(frozen=True, extra="forbid", strict=True)


class UnresolvedAnnexReference(_Model):
    sheet: str
    row: int
    edition: str
    code: str


class AnnexReconciliation(_Model):
    annex_sha256: str
    old_serving_sha256: str
    new_serving_sha256: str
    sheet_counts: dict[str, int]
    checked_rows: int
    unresolved: tuple[UnresolvedAnnexReference, ...]


def _matches(code: str, available: set[str]) -> bool:
    if "_" in code:
        prefix, behaviour = code.split("/", 1)
        return any(
            candidate.startswith(prefix.removesuffix("_"))
            and candidate.endswith(f"/{behaviour}")
            for candidate in available
        )
    return code in available


def _references(sheet: str, row: tuple[object, ...]) -> tuple[tuple[str, str], ...]:
    values = tuple(str(value).strip() if value is not None else "" for value in row)
    rules: dict[str, tuple[tuple[str, int], ...]] = {
        "New morphology codes (4 digits)": (("4.0", 0),),
        "New morphology codes (5 digits)": (("4.0", 0),),
        "Morphology code changes": (("4.0", 0), ("3.2", 3)),
        "Deleted morphology codes": (("3.2", 0),),
        "Behaviour code changes": (("4.0", 0), ("3.2", 3)),
        "New morphology terms": (("4.0", 0),),
        "Morphology term changes": (("4.0", 0), ("3.2", 3)),
        "Deleted morphology terms": (("3.2", 0),),
    }
    return tuple(
        (edition, values[index])
        for edition, index in rules.get(sheet, ())
        if index < len(values) and values[index]
    )


def _unresolved(
    sheet: str,
    row: int,
    refs: tuple[tuple[str, str], ...],
    available: dict[str, set[str]],
) -> list[UnresolvedAnnexReference]:
    return [
        UnresolvedAnnexReference(sheet=sheet, row=row, edition=edition, code=code)
        for edition, code in refs
        if not _matches(code, available[edition])
    ]


def reconcile_morphology_annex(
    path: Path, *, old: CanonicalDataset, new: CanonicalDataset
) -> AnnexReconciliation:
    payload = path.read_bytes()
    workbook = load_workbook(path, read_only=True, data_only=True)
    available = {
        "3.2": {record.code for record in old.records},
        "4.0": {record.code for record in new.records},
    }
    counts: dict[str, int] = {}
    unresolved: list[UnresolvedAnnexReference] = []
    checked = 0
    for sheet in workbook:
        count = 0
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=3, values_only=True), 3
        ):
            refs = _references(sheet.title, row)
            if not refs:
                continue
            count += 1
            checked += 1
            unresolved.extend(_unresolved(sheet.title, row_number, refs, available))
        counts[sheet.title] = count
    return AnnexReconciliation(
        annex_sha256=hashlib.sha256(payload).hexdigest(),
        old_serving_sha256=canonical_sha256(old),
        new_serving_sha256=canonical_sha256(new),
        sheet_counts=counts,
        checked_rows=checked,
        unresolved=tuple(unresolved),
    )
