"""Offline, source-bound manual SME review boundary for three R103 assertions."""

from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from collections import Counter
from contextlib import suppress
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal, Self, cast
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from defusedxml.ElementTree import iterparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, ConfigDict, Field, model_validator
from pydantic_core import to_jsonable_python

from ontolib.decomposition.axes import is_generic_filler, is_unsupported_filler
from ontolib.decomposition.models import (
    canonical_definition_fact_id,
    canonical_definition_group_id,
    canonical_source_occurrence_id,
)
from ontolib.decomposition.proposal_registry import load_proposal_registry

_NCIT = "http://ncicb.nci.nih.gov/xml/owl/EVS/Thesaurus.owl#"
_RDF = "http://www.w3.org/1999/02/22-rdf-syntax-ns#"
_RDFS = "http://www.w3.org/2000/01/rdf-schema#"
_OWL = "http://www.w3.org/2002/07/owl#"
_ABOUT = f"{{{_RDF}}}about"
_RESOURCE = f"{{{_RDF}}}resource"
_CLASS = f"{{{_OWL}}}Class"
_INTERSECTION = f"{{{_OWL}}}intersectionOf"
_RESTRICTION = f"{{{_OWL}}}Restriction"
_ON_PROPERTY = f"{{{_OWL}}}onProperty"
_SOME_VALUES = f"{{{_OWL}}}someValuesFrom"
_LABEL = f"{{{_RDFS}}}label"
_P97 = f"{{{_NCIT}}}P97"
_SHA256 = r"^[0-9a-f]{64}$"
_CODE = r"^C[0-9]+$"
_ROLE = r"^R[0-9]+$"
_SCHEMA_VERSION = 1
_WORKBOOK_MAX_ROW = 4
_TOOL_VERSION = "ontoprism-r103-offline-rdfxml-v1"
_INVENTORY = (
    ("C2860", "R103", "C12950"),
    ("C3264", "R103", "C12950"),
    ("C3716", "R103", "C34228"),
)
_METHOD = ("C3708", "R103", "C54105")
_OUTCOMES = (
    "source-supported",
    "correction-proposal",
    "concept-scoped-accuracy-exclusion",
    "review-required",
)
R103Outcome = Literal[
    "source-supported",
    "correction-proposal",
    "concept-scoped-accuracy-exclusion",
    "review-required",
]
_QUERY_CONTRACT = {
    "candidate_assertions": _INVENTORY,
    "method_reference": _METHOD,
    "source": "stated RDF/XML equivalentClass/intersectionOf",
    "source_passes": 2,
    "labels": "rdfs:label",
    "definitions": "P97",
    "restrictions": "owl:onProperty/owl:someValuesFrom",
    "ordering": "fixed assertion inventory, then source member position",
}


class R103ReviewValidationError(ValueError):
    """Evidence cannot support the exact source-bound R103 review contract."""


class _StrictModel(BaseModel):
    model_config = ConfigDict(frozen=True, extra="forbid", strict=True)


def _canonical(value: object) -> bytes:
    return json.dumps(
        to_jsonable_python(value),
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    ).encode("ascii")


def _identity(value: object) -> str:
    return hashlib.sha256(_canonical(value)).hexdigest()


def _tool_identity() -> str:
    return _identity(
        {
            "tool": _TOOL_VERSION,
            "module_sha256": hashlib.sha256(Path(__file__).read_bytes()).hexdigest(),
        }
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _load_json(path: Path) -> object:
    def reject_duplicates(pairs: list[tuple[str, object]]) -> dict[str, object]:
        value: dict[str, object] = {}
        for key, item in pairs:
            if key in value:
                raise R103ReviewValidationError(f"duplicate JSON key: {key}")
            value[key] = item
        return value

    try:
        return json.loads(
            path.read_text(encoding="utf-8"), object_pairs_hook=reject_duplicates
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
        raise R103ReviewValidationError("invalid JSON evidence") from error


class SourceMember(_StrictModel):
    kind: Literal["genus", "restriction"]
    member_position: int = Field(ge=0)
    genus_code: str | None = Field(default=None, pattern=_CODE)
    role_code: str | None = Field(default=None, pattern=_ROLE)
    filler_code: str | None = Field(default=None, pattern=_CODE)
    source_fact_identity: str = Field(pattern=_SHA256)


class CoassertedFact(_StrictModel):
    role_code: Literal["R103", "R104"]
    filler_code: str = Field(pattern=_CODE)
    member_position: int = Field(ge=0)
    source_fact_identity: str = Field(pattern=_SHA256)


class GenusFact(_StrictModel):
    genus_code: str = Field(pattern=_CODE)
    genus_label: str = Field(min_length=1)
    member_position: int = Field(ge=0)
    source_fact_identity: str = Field(pattern=_SHA256)


class R103EvidenceRow(_StrictModel):
    row_identity: str = Field(pattern=_SHA256)
    subject_code: str = Field(pattern=_CODE)
    role_code: Literal["R103"]
    filler_code: str = Field(pattern=_CODE)
    subject_label: str = Field(min_length=1)
    role_label: str = Field(min_length=1)
    filler_label: str = Field(min_length=1)
    subject_p97_definition: str = Field(min_length=1)
    role_p97_definition: str = Field(min_length=1)
    filler_p97_definition: str = Field(min_length=1)
    complete_definition_identity: str = Field(pattern=_SHA256)
    source_fact_identity: str = Field(pattern=_SHA256)
    source_group_identity: str = Field(pattern=_SHA256)
    source_occurrence_identity: str = Field(pattern=_SHA256)
    anchor_code: str = Field(pattern=_CODE)
    depth: int = Field(ge=0)
    structural_path: tuple[int, ...] = Field(min_length=1)
    member_position: int = Field(ge=0)
    source_citation: str = Field(min_length=1)
    genus_facts: tuple[GenusFact, ...] = Field(min_length=1)
    coasserted_facts: tuple[CoassertedFact, ...] = Field(min_length=1)
    role_is_non_defining: Literal[True]
    current_state: Literal["projected", "suppressed", "review-required"]
    inherited_roots: tuple[str, ...]
    impact_concepts: tuple[str, ...] = Field(min_length=1)
    machine_evidence: str = Field(min_length=1)
    contrast_to_method: str = Field(min_length=1)

    @model_validator(mode="after")
    def _validate_row(self) -> Self:
        if self.structural_path[-1] != self.member_position:
            raise ValueError("member position must terminate structural path")
        if self.anchor_code != self.subject_code:
            raise ValueError("direct review row anchor must equal subject")
        if self.row_identity != _identity(self.model_dump(exclude={"row_identity"})):
            raise ValueError("row identity differs")
        return self


class MethodReference(_StrictModel):
    subject_code: Literal["C3708"]
    role_code: Literal["R103"]
    filler_code: Literal["C54105"]
    subject_label: str = Field(min_length=1)
    filler_label: str = Field(min_length=1)
    subject_p97_definition: str = Field(min_length=1)
    filler_p97_definition: str = Field(min_length=1)
    complete_definition_identity: str = Field(pattern=_SHA256)
    source_fact_identity: str = Field(pattern=_SHA256)
    source_group_identity: str = Field(pattern=_SHA256)
    source_occurrence_identity: str = Field(pattern=_SHA256)
    source_citation: str = Field(min_length=1)
    genus_facts: tuple[GenusFact, ...] = Field(min_length=1)
    coasserted_facts: tuple[CoassertedFact, ...] = Field(min_length=1)
    is_decision_row: Literal[False]
    comparison_scope: Literal["method comparison evidence only"]


class R103ReviewPacket(_StrictModel):
    schema_version: Literal[1]
    source_release: Literal["26.07d"]
    source_identity: str = Field(pattern=_SHA256)
    source_artifact_sha256: str = Field(pattern=_SHA256)
    source_artifact_size: int = Field(gt=0)
    candidate_manifest_identity: str = Field(pattern=_SHA256)
    proposal_registry_identity: str = Field(pattern=_SHA256)
    query_contract_identity: str = Field(pattern=_SHA256)
    tool_identity: str = Field(pattern=_SHA256)
    source_pass_count: Literal[2]
    rows: tuple[R103EvidenceRow, R103EvidenceRow, R103EvidenceRow]
    method_reference: MethodReference
    packet_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_packet(self) -> Self:
        inventory = tuple(
            (row.subject_code, row.role_code, row.filler_code) for row in self.rows
        )
        if inventory != _INVENTORY:
            raise ValueError(
                "packet must contain the exact ordered assertion inventory"
            )
        if self.query_contract_identity != _identity(_QUERY_CONTRACT):
            raise ValueError("query contract identity differs")
        if self.tool_identity != _tool_identity():
            raise ValueError("tool identity differs")
        expected = _identity(self.model_dump(exclude={"packet_identity"}))
        if self.packet_identity != expected:
            raise ValueError("packet identity differs")
        return self


class _Definition:
    def __init__(self, code: str, members: tuple[tuple[str, ...], ...]) -> None:
        self.code = code
        self.members = members
        signatures = tuple(":".join(member) for member in members)
        self.group_id = canonical_definition_group_id(code, signatures)
        self.fact_ids = tuple(
            _member_fact_id(code, self.group_id, member) for member in members
        )
        self.identity = _identity(
            {
                "root_code": code,
                "group_id": self.group_id,
                "members": tuple(zip(members, self.fact_ids, strict=True)),
            }
        )


def _member_fact_id(code: str, group_id: str, member: tuple[str, ...]) -> str:
    if member[0] == "genus":
        return canonical_definition_fact_id(code, group_id, "genus", *member[1:])
    return canonical_definition_fact_id(code, group_id, "restriction", *member[1:])


def _code(iri: str | None, pattern: str, description: str) -> str:
    if not isinstance(iri, str) or not iri.startswith(_NCIT):
        raise R103ReviewValidationError(f"{description} is not an NCIt IRI")
    value = iri.removeprefix(_NCIT)
    if re.fullmatch(pattern, value) is None:
        raise R103ReviewValidationError(f"{description} is malformed")
    return value


def _definition_from_class(code: str, element: object) -> _Definition | None:
    # ElementTree's runtime element API is intentionally used without retaining nodes.
    intersections = list(element.iter(_INTERSECTION))  # type: ignore[attr-defined]
    if not intersections:
        return None
    if len(intersections) != 1:
        raise R103ReviewValidationError(f"{code} has multiple definition intersections")
    members = [_source_member(code, member) for member in list(intersections[0])]
    if not members:
        raise R103ReviewValidationError(f"{code} has an empty definition")
    return _Definition(code, tuple(members))


def _source_member(code: str, member: Any) -> tuple[str, ...]:
    if member.tag == f"{{{_RDF}}}Description":
        return ("genus", _code(member.get(_ABOUT), _CODE, "genus"), "primitive")
    if member.tag != _RESTRICTION:
        raise R103ReviewValidationError(f"{code} has unsupported definition member")
    properties = list(member.findall(_ON_PROPERTY))
    fillers = list(member.findall(_SOME_VALUES))
    if len(properties) != 1 or len(fillers) != 1:
        raise R103ReviewValidationError(f"{code} has malformed restriction")
    return (
        "restriction",
        _code(properties[0].get(_RESOURCE), _ROLE, "role"),
        _code(fillers[0].get(_RESOURCE), _CODE, "filler"),
    )


def _scan_definitions(path: Path) -> dict[str, _Definition]:
    wanted = {item[0] for item in (*_INVENTORY, _METHOD)}
    found: dict[str, _Definition] = {}
    depth = 0
    for event, element in iterparse(path, events=("start", "end")):
        if element.tag != _CLASS:
            continue
        if event == "start":
            depth += 1
            continue
        if depth == 1:
            _record_definition(element, wanted, found)
            element.clear()
        depth -= 1
    missing = wanted - found.keys()
    if missing:
        raise R103ReviewValidationError(
            f"missing source definitions: {sorted(missing)}"
        )
    return found


def _record_definition(
    element: Any, wanted: set[str], found: dict[str, _Definition]
) -> None:
    code = _source_entity_code(element)
    if code is None:
        return
    if code not in wanted:
        return
    definition = _definition_from_class(code, element)
    if definition is None:
        raise R103ReviewValidationError(f"{code} lacks a complete definition")
    if code in found:
        raise R103ReviewValidationError(f"duplicate source class {code}")
    found[code] = definition


def _scan_text(
    path: Path, codes: set[str], required_definitions: set[str]
) -> dict[str, tuple[str, str]]:
    found: dict[str, tuple[str, str]] = {}
    for _event, element in iterparse(path, events=("end",)):
        _record_source_text(element, codes, required_definitions, found)
    missing = codes - found.keys()
    if missing:
        raise R103ReviewValidationError(f"missing source text: {sorted(missing)}")
    return found


def _record_source_text(
    element: Any,
    codes: set[str],
    required_definitions: set[str],
    found: dict[str, tuple[str, str]],
) -> None:
    code = _requested_entity_code(element, codes)
    if code is None:
        return
    labels, definitions = _source_text_values(element)
    if not labels and not definitions:
        element.clear()
        return
    _validate_source_text(code, labels, definitions, required_definitions)
    if code in found:
        raise R103ReviewValidationError(f"duplicate source entity {code}")
    found[code] = (labels[0], definitions[0] if definitions else "")
    element.clear()


def _source_entity_code(element: Any) -> str | None:
    about = element.get(_ABOUT)
    if not isinstance(about, str) or not about.startswith(_NCIT):
        return None
    return about.removeprefix(_NCIT)


def _requested_entity_code(element: Any, codes: set[str]) -> str | None:
    code = _source_entity_code(element)
    if code is None:
        return None
    if code not in codes:
        element.clear()
        return None
    return code


def _source_text_values(element: Any) -> tuple[tuple[str, ...], tuple[str, ...]]:
    labels = tuple((child.text or "").strip() for child in element.findall(_LABEL))
    definitions = tuple((child.text or "").strip() for child in element.findall(_P97))
    return labels, definitions


def _validate_source_text(
    code: str,
    labels: tuple[str, ...],
    definitions: tuple[str, ...],
    required_definitions: set[str],
) -> None:
    if len(labels) != 1 or not labels[0]:
        raise R103ReviewValidationError(f"{code} lacks one exact label")
    if code in required_definitions and (len(definitions) != 1 or not definitions[0]):
        raise R103ReviewValidationError(f"{code} lacks one exact P97 definition")
    if len(definitions) > 1:
        raise R103ReviewValidationError(f"{code} has multiple P97 definitions")


def _manifest_binding(owl_path: Path, manifest_path: Path) -> tuple[str, str, int, str]:
    manifest = _load_json(manifest_path)
    if not isinstance(manifest, dict):
        raise R103ReviewValidationError("candidate manifest must be an object")
    source_identity, expected_sha, expected_size = _manifest_fields(manifest)
    actual_sha = _file_sha256(owl_path)
    actual_size = owl_path.stat().st_size
    if expected_sha != actual_sha or expected_size != actual_size:
        raise R103ReviewValidationError("stale source artifact")
    return source_identity, actual_sha, actual_size, _identity(manifest)


def _manifest_fields(manifest: dict[str, object]) -> tuple[str, object, object]:
    stated = manifest.get("stated_artifact")
    if not isinstance(stated, dict):
        raise R103ReviewValidationError("candidate manifest lacks stated artifact")
    if manifest.get("ontology_version") != "26.07d":
        raise R103ReviewValidationError("source release must be NCIt 26.07d")
    source_identity = manifest.get("source_identity")
    if (
        not isinstance(source_identity, str)
        or re.fullmatch(_SHA256, source_identity) is None
    ):
        raise R103ReviewValidationError("candidate source identity is invalid")
    return source_identity, stated.get("sha256"), stated.get("size_bytes")


def _relevant_codes(definitions: dict[str, _Definition]) -> set[str]:
    codes = {"R103"}
    for definition in definitions.values():
        codes.add(definition.code)
        for member in definition.members:
            if member[0] == "genus":
                codes.add(member[1])
    codes.update(item[2] for item in (*_INVENTORY, _METHOD))
    return codes


def _definition_codes() -> set[str]:
    return {
        *(item[0] for item in _INVENTORY),
        *(item[2] for item in _INVENTORY),
        _METHOD[0],
        "R103",
    }


def _facts(
    definition: _Definition, text: dict[str, tuple[str, str]]
) -> tuple[tuple[GenusFact, ...], tuple[CoassertedFact, ...]]:
    genera = tuple(
        GenusFact(
            genus_code=member[1],
            genus_label=text[member[1]][0],
            member_position=position,
            source_fact_identity=definition.fact_ids[position],
        )
        for position, member in enumerate(definition.members)
        if member[0] == "genus"
    )
    coasserted = tuple(
        CoassertedFact(
            role_code=member[1],  # type: ignore[arg-type]
            filler_code=member[2],
            member_position=position,
            source_fact_identity=definition.fact_ids[position],
        )
        for position, member in enumerate(definition.members)
        if member[0] == "restriction" and member[1] in {"R103", "R104"}
    )
    return genera, coasserted


def _restriction_position(definition: _Definition, role: str, filler: str) -> int:
    positions = tuple(
        position
        for position, member in enumerate(definition.members)
        if member == ("restriction", role, filler)
    )
    if len(positions) != 1:
        raise R103ReviewValidationError(
            f"{definition.code}/{role}/{filler} must occur exactly once in source"
        )
    return positions[0]


def _current_state(subject: str, role: str, filler: str) -> str:
    if is_unsupported_filler(subject, role, filler):
        return "review-required"
    if is_generic_filler(role, filler):
        return "suppressed"
    return "projected"


def _evidence_row(
    assertion: tuple[str, str, str],
    definitions: dict[str, _Definition],
    text: dict[str, tuple[str, str]],
) -> R103EvidenceRow:
    subject, role, filler = assertion
    definition = definitions[subject]
    position = _restriction_position(definition, role, filler)
    fact_id = definition.fact_ids[position]
    occurrence_id = canonical_source_occurrence_id(subject, fact_id, (position,))
    genera, coasserted = _facts(definition, text)
    payload: dict[str, object] = {
        "subject_code": subject,
        "role_code": role,
        "filler_code": filler,
        "subject_label": text[subject][0],
        "role_label": text[role][0],
        "filler_label": text[filler][0],
        "subject_p97_definition": text[subject][1],
        "role_p97_definition": text[role][1],
        "filler_p97_definition": text[filler][1],
        "complete_definition_identity": definition.identity,
        "source_fact_identity": fact_id,
        "source_group_identity": definition.group_id,
        "source_occurrence_identity": occurrence_id,
        "anchor_code": subject,
        "depth": 0,
        "structural_path": (position,),
        "member_position": position,
        "source_citation": (
            f"Thesaurus-stated.owl#{subject} equivalentClass member {position}"
        ),
        "genus_facts": genera,
        "coasserted_facts": coasserted,
        "role_is_non_defining": True,
        "current_state": _current_state(subject, role, filler),
        "inherited_roots": (),
        "impact_concepts": (subject,),
        "machine_evidence": (
            "Mechanical evidence only; the stated restriction, definition text, "
            "genera, "
            "and coasserted R103/R104 facts are shown without a clinical verdict."
        ),
        "contrast_to_method": (
            "Comparison only; C3708/R103/C54105 demonstrates the same source-query and "
            "evidence method and is neither a recommendation nor a rationale."
        ),
    }
    return R103EvidenceRow.model_validate(
        {"row_identity": _identity(payload), **payload}
    )


def _method_reference(
    definitions: dict[str, _Definition], text: dict[str, tuple[str, str]]
) -> MethodReference:
    subject, role, filler = _METHOD
    definition = definitions[subject]
    position = _restriction_position(definition, role, filler)
    fact_id = definition.fact_ids[position]
    genera, coasserted = _facts(definition, text)
    return MethodReference(
        subject_code="C3708",
        role_code="R103",
        filler_code="C54105",
        subject_label=text[subject][0],
        filler_label=text[filler][0],
        subject_p97_definition=text[subject][1],
        filler_p97_definition=(
            text[filler][1]
            or "No P97 definition is present in the pinned stated source."
        ),
        complete_definition_identity=definition.identity,
        source_fact_identity=fact_id,
        source_group_identity=definition.group_id,
        source_occurrence_identity=canonical_source_occurrence_id(
            subject, fact_id, (position,)
        ),
        source_citation=(
            f"Thesaurus-stated.owl#{subject} equivalentClass member {position}"
        ),
        genus_facts=genera,
        coasserted_facts=coasserted,
        is_decision_row=False,
        comparison_scope="method comparison evidence only",
    )


def build_r103_review_packet(
    owl_path: Path, candidate_manifest_path: Path, proposal_registry_path: Path
) -> R103ReviewPacket:
    """Build exactly three R103 review rows in two bounded streaming XML passes."""
    source_identity, source_sha, source_size, manifest_identity = _manifest_binding(
        owl_path, candidate_manifest_path
    )
    proposal_registry = load_proposal_registry(proposal_registry_path)
    if proposal_registry.ontology_version != "26.07d":
        raise R103ReviewValidationError(
            "proposal registry release does not bind source"
        )
    definitions = _scan_definitions(owl_path)
    text = _scan_text(owl_path, _relevant_codes(definitions), _definition_codes())
    payload: dict[str, object] = {
        "schema_version": _SCHEMA_VERSION,
        "source_release": "26.07d",
        "source_identity": source_identity,
        "source_artifact_sha256": source_sha,
        "source_artifact_size": source_size,
        "candidate_manifest_identity": manifest_identity,
        "proposal_registry_identity": proposal_registry.registry_identity,
        "query_contract_identity": _identity(_QUERY_CONTRACT),
        "tool_identity": _tool_identity(),
        "source_pass_count": 2,
        "rows": tuple(_evidence_row(item, definitions, text) for item in _INVENTORY),
        "method_reference": _method_reference(definitions, text),
    }
    return R103ReviewPacket.model_validate(
        {**payload, "packet_identity": _identity(payload)}
    )


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = (
        json.dumps(
            to_jsonable_python(value), sort_keys=True, indent=2, ensure_ascii=True
        ).encode("ascii")
        + b"\n"
    )
    fd, staging = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as stream:
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(staging, path)
    except BaseException:
        with suppress(FileNotFoundError):
            os.unlink(staging)
        raise


def write_r103_review_packet(path: Path, packet: R103ReviewPacket) -> None:
    _write_json(path, packet.model_dump(mode="json"))


def load_r103_review_packet(path: Path) -> R103ReviewPacket:
    try:
        return R103ReviewPacket.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


_EVIDENCE_HEADERS = (
    "Subject Code",
    "Role Code",
    "Filler Code",
    "Subject Label",
    "Role Label",
    "Filler Label",
    "Subject P97 Definition",
    "Role P97 Definition",
    "Filler P97 Definition",
    "Complete Definition Identity",
    "Source Fact Identity",
    "Source Group Identity",
    "Source Occurrence Identity",
    "Anchor Code",
    "Depth",
    "Structural Path",
    "Member Position",
    "Source Citation",
    "Genus Facts",
    "Coasserted R103/R104 Facts",
    "Role Modality",
    "Current State",
    "Inherited Roots",
    "Impact Concepts",
    "Machine Evidence",
    "C3708 Method Contrast",
)
_HUMAN_HEADERS = ("Outcome", "Rationale", "Reviewer", "Date")
_HEADERS = (*_EVIDENCE_HEADERS, *_HUMAN_HEADERS)
_BINDINGS = (
    "packet_identity",
    "source_identity",
    "source_release",
    "source_artifact_sha256",
    "candidate_manifest_identity",
    "proposal_registry_identity",
    "query_contract_identity",
    "tool_identity",
)


def _genus_text(facts: tuple[GenusFact, ...]) -> str:
    return " | ".join(
        f"{item.genus_label} ({item.genus_code}) @ {item.member_position} "
        f"[{item.source_fact_identity}]"
        for item in facts
    )


def _coasserted_text(facts: tuple[CoassertedFact, ...]) -> str:
    return " | ".join(
        f"{item.role_code}/{item.filler_code} @ {item.member_position} "
        f"[{item.source_fact_identity}]"
        for item in facts
    )


def _row_values(row: R103EvidenceRow) -> tuple[object, ...]:
    return (
        row.subject_code,
        row.role_code,
        row.filler_code,
        row.subject_label,
        row.role_label,
        row.filler_label,
        row.subject_p97_definition,
        row.role_p97_definition,
        row.filler_p97_definition,
        row.complete_definition_identity,
        row.source_fact_identity,
        row.source_group_identity,
        row.source_occurrence_identity,
        row.anchor_code,
        row.depth,
        "/".join(map(str, row.structural_path)),
        row.member_position,
        row.source_citation,
        _genus_text(row.genus_facts),
        _coasserted_text(row.coasserted_facts),
        "non-defining",
        row.current_state,
        " | ".join(row.inherited_roots) or None,
        " | ".join(row.impact_concepts),
        row.machine_evidence,
        row.contrast_to_method,
    )


def _style_sheet(sheet: object) -> None:
    for cell in sheet[1]:  # type: ignore[index]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"  # type: ignore[attr-defined]
    sheet.auto_filter.ref = sheet.dimensions  # type: ignore[attr-defined]
    sheet.protection.sheet = True  # type: ignore[attr-defined]


def _normalize_xlsx(source: Path, destination: Path) -> None:
    with (
        ZipFile(source) as original,
        ZipFile(
            destination, "w", compression=ZIP_DEFLATED, compresslevel=9
        ) as normalized,
    ):
        for name in sorted(original.namelist()):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            content = original.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:(created|modified)[^>]*>.*?</dcterms:\1>",
                    rb'<dcterms:\1 xsi:type="dcterms:W3CDTF">'
                    rb"2000-01-01T00:00:00Z</dcterms:\1>",
                    content,
                )
            elif name == "xl/workbook.xml":
                content = re.sub(rb' calcId="[0-9]+"', b"", content)
            normalized.writestr(info, content)


def write_r103_review_workbook(path: Path, packet: R103ReviewPacket) -> None:
    """Write a deterministic blank workbook; software supplies no SME decision."""
    book = Workbook()
    instructions = book.active
    if instructions is None:
        raise R103ReviewValidationError("workbook has no active sheet")
    instructions.title = "Instructions"
    instructions.append(
        ("R103 manual SME review", "No outcome or rationale is prefilled.")
    )
    instructions.append(
        (
            "Scope",
            "Choose one closed outcome for each of exactly three source-bound "
            "assertions.",
        )
    )
    instructions.append(
        (
            "Method reference",
            "C3708/R103/C54105 is comparison evidence only and is not a decision row.",
        )
    )
    instructions.append(
        (
            "Correction proposal",
            "For correction-proposal only, Rationale must be a JSON object containing "
            "human_rationale, proposed_correction, duplicate_search_evidence, "
            "provenance, and lifecycle='proposed'. Import creates only a preview; any "
            "later accepted proposal extends the existing ProposalRegistry.",
        )
    )
    instructions.append(
        (
            "Exclusion scope",
            "A concept-scoped accuracy exclusion remains exact "
            "subject+role+filler+source+release and is never global.",
        )
    )
    _style_sheet(instructions)
    review = book.create_sheet("R103 Review")
    review.append(_HEADERS)
    for row in packet.rows:
        review.append((*_row_values(row), None, None, None, None))
    for row in review.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row[-4:]:
            cell.protection = Protection(locked=False)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
    outcome_column = len(_EVIDENCE_HEADERS) + 1
    validation = DataValidation(type="list", formula1='"' + ",".join(_OUTCOMES) + '"')
    validation.error = "Choose one listed outcome"
    validation.showErrorMessage = True
    review.add_data_validation(validation)
    validation.add(
        f"{review.cell(2, outcome_column).coordinate}:"
        f"{review.cell(_WORKBOOK_MAX_ROW, outcome_column).coordinate}"
    )
    _style_sheet(review)
    bindings = book.create_sheet("Bindings")
    binding_values = (
        packet.packet_identity,
        packet.source_identity,
        packet.source_release,
        packet.source_artifact_sha256,
        packet.candidate_manifest_identity,
        packet.proposal_registry_identity,
        packet.query_contract_identity,
        packet.tool_identity,
    )
    bindings.append(("Binding", "Value"))
    for name, value in zip(_BINDINGS, binding_values, strict=True):
        bindings.append((name, value))
    bindings.sheet_state = "veryHidden"
    bindings.protection.sheet = True
    book.calculation.calcMode = "auto"
    book.calculation.fullCalcOnLoad = False
    book.calculation.forceFullCalc = False
    fixed = datetime(2000, 1, 1)
    book.properties.created = fixed
    book.properties.modified = fixed
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, staging_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(fd)
    try:
        book.save(staging_name)
        _normalize_xlsx(Path(staging_name), path)
    finally:
        with suppress(FileNotFoundError):
            os.unlink(staging_name)


class CorrectionProposalPreview(_StrictModel):
    subject_code: str = Field(pattern=_CODE)
    role_code: Literal["R103"]
    filler_code: str = Field(pattern=_CODE)
    proposed_correction: str = Field(min_length=1)
    duplicate_search_evidence: str = Field(min_length=1)
    provenance: str = Field(min_length=1)
    lifecycle: Literal["proposed"]


class ConceptScopedExclusionPreview(_StrictModel):
    subject_code: str = Field(pattern=_CODE)
    role_code: Literal["R103"]
    filler_code: str = Field(pattern=_CODE)
    source_identity: str = Field(pattern=_SHA256)
    source_release: Literal["26.07d"]


class R103Decision(_StrictModel):
    decision_identity: str = Field(pattern=_SHA256)
    row_identity: str = Field(pattern=_SHA256)
    packet_identity: str = Field(pattern=_SHA256)
    workbook_identity: str = Field(pattern=_SHA256)
    source_identity: str = Field(pattern=_SHA256)
    source_release: Literal["26.07d"]
    subject_code: str = Field(pattern=_CODE)
    role_code: Literal["R103"]
    filler_code: str = Field(pattern=_CODE)
    outcome: R103Outcome
    rationale: str = Field(min_length=1)
    reviewer: str = Field(min_length=1)
    review_date: str = Field(pattern=r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")

    @model_validator(mode="after")
    def _validate_identity(self) -> Self:
        if self.decision_identity != _identity(
            self.model_dump(exclude={"decision_identity"})
        ):
            raise ValueError("decision identity differs")
        return self


class R103DecisionRegistry(_StrictModel):
    schema_version: Literal[1]
    packet_identity: str = Field(pattern=_SHA256)
    workbook_identity: str = Field(pattern=_SHA256)
    source_identity: str = Field(pattern=_SHA256)
    source_release: Literal["26.07d"]
    decisions: tuple[R103Decision, R103Decision, R103Decision]
    proposal_preview: tuple[CorrectionProposalPreview, ...]
    exclusion_preview: tuple[ConceptScopedExclusionPreview, ...]
    registry_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_registry(self) -> Self:
        if (
            tuple(
                (row.subject_code, row.role_code, row.filler_code)
                for row in self.decisions
            )
            != _INVENTORY
        ):
            raise ValueError("registry must contain exact ordered decisions")
        if self.registry_identity != _identity(
            self.model_dump(exclude={"registry_identity"})
        ):
            raise ValueError("registry identity differs")
        return self


def load_r103_decision_registry(path: Path) -> R103DecisionRegistry:
    try:
        return R103DecisionRegistry.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def _reject_archive_features(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            names = archive.namelist()
            if any(name.casefold().endswith("vbaproject.bin") for name in names):
                raise R103ReviewValidationError("workbook contains a macro")
            if any(name.startswith("xl/externalLinks/") for name in names):
                raise R103ReviewValidationError("workbook contains an external link")
    except R103ReviewValidationError:
        raise
    except Exception as error:
        raise R103ReviewValidationError("invalid workbook archive") from error


def _validate_workbook_structure(book: Any, packet: R103ReviewPacket) -> Any:
    review = _validate_workbook_container(book)
    _validate_workbook_bindings(book, packet)
    _validate_formula_free(book)
    _validate_immutable_rows(review, packet)
    return review


def _validate_workbook_container(book: Any) -> Any:
    if book.sheetnames != ["Instructions", "R103 Review", "Bindings"]:  # type: ignore[attr-defined]
        raise R103ReviewValidationError("workbook sheet inventory differs")
    if book["Bindings"].sheet_state != "veryHidden":  # type: ignore[index]
        raise R103ReviewValidationError("bindings visibility differs")
    review = book["R103 Review"]  # type: ignore[index]
    _validate_review_sheet_shape(review)
    return review


def _validate_review_sheet_shape(review: Any) -> None:
    if review.max_row != _WORKBOOK_MAX_ROW:
        raise R103ReviewValidationError(
            "workbook must contain exactly three decision rows"
        )
    if tuple(cell.value for cell in review[1]) != _HEADERS:
        raise R103ReviewValidationError("review headers differ")
    if any(
        review.row_dimensions[index].hidden for index in range(1, _WORKBOOK_MAX_ROW + 1)
    ):
        raise R103ReviewValidationError("workbook contains a hidden row")


def _validate_workbook_bindings(book: Any, packet: R103ReviewPacket) -> None:
    expected_bindings = tuple(
        zip(
            _BINDINGS,
            (
                packet.packet_identity,
                packet.source_identity,
                packet.source_release,
                packet.source_artifact_sha256,
                packet.candidate_manifest_identity,
                packet.proposal_registry_identity,
                packet.query_contract_identity,
                packet.tool_identity,
            ),
            strict=True,
        )
    )
    actual_bindings = tuple(
        (row[0].value, row[1].value)
        for row in book["Bindings"].iter_rows(min_row=2)  # type: ignore[index]
    )
    if actual_bindings != expected_bindings:
        raise R103ReviewValidationError("workbook bindings differ")


def _validate_formula_free(book: Any) -> None:
    for sheet in book.worksheets:  # type: ignore[attr-defined]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    raise R103ReviewValidationError("workbook contains a formula")


def _validate_immutable_rows(review: Any, packet: R103ReviewPacket) -> None:
    for row_number, evidence in enumerate(packet.rows, start=2):
        actual = tuple(
            review.cell(row_number, column).value
            for column in range(1, len(_EVIDENCE_HEADERS) + 1)
        )
        if actual != _row_values(evidence):
            raise R103ReviewValidationError("immutable row evidence differs")


def _human_values(review: Any, row_number: int) -> tuple[str, str, str, str]:
    start = len(_EVIDENCE_HEADERS) + 1
    values = tuple(
        review.cell(row_number, column).value for column in range(start, start + 4)
    )
    outcome, rationale, reviewer, review_date = _required_human_strings(values)
    if outcome not in _OUTCOMES:
        raise R103ReviewValidationError("outcome is not a closed outcome")
    _validate_iso_date(review_date)
    return outcome, rationale, reviewer, review_date


def _required_human_strings(values: tuple[object, ...]) -> tuple[str, str, str, str]:
    if any(not isinstance(value, str) or not value.strip() for value in values):
        raise R103ReviewValidationError("required human field is blank")
    return cast("tuple[str, str, str, str]", values)


def _validate_iso_date(review_date: str) -> None:
    if re.fullmatch(r"[0-9]{4}-[0-9]{2}-[0-9]{2}", review_date) is None:
        raise R103ReviewValidationError("review date is not an ISO date")
    try:
        if date.fromisoformat(review_date).isoformat() != review_date:
            raise ValueError
    except ValueError as error:
        raise R103ReviewValidationError("review date is not an ISO date") from error


def _proposal_preview(
    row: R103EvidenceRow, rationale: str
) -> CorrectionProposalPreview:
    value = _correction_payload(rationale)
    validated = _validated_correction_fields(value)
    return _correction_preview_model(row, validated)


def _correction_payload(rationale: str) -> dict[str, object]:
    try:
        value = json.loads(rationale)
    except json.JSONDecodeError as error:
        raise R103ReviewValidationError(
            "correction-proposal rationale must be JSON"
        ) from error
    if not isinstance(value, dict):
        raise R103ReviewValidationError(
            "correction-proposal rationale must be an object"
        )
    return value


def _validated_correction_fields(value: dict[str, object]) -> dict[str, str]:
    required = (
        "human_rationale",
        "proposed_correction",
        "duplicate_search_evidence",
        "provenance",
        "lifecycle",
    )
    for field in required:
        item = value.get(field)
        if not isinstance(item, str) or not item.strip():
            raise R103ReviewValidationError(f"correction-proposal requires {field}")
    if set(value) != set(required):
        raise R103ReviewValidationError("correction-proposal contains unknown fields")
    return cast("dict[str, str]", value)


def _correction_preview_model(
    row: R103EvidenceRow, value: dict[str, str]
) -> CorrectionProposalPreview:
    try:
        return CorrectionProposalPreview.model_validate(
            {
                "subject_code": row.subject_code,
                "role_code": "R103",
                "filler_code": row.filler_code,
                "proposed_correction": value["proposed_correction"],
                "duplicate_search_evidence": value["duplicate_search_evidence"],
                "provenance": value["provenance"],
                "lifecycle": value["lifecycle"],
            }
        )
    except ValueError as error:
        raise R103ReviewValidationError(
            "correction-proposal lifecycle must be proposed"
        ) from error


def import_r103_review_decisions(
    packet: R103ReviewPacket, workbook_path: Path, output_path: Path
) -> R103DecisionRegistry:
    """Fail closed and write a distinct review-evidence registry atomically."""
    try:
        R103ReviewPacket.model_validate(packet.model_dump(mode="python"))
    except ValueError as error:
        raise R103ReviewValidationError(
            f"packet identity validation failed: {error}"
        ) from error
    _reject_archive_features(workbook_path)
    try:
        book = load_workbook(workbook_path, data_only=False, keep_links=False)
    except Exception as error:
        raise R103ReviewValidationError("invalid review workbook") from error
    review = _validate_workbook_structure(book, packet)
    humans = tuple(_human_values(review, row_number) for row_number in range(2, 5))
    workbook_identity = _identity(
        {"packet_identity": packet.packet_identity, "human_rows": humans}
    )
    decisions: list[R103Decision] = []
    proposals: list[CorrectionProposalPreview] = []
    exclusions: list[ConceptScopedExclusionPreview] = []
    for evidence, (outcome, rationale, reviewer, review_date) in zip(
        packet.rows, humans, strict=True
    ):
        if outcome == "correction-proposal":
            proposals.append(_proposal_preview(evidence, rationale))
        if outcome == "concept-scoped-accuracy-exclusion":
            exclusions.append(
                ConceptScopedExclusionPreview(
                    subject_code=evidence.subject_code,
                    role_code="R103",
                    filler_code=evidence.filler_code,
                    source_identity=packet.source_identity,
                    source_release=packet.source_release,
                )
            )
        payload = {
            "row_identity": evidence.row_identity,
            "packet_identity": packet.packet_identity,
            "workbook_identity": workbook_identity,
            "source_identity": packet.source_identity,
            "source_release": packet.source_release,
            "subject_code": evidence.subject_code,
            "role_code": "R103",
            "filler_code": evidence.filler_code,
            "outcome": outcome,
            "rationale": rationale,
            "reviewer": reviewer,
            "review_date": review_date,
        }
        decisions.append(
            R103Decision.model_validate(
                {"decision_identity": _identity(payload), **payload}
            )
        )
    payload = {
        "schema_version": 1,
        "packet_identity": packet.packet_identity,
        "workbook_identity": workbook_identity,
        "source_identity": packet.source_identity,
        "source_release": packet.source_release,
        "decisions": tuple(decisions),
        "proposal_preview": tuple(proposals),
        "exclusion_preview": tuple(exclusions),
    }
    registry = R103DecisionRegistry.model_validate(
        {**payload, "registry_identity": _identity(payload)}
    )
    _write_json(output_path, registry.model_dump(mode="json"))
    return registry


class R103ReviewDryRun(_StrictModel):
    writes_performed: Literal[False]
    outcome_counts: dict[str, int]
    proposal_previews: tuple[CorrectionProposalPreview, ...]
    exclusion_previews: tuple[ConceptScopedExclusionPreview, ...]
    unresolved: int = Field(ge=0)
    readiness: Literal["ready-for-separate-application", "review-incomplete"]
    oracle_identity_before: str = Field(pattern=_SHA256)
    oracle_identity_after: str = Field(pattern=_SHA256)
    proposal_registry_identity_before: str = Field(pattern=_SHA256)
    proposal_registry_identity_after: str = Field(pattern=_SHA256)


def dry_run_r103_review(
    packet: R103ReviewPacket,
    registry: R103DecisionRegistry,
    *,
    oracle_path: Path,
    proposal_registry_path: Path,
) -> R103ReviewDryRun:
    """Preview consequences without modifying the oracle or ProposalRegistry."""
    _validate_dry_run_binding(packet, registry)
    oracle_before, proposal_before = _dry_run_input_identities(
        oracle_path, proposal_registry_path
    )
    unresolved = sum(row.outcome == "review-required" for row in registry.decisions)
    return R103ReviewDryRun(
        writes_performed=False,
        outcome_counts=dict(Counter(row.outcome for row in registry.decisions)),
        proposal_previews=registry.proposal_preview,
        exclusion_previews=registry.exclusion_preview,
        unresolved=unresolved,
        readiness=(
            "review-incomplete" if unresolved else "ready-for-separate-application"
        ),
        oracle_identity_before=oracle_before,
        oracle_identity_after=oracle_before,
        proposal_registry_identity_before=proposal_before,
        proposal_registry_identity_after=proposal_before,
    )


def _validate_dry_run_binding(
    packet: R103ReviewPacket, registry: R103DecisionRegistry
) -> None:
    if (
        registry.packet_identity != packet.packet_identity
        or registry.source_identity != packet.source_identity
        or registry.source_release != packet.source_release
    ):
        raise R103ReviewValidationError("decision registry does not bind packet source")


def _dry_run_input_identities(
    oracle_path: Path, proposal_registry_path: Path
) -> tuple[str, str]:
    oracle_before = _file_sha256(oracle_path)
    proposal_before = _file_sha256(proposal_registry_path)
    load_proposal_registry(proposal_registry_path)
    oracle_after = _file_sha256(oracle_path)
    proposal_after = _file_sha256(proposal_registry_path)
    if oracle_after != oracle_before or proposal_after != proposal_before:
        raise R103ReviewValidationError("dry-run input changed during evaluation")
    return oracle_before, proposal_before


def write_r103_review_dry_run(path: Path, result: R103ReviewDryRun) -> None:
    _write_json(path, result.model_dump(mode="json"))
