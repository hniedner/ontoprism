"""Durable promotion of the complete, write-free R103 local-SME review state."""

from __future__ import annotations

import json
from typing import TYPE_CHECKING, Literal, Self

from openpyxl import load_workbook
from pydantic import Field, model_validator
from pydantic_core import to_jsonable_python

from ontolib.decomposition.proposal_registry import load_proposal_registry
from ontolib.decomposition.r103_review import (
    R103Decision,
    R103DecisionRegistry,
    R103ReviewDryRun,
    R103ReviewPacket,
    R103ReviewValidationError,
    _canonical,
    _file_sha256,
    _identity,
    _load_json,
    _StrictModel,
    _validate_dry_run_binding,
    _validate_workbook_structure,
    _write_json,
    dry_run_r103_review,
    import_r103_review_decisions,
    load_r103_decision_registry,
    load_r103_review_packet,
    write_r103_review_dry_run,
    write_r103_review_workbook,
)

if TYPE_CHECKING:
    from pathlib import Path
    from typing import Any

_SHA256 = r"^[0-9a-f]{64}$"
_EXPECTED_REVIEWS = (
    (
        "source-supported",
        "C12950 is supported by C2860’s stated derivation from adrenal embryonic "  # noqa: RUF001
        "rest cells. It is anatomically broad but is the most specific currently "
        "available NCIt tissue-origin filler; no range-compatible, more specific "
        "NCIt replacement has been identified.",
        "R. Hannes Niedner, M.D.",
        "2026-08-26",
    ),
    (
        "review-required",
        "Fetal-tissue resemblance describes morphology, not necessarily normal "
        "tissue of origin, so the R103 assertion should not be projected without "
        "stronger evidence.",
        "R. Hannes Niedner, M.D.",
        "2026-08-26",
    ),
    (
        "source-supported",
        "C3716 explicitly states origin in neuroectoderm, and C34228 defines that "
        "embryologic tissue; the R103 assertion directly represents the stated "
        "normal tissue of origin while R104 separately preserves cell-origin context.",
        "R. Hannes Niedner, M.D.",
        "2026-08-26",
    ),
)
_REVISION_ASSERTION = ("C3264", "R103", "C12950")
_REVISION_OUTCOME = "concept-scoped-accuracy-exclusion"
_REVISION_REVIEWER = "R. Hannes Niedner, M.D."
_REVISION_DATE = "2026-08-28"
_REVISION_RATIONALE = (
    "My Recommendation: Concept-scoped exclusion\n\n"
    "Rationale from Scientific Literature: The modern understanding of CNS and "
    "peripheral embryonal tumors (such as medulloblastoma, atypical "
    "teratoid/rhabdoid tumor, and ETMR) has shifted significantly with molecular "
    "profiling. Literature confirms that these tumors arise from the transformation "
    "of very specific local progenitor populations or stem cells (e.g., transitional "
    "cerebellar progenitors in the rhombic lip for certain medulloblastomas) whose "
    'developmental program stalls—not from a generalized pool of "embryonic tissue."'
    '\n\nThe term "embryonal" in oncology refers primarily to the primitive, '
    "undifferentiated morphologic appearance of the tumor cells (small round blue "
    "cells resembling those in a developing embryo) rather than a literal derivation "
    "from generic embryonic tissue. Because R103 represents a strict causal origin "
    '(Disease_Has_Normal_Tissue_Origin), applying a broad "Embryonic Tissue" filler '
    "across this entire umbrella misrepresents the biology.\n\n"
    "Therefore, selecting Concept-scoped exclusion is the correct semantic action. "
    "It safely prevents this morphologic resemblance from being falsely projected as "
    "a strict anatomic origin in downstream reasoning, while fully preserving the "
    "original NCIt source assertion and its provenance in the gr" + "aph."
)
R103_REVISION_MACHINE_QUALIFICATION = (
    "R103 is non-defining; this exclusion applies exactly to the C3264/R103/C12950 "
    "source assertion, and individual descendants may have specific embryonic or "
    "fetal origins."
)
_REVISION_BINDING_AUTHORITY = "explicit-human-instruction"
_CORROBORATION_CITATIONS = (
    (
        "Gibson P",
        "Subtypes of medulloblastoma have distinct developmental origins.",
        "10.1038/nature09587",
        "21150899",
    ),
    (
        "Vladoiu MC",
        "Childhood cerebellar tumours mirror conserved fetal transcriptional programs.",
        "10.1038/s41586-019-1158-7",
        "31043743",
    ),
    (
        "Meredith DM",
        "Embryonal and non-meningothelial mesenchymal tumors of the central nervous "
        "system - Advances in diagnosis and prognostication.",
        "10.1111/bpa.13059",
        "35266242",
    ),
    (
        "Zeineldin M",
        "Neuroblastoma: When differentiation goes awry.",
        "10.1016/j.neuron.2022.07.012",
        "35985323",
    ),
    (
        "Li H",
        "Embryonic Kidney Development, Stem Cells and the Origin of Wilms Tumor.",
        "10.3390/genes12020318",
        "33672414",
    ),
)


class R103PromotedReviewState(_StrictModel):
    """Complete durable local-SME state with file and semantic bindings."""

    schema_version: Literal[1]
    packet: R103ReviewPacket
    registry: R103DecisionRegistry
    dry_run: R103ReviewDryRun
    oracle_file_sha256: str = Field(pattern=_SHA256)
    proposal_registry_identity: str = Field(pattern=_SHA256)
    proposal_registry_file_sha256: str = Field(pattern=_SHA256)
    artifact_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_promoted_state(self) -> Self:
        _validate_promoted_cross_bindings(self)
        expected = _identity(self.model_dump(exclude={"artifact_identity"}))
        if self.artifact_identity != expected:
            raise ValueError("promoted artifact identity differs")
        return self


def _validate_promoted_cross_bindings(state: R103PromotedReviewState) -> None:
    _validate_registry_bindings(state.packet, state.registry)
    _validate_review_values(state.packet, state.registry)
    _validate_dry_run_state(state)
    if state.proposal_registry_identity != state.packet.proposal_registry_identity:
        raise ValueError("proposal registry semantic binding differs")


def _validate_registry_bindings(
    packet: R103ReviewPacket, registry: R103DecisionRegistry
) -> None:
    _validate_dry_run_binding(packet, registry)
    if registry.workbook_identity != registry.decisions[0].workbook_identity:
        raise ValueError("decision registry workbook or source binding differs")
    for decision in registry.decisions:
        _validate_decision_binding(packet, registry, decision)
    packet_rows = tuple(
        (row.row_identity, row.subject_code, row.role_code, row.filler_code)
        for row in packet.rows
    )
    decision_rows = tuple(
        (row.row_identity, row.subject_code, row.role_code, row.filler_code)
        for row in registry.decisions
    )
    if decision_rows != packet_rows:
        raise ValueError("packet and decision row joins differ")


def _validate_decision_binding(
    packet: R103ReviewPacket,
    registry: R103DecisionRegistry,
    decision: R103Decision,
) -> None:
    observed = (
        decision.workbook_identity,
        decision.packet_identity,
        decision.source_identity,
        decision.source_release,
    )
    expected = (
        registry.workbook_identity,
        packet.packet_identity,
        packet.source_identity,
        packet.source_release,
    )
    if observed != expected:
        raise ValueError("decision registry workbook or source binding differs")


def _validate_review_values(
    packet: R103ReviewPacket, registry: R103DecisionRegistry
) -> None:
    human_rows = tuple(
        (row.outcome, row.rationale, row.reviewer, row.review_date)
        for row in registry.decisions
    )
    if human_rows != _EXPECTED_REVIEWS:
        raise ValueError("promoted human review values differ")
    expected_workbook_identity = _identity(
        {"packet_identity": packet.packet_identity, "human_rows": human_rows}
    )
    if registry.workbook_identity != expected_workbook_identity:
        raise ValueError("promoted workbook identity differs")
    outcomes = tuple((row.subject_code, row.outcome) for row in registry.decisions)
    if outcomes != (
        ("C2860", "source-supported"),
        ("C3264", "review-required"),
        ("C3716", "source-supported"),
    ):
        raise ValueError("promoted decision vector differs")


def _validate_dry_run_state(state: R103PromotedReviewState) -> None:
    dry_run = state.dry_run
    observed = (
        dry_run.writes_performed,
        dry_run.outcome_counts,
        dry_run.proposal_previews,
        dry_run.exclusion_previews,
        dry_run.unresolved,
        dry_run.readiness,
        dry_run.oracle_identity_before,
        dry_run.oracle_identity_after,
        dry_run.proposal_registry_identity_before,
        dry_run.proposal_registry_identity_after,
    )
    expected = (
        False,
        {"source-supported": 2, "review-required": 1},
        (),
        (),
        1,
        "review-incomplete",
        state.oracle_file_sha256,
        state.oracle_file_sha256,
        state.proposal_registry_file_sha256,
        state.proposal_registry_file_sha256,
    )
    if observed != expected:
        raise ValueError("promoted dry-run boundary differs")


def _load_r103_review_dry_run(path: Path) -> R103ReviewDryRun:
    try:
        return R103ReviewDryRun.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def load_r103_promoted_review_state(path: Path) -> R103PromotedReviewState:
    """Parse a strict promotion without reading its vanished staging inputs."""
    try:
        return R103PromotedReviewState.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def _render_json(value: object) -> bytes:
    return (
        json.dumps(
            to_jsonable_python(value), sort_keys=True, indent=2, ensure_ascii=True
        ).encode("ascii")
        + b"\n"
    )


def _load_bound_proposal_registry(path: Path, packet: R103ReviewPacket):
    try:
        proposal_registry = load_proposal_registry(path)
    except (OSError, UnicodeDecodeError, json.JSONDecodeError, ValueError) as error:
        raise R103ReviewValidationError("invalid proposal registry") from error
    if (
        proposal_registry.registry_identity != packet.proposal_registry_identity
        or proposal_registry.ontology_version != packet.source_release
    ):
        raise R103ReviewValidationError(
            "proposal registry semantic binding differs from packet"
        )
    return proposal_registry


def _require_recomputed_dry_run(
    packet: R103ReviewPacket,
    registry: R103DecisionRegistry,
    persisted: R103ReviewDryRun,
    oracle_path: Path,
    proposal_registry_path: Path,
) -> None:
    recomputed = dry_run_r103_review(
        packet,
        registry,
        oracle_path=oracle_path,
        proposal_registry_path=proposal_registry_path,
    )
    if persisted != recomputed:
        raise R103ReviewValidationError("persisted dry-run differs from recomputation")


def _promoted_model(payload: dict[str, object]) -> R103PromotedReviewState:
    try:
        return R103PromotedReviewState.model_validate(
            {**payload, "artifact_identity": _identity(payload)}
        )
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def _write_absent_or_equal(
    output_path: Path, promoted: R103PromotedReviewState
) -> None:
    content = _render_json(promoted.model_dump(mode="json"))
    if not output_path.exists():
        _write_json(output_path, promoted.model_dump(mode="json"))
        return
    try:
        existing = output_path.read_bytes()
    except OSError as error:
        raise R103ReviewValidationError("promotion output cannot be read") from error
    if existing != content:
        raise R103ReviewValidationError("promotion output conflict")


def promote_r103_review_state(
    *,
    packet_path: Path,
    registry_path: Path,
    dry_run_path: Path,
    oracle_path: Path,
    proposal_registry_path: Path,
    output_path: Path,
) -> R103PromotedReviewState:
    """Validate and compose the current incomplete, write-free review state."""
    packet = load_r103_review_packet(packet_path)
    registry = load_r103_decision_registry(registry_path)
    persisted_dry_run = _load_r103_review_dry_run(dry_run_path)
    proposal_registry = _load_bound_proposal_registry(proposal_registry_path, packet)
    oracle_sha256 = _file_sha256(oracle_path)
    proposal_file_sha256 = _file_sha256(proposal_registry_path)
    _require_recomputed_dry_run(
        packet,
        registry,
        persisted_dry_run,
        oracle_path=oracle_path,
        proposal_registry_path=proposal_registry_path,
    )
    payload: dict[str, object] = {
        "schema_version": 1,
        "packet": packet,
        "registry": registry,
        "dry_run": persisted_dry_run,
        "oracle_file_sha256": oracle_sha256,
        "proposal_registry_identity": proposal_registry.registry_identity,
        "proposal_registry_file_sha256": proposal_file_sha256,
    }
    promoted = _promoted_model(payload)
    _write_absent_or_equal(output_path, promoted)
    return promoted


class R103RevisionTranscription(_StrictModel):
    actor: Literal["software-transcriber"]
    authority: Literal["explicit-human-instruction"]
    authorship_claimed: Literal[False]
    assertion: tuple[Literal["C3264"], Literal["R103"], Literal["C12950"]]
    workbook_identity: str = Field(pattern=_SHA256)


class R103PromotedReviewRevision(_StrictModel):
    """Append-only terminal revision of the historical promoted review state."""

    schema_version: Literal[2]
    predecessor: R103PromotedReviewState
    predecessor_artifact_identity: str = Field(pattern=_SHA256)
    registry: R103DecisionRegistry
    dry_run: R103ReviewDryRun
    oracle_file_sha256: str = Field(pattern=_SHA256)
    proposal_registry_identity: str = Field(pattern=_SHA256)
    proposal_registry_file_sha256: str = Field(pattern=_SHA256)
    transcription: R103RevisionTranscription
    machine_qualification: str = Field(min_length=1)
    artifact_identity: str = Field(pattern=_SHA256)

    @property
    def packet(self) -> R103ReviewPacket:
        return self.predecessor.packet

    @model_validator(mode="after")
    def _validate_revision(self) -> Self:
        if self.predecessor_artifact_identity != self.predecessor.artifact_identity:
            raise ValueError("revision predecessor identity differs")
        _validate_registry_bindings(self.packet, self.registry)
        _validate_revision_values(self.registry)
        _validate_revision_dry_run(self)
        if self.proposal_registry_identity != self.packet.proposal_registry_identity:
            raise ValueError("revision proposal registry semantic binding differs")
        if self.transcription.workbook_identity != self.registry.workbook_identity:
            raise ValueError("revision transcription workbook binding differs")
        if self.machine_qualification != R103_REVISION_MACHINE_QUALIFICATION:
            raise ValueError("revision machine qualification differs")
        expected = _identity(self.model_dump(exclude={"artifact_identity"}))
        if self.artifact_identity != expected:
            raise ValueError("revision artifact identity differs")
        return self


def _validate_revision_values(registry: R103DecisionRegistry) -> None:
    expected = (
        _EXPECTED_REVIEWS[0],
        (
            _REVISION_OUTCOME,
            _REVISION_RATIONALE,
            _REVISION_REVIEWER,
            _REVISION_DATE,
        ),
        _EXPECTED_REVIEWS[2],
    )
    observed = tuple(
        (row.outcome, row.rationale, row.reviewer, row.review_date)
        for row in registry.decisions
    )
    if observed != expected:
        raise ValueError("revision human review values differ")
    _validate_revision_vector(registry)


def _validate_revision_vector(registry: R103DecisionRegistry) -> None:
    outcomes = tuple((row.subject_code, row.outcome) for row in registry.decisions)
    if outcomes != (
        ("C2860", "source-supported"),
        ("C3264", _REVISION_OUTCOME),
        ("C3716", "source-supported"),
    ):
        raise ValueError("revision decision vector differs")
    if registry.proposal_preview or tuple(
        (item.subject_code, item.role_code, item.filler_code)
        for item in registry.exclusion_preview
    ) != (_REVISION_ASSERTION,):
        raise ValueError("revision preview state differs")


def _validate_revision_dry_run(state: R103PromotedReviewRevision) -> None:
    dry_run = state.dry_run
    observed = (
        dry_run.writes_performed,
        dry_run.outcome_counts,
        dry_run.proposal_previews,
        dry_run.exclusion_previews,
        dry_run.unresolved,
        dry_run.readiness,
        dry_run.oracle_identity_before,
        dry_run.oracle_identity_after,
        dry_run.proposal_registry_identity_before,
        dry_run.proposal_registry_identity_after,
    )
    expected = (
        False,
        {"source-supported": 2, _REVISION_OUTCOME: 1},
        (),
        state.registry.exclusion_preview,
        0,
        "ready-for-separate-application",
        state.oracle_file_sha256,
        state.oracle_file_sha256,
        state.proposal_registry_file_sha256,
        state.proposal_registry_file_sha256,
    )
    if observed != expected:
        raise ValueError("revision dry-run boundary differs")


def prepare_r103_review_revision(
    *, predecessor_path: Path, output_workbook_path: Path
) -> None:
    """Reconstruct a blank revision workbook from the predecessor's embedded packet."""
    predecessor = load_r103_promoted_review_state(predecessor_path)
    write_r103_review_workbook(output_workbook_path, predecessor.packet)


def transcribe_r103_review_revision(
    *,
    predecessor_path: Path,
    blank_workbook_path: Path,
    output_workbook_path: Path,
    assertion: tuple[str, str, str],
    outcome: str,
    rationale: str,
    reviewer: str,
    review_date: str,
) -> None:
    """Transcribe explicit human values without asserting software authorship."""
    predecessor = load_r103_promoted_review_state(predecessor_path)
    _require_explicit_revision_values(
        assertion, outcome, rationale, reviewer, review_date
    )
    book = _load_revision_workbook(blank_workbook_path)
    review = _validate_workbook_structure(book, predecessor.packet)
    human_start = review.max_column - 3
    _require_blank_human_fields(review, human_start)
    _fill_effective_human_values(review, human_start, predecessor)
    for column, value in enumerate(
        (outcome, rationale, reviewer, review_date), start=human_start
    ):
        review.cell(3, column, value)
    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_workbook_path)


def _require_explicit_revision_values(
    assertion: tuple[str, str, str],
    outcome: str,
    rationale: str,
    reviewer: str,
    review_date: str,
) -> None:
    observed = (assertion, outcome, rationale, reviewer, review_date)
    expected = (
        _REVISION_ASSERTION,
        _REVISION_OUTCOME,
        _REVISION_RATIONALE,
        _REVISION_REVIEWER,
        _REVISION_DATE,
    )
    if observed != expected:
        raise R103ReviewValidationError("explicit revision values differ")


def _load_revision_workbook(path: Path) -> Any:
    try:
        return load_workbook(path, data_only=False, keep_links=False)
    except Exception as error:
        raise R103ReviewValidationError("invalid review workbook") from error


def _require_blank_human_fields(review: Any, human_start: int) -> None:
    if any(
        review.cell(row, column).value is not None
        for row in range(2, 5)
        for column in range(human_start, review.max_column + 1)
    ):
        raise R103ReviewValidationError("revision input workbook is not blank")


def _fill_effective_human_values(
    review: Any, human_start: int, predecessor: R103PromotedReviewState
) -> None:
    for row_number, decision in enumerate(predecessor.registry.decisions, start=2):
        values = (
            decision.outcome,
            decision.rationale,
            decision.reviewer,
            decision.review_date,
        )
        for column, value in enumerate(values, start=human_start):
            review.cell(row_number, column, value)


def _revision_model(payload: dict[str, object]) -> R103PromotedReviewRevision:
    try:
        return R103PromotedReviewRevision.model_validate(
            {**payload, "artifact_identity": _identity(payload)}
        )
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def load_r103_promoted_review_revision(path: Path) -> R103PromotedReviewRevision:
    """Strictly parse only the terminal schema-v2 R103 revision."""
    try:
        return R103PromotedReviewRevision.model_validate_json(
            _canonical(_load_json(path))
        )
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error


def promote_r103_review_revision(
    *,
    predecessor_path: Path,
    reviewed_workbook_path: Path,
    oracle_path: Path,
    proposal_registry_path: Path,
    qualification: str,
    output_registry_path: Path,
    output_dry_run_path: Path,
    output_path: Path,
) -> R103PromotedReviewRevision:
    """Import and append a terminal, write-free revision from its governed workbook."""
    predecessor = load_r103_promoted_review_state(predecessor_path)
    if qualification != R103_REVISION_MACHINE_QUALIFICATION:
        raise R103ReviewValidationError("revision machine qualification differs")
    registry = import_r103_review_decisions(
        predecessor.packet,
        reviewed_workbook_path,
        output_registry_path,
    )
    dry_run = dry_run_r103_review(
        predecessor.packet,
        registry,
        oracle_path=oracle_path,
        proposal_registry_path=proposal_registry_path,
    )
    write_r103_review_dry_run(output_dry_run_path, dry_run)
    proposal_registry = _load_bound_proposal_registry(
        proposal_registry_path, predecessor.packet
    )
    payload: dict[str, object] = {
        "schema_version": 2,
        "predecessor": predecessor,
        "predecessor_artifact_identity": predecessor.artifact_identity,
        "registry": registry,
        "dry_run": dry_run,
        "oracle_file_sha256": _file_sha256(oracle_path),
        "proposal_registry_identity": proposal_registry.registry_identity,
        "proposal_registry_file_sha256": _file_sha256(proposal_registry_path),
        "transcription": {
            "actor": "software-transcriber",
            "authority": _REVISION_BINDING_AUTHORITY,
            "authorship_claimed": False,
            "assertion": _REVISION_ASSERTION,
            "workbook_identity": registry.workbook_identity,
        },
        "machine_qualification": qualification,
    }
    revision = _revision_model(payload)
    content = _render_json(revision.model_dump(mode="json"))
    if output_path.exists() and output_path.read_bytes() != content:
        raise R103ReviewValidationError("revision promotion output conflict")
    if not output_path.exists():
        _write_json(output_path, revision.model_dump(mode="json"))
    return revision


class R103CorroborationCitation(_StrictModel):
    first_author: str = Field(min_length=1)
    title: str = Field(min_length=1)
    doi: str = Field(min_length=1)
    pmid: str = Field(pattern=r"^[0-9]+$")


class R103DecisionCorroboration(_StrictModel):
    schema_version: Literal[1]
    effective_decision_identity: str = Field(pattern=_SHA256)
    relationship: Literal["corroboration-not-proof"]
    scope_qualification: str = Field(min_length=1)
    authoritative_metadata_source: Literal["NCBI PubMed ESummary"]
    verified_date: Literal["2026-08-28"]
    citations: tuple[
        R103CorroborationCitation,
        R103CorroborationCitation,
        R103CorroborationCitation,
        R103CorroborationCitation,
        R103CorroborationCitation,
    ]
    corroboration_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_corroboration(self) -> Self:
        observed = tuple(
            (item.first_author, item.title, item.doi, item.pmid)
            for item in self.citations
        )
        if observed != _CORROBORATION_CITATIONS:
            raise ValueError("corroboration citation metadata differs")
        if self.scope_qualification != R103_REVISION_MACHINE_QUALIFICATION:
            raise ValueError("corroboration scope qualification differs")
        expected = _identity(self.model_dump(exclude={"corroboration_identity"}))
        if self.corroboration_identity != expected:
            raise ValueError("corroboration identity differs")
        return self


def build_r103_corroboration(
    revision: R103PromotedReviewRevision,
) -> R103DecisionCorroboration:
    payload: dict[str, object] = {
        "schema_version": 1,
        "effective_decision_identity": revision.registry.decisions[1].decision_identity,
        "relationship": "corroboration-not-proof",
        "scope_qualification": R103_REVISION_MACHINE_QUALIFICATION,
        "authoritative_metadata_source": "NCBI PubMed ESummary",
        "verified_date": "2026-08-28",
        "citations": tuple(
            {
                "first_author": first_author,
                "title": title,
                "doi": doi,
                "pmid": pmid,
            }
            for first_author, title, doi, pmid in _CORROBORATION_CITATIONS
        ),
    }
    return R103DecisionCorroboration.model_validate(
        {**payload, "corroboration_identity": _identity(payload)}
    )


def write_r103_corroboration(
    path: Path, corroboration: R103DecisionCorroboration
) -> None:
    _write_json(path, corroboration.model_dump(mode="json"))


def load_r103_corroboration(
    path: Path, *, revision: R103PromotedReviewRevision
) -> R103DecisionCorroboration:
    try:
        value = R103DecisionCorroboration.model_validate_json(
            _canonical(_load_json(path))
        )
    except ValueError as error:
        raise R103ReviewValidationError(str(error)) from error
    if (
        value.effective_decision_identity
        != revision.registry.decisions[1].decision_identity
    ):
        raise R103ReviewValidationError("corroboration decision binding differs")
    return value
