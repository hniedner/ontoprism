"""Human-centered, packet-bound review of R101 projection coverage."""

# ruff: noqa: E501 - reviewer-facing sentences are deliberately kept whole.

from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from collections import Counter, defaultdict
from contextlib import suppress
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Literal, Protocol, Self, cast
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, ConfigDict, Field, model_validator
from pydantic_core import to_jsonable_python

from ontolib.terminologies.ncit.owl_load import STATED_GRAPH_IRI
from ontolib.terminologies.ncit.sibling_store import validate_ncit_sibling_manifest
from ontolib.terminologies.sparql_transport import safe_iri

if TYPE_CHECKING:
    from collections.abc import Collection, Iterable

    from openpyxl.cell import Cell, MergedCell
    from openpyxl.workbook.workbook import Workbook as WorkbookType
    from openpyxl.worksheet.worksheet import Worksheet

    from ontolib.decomposition.r101_conservation import (
        LedgerOccurrence,
        R101ConservationReport,
    )

_SHA256 = r"^[0-9a-f]{64}$"
_CODE = r"^C[0-9]+$"
_NCIT = "http://ncicb.nci.nih.gov/xml/owl/EVS/Thesaurus.owl#"
_SCOPE = "non-exclusive projection coverage"
_SCHEMA_VERSION = 3
_EXPECTED_PATTERNS = 162
_EXPECTED_DISEASE_PROPOSITIONS = 2800
_EXPECTED_OCCURRENCES = 3291
_LABEL_BATCH_SIZE = 500
_SMALL_PATTERN_DISEASE_LIMIT = 5
_SMALL_PATTERN_TEXT_LIMIT = 900
_EXCEL_SAFE_TEXT_LIMIT = 30_000
_DENIAL_VALUES: dict[str, bool] = {
    "scope_non_exclusive": True,
    "source_preserved": True,
    "not_equivalent": True,
    "not_universal": True,
    "not_exclusive": True,
}

APPROVE = "Approve non-exclusive coverage except marked exceptions"
REJECT = "Reject; retain broader site in projection"
INDIVIDUAL = "Require individual disease review"
ABSTAIN = "Abstain / escalate"
ReviewDecision = Literal[
    "Approve non-exclusive coverage except marked exceptions",
    "Reject; retain broader site in projection",
    "Require individual disease review",
    "Abstain / escalate",
]
_DECISIONS = (APPROVE, REJECT, INDIVIDUAL, ABSTAIN)

_PATTERN_HEADERS = (
    "Pattern Number",
    "Review Proposition",
    "Broader Site",
    "Retained More-Specific Site",
    "Human-readable R82 path(s)",
    "Affected Disease Count",
    "Affected Diseases",
    "Source Occurrence Count",
    "One-step Count",
    "Transitive Count",
    "Min Path Length",
    "Max Path Length",
    "Context/Risk Summary",
    "Fixed Scope",
    "Decision",
    "Rationale",
    "Reviewer Identity",
    "Review Date",
)
_DISEASE_HEADERS = (
    "Pattern Number",
    "Disease",
    "Specific Proposition",
    "Broader Site",
    "Retained Site",
    "Readable R82 path",
    "Source Occurrence Count",
    "Context Summary",
    "Review Priority / Risk Flags",
    "Exception?",
    "Exception Rationale",
)
_PATTERN_EDITABLE = frozenset(
    {"Decision", "Rationale", "Reviewer Identity", "Review Date"}
)
_DISEASE_EDITABLE = frozenset({"Exception?", "Exception Rationale"})
_BINDING_NAMES = (
    "packet_identity",
    "guidance_identity",
    "visible_rows_identity",
    "membership_identity",
    "schema_version",
    "source_release_id",
)


class R101ReviewValidationError(ValueError):
    """The review artifact cannot support a packet-bound human decision."""


class _StrictModel(BaseModel):
    model_config = ConfigDict(frozen=True, extra="forbid", strict=True)


def _canonical(value: object) -> bytes:
    return json.dumps(
        to_jsonable_python(value),
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    ).encode("ascii")


def _identity(value: object) -> str:
    return hashlib.sha256(_canonical(value)).hexdigest()


class ReviewBindings(_StrictModel):
    report_identity: str = Field(pattern=_SHA256)
    json_identity: str = Field(pattern=_SHA256)
    tsv_identity: str = Field(pattern=_SHA256)
    source_identity: str = Field(pattern=_SHA256)
    source_release_id: str = Field(min_length=1)
    old_run_id: str = Field(min_length=1)
    old_run_fingerprint_identity: str = Field(pattern=_SHA256)
    old_representation_identity: str = Field(pattern=_SHA256)
    old_baseline_identity: str = Field(pattern=_SHA256)
    new_run_id: str = Field(min_length=1)
    new_run_fingerprint_identity: str = Field(pattern=_SHA256)
    new_representation_identity: str = Field(pattern=_SHA256)
    detector_identity: str = Field(pattern=_SHA256)
    pre_resume_proof_identity: str = Field(pattern=_SHA256)
    resume_dry_run_identity: str = Field(pattern=_SHA256)
    mixed_cohort_identity: str = Field(pattern=_SHA256)
    proof_identity: str = Field(pattern=_SHA256)


class DenialFlags(_StrictModel):
    scope_non_exclusive: Literal[True] = True
    source_preserved: Literal[True] = True
    not_equivalent: Literal[True] = True
    not_universal: Literal[True] = True
    not_exclusive: Literal[True] = True


class EvidenceKindCounts(_StrictModel):
    one_step: int = Field(ge=0)
    transitive: int = Field(ge=0)


class ReviewPath(_StrictModel):
    path_identity: str = Field(pattern=_SHA256)
    code_path: tuple[str, ...] = Field(min_length=2)
    labels: tuple[str, ...] = Field(min_length=2)
    fact_identities: tuple[str, ...] = Field(min_length=1)
    source_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_shape(self) -> Self:
        if len(self.labels) != len(self.code_path):
            raise ValueError("path labels do not match code path")
        if len(self.fact_identities) + 1 != len(self.code_path):
            raise ValueError("path facts do not match code path")
        if self.path_identity != _identity(self.model_dump(exclude={"path_identity"})):
            raise ValueError("path identity differs")
        return self


class ReviewPattern(DenialFlags):
    pattern_id: str = Field(pattern=r"^r101-[0-9a-f]{16}$")
    row_identity: str = Field(pattern=_SHA256)
    pattern_number: int = Field(ge=1)
    review_proposition: str = Field(min_length=1)
    broader_code: str = Field(pattern=_CODE)
    broader_label: str = Field(min_length=1)
    retained_code: str = Field(pattern=_CODE)
    retained_label: str = Field(min_length=1)
    axis: Literal["op:PrimarySite"]
    occurrence_count: int = Field(gt=0)
    evidence_kind_counts: EvidenceKindCounts
    min_path_length: int = Field(ge=1, le=8)
    max_path_length: int = Field(ge=1, le=8)
    paths: tuple[ReviewPath, ...] = Field(min_length=1)
    disease_codes: tuple[str, ...] = Field(min_length=1)
    context_risk_summary: str = Field(min_length=1)
    fixed_scope: Literal["non-exclusive projection coverage"] = _SCOPE

    @model_validator(mode="after")
    def _validate_counts_and_identity(self) -> Self:
        if (
            self.occurrence_count
            != self.evidence_kind_counts.one_step + self.evidence_kind_counts.transitive
        ):
            raise ValueError("pattern evidence counts differ")
        if len(self.disease_codes) != len(set(self.disease_codes)):
            raise ValueError("pattern disease membership is not unique")
        if self.row_identity != _identity(self.model_dump(exclude={"row_identity"})):
            raise ValueError("pattern row identity differs")
        return self


class DiseaseProposition(DenialFlags):
    proposition_identity: str = Field(pattern=_SHA256)
    pattern_id: str = Field(pattern=r"^r101-[0-9a-f]{16}$")
    pattern_number: int = Field(ge=1)
    disease_code: str = Field(pattern=_CODE)
    disease_label: str = Field(min_length=1)
    specific_proposition: str = Field(min_length=1)
    broader_code: str = Field(pattern=_CODE)
    broader_label: str = Field(min_length=1)
    retained_code: str = Field(pattern=_CODE)
    retained_label: str = Field(min_length=1)
    paths: tuple[ReviewPath, ...] = Field(min_length=1)
    occurrence_count: int = Field(ge=1, le=3)
    context_summary: str = Field(min_length=1)
    risk_flags: str = Field(min_length=1)

    @model_validator(mode="after")
    def _validate_identity(self) -> Self:
        if self.proposition_identity != _identity(
            self.model_dump(exclude={"proposition_identity"})
        ):
            raise ValueError("disease proposition identity differs")
        return self


class ReviewOccurrence(_StrictModel):
    pattern_id: str = Field(pattern=r"^r101-[0-9a-f]{16}$")
    pattern_number: int = Field(ge=1)
    disease_code: str = Field(pattern=_CODE)
    occurrence_id: str = Field(pattern=_SHA256)
    source_fact_id: str = Field(pattern=_SHA256)
    source_group_id: str = Field(pattern=_SHA256)
    anchor_code: str = Field(pattern=_CODE)
    depth: int = Field(ge=0)
    structural_path: tuple[int, ...] = Field(min_length=1)
    member_position: int = Field(ge=0)
    evidence_kind: Literal["one-step", "closure-only"]
    path_length: int = Field(ge=1, le=8)
    path_identity: str = Field(pattern=_SHA256)
    broader_code: str = Field(pattern=_CODE)
    retained_code: str = Field(pattern=_CODE)


class ReviewMembership(_StrictModel):
    membership_identity: str = Field(pattern=_SHA256)
    pattern_id: str = Field(pattern=r"^r101-[0-9a-f]{16}$")
    pattern_number: int = Field(ge=1)
    disease_code: str = Field(pattern=_CODE)
    occurrence_ids: tuple[str, ...] = Field(min_length=1, max_length=3)

    @model_validator(mode="after")
    def _validate_identity(self) -> Self:
        if self.membership_identity != _identity(
            self.model_dump(exclude={"membership_identity"})
        ):
            raise ValueError("membership row identity differs")
        return self


def _path_text(path: ReviewPath) -> str:
    return " → ".join(
        f"{label} ({code})"
        for code, label in zip(path.code_path, path.labels, strict=True)
    )


def _paths_text(paths: Iterable[ReviewPath]) -> str:
    return "\n".join(_path_text(path) for path in paths)


def _site(label: str, code: str) -> str:
    return f"{label} ({code})"


def _specific_proposition(disease: str, retained: str, broader: str) -> str:
    return (
        f"For {disease}, {retained} is valid, more-precise primary-site coverage of {broader} "
        "in the curated projection. The broader source assertions remain preserved. "
        f"This does not mean every case occurs only in {retained} or that {retained} is the only valid site."
    )


def _pattern_proposition(
    diseases: tuple[str, ...], pattern_number: int, retained: str, broader: str
) -> str:
    subject = (
        "For " + " and ".join(diseases)
        if len(diseases) <= _SMALL_PATTERN_DISEASE_LIMIT
        and len(" and ".join(diseases)) <= _SMALL_PATTERN_TEXT_LIMIT
        else f"For the {len(diseases)} diseases listed under Pattern {pattern_number}"
    )
    return (
        f"{subject}, {retained} is valid, more-precise primary-site coverage of {broader} "
        "in the curated projection. The broader source assertions remain preserved. "
        f"This does not mean every case occurs only in {retained} or that {retained} is the only valid site."
    )


def _instruction_rows(source_release: str) -> tuple[tuple[str, str], ...]:
    return (
        ("R101 projection-coverage review", "Read this sheet before reviewing."),
        (
            "Purpose",
            "Review whether each retained more-specific primary site provides non-exclusive projection coverage of the omitted broader site for the exact listed disease and source occurrences.",
        ),
        (
            "What approval means",
            "The reviewer is approving projection coverage, not disease exclusivity. Approval is limited to non-exclusive projection coverage in this curated projection for the frozen packet membership.",
        ),
        (
            "What approval does not mean",
            "Approval does not assert equivalence, universality, completeness, exclusivity, that every case occurs only at the retained site, or that it is the only valid site. Multiple valid narrower sites remain independent.",
        ),
        (
            "Source preservation",
            "The broader source assertions remain preserved. Review changes a curated projection policy only; it does not delete or alter source assertions.",
        ),
        (
            "How generated",
            f"The mechanically validated NCIt {source_release} conservation report supplied 3,291 exact R101 source occurrences. They are frozen into 162 endpoint patterns and 2,800 disease propositions; software generated no human decisions.",
        ),
        (
            "Review procedure",
            "Review Pattern Review first. Inspect the plain proposition, both sites, every readable R82 path, counts, affected diseases, and risk summary. Use Disease Propositions for disease-specific context. Choose one closed decision and enter rationale, reviewer identity, and date.",
        ),
        (
            "Pattern versus atomic scope",
            "The primary human decision concerns an endpoint pattern. Import expands it to every frozen disease and exact source occurrence in the separate packet. Each disease row starts with a generated No scope default, which has no effect until its pattern is approved. A disease changed to Yes is excluded from an approved pattern only when its exception rationale is nonempty.",
        ),
        (
            "Decision meanings",
            f"{APPROVE}: approve the limited scope except explicit disease exceptions. {REJECT}: preserve broader-site projection for every member. {INDIVIDUAL}: create follow-up records, not approvals. {ABSTAIN}: create escalation records, not approvals.",
        ),
        (
            "Exception editing policy",
            "Every Disease Propositions row is generated with Exception? set to No and a blank rationale. Change only true exceptions to Yes and supply a disease-specific rationale. For an approved pattern, generated No or a justified Yes is required; a missing, invalid, or mismatched value refuses import. For every non-approve decision, all rows must remain No with blank rationale.",
        ),
        (
            "SEER / ICD-O pilot",
            "The SEER/ICD-O pilot found zero strict rule-eligible cases, so it supports no automation and no safe workload reduction. It is context only; there are no SEER decision fields in this workbook.",
        ),
        (
            "Error signs",
            "Stop and escalate for surprising labels, wrong path direction, clinically distinct components, mixed disease contexts, count disagreement, missing rows, or wording that suggests equivalence, universality, exclusivity, or source deletion.",
        ),
        (
            "Integrity boundary",
            "Hiddenness is not security. Bindings is veryHidden only to reduce accidental edits. Import regenerates and compares every immutable visible cell with the canonical packet; edited, stale, missing, duplicate, or extra rows refuse the whole import with no partial writes.",
        ),
        (
            "Saving",
            "Excel or openpyxl may rewrite container bytes. Formula-free cells and the packet binding, not byte-for-byte XLSX identity, define acceptance.",
        ),
    )


_DEFINITION_HEADERS = (
    "Sheet",
    "Header",
    "Plain-language definition",
    "Source / procedure",
    "Warning",
)


def _definition_rows() -> tuple[tuple[str, str, str, str, str], ...]:
    pattern_definitions = {
        "Pattern Number": "Human cross-reference numbered 1 through 162.",
        "Review Proposition": "Frozen plain-language scope proposed for this endpoint pattern.",
        "Broader Site": "Broader source site as release-bound Label (C…).",
        "Retained More-Specific Site": "More-specific retained site as release-bound Label (C…).",
        "Human-readable R82 path(s)": "Every distinct directed part-to-whole path in readable Label (C…) form.",
        "Affected Disease Count": "Number of frozen diseases represented by the pattern.",
        "Affected Diseases": "Readable disease list, or a concise list directing the reviewer to filtered disease rows.",
        "Source Occurrence Count": "Number of exact R101 source occurrences represented.",
        "One-step Count": "Occurrences supported by one stated R82 edge.",
        "Transitive Count": "Occurrences supported by a stated R82 path longer than one edge.",
        "Min Path Length": "Smallest edge count among represented paths.",
        "Max Path Length": "Largest edge count among represented paths.",
        "Context/Risk Summary": "Plain-language scale and path-risk cues; not a decision.",
        "Fixed Scope": "The only approval scope: non-exclusive projection coverage.",
        "Decision": "Required closed human decision for the whole frozen pattern.",
        "Rationale": "Required human explanation for the decision.",
        "Reviewer Identity": "Required accountable reviewer name or approved identifier.",
        "Review Date": "Required ISO date (YYYY-MM-DD).",
    }
    disease_definitions = {
        "Pattern Number": "Human cross-reference to Pattern Review.",
        "Disease": "Release-bound disease Label (C…).",
        "Specific Proposition": "Frozen proposition limited to this disease and packet membership.",
        "Broader Site": "Broader source site in Label (C…) form.",
        "Retained Site": "Retained more-specific site in Label (C…) form.",
        "Readable R82 path": "Distinct directed path or paths represented for this disease.",
        "Source Occurrence Count": "Exact source occurrence multiplicity, from 1 through 3.",
        "Context Summary": "Human-readable disease and site context.",
        "Review Priority / Risk Flags": "Plain-language cues for additional scrutiny.",
        "Exception?": "Generated as No for every disease. It has no effect until the pattern is approved; change only a true exception to Yes. Missing or other values refuse import, and non-approve patterns must remain No.",
        "Exception Rationale": "Required only when Exception? is changed to Yes; must remain blank for generated No and every non-approve pattern.",
    }
    rows: list[tuple[str, str, str, str, str]] = []
    for sheet, headers, definitions in (
        ("Pattern Review", _PATTERN_HEADERS, pattern_definitions),
        ("Disease Propositions", _DISEASE_HEADERS, disease_definitions),
    ):
        for header in headers:
            rows.append(
                (
                    sheet,
                    header,
                    definitions[header],
                    "Generated from and revalidated against the canonical packet.",
                    "Any disagreement, blank required decision, or semantic overclaim.",
                )
            )
    return tuple(rows)


def _example_rows() -> tuple[tuple[str, str, str, str], ...]:
    return (
        (
            "ILLUSTRATIVE ONLY — not packet evidence",
            "Situation",
            "Illustrative response",
            "Why",
        ),
        (
            "Approve",
            "A component is a valid more-precise projection site for every listed member while broader source assertions remain preserved.",
            APPROVE,
            "Coverage is non-exclusive and packet-limited.",
        ),
        (
            "Reject",
            "The broader site adds clinically important projection meaning that the retained component does not cover.",
            REJECT,
            "Retain the broader site in projection.",
        ),
        (
            "Disease exception",
            "One disease has a distinct context although the endpoint pattern is otherwise acceptable.",
            "Approve pattern; change only that disease's generated No to Yes and add rationale",
            "All other generated No values need no manual edit; only that disease's frozen occurrences are excluded from approval.",
        ),
        (
            "Exclusivity counterexample",
            "The reviewer believes the retained site is the only site where this disease can occur.",
            "Do not record that claim; abstain or escalate",
            "The workbook never asks for universal or exclusive disease claims.",
        ),
    )


def _guidance_payload(source_release: str) -> dict[str, object]:
    return {
        "instructions": _instruction_rows(source_release),
        "definitions": (_DEFINITION_HEADERS, *_definition_rows()),
        "examples": _example_rows(),
    }


def _bindings(report: R101ConservationReport) -> ReviewBindings:
    return ReviewBindings.model_validate(
        {name: getattr(report, name) for name in ReviewBindings.model_fields},
        strict=True,
    )


def _pattern_values(
    row: ReviewPattern, disease_labels: dict[str, str]
) -> tuple[str | int, ...]:
    diseases = tuple(_site(disease_labels[code], code) for code in row.disease_codes)
    joined = "; ".join(diseases)
    if len(joined) > _EXCEL_SAFE_TEXT_LIMIT:
        joined = (
            "; ".join(diseases[:20])
            + f"; See Disease Propositions filtered to Pattern {row.pattern_number}"
        )
    return (
        row.pattern_number,
        row.review_proposition,
        _site(row.broader_label, row.broader_code),
        _site(row.retained_label, row.retained_code),
        _paths_text(row.paths),
        len(row.disease_codes),
        joined,
        row.occurrence_count,
        row.evidence_kind_counts.one_step,
        row.evidence_kind_counts.transitive,
        row.min_path_length,
        row.max_path_length,
        row.context_risk_summary,
        row.fixed_scope,
    )


def _disease_values(row: DiseaseProposition) -> tuple[str | int, ...]:
    return (
        row.pattern_number,
        _site(row.disease_label, row.disease_code),
        row.specific_proposition,
        _site(row.broader_label, row.broader_code),
        _site(row.retained_label, row.retained_code),
        _paths_text(row.paths),
        row.occurrence_count,
        row.context_summary,
        row.risk_flags,
    )


def _generated_disease_values(row: DiseaseProposition) -> tuple[str | int | None, ...]:
    return (*_disease_values(row), "No", None)


def _visible_payload(
    patterns: tuple[ReviewPattern, ...], propositions: tuple[DiseaseProposition, ...]
) -> dict[str, object]:
    disease_labels = {row.disease_code: row.disease_label for row in propositions}
    return {
        "pattern_headers": _PATTERN_HEADERS,
        "pattern_rows": tuple(_pattern_values(row, disease_labels) for row in patterns),
        "disease_headers": _DISEASE_HEADERS,
        "disease_rows": tuple(_generated_disease_values(row) for row in propositions),
    }


class R101ReviewPacket(_StrictModel):
    schema_version: Literal[3]
    review_scope: Literal["non-exclusive projection coverage"]
    bindings: ReviewBindings
    guidance_identity: str = Field(pattern=_SHA256)
    visible_rows_identity: str = Field(pattern=_SHA256)
    membership_identity: str = Field(pattern=_SHA256)
    patterns: tuple[ReviewPattern, ...]
    disease_propositions: tuple[DiseaseProposition, ...]
    occurrences: tuple[ReviewOccurrence, ...]
    membership: tuple[ReviewMembership, ...]
    packet_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_complete_packet(self) -> Self:
        _validate_packet_inventories(self)
        _validate_packet_membership(self)
        if self.guidance_identity != _identity(
            _guidance_payload(self.bindings.source_release_id)
        ):
            raise ValueError("guidance identity differs")
        if self.visible_rows_identity != _identity(
            _visible_payload(self.patterns, self.disease_propositions)
        ):
            raise ValueError("visible rows identity differs")
        if self.membership_identity != _identity(self.membership):
            raise ValueError("membership identity differs")
        if self.packet_identity != _identity(
            self.model_dump(exclude={"packet_identity"})
        ):
            raise ValueError("packet identity differs")
        return self


def _validate_packet_inventories(packet: R101ReviewPacket) -> None:
    inventories = (
        (len(packet.patterns), _EXPECTED_PATTERNS, "162 patterns"),
        (
            len(packet.disease_propositions),
            _EXPECTED_DISEASE_PROPOSITIONS,
            "2800 disease propositions",
        ),
        (len(packet.occurrences), _EXPECTED_OCCURRENCES, "3291 occurrences"),
        (
            len(packet.membership),
            _EXPECTED_DISEASE_PROPOSITIONS,
            "2800 membership rows",
        ),
    )
    for actual, expected, description in inventories:
        if actual != expected:
            raise ValueError(f"packet must contain exactly {description}")
    if tuple(row.pattern_number for row in packet.patterns) != tuple(
        range(1, _EXPECTED_PATTERNS + 1)
    ):
        raise ValueError("pattern numbers are not canonical")


def _validate_packet_membership(packet: R101ReviewPacket) -> None:
    proposition_keys = tuple(
        (row.pattern_number, row.disease_code) for row in packet.disease_propositions
    )
    membership_keys = tuple(
        (row.pattern_number, row.disease_code) for row in packet.membership
    )
    if proposition_keys != tuple(sorted(set(proposition_keys))):
        raise ValueError("disease proposition membership differs")
    if proposition_keys != membership_keys:
        raise ValueError("disease proposition membership differs")
    _validate_occurrence_membership(packet)


def _validate_occurrence_membership(packet: R101ReviewPacket) -> None:
    occurrence_ids = {row.occurrence_id for row in packet.occurrences}
    member_ids = [item for row in packet.membership for item in row.occurrence_ids]
    if len(member_ids) != _EXPECTED_OCCURRENCES:
        raise ValueError("membership does not exhaust occurrence audit records")
    if set(member_ids) != occurrence_ids:
        raise ValueError("membership does not exhaust occurrence audit records")


class ReviewLabelSource(Protocol):
    async def labels_for_review(
        self, codes: tuple[str, ...]
    ) -> dict[str, tuple[str, ...]]: ...


class ReviewSelectClient(Protocol):
    async def select(
        self, query: str, *, required_variables: Collection[str] = ()
    ) -> list[dict[str, str]]: ...


class QLeverReviewLabels:
    """Bounded-batch stated-graph NCIt preferred-label reader."""

    def __init__(
        self, client: ReviewSelectClient, *, max_batch_size: int = _LABEL_BATCH_SIZE
    ) -> None:
        if max_batch_size < 1 or max_batch_size > _LABEL_BATCH_SIZE:
            raise ValueError(
                f"max_batch_size must be between 1 and {_LABEL_BATCH_SIZE}"
            )
        self._client = client
        self._max_batch_size = max_batch_size
        self.query_count = 0
        self.requested_codes: tuple[str, ...] = ()

    async def labels_for_review(
        self, codes: tuple[str, ...]
    ) -> dict[str, tuple[str, ...]]:
        self.requested_codes = codes
        result: dict[str, set[str]] = defaultdict(set)
        requested = set(codes)
        for start in range(0, len(codes), self._max_batch_size):
            batch = codes[start : start + self._max_batch_size]
            values = " ".join(f"<{safe_iri(code, _NCIT)}>" for code in batch)
            query = (
                "PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#> "
                "SELECT DISTINCT ?c ?label WHERE { VALUES ?c { "
                f"{values} }} GRAPH <{STATED_GRAPH_IRI}> {{ ?c rdfs:label ?raw_label . }} "
                "BIND(STR(?raw_label) AS ?label) } ORDER BY ?c ?label"
            )
            self.query_count += 1
            rows = await self._client.select(query, required_variables=("c", "label"))
            for row in rows:
                code, label = _validated_label_row(row)
                if code not in requested:
                    raise R101ReviewValidationError("unexpected label code binding")
                result[code].add(label)
        return {code: tuple(sorted(values)) for code, values in result.items()}


def _validated_label_row(row: dict[str, str]) -> tuple[str, str]:
    iri, label = row.get("c"), row.get("label")
    if not isinstance(iri, str) or not iri.startswith(_NCIT):
        raise R101ReviewValidationError("malformed label code binding")
    if not isinstance(label, str):
        raise R101ReviewValidationError("malformed label")
    return iri.removeprefix(_NCIT), label


def _review_groups(
    report: R101ConservationReport,
) -> tuple[
    tuple[LedgerOccurrence, ...], dict[tuple[str, str, str], list[LedgerOccurrence]]
]:
    covered = tuple(
        row
        for row in report.occurrences
        if row.disposition == "covered-by-retained-r82"
    )
    groups: dict[tuple[str, str, str], list[LedgerOccurrence]] = defaultdict(list)
    for row in covered:
        groups[_review_group_key(row)].append(row)
    _validate_review_group_inventory(report, covered, groups)
    return covered, groups


def _review_group_key(row: LedgerOccurrence) -> tuple[str, str, str]:
    if len(row.old_links) != 1 or row.retained_r82_target is None:
        raise R101ReviewValidationError("covered occurrence is not review-groupable")
    old, retained = row.old_links[0], row.retained_r82_target
    if old.axis != retained.axis:
        raise R101ReviewValidationError("cross-axis review pattern")
    if old.axis != "op:PrimarySite":
        raise R101ReviewValidationError("cross-axis review pattern")
    return old.axis, old.filler_code, retained.filler_code


def _validate_review_group_inventory(
    report: R101ConservationReport,
    covered: tuple[LedgerOccurrence, ...],
    groups: dict[tuple[str, str, str], list[LedgerOccurrence]],
) -> None:
    presentation = {
        (row.old_filler_code, row.retained_filler_code): row.occurrence_count
        for row in report.grouping_presentation
    }
    derived = {
        (old, retained): len(rows) for (_, old, retained), rows in groups.items()
    }
    if len(covered) != report.counts.covered_by_retained_r82:
        raise R101ReviewValidationError("grouping does not exhaust covered occurrences")
    if derived != presentation:
        raise R101ReviewValidationError("grouping does not exhaust covered occurrences")


def _occurrence_codes(row: LedgerOccurrence) -> tuple[str, ...]:
    target = row.retained_r82_target
    if target is None:
        raise R101ReviewValidationError("covered occurrence has no retained target")
    return (
        row.concept_code,
        row.old_links[0].filler_code,
        target.filler_code,
        row.r82_path[0].part_code,
        *(edge.whole_code for edge in row.r82_path),
    )


def _validate_label(code: str, values: tuple[str, ...]) -> str:
    if not values:
        raise R101ReviewValidationError(f"missing label for {code}")
    if len(values) != 1:
        raise R101ReviewValidationError(f"multiple labels for {code}")
    label = values[0]
    if (
        not label.strip()
        or label != label.strip()
        or any(char in label for char in "\r\n\t")
    ):
        raise R101ReviewValidationError(f"malformed label for {code}")
    return label


async def _review_labels(
    covered: tuple[LedgerOccurrence, ...], source: ReviewLabelSource
) -> dict[str, str]:
    codes = tuple(sorted({code for row in covered for code in _occurrence_codes(row)}))
    raw = await source.labels_for_review(codes)
    return {code: _validate_label(code, raw.get(code, ())) for code in codes}


def _path_payload(
    row: LedgerOccurrence, labels: dict[str, str], source_identity: str
) -> dict[str, object]:
    codes = (row.r82_path[0].part_code, *(edge.whole_code for edge in row.r82_path))
    payload: dict[str, object] = {
        "code_path": codes,
        "labels": tuple(labels[code] for code in codes),
        "fact_identities": tuple(edge.fact_identity for edge in row.r82_path),
        "source_identity": source_identity,
    }
    return {"path_identity": _identity(payload), **payload}


def _review_paths(
    rows: Iterable[LedgerOccurrence], labels: dict[str, str], source_identity: str
) -> tuple[ReviewPath, ...]:
    unique: dict[str, dict[str, object]] = {}
    for row in rows:
        payload = _path_payload(row, labels, source_identity)
        unique[cast("str", payload["path_identity"])] = payload
    return tuple(
        ReviewPath.model_validate(payload) for _, payload in sorted(unique.items())
    )


def _pattern_id(report: R101ConservationReport, key: tuple[str, str, str]) -> str:
    axis, broader, retained = key
    return (
        "r101-"
        + _identity(
            {
                "axis": axis,
                "broader": broader,
                "retained": retained,
                "report_identity": report.report_identity,
                "source_identity": report.source_identity,
            }
        )[:16]
    )


def _risk_summary(rows: list[LedgerOccurrence], paths: tuple[ReviewPath, ...]) -> str:
    diseases = len({row.concept_code for row in rows})
    longest = max(row.path_length for row in rows)
    transitive = sum(row.path_length > 1 for row in rows)
    return f"{diseases} diseases; {len(rows)} source occurrences; {len(paths)} distinct path(s); longest path {longest}; {transitive} transitive occurrence(s). Review transitive and mixed-context evidence carefully."


def _build_rows(
    report: R101ConservationReport,
    groups: dict[tuple[str, str, str], list[LedgerOccurrence]],
    labels: dict[str, str],
) -> tuple[
    list[ReviewPattern],
    list[DiseaseProposition],
    list[ReviewOccurrence],
    list[ReviewMembership],
]:
    patterns: list[ReviewPattern] = []
    propositions: list[DiseaseProposition] = []
    occurrences: list[ReviewOccurrence] = []
    membership: list[ReviewMembership] = []
    for pattern_number, (key, rows) in enumerate(sorted(groups.items()), start=1):
        pattern_id = _pattern_id(report, key)
        paths = _review_paths(rows, labels, report.source_identity)
        patterns.append(
            _build_pattern(pattern_id, pattern_number, key, rows, paths, labels)
        )
        group_occurrences, by_disease = _build_occurrences(
            pattern_id, pattern_number, key, rows, paths
        )
        occurrences.extend(group_occurrences)
        for disease_code, disease_rows in sorted(by_disease.items()):
            proposition, member = _build_disease_proposition(
                report,
                pattern_id,
                pattern_number,
                key,
                disease_code,
                disease_rows,
                labels,
            )
            propositions.append(proposition)
            membership.append(member)
    propositions.sort(key=lambda row: (row.pattern_number, row.disease_code))
    occurrences.sort(
        key=lambda row: (row.pattern_number, row.disease_code, row.occurrence_id)
    )
    membership.sort(key=lambda row: (row.pattern_number, row.disease_code))
    return patterns, propositions, occurrences, membership


def _build_pattern(
    pattern_id: str,
    pattern_number: int,
    key: tuple[str, str, str],
    rows: list[LedgerOccurrence],
    paths: tuple[ReviewPath, ...],
    labels: dict[str, str],
) -> ReviewPattern:
    axis, broader, retained = key
    disease_codes = tuple(sorted({row.concept_code for row in rows}))
    disease_sites = tuple(_site(labels[code], code) for code in disease_codes)
    broader_site = _site(labels[broader], broader)
    retained_site = _site(labels[retained], retained)
    evidence = Counter(row.r82_evidence_kind for row in rows)
    payload: dict[str, object] = {
        **_DENIAL_VALUES,
        "pattern_id": pattern_id,
        "pattern_number": pattern_number,
        "review_proposition": _pattern_proposition(
            disease_sites, pattern_number, retained_site, broader_site
        ),
        "broader_code": broader,
        "broader_label": labels[broader],
        "retained_code": retained,
        "retained_label": labels[retained],
        "axis": axis,
        "occurrence_count": len(rows),
        "evidence_kind_counts": {
            "one_step": evidence["one-step"],
            "transitive": evidence["closure-only"],
        },
        "min_path_length": min(row.path_length for row in rows),
        "max_path_length": max(row.path_length for row in rows),
        "paths": paths,
        "disease_codes": disease_codes,
        "context_risk_summary": _risk_summary(rows, paths),
        "fixed_scope": _SCOPE,
    }
    return ReviewPattern.model_validate({"row_identity": _identity(payload), **payload})


def _build_occurrences(
    pattern_id: str,
    pattern_number: int,
    key: tuple[str, str, str],
    rows: list[LedgerOccurrence],
    paths: tuple[ReviewPath, ...],
) -> tuple[list[ReviewOccurrence], dict[str, list[LedgerOccurrence]]]:
    _, broader, retained = key
    path_by_codes = {path.code_path: path for path in paths}
    by_disease: dict[str, list[LedgerOccurrence]] = defaultdict(list)
    occurrences: list[ReviewOccurrence] = []
    for row in rows:
        by_disease[row.concept_code].append(row)
        occurrences.append(
            _build_occurrence(
                pattern_id,
                pattern_number,
                broader,
                retained,
                row,
                path_by_codes,
            )
        )
    return occurrences, by_disease


def _build_occurrence(
    pattern_id: str,
    pattern_number: int,
    broader: str,
    retained: str,
    row: LedgerOccurrence,
    path_by_codes: dict[tuple[str, ...], ReviewPath],
) -> ReviewOccurrence:
    code_path = (
        row.r82_path[0].part_code,
        *(edge.whole_code for edge in row.r82_path),
    )
    return ReviewOccurrence(
        pattern_id=pattern_id,
        pattern_number=pattern_number,
        disease_code=row.concept_code,
        occurrence_id=row.occurrence_id,
        source_fact_id=row.source_fact_id,
        source_group_id=row.source_group_id,
        anchor_code=row.anchor_code,
        depth=row.depth,
        structural_path=row.structural_path,
        member_position=row.member_position,
        evidence_kind=row.r82_evidence_kind,  # type: ignore[arg-type]
        path_length=row.path_length,
        path_identity=path_by_codes[code_path].path_identity,
        broader_code=broader,
        retained_code=retained,
    )


def _build_disease_proposition(
    report: R101ConservationReport,
    pattern_id: str,
    pattern_number: int,
    key: tuple[str, str, str],
    disease_code: str,
    disease_rows: list[LedgerOccurrence],
    labels: dict[str, str],
) -> tuple[DiseaseProposition, ReviewMembership]:
    _, broader, retained = key
    paths = _review_paths(disease_rows, labels, report.source_identity)
    disease_site = _site(labels[disease_code], disease_code)
    broader_site = _site(labels[broader], broader)
    retained_site = _site(labels[retained], retained)
    payload: dict[str, object] = {
        **_DENIAL_VALUES,
        "pattern_id": pattern_id,
        "pattern_number": pattern_number,
        "disease_code": disease_code,
        "disease_label": labels[disease_code],
        "specific_proposition": _specific_proposition(
            disease_site, retained_site, broader_site
        ),
        "broader_code": broader,
        "broader_label": labels[broader],
        "retained_code": retained,
        "retained_label": labels[retained],
        "paths": paths,
        "occurrence_count": len(disease_rows),
        "context_summary": f"{disease_site}: {retained_site} provides proposed non-exclusive projection coverage of {broader_site} for {len(disease_rows)} frozen source occurrence(s).",
        "risk_flags": _disease_risk(disease_rows),
    }
    proposition = DiseaseProposition.model_validate(
        {"proposition_identity": _identity(payload), **payload}
    )
    member_payload: dict[str, object] = {
        "pattern_id": pattern_id,
        "pattern_number": pattern_number,
        "disease_code": disease_code,
        "occurrence_ids": tuple(sorted(row.occurrence_id for row in disease_rows)),
    }
    member = ReviewMembership.model_validate(
        {"membership_identity": _identity(member_payload), **member_payload}
    )
    return proposition, member


def _disease_risk(rows: list[LedgerOccurrence]) -> str:
    if any(row.path_length > 1 for row in rows):
        return "Transitive path — extra scrutiny"
    return "One-step path; still review disease context"


async def build_r101_review_packet(
    report: R101ConservationReport,
    source_manifest: Path,
    label_source: ReviewLabelSource,
) -> R101ReviewPacket:
    """Build schema-v3 review evidence after exhaustive source reconciliation."""
    manifest = validate_ncit_sibling_manifest(source_manifest)
    if (
        manifest.source_identity != report.source_identity
        or manifest.ontology_version != report.source_release_id
    ):
        raise R101ReviewValidationError("source manifest does not bind report")
    if report.mechanical_status != "complete" or report.counts.unresolved:
        raise R101ReviewValidationError("report mechanics are incomplete")
    covered, groups = _review_groups(report)
    labels = await _review_labels(covered, label_source)
    patterns, propositions, occurrences, membership = _build_rows(
        report, groups, labels
    )
    patterns_tuple = tuple(patterns)
    propositions_tuple = tuple(propositions)
    membership_tuple = tuple(membership)
    payload: dict[str, object] = {
        "schema_version": _SCHEMA_VERSION,
        "review_scope": _SCOPE,
        "bindings": _bindings(report),
        "guidance_identity": _identity(_guidance_payload(report.source_release_id)),
        "visible_rows_identity": _identity(
            _visible_payload(patterns_tuple, propositions_tuple)
        ),
        "membership_identity": _identity(membership_tuple),
        "patterns": patterns_tuple,
        "disease_propositions": propositions_tuple,
        "occurrences": tuple(occurrences),
        "membership": membership_tuple,
    }
    return R101ReviewPacket.model_validate(
        {**payload, "packet_identity": _identity(payload)}
    )


def _write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = (
        json.dumps(
            to_jsonable_python(payload), sort_keys=True, indent=2, ensure_ascii=True
        ).encode("ascii")
        + b"\n"
    )
    fd, staging = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as stream:
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(staging, path)
    except BaseException:
        with suppress(FileNotFoundError):
            os.unlink(staging)
        raise


def write_r101_review_packet(path: Path, packet: R101ReviewPacket) -> None:
    _write_json(path, packet.model_dump(mode="json"))


def _load_json(path: Path) -> object:
    def reject_duplicate(pairs: list[tuple[str, object]]) -> dict[str, object]:
        result: dict[str, object] = {}
        for key, value in pairs:
            if key in result:
                raise R101ReviewValidationError(f"duplicate JSON key: {key}")
            result[key] = value
        return result

    try:
        return json.loads(path.read_text(), object_pairs_hook=reject_duplicate)
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
        raise R101ReviewValidationError("invalid review JSON") from error


def load_r101_review_packet(path: Path) -> R101ReviewPacket:
    try:
        return R101ReviewPacket.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R101ReviewValidationError(str(error)) from error


def _style_header(cell: Cell | MergedCell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_sheet(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    for cell in sheet[1]:
        _style_header(cell)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, width in enumerate(widths, start=1):
        header_cell = cast("Cell", sheet.cell(1, index))
        sheet.column_dimensions[header_cell.column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True


def _fill_guidance(book: WorkbookType, packet: R101ReviewPacket) -> None:
    instructions = book.active
    if instructions is None:
        raise R101ReviewValidationError("workbook has no active sheet")
    instructions.title = "Instructions and Semantics"
    for row in _instruction_rows(packet.bindings.source_release_id):
        instructions.append(row)
    _style_sheet(instructions, (34, 116))
    for row in instructions.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color="1F4E78")

    definitions = book.create_sheet("Column Definitions")
    definitions.append(_DEFINITION_HEADERS)
    for row in _definition_rows():
        definitions.append(row)
    _style_sheet(definitions, (24, 36, 68, 54, 52))

    examples = book.create_sheet("Review Examples")
    for row in _example_rows():
        examples.append(row)
    _style_sheet(examples, (34, 76, 62, 70))


def _add_pattern_review(book: WorkbookType, packet: R101ReviewPacket) -> None:
    sheet = book.create_sheet("Pattern Review")
    sheet.append(_PATTERN_HEADERS)
    disease_labels = {
        row.disease_code: row.disease_label for row in packet.disease_propositions
    }
    for pattern in packet.patterns:
        sheet.append(
            (*_pattern_values(pattern, disease_labels), None, None, None, None)
        )
    _mark_editable_cells(sheet, _PATTERN_HEADERS, _PATTERN_EDITABLE)
    decision_column = _PATTERN_HEADERS.index("Decision") + 1
    validation = DataValidation(type="list", formula1='"' + ",".join(_DECISIONS) + '"')
    validation.error = "Choose one listed decision"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(2, decision_column).coordinate}:{sheet.cell(sheet.max_row, decision_column).coordinate}"
    )
    widths = (16, 86, 34, 38, 68, 18, 72, 18, 15, 15, 15, 15, 68, 32, 42, 56, 28, 18)
    _style_sheet(sheet, widths)
    for header in ("Review Proposition", "Decision", "Fixed Scope"):
        sheet.cell(1, _PATTERN_HEADERS.index(header) + 1).comment = Comment(
            next(
                row[2]
                for row in _definition_rows()
                if row[0] == "Pattern Review" and row[1] == header
            ),
            "OntoPrism",
        )


def _mark_editable_cells(
    sheet: Worksheet, headers: tuple[str, ...], editable: frozenset[str]
) -> None:
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in sheet.iter_rows(min_row=2):
        for cell, header in zip(row, headers, strict=True):
            cell.protection = Protection(locked=header not in editable)
            if header in editable:
                cell.fill = editable_fill


def _add_disease_propositions(book: WorkbookType, packet: R101ReviewPacket) -> None:
    sheet = book.create_sheet("Disease Propositions")
    sheet.append(_DISEASE_HEADERS)
    for proposition in packet.disease_propositions:
        sheet.append(_generated_disease_values(proposition))
    _mark_editable_cells(sheet, _DISEASE_HEADERS, _DISEASE_EDITABLE)
    exception_column = _DISEASE_HEADERS.index("Exception?") + 1
    validation = DataValidation(type="list", formula1='"Yes,No"')
    validation.error = "Keep generated No or choose Yes for a justified exception"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(
        f"{sheet.cell(2, exception_column).coordinate}:{sheet.cell(sheet.max_row, exception_column).coordinate}"
    )
    _style_sheet(sheet, (16, 42, 88, 34, 34, 68, 18, 70, 44, 16, 58))


def _add_bindings(book: WorkbookType, packet: R101ReviewPacket) -> None:
    sheet = book.create_sheet("Bindings")
    sheet.append(("Binding", "Value"))
    values = (
        packet.packet_identity,
        packet.guidance_identity,
        packet.visible_rows_identity,
        packet.membership_identity,
        packet.schema_version,
        packet.bindings.source_release_id,
    )
    for name, value in zip(_BINDING_NAMES, values, strict=True):
        sheet.append((name, value))
    sheet.sheet_state = "veryHidden"
    sheet.protection.sheet = True


def write_r101_review_workbook(path: Path, packet: R101ReviewPacket) -> None:
    """Write a formula-free workbook containing no technical packet row identifiers."""
    book = Workbook()
    _fill_guidance(book, packet)
    _add_pattern_review(book, packet)
    _add_disease_propositions(book, packet)
    # Keep the requested reviewer-facing order via openpyxl's public move API.
    book.move_sheet(book["Pattern Review"], offset=-2)
    book.move_sheet(book["Disease Propositions"], offset=-2)
    _add_bindings(book, packet)
    for sheet in book.worksheets:
        sheet.sheet_view.showGridLines = False
    book.calculation.calcMode = "auto"
    book.calculation.fullCalcOnLoad = False
    book.calculation.forceFullCalc = False
    fixed = datetime(2000, 1, 1)
    book.properties.created = fixed
    book.properties.modified = fixed
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, staging = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(fd)
    try:
        book.save(staging)
        _normalize_xlsx_archive(Path(staging), path)
    finally:
        with suppress(FileNotFoundError):
            os.unlink(staging)


def _normalize_xlsx_archive(source: Path, destination: Path) -> None:
    fd, normalized_name = tempfile.mkstemp(
        prefix=f".{destination.name}.normalized.", dir=destination.parent
    )
    os.close(fd)
    try:
        with (
            ZipFile(source) as original,
            ZipFile(
                normalized_name, "w", compression=ZIP_DEFLATED, compresslevel=9
            ) as normalized,
        ):
            for name in sorted(original.namelist()):
                info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o600 << 16
                content = original.read(name)
                if name == "docProps/core.xml":
                    content = re.sub(
                        rb"<dcterms:(created|modified)[^>]*>.*?</dcterms:\1>",
                        rb'<dcterms:\1 xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:\1>',
                        content,
                    )
                elif name == "xl/workbook.xml":
                    content = re.sub(rb' calcId="[0-9]+"', b"", content)
                normalized.writestr(info, content, compress_type=ZIP_DEFLATED)
        os.replace(normalized_name, destination)
    finally:
        with suppress(FileNotFoundError):
            os.unlink(normalized_name)


class PatternDecision(_StrictModel):
    pattern_number: int = Field(ge=1, le=_EXPECTED_PATTERNS)
    decision: ReviewDecision
    rationale: str = Field(min_length=1)
    reviewer_identity: str = Field(min_length=1)
    review_date: str = Field(pattern=r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")


class DiseaseException(_StrictModel):
    pattern_number: int = Field(ge=1, le=_EXPECTED_PATTERNS)
    disease_code: str = Field(pattern=_CODE)
    is_exception: bool
    rationale: str | None


AtomicOutcome = Literal[
    "approved-non-exclusive-coverage",
    "disease-exception",
    "rejected-retain-broader",
    "individual-review-required",
    "escalated",
]


class AtomicDecision(DenialFlags):
    atomic_decision_identity: str = Field(pattern=_SHA256)
    pattern_number: int = Field(ge=1, le=_EXPECTED_PATTERNS)
    disease_code: str = Field(pattern=_CODE)
    occurrence_id: str = Field(pattern=_SHA256)
    broader_code: str = Field(pattern=_CODE)
    retained_code: str = Field(pattern=_CODE)
    outcome: AtomicOutcome
    proposition_text: str = Field(min_length=1)
    rationale: str = Field(min_length=1)
    reviewer_identity: str = Field(min_length=1)
    review_date: str = Field(pattern=r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")
    source_packet_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_identity(self) -> Self:
        if self.atomic_decision_identity != _identity(
            self.model_dump(exclude={"atomic_decision_identity"})
        ):
            raise ValueError("atomic decision identity differs")
        return self


class R101DecisionRegistry(_StrictModel):
    schema_version: Literal[3]
    status: Literal["proposed"]
    provenance: Literal["sme", "test-only"]
    packet_identity: str = Field(pattern=_SHA256)
    report_identity: str = Field(pattern=_SHA256)
    source_identity: str = Field(pattern=_SHA256)
    pattern_decisions: tuple[PatternDecision, ...]
    disease_exceptions: tuple[DiseaseException, ...]
    atomic_decisions: tuple[AtomicDecision, ...]
    registry_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_registry(self) -> Self:
        if len(self.pattern_decisions) != _EXPECTED_PATTERNS:
            raise ValueError("registry must contain 162 pattern decisions")
        if len(self.disease_exceptions) != _EXPECTED_DISEASE_PROPOSITIONS:
            raise ValueError("registry must contain 2800 disease exception rows")
        if len(self.atomic_decisions) != _EXPECTED_OCCURRENCES:
            raise ValueError("registry must contain 3291 atomic decisions")
        if self.registry_identity != _identity(
            self.model_dump(exclude={"registry_identity"})
        ):
            raise ValueError("registry identity differs")
        return self


def _validate_archive(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            names = archive.namelist()
    except (OSError, BadZipFile) as error:
        raise R101ReviewValidationError("invalid XLSX archive") from error
    if any(name.casefold().endswith(("vbaproject.bin", ".vba")) for name in names):
        raise R101ReviewValidationError("macro content is forbidden")
    if any(name.startswith("xl/externalLinks/") for name in names):
        raise R101ReviewValidationError("external link content is forbidden")


def _validate_no_formulas(book: WorkbookType) -> None:
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    raise R101ReviewValidationError("formula cells are forbidden")


def _expected_binding_rows(packet: R101ReviewPacket) -> list[tuple[object, object]]:
    values = (
        packet.packet_identity,
        packet.guidance_identity,
        packet.visible_rows_identity,
        packet.membership_identity,
        packet.schema_version,
        packet.bindings.source_release_id,
    )
    return [("Binding", "Value"), *zip(_BINDING_NAMES, values, strict=True)]


def _validate_workbook_structure(book: WorkbookType, packet: R101ReviewPacket) -> None:
    expected_sheets = [
        "Instructions and Semantics",
        "Pattern Review",
        "Disease Propositions",
        "Column Definitions",
        "Review Examples",
        "Bindings",
    ]
    if book.sheetnames != expected_sheets:
        raise R101ReviewValidationError(
            "workbook sheet contract differs; stale v2 workbooks are not accepted"
        )
    _validate_binding_sheet(book["Bindings"], packet)
    _validate_guidance_sheets(book, packet)


def _validate_binding_sheet(sheet: Worksheet, packet: R101ReviewPacket) -> None:
    if sheet.sheet_state != "veryHidden":
        raise R101ReviewValidationError("binding sheet visibility differs")
    actual = [tuple(cell.value for cell in row[:2]) for row in sheet.iter_rows()]
    if actual != _expected_binding_rows(packet):
        raise R101ReviewValidationError("workbook binding cells differ")


def _validate_guidance_sheets(book: WorkbookType, packet: R101ReviewPacket) -> None:
    guidance = _guidance_payload(packet.bindings.source_release_id)
    expected = {
        "Instructions and Semantics": guidance["instructions"],
        "Column Definitions": guidance["definitions"],
        "Review Examples": guidance["examples"],
    }
    for name, rows in expected.items():
        actual_rows = tuple(
            tuple(cell.value for cell in row) for row in book[name].iter_rows()
        )
        if actual_rows != rows:
            raise R101ReviewValidationError(f"workbook guidance differs in {name}")


def _required_text(value: object, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise R101ReviewValidationError(f"{field} is required")
    return value.strip()


def _required_decision(value: object) -> ReviewDecision:
    decision = _required_text(value, "decision")
    if decision not in _DECISIONS:
        raise R101ReviewValidationError("decision is not one of the closed values")
    return cast("ReviewDecision", decision)


def _read_pattern_decisions(
    book: WorkbookType, packet: R101ReviewPacket
) -> tuple[PatternDecision, ...]:
    sheet = book["Pattern Review"]
    if tuple(cell.value for cell in sheet[1]) != _PATTERN_HEADERS:
        raise R101ReviewValidationError("pattern headers differ")
    if sheet.max_row != _EXPECTED_PATTERNS + 1:
        raise R101ReviewValidationError(
            "workbook must contain exactly 162 pattern rows"
        )
    disease_labels = {
        row.disease_code: row.disease_label for row in packet.disease_propositions
    }
    result: list[PatternDecision] = []
    immutable_count = len(_PATTERN_HEADERS) - len(_PATTERN_EDITABLE)
    for row_number, (cells, pattern) in enumerate(
        zip(sheet.iter_rows(min_row=2), packet.patterns, strict=True), start=2
    ):
        result.append(
            _read_pattern_decision_row(
                cells, pattern, disease_labels, immutable_count, row_number
            )
        )
    return tuple(result)


def _read_pattern_decision_row(
    cells: tuple[Cell | MergedCell, ...],
    pattern: ReviewPattern,
    disease_labels: dict[str, str],
    immutable_count: int,
    row_number: int,
) -> PatternDecision:
    if tuple(cell.value for cell in cells[:immutable_count]) != _pattern_values(
        pattern, disease_labels
    ):
        raise R101ReviewValidationError(
            f"pattern visible row differs at row {row_number}"
        )
    decision, rationale, reviewer, reviewed_at = (
        cell.value for cell in cells[immutable_count:]
    )
    try:
        return PatternDecision(
            pattern_number=pattern.pattern_number,
            decision=_required_decision(decision),
            rationale=_required_text(rationale, "rationale"),
            reviewer_identity=_required_text(reviewer, "reviewer identity"),
            review_date=_required_text(reviewed_at, "review date"),
        )
    except ValueError as error:
        raise R101ReviewValidationError(
            f"invalid pattern decision at row {row_number}: {error}"
        ) from error


def _disease_code_from_cell(value: object) -> str:
    if not isinstance(value, str):
        return ""
    match = re.search(r"\((C[0-9]+)\)$", value)
    return match.group(1) if match else ""


def _read_disease_exceptions(
    book: WorkbookType,
    packet: R101ReviewPacket,
    decisions: tuple[PatternDecision, ...],
) -> tuple[DiseaseException, ...]:
    sheet = book["Disease Propositions"]
    if tuple(cell.value for cell in sheet[1]) != _DISEASE_HEADERS:
        raise R101ReviewValidationError("disease proposition headers differ")
    if sheet.max_row != _EXPECTED_DISEASE_PROPOSITIONS + 1:
        raise R101ReviewValidationError(
            "workbook must contain exactly 2800 disease rows"
        )
    by_pattern = {row.pattern_number: row for row in decisions}
    result: list[DiseaseException] = []
    immutable_count = len(_DISEASE_HEADERS) - len(_DISEASE_EDITABLE)
    seen: set[tuple[int, str]] = set()
    for row_number, (cells, proposition) in enumerate(
        zip(sheet.iter_rows(min_row=2), packet.disease_propositions, strict=True),
        start=2,
    ):
        result.append(
            _read_disease_exception_row(
                cells,
                proposition,
                by_pattern[proposition.pattern_number].decision,
                immutable_count,
                row_number,
                seen,
            )
        )
    return tuple(result)


def _read_disease_exception_row(
    cells: tuple[Cell | MergedCell, ...],
    proposition: DiseaseProposition,
    decision: ReviewDecision,
    immutable_count: int,
    row_number: int,
    seen: set[tuple[int, str]],
) -> DiseaseException:
    if tuple(cell.value for cell in cells[:immutable_count]) != _disease_values(
        proposition
    ):
        raise R101ReviewValidationError(
            f"disease proposition differs at row {row_number}"
        )
    key = (cast("int", cells[0].value), _disease_code_from_cell(cells[1].value))
    if key in seen:
        raise R101ReviewValidationError("duplicate disease proposition row")
    seen.add(key)
    is_exception, rationale = _parse_exception_values(
        decision,
        cells[immutable_count].value,
        cells[immutable_count + 1].value,
    )
    return DiseaseException(
        pattern_number=proposition.pattern_number,
        disease_code=proposition.disease_code,
        is_exception=is_exception,
        rationale=rationale,
    )


def _parse_exception_values(
    decision: ReviewDecision, exception_value: object, rationale_value: object
) -> tuple[bool, str | None]:
    if decision != APPROVE:
        if exception_value != "No" or rationale_value not in (None, ""):
            raise R101ReviewValidationError(
                "exceptions are invalid for a non-approve decision"
            )
        return False, None
    if exception_value not in ("Yes", "No"):
        raise R101ReviewValidationError(
            "approved patterns require generated No or a justified Yes for every disease exception"
        )
    if exception_value == "Yes":
        return True, _required_text(rationale_value, "exception rationale")
    if rationale_value not in (None, ""):
        raise R101ReviewValidationError(
            "exception rationale is allowed only when Exception? is Yes"
        )
    return False, None


def _atomic_outcome(decision: ReviewDecision, exception: bool) -> AtomicOutcome:
    if decision == APPROVE:
        return "disease-exception" if exception else "approved-non-exclusive-coverage"
    if decision == REJECT:
        return "rejected-retain-broader"
    if decision == INDIVIDUAL:
        return "individual-review-required"
    return "escalated"


def _expand_atomic(
    packet: R101ReviewPacket,
    decisions: tuple[PatternDecision, ...],
    exceptions: tuple[DiseaseException, ...],
) -> tuple[AtomicDecision, ...]:
    decision_by_pattern = {row.pattern_number: row for row in decisions}
    exception_by_key = {
        (row.pattern_number, row.disease_code): row for row in exceptions
    }
    proposition_by_key = {
        (row.pattern_number, row.disease_code): row
        for row in packet.disease_propositions
    }
    occurrence_by_id = {row.occurrence_id: row for row in packet.occurrences}
    result: list[AtomicDecision] = []
    for member in packet.membership:
        result.extend(
            _expand_member(
                packet,
                member,
                decision_by_pattern[member.pattern_number],
                exception_by_key[(member.pattern_number, member.disease_code)],
                proposition_by_key[(member.pattern_number, member.disease_code)],
                occurrence_by_id,
            )
        )
    return tuple(
        sorted(
            result,
            key=lambda row: (row.pattern_number, row.disease_code, row.occurrence_id),
        )
    )


def _expand_member(
    packet: R101ReviewPacket,
    member: ReviewMembership,
    decision: PatternDecision,
    exception: DiseaseException,
    proposition: DiseaseProposition,
    occurrence_by_id: dict[str, ReviewOccurrence],
) -> list[AtomicDecision]:
    outcome = _atomic_outcome(decision.decision, exception.is_exception)
    rationale = exception.rationale if exception.is_exception else decision.rationale
    if rationale is None:
        raise R101ReviewValidationError("exception rationale is required")
    return [
        _atomic_decision(
            packet,
            member,
            decision,
            proposition,
            occurrence_by_id[occurrence_id],
            outcome,
            rationale,
        )
        for occurrence_id in member.occurrence_ids
    ]


def _atomic_decision(
    packet: R101ReviewPacket,
    member: ReviewMembership,
    decision: PatternDecision,
    proposition: DiseaseProposition,
    occurrence: ReviewOccurrence,
    outcome: AtomicOutcome,
    rationale: str,
) -> AtomicDecision:
    payload: dict[str, object] = {
        **_DENIAL_VALUES,
        "pattern_number": member.pattern_number,
        "disease_code": member.disease_code,
        "occurrence_id": occurrence.occurrence_id,
        "broader_code": occurrence.broader_code,
        "retained_code": occurrence.retained_code,
        "outcome": outcome,
        "proposition_text": proposition.specific_proposition,
        "rationale": rationale,
        "reviewer_identity": decision.reviewer_identity,
        "review_date": decision.review_date,
        "source_packet_identity": packet.packet_identity,
    }
    return AtomicDecision.model_validate(
        {"atomic_decision_identity": _identity(payload), **payload}
    )


def import_r101_review_decisions(
    packet: R101ReviewPacket,
    workbook_path: Path,
    output_path: Path,
    *,
    provenance: Literal["sme", "test-only"],
) -> R101DecisionRegistry:
    """Validate every cell and atomically write a proposed occurrence registry."""
    _validate_archive(workbook_path)
    try:
        book = load_workbook(workbook_path, data_only=False, keep_links=False)
    except (OSError, ValueError, BadZipFile) as error:
        raise R101ReviewValidationError("invalid review workbook") from error
    _validate_no_formulas(book)
    _validate_workbook_structure(book, packet)
    decisions = _read_pattern_decisions(book, packet)
    exceptions = _read_disease_exceptions(book, packet, decisions)
    atomic = _expand_atomic(packet, decisions, exceptions)
    payload: dict[str, object] = {
        "schema_version": _SCHEMA_VERSION,
        "status": "proposed",
        "provenance": provenance,
        "packet_identity": packet.packet_identity,
        "report_identity": packet.bindings.report_identity,
        "source_identity": packet.bindings.source_identity,
        "pattern_decisions": decisions,
        "disease_exceptions": exceptions,
        "atomic_decisions": atomic,
    }
    registry = R101DecisionRegistry.model_validate(
        {**payload, "registry_identity": _identity(payload)}
    )
    _write_json(output_path, registry.model_dump(mode="json"))
    return registry


def load_r101_decision_registry(path: Path) -> R101DecisionRegistry:
    try:
        return R101DecisionRegistry.model_validate_json(_canonical(_load_json(path)))
    except ValueError as error:
        raise R101ReviewValidationError(str(error)) from error


class DecisionExpansionDryRun(_StrictModel):
    schema_version: Literal[3]
    verdict: Literal["validated-proposed-registry"]
    report_identity: str = Field(pattern=_SHA256)
    packet_identity: str = Field(pattern=_SHA256)
    registry_identity: str = Field(pattern=_SHA256)
    provenance: Literal["sme", "test-only"]
    pattern_decisions: int = Field(ge=0)
    atomic_decisions: int = Field(ge=0)
    approved_occurrences: int = Field(ge=0)
    rejected_occurrences: int = Field(ge=0)
    follow_up_occurrences: int = Field(ge=0)
    escalated_occurrences: int = Field(ge=0)
    exception_occurrences: int = Field(ge=0)
    writes_performed: Literal[False]


def _validate_registry_bindings(
    report: R101ConservationReport,
    packet: R101ReviewPacket,
    registry: R101DecisionRegistry,
) -> None:
    checks = (
        (packet.bindings, _bindings(report), "packet bindings do not match report"),
        (
            registry.packet_identity,
            packet.packet_identity,
            "registry packet identity differs",
        ),
        (
            registry.report_identity,
            report.report_identity,
            "registry report identity differs",
        ),
        (
            registry.source_identity,
            report.source_identity,
            "registry source identity differs",
        ),
    )
    for actual, expected, message in checks:
        if actual != expected:
            raise R101ReviewValidationError(message)
    expected_atomic = _expand_atomic(
        packet, registry.pattern_decisions, registry.disease_exceptions
    )
    if registry.atomic_decisions != expected_atomic:
        raise R101ReviewValidationError(
            "registry atomic expansion differs from frozen packet membership"
        )


def dry_run_r101_decision_expansion(
    report: R101ConservationReport,
    packet: R101ReviewPacket,
    registry: R101DecisionRegistry,
) -> DecisionExpansionDryRun:
    """Re-run exact expansion and validation without authorization or persistence."""
    _validate_registry_bindings(report, packet, registry)
    counts = Counter(row.outcome for row in registry.atomic_decisions)
    return DecisionExpansionDryRun(
        schema_version=3,
        verdict="validated-proposed-registry",
        report_identity=report.report_identity,
        packet_identity=packet.packet_identity,
        registry_identity=registry.registry_identity,
        provenance=registry.provenance,
        pattern_decisions=len(registry.pattern_decisions),
        atomic_decisions=len(registry.atomic_decisions),
        approved_occurrences=counts["approved-non-exclusive-coverage"],
        rejected_occurrences=counts["rejected-retain-broader"],
        follow_up_occurrences=counts["individual-review-required"],
        escalated_occurrences=counts["escalated"],
        exception_occurrences=counts["disease-exception"],
        writes_performed=False,
    )


def write_r101_decision_expansion_dry_run(
    path: Path, result: DecisionExpansionDryRun
) -> None:
    _write_json(path, result.model_dump(mode="json"))
