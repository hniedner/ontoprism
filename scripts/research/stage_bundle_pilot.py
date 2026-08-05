"""Generate the M1 v14 cancer-stage semantic-bundle review candidate."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal, cast

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ontolib.decomposition.semantic_bundles import (
    AdjudicatedSemanticContext,
    BundleAxis,
    BundleKind,
    EvidenceClaim,
    EvidenceClaimKind,
    EvidenceClaimTarget,
    EvidenceRegistry,
    MemberRole,
    ProjectedConstituentEvidence,
    SemanticBundleCandidate,
    SemanticBundleMember,
    SourceOccurrence,
    StageClassification,
    canonical_restriction_fact_id,
    evaluate_pair_availability,
    validate_candidate_evidence,
)

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from ontolib.decomposition.semantic_bundles import ProjectionAxisSource

_NCIT_RELEASE = "26.07d"
_WORKBOOK_SHA256 = "a538de0772df786da39f0eaeb9c374e837b71e58b567d6f045f126225e759cc8"
_SOURCE_AUDIT_SHA256 = (
    "aec3910bc5a72c5132c996c52861534c615912f5001928b1cd374f7d566fadfb"
)
_ENGINE_EVIDENCE_SHA256 = (
    "42e33238c7b18985263f780a165ad42d1230bb620a2aac8edf11748cf661f74f"
)
_SOURCE_IDENTITY = "f54dd2910a31245a30cea094dc72ce6a5c8d7b5a9c4e484007a35a1c343624c8"
_STRUCTURE_CLAIM_ID = "mcode-4.0.0-cancer-stage-structure"
_METHOD_CLAIM_ID = "mcode-4.0.0-valg-method"
_SOURCE_OCCURRENCE_COUNT = 304
_ANCHOR_PART_COUNT = 2
_R101_SAME_AXIS_R82_COLLAPSES = {
    ("C100051", "C12810"): "C12413",
    ("C101539", "C12418"): "C13063",
    ("C162226", "C12810"): "C12402",
    ("C181564", "C12810"): "C12402",
    ("C186620", "C12810"): "C12402",
    ("C206219", "C12810"): "C12402",
    ("C4791", "C12727"): "C13004",
    ("C6135", "C12418"): "C13063",
}


@dataclass(frozen=True, slots=True, kw_only=True)
class EvidenceSource:
    evidence_id: str
    title: str
    uri: str
    access: str
    supports: str


@dataclass(frozen=True, slots=True, kw_only=True)
class ConstituentCorrection:
    action: Literal["add", "remove"]
    concept_code: str
    axis: str
    filler_code: str
    rationale: str

    def __post_init__(self) -> None:
        if self.action not in {"add", "remove"}:
            raise ValueError("constituent correction action is invalid")
        for value, field in (
            (self.concept_code, "concept_code"),
            (self.axis, "axis"),
            (self.filler_code, "filler_code"),
            (self.rationale, "rationale"),
        ):
            if not value.strip():
                raise ValueError(f"constituent correction {field} must not be empty")


EVIDENCE_SOURCES = (
    EvidenceSource(
        evidence_id="ncit-26.07d-role-audit",
        title="M1 #57 complete contracted-role audit for NCIt 26.07d",
        uri="M1-57_Complete_Role_Audit_2607d_v1.json",
        access="local hash-pinned evidence",
        supports="Exact R88 source occurrences, anchors, depths, and canonical groups.",
    ),
    EvidenceSource(
        evidence_id="mcode-4.0.0-cancer-stage",
        title="mCODE 4.0.0 Cancer Stage Profile",
        uri=(
            "https://hl7.org/fhir/us/mcode/STU4/"
            "StructureDefinition-mcode-cancer-stage.html"
        ),
        access="public",
        supports=(
            "A cancer-stage observation has a required stage type, optional method, "
            "and at most one stage value."
        ),
    ),
    EvidenceSource(
        evidence_id="mcode-4.0.0-stage-type",
        title="mCODE 4.0.0 Cancer Stage Type Value Set",
        uri=(
            "https://hl7.org/fhir/us/mcode/STU4/"
            "ValueSet-mcode-cancer-stage-type-vs.html"
        ),
        access="public; terminology licenses still apply",
        supports=(
            "NCIt AJCC v6-v9 and FIGO 2009/2018 stage-type codes used by this pilot."
        ),
    ),
    EvidenceSource(
        evidence_id="mcode-4.0.0-stage-method",
        title="mCODE 4.0.0 Cancer Staging Method Value Set",
        uri=(
            "https://hl7.org/fhir/us/mcode/STU4/"
            "ValueSet-mcode-cancer-staging-method-vs.html"
        ),
        access="public; terminology licenses still apply",
        supports="NCIt C141685 as the VALG staging method.",
    ),
    EvidenceSource(
        evidence_id="ajcc-official-staging-system",
        title="AJCC Cancer Staging System",
        uri=(
            "https://www.facs.org/quality-programs/cancer-programs/"
            "american-joint-committee-on-cancer/version-9/"
        ),
        access="licensed/restricted AJCC content",
        supports=(
            "AJCC authority, edition/version distinction, and the cervical Version 9 "
            "protocol. No restricted staging tables are reproduced."
        ),
    ),
    EvidenceSource(
        evidence_id="figo-cervix-2009",
        title="Revised FIGO staging for carcinoma of the cervix",
        uri="https://doi.org/10.1016/j.ijgo.2009.02.009",
        access="publisher terms apply",
        supports="The 2009 FIGO cervical staging revision.",
    ),
    EvidenceSource(
        evidence_id="figo-cervix-2018",
        title="Revised FIGO staging for carcinoma of the cervix uteri",
        uri="https://doi.org/10.1002/ijgo.12749",
        access="publisher terms apply",
        supports="The 2018 FIGO cervical staging revision.",
    ),
    EvidenceSource(
        evidence_id="figo-endometrial-2023",
        title="FIGO staging of endometrial cancer: 2023",
        uri="https://doi.org/10.1002/ijgo.14923",
        access="publisher terms apply",
        supports="The 2023 FIGO endometrial staging revision.",
    ),
    EvidenceSource(
        evidence_id="nci-pdq-sclc",
        title="Small Cell Lung Cancer Treatment (PDQ), Health Professional Version",
        uri="https://www.cancer.gov/types/lung/hp/small-cell-lung-treatment-pdq",
        access="public",
        supports=(
            "VALG as an SCLC staging system and the limited/extensive-stage "
            "distinction."
        ),
    ),
)


_STRUCTURE_ASSERTION = (
    "ONTOPrism's stage contract maps one typed stage framework or method and one "
    "stage value to the mCODE Cancer Stage observation structure; it does not claim "
    "direct mCODE conformance."
)
_STRUCTURE_URI = (
    "https://hl7.org/fhir/us/mcode/STU4/StructureDefinition-mcode-cancer-stage.html"
)
_METHOD_ASSERTION = "NCIt C141685 denotes the VALG cancer staging method."
_METHOD_URI = (
    "https://hl7.org/fhir/us/mcode/STU4/ValueSet-mcode-cancer-staging-method-vs.html"
)
EVIDENCE_REGISTRY = EvidenceRegistry(
    claims=(
        EvidenceClaim(
            claim_id=_STRUCTURE_CLAIM_ID,
            kind=EvidenceClaimKind.STRUCTURE,
            source_id="mcode-cancer-stage",
            source_version="4.0.0",
            uri=_STRUCTURE_URI,
            assertion=_STRUCTURE_ASSERTION,
        ),
        EvidenceClaim(
            claim_id=_METHOD_CLAIM_ID,
            kind=EvidenceClaimKind.MEMBER,
            source_id="mcode-cancer-staging-method",
            source_version="4.0.0",
            uri=_METHOD_URI,
            assertion=_METHOD_ASSERTION,
            target=EvidenceClaimTarget(
                subject_code=None,
                role=MemberRole.STAGING_METHOD,
                filler_code="C141685",
            ),
        ),
    )
)


def _source_member(
    role: str,
    filler_code: str,
    *,
    root_code: str,
    anchor_code: str,
    depth: int,
    group_id: str,
) -> SemanticBundleMember:
    typed_role = MemberRole(role)
    axis = (
        BundleAxis.STAGE_VALUE
        if typed_role is MemberRole.STAGE_VALUE
        else BundleAxis.STAGE_SYSTEM
    )
    fact_id = canonical_restriction_fact_id(
        anchor_code,
        group_id,
        "R88",
        filler_code,
    )
    return SemanticBundleMember(
        role=typed_role,
        axis=axis,
        filler_code=filler_code,
        source_occurrences=(
            SourceOccurrence(
                source_identity=_SOURCE_IDENTITY,
                ncit_release=_NCIT_RELEASE,
                root_code=root_code,
                fact_id=fact_id,
                source_role="R88",
                filler_code=filler_code,
                anchor_code=anchor_code,
                depth=depth,
                source_group_id=group_id,
            ),
        ),
    )


def _external_method(
    filler_code: str, evidence_ids: tuple[str, ...]
) -> SemanticBundleMember:
    if "mcode-4.0.0-stage-method" not in evidence_ids:
        raise ValueError("external VALG method requires mCODE method evidence")
    return SemanticBundleMember(
        role=MemberRole.STAGING_METHOD,
        axis=BundleAxis.STAGE_SYSTEM,
        filler_code=filler_code,
        source_occurrences=(),
        evidence_claim_ids=(_METHOD_CLAIM_ID,),
    )


def _rule(
    rule_id: str,
    subject_code: str,
    name: str,
    system: SemanticBundleMember,
    value: SemanticBundleMember,
    *,
    authority: str,
    version: str,
    evidence_ids: tuple[str, ...],
) -> SemanticBundleCandidate:
    known_evidence_ids = {source.evidence_id for source in EVIDENCE_SOURCES}
    if unknown := set(evidence_ids) - known_evidence_ids:
        raise ValueError(f"unknown candidate reference: {min(unknown)}")
    return SemanticBundleCandidate(
        candidate_id=rule_id,
        subject_code=subject_code,
        name=name,
        kind=BundleKind.CANCER_STAGE,
        classification=StageClassification(authority=authority, version=version),
        members=(system, value),
        evidence_claim_ids=(_STRUCTURE_CLAIM_ID,),
        evidence_source_ids=evidence_ids,
    )


_AUDIT = "ncit-26.07d-role-audit"
_MCODE = "mcode-4.0.0-cancer-stage"
_MCODE_TYPE = "mcode-4.0.0-stage-type"
_AJCC = "ajcc-official-staging-system"

_C115057_VALUE = _source_member(
    "stage-value",
    "C27966",
    root_code="C115057",
    anchor_code="C8033",
    depth=1,
    group_id="aae21beea65bfffe56c6588e65262894bf6edcf8f9f2bf47eac6b70f2cac3726",
)
_C27787_VALUE = _source_member(
    "stage-value",
    "C27970",
    root_code="C27787",
    anchor_code="C9074",
    depth=1,
    group_id="73f53459b912b33478b591c3ac93958b7f413cb5f39cf6e11e7102eef18309ac",
)

STAGE_BUNDLE_CANDIDATES = (
    _rule(
        "stage-c115057-ajcc-v6",
        "C115057",
        "AJCC v6 Stage I lip and oral cavity squamous cell carcinoma",
        _source_member(
            "stage-type",
            "C90529",
            root_code="C115057",
            anchor_code="C132736",
            depth=2,
            group_id=(
                "247877db9643ab31d55755ee76369cdcac320a5573ca31b3b7971f53998ad454"
            ),
        ),
        _C115057_VALUE,
        authority="AJCC",
        version="6",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c115057-ajcc-v7",
        "C115057",
        "AJCC v7 Stage I lip and oral cavity squamous cell carcinoma",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C115057",
            anchor_code="C132736",
            depth=2,
            group_id=(
                "247877db9643ab31d55755ee76369cdcac320a5573ca31b3b7971f53998ad454"
            ),
        ),
        _C115057_VALUE,
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c101539-ajcc-v7",
        "C101539",
        "AJCC v7 Stage I differentiated thyroid carcinoma under 45 years",
        _source_member(
            "stage-type",
            "C140961",
            root_code="C101539",
            anchor_code="C101539",
            depth=0,
            group_id=(
                "2c871f775d0fad5572b10f0df1f62aa9be26d04a5f023de0f529105184690cce"
            ),
        ),
        _source_member(
            "stage-value",
            "C27966",
            root_code="C101539",
            anchor_code="C87543",
            depth=1,
            group_id=(
                "4e04d8d24287f1200e60bc5fb443fd569b06864d083e25299398e2ab6914e66c"
            ),
        ),
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _AJCC),
    ),
    _rule(
        "stage-c132677-ajcc-v8",
        "C132677",
        "AJCC v8 Stage III unknown primary tumor with metastatic cervical adenopathy",
        _source_member(
            "stage-type",
            "C132248",
            root_code="C132677",
            anchor_code="C132676",
            depth=1,
            group_id=(
                "b1b92d49c5fb310194984326ad672d6ecd31068739420acafaeb2dfc5cad3802"
            ),
        ),
        _source_member(
            "stage-value",
            "C27970",
            root_code="C132677",
            anchor_code="C132677",
            depth=0,
            group_id=(
                "c4ba770a80dc43b78a3b9f7a66b78e3d34eacaed1f6d7ab879870a6b63be5bd5"
            ),
        ),
        authority="AJCC",
        version="8",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c181564-ajcc-v9",
        "C181564",
        "AJCC v9 Stage I cervical cancer",
        _source_member(
            "stage-type",
            "C180901",
            root_code="C181564",
            anchor_code="C181562",
            depth=1,
            group_id=(
                "7f10ca575c6631d1d39f940924f7cdfd555c0df4aee153b54e17f8605e28f754"
            ),
        ),
        _source_member(
            "stage-value",
            "C27966",
            root_code="C181564",
            anchor_code="C181564",
            depth=0,
            group_id=(
                "2295ee2a6aa7b2f09a0c61ad1feb622edf8ad54439a9459a4b4c8dcc1dd27cb8"
            ),
        ),
        authority="AJCC",
        version="9",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c186620-figo-2009",
        "C186620",
        "FIGO 2009 Stage I cervical cancer",
        _source_member(
            "stage-type",
            "C186618",
            root_code="C186620",
            anchor_code="C186619",
            depth=1,
            group_id=(
                "b7b9600ed1d8ea5b0ca11c21b03d4189a4223293763787bcad7a92fb9fee4e0c"
            ),
        ),
        _source_member(
            "stage-value",
            "C27966",
            root_code="C186620",
            anchor_code="C186620",
            depth=0,
            group_id=(
                "63391dd6710a39cd789c2bca0600a15bb9bd9bf6dd729bab957c458d95e1f29a"
            ),
        ),
        authority="FIGO",
        version="2009",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, "figo-cervix-2009"),
    ),
    _rule(
        "stage-c162226-figo-2018",
        "C162226",
        "FIGO 2018 Stage I cervical cancer",
        _source_member(
            "stage-type",
            "C186617",
            root_code="C162226",
            anchor_code="C162225",
            depth=1,
            group_id=(
                "d5ab55bccdef1e0cecf9d5ef3a1a2bf588319abb319c373589b5221b37b5885a"
            ),
        ),
        _source_member(
            "stage-value",
            "C96244",
            root_code="C162226",
            anchor_code="C162226",
            depth=0,
            group_id=(
                "5a495fb01b3c8576e7b951f2a9de8a11c8bd58d854a04ae92120fac5f11cd9a2"
            ),
        ),
        authority="FIGO",
        version="2018",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, "figo-cervix-2018"),
    ),
    _rule(
        "stage-c206219-figo-2023",
        "C206219",
        "FIGO 2023 Stage I endometrial cancer",
        _source_member(
            "stage-type",
            "C206211",
            root_code="C206219",
            anchor_code="C206217",
            depth=1,
            group_id=(
                "24a061ad6be30f67c8ab0fdbd73987f4baf0be8c8805b52a755e6a77495a184b"
            ),
        ),
        _source_member(
            "stage-value",
            "C96244",
            root_code="C206219",
            anchor_code="C206219",
            depth=0,
            group_id=(
                "558cfd1c54a0ec108027aee0739c6f50150af2fceaa017781ebe207f8244eb75"
            ),
        ),
        authority="FIGO",
        version="2023",
        evidence_ids=(_AUDIT, _MCODE, "figo-endometrial-2023"),
    ),
    _rule(
        "stage-c6135-ajcc-v7",
        "C6135",
        "AJCC v7 Stage III thyroid gland medullary carcinoma",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C6135",
            anchor_code="C141041",
            depth=1,
            group_id=(
                "6f9a00ec0975f63a9213951a287752ee720de86ae997c95f5c0655d1e8551c45"
            ),
        ),
        _source_member(
            "stage-value",
            "C27970",
            root_code="C6135",
            anchor_code="C6135",
            depth=0,
            group_id=(
                "71b0af99dd55fb44e10b66a6b17e6efa926455cf2b4ff7788680968fa28c5723"
            ),
        ),
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c35756-ajcc-v7",
        "C35756",
        "AJCC v7 Stage IIIB lung small cell carcinoma with pleural effusion",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C35756",
            anchor_code="C91232",
            depth=4,
            group_id=(
                "a1aaa29d4b372c7e7138d244e3ebcb6637f9b86926dfa0486140c67ddcd0e3e7"
            ),
        ),
        _source_member(
            "stage-value",
            "C27978",
            root_code="C35756",
            anchor_code="C5647",
            depth=2,
            group_id=(
                "dd345e6d5490e8217b1453e6549a6de96c6a41d3c3851cc67b4606f224d036dd"
            ),
        ),
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c35756-valg-extensive",
        "C35756",
        "VALG extensive-stage lung small cell carcinoma with pleural effusion",
        _external_method(
            "C141685",
            ("mcode-4.0.0-stage-method", "nci-pdq-sclc"),
        ),
        _source_member(
            "stage-value",
            "C28064",
            root_code="C35756",
            anchor_code="C9049",
            depth=1,
            group_id=(
                "5a86123986099e1e3e3223f922abfc601887daeb2f2b5046f9589d477300d6e0"
            ),
        ),
        authority="VALG",
        version="limited-extensive",
        evidence_ids=(
            _AUDIT,
            _MCODE,
            "mcode-4.0.0-stage-method",
            "nci-pdq-sclc",
        ),
    ),
    _rule(
        "stage-c89995-ajcc-v7",
        "C89995",
        "AJCC v7 Stage III colon cancer",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C89995",
            anchor_code="C91223",
            depth=2,
            group_id=(
                "f6b9daa0a913a0de40a4ae40520bf4fbda77b8a5c6ff26892513aeb79eaac7ac"
            ),
        ),
        _source_member(
            "stage-value",
            "C27970",
            root_code="C89995",
            anchor_code="C89994",
            depth=1,
            group_id=(
                "363cefec5569b1888778f6d9d210df0acc474b68c5038322c61119fea994ea47"
            ),
        ),
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c27787-ajcc-v6",
        "C27787",
        "AJCC v6 Stage III testicular non-seminomatous germ cell tumor",
        _source_member(
            "stage-type",
            "C90529",
            root_code="C27787",
            anchor_code="C140241",
            depth=2,
            group_id=(
                "0214f14693f059aae2d41b1b8426552f57dabe5cb2633ba455f21a52a5c4216f"
            ),
        ),
        _C27787_VALUE,
        authority="AJCC",
        version="6",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c27787-ajcc-v7",
        "C27787",
        "AJCC v7 Stage III testicular non-seminomatous germ cell tumor",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C27787",
            anchor_code="C140241",
            depth=2,
            group_id=(
                "0214f14693f059aae2d41b1b8426552f57dabe5cb2633ba455f21a52a5c4216f"
            ),
        ),
        _C27787_VALUE,
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
    _rule(
        "stage-c115118-ajcc-v7",
        "C115118",
        "AJCC v7 Stage IB esophageal cancer",
        _source_member(
            "stage-type",
            "C90530",
            root_code="C115118",
            anchor_code="C91221",
            depth=2,
            group_id=(
                "9456e4903fb8cdf39f626ddd74a95be4b269ee7e379156c29099aa8c671aacaa"
            ),
        ),
        _source_member(
            "stage-value",
            "C27976",
            root_code="C115118",
            anchor_code="C115118",
            depth=0,
            group_id=(
                "95549abf4674043e4470fd7002ab0fcdc5030b5786eb5080122a75670f5f4869"
            ),
        ),
        authority="AJCC",
        version="7",
        evidence_ids=(_AUDIT, _MCODE, _MCODE_TYPE, _AJCC),
    ),
)

_RULE_SOURCE_VALUE_GROUP = {
    "stage-c115057-ajcc-v6": "stage-ajcc-v6-and-v7",
    "stage-c115057-ajcc-v7": "stage-ajcc-v6-and-v7",
    "stage-c101539-ajcc-v7": "stage-ajcc-v7-dtc-under45",
    "stage-c132677-ajcc-v8": "stage-ajcc-v8",
    "stage-c181564-ajcc-v9": "stage-ajcc-v9",
    "stage-c186620-figo-2009": "stage-figo-2009",
    "stage-c162226-figo-2018": "stage-figo-2018",
    "stage-c206219-figo-2023": "stage-figo-2023",
    "stage-c6135-ajcc-v7": "stage-ajcc-v7",
    "stage-c35756-ajcc-v7": "stage-ajcc-v7",
    "stage-c35756-valg-extensive": "stage-sclc-extensive-vs-limited",
    "stage-c89995-ajcc-v7": "stage-ajcc-v7",
    "stage-c27787-ajcc-v6": "stage-ajcc-v6-and-v7",
    "stage-c27787-ajcc-v7": "stage-ajcc-v6-and-v7",
    "stage-c115118-ajcc-v7": "stage-ajcc-v7",
}

_CONTEXT_ONLY = AdjudicatedSemanticContext(
    context_id="context-c198031-toronto",
    subject_code="C198031",
    name="Toronto classification context",
    member=_source_member(
        "classification-context",
        "C198023",
        root_code="C198031",
        anchor_code="C198031",
        depth=0,
        group_id=("0d414f8ad31ecc05baa4617d99f8aa622c9c1a684f55f49120ffe79e78b594cf"),
    ),
    rationale="C198023 identifies context but no stage value is asserted.",
)


def _context_dict(context: AdjudicatedSemanticContext) -> dict[str, object]:
    return {
        "context_id": context.context_id,
        "subject_code": context.subject_code,
        "name": context.name,
        "member": _member_dict(context.member),
        "rationale": context.rationale,
    }


def _source_fact_dict(fact: SourceOccurrence) -> dict[str, object]:
    return {
        "source_identity": fact.source_identity,
        "ncit_release": fact.ncit_release,
        "root_code": fact.root_code,
        "fact_id": fact.fact_id,
        "role_code": fact.source_role,
        "filler_code": fact.filler_code,
        "anchor_code": fact.anchor_code,
        "depth": fact.depth,
        "source_group_id": fact.source_group_id,
    }


def _member_dict(member: SemanticBundleMember) -> dict[str, object]:
    return {
        "role": member.role.value,
        "axis": member.axis.value,
        "filler_code": member.filler_code,
        "source_occurrences": [
            _source_fact_dict(fact) for fact in member.source_occurrences
        ],
        "evidence_claim_ids": list(member.evidence_claim_ids),
    }


def _rule_dict(rule: SemanticBundleCandidate) -> dict[str, object]:
    return {
        "candidate_id": rule.candidate_id,
        "semantic_identity": rule.semantic_identity,
        "subject_code": rule.subject_code,
        "kind": rule.kind.value,
        "name": rule.name,
        "source_value_group": _RULE_SOURCE_VALUE_GROUP[rule.candidate_id],
        "classification": {
            "authority": rule.classification.authority,
            "version": rule.classification.version,
        },
        "members": [_member_dict(member) for member in rule.members],
        "evidence_claim_ids": list(rule.evidence_claim_ids),
        "evidence_source_ids": list(rule.evidence_source_ids),
    }


def _source_value_groups() -> list[dict[str, object]]:
    grouped: defaultdict[tuple[str, str], list[SemanticBundleCandidate]] = defaultdict(
        list
    )
    for rule in STAGE_BUNDLE_CANDIDATES:
        group_id = _RULE_SOURCE_VALUE_GROUP[rule.candidate_id]
        grouped[(rule.subject_code, group_id)].append(rule)
    result: list[dict[str, object]] = []
    for (subject_code, group_id), rules in grouped.items():
        members = {
            member.semantic_key: member for rule in rules for member in rule.members
        }
        result.append(
            {
                "subject_code": subject_code,
                "source_value_group": group_id,
                "members": [
                    _member_dict(members[key]) for key in sorted(members.keys())
                ],
                "semantic_candidate_ids": sorted(rule.candidate_id for rule in rules),
            }
        )
    return sorted(
        result,
        key=lambda item: (
            str(item["subject_code"]),
            str(item["source_value_group"]),
        ),
    )


def _audit_fact_key(value: object) -> tuple[str, str, str, str, int, str]:
    if not isinstance(value, dict):
        raise ValueError("source audit facts must be objects")
    fields = (
        value.get("root_code"),
        value.get("role_code"),
        value.get("filler_code"),
        value.get("anchor_code"),
        value.get("depth"),
        value.get("group_id"),
    )
    if (
        not all(isinstance(item, str) for item in fields[:4])
        or not isinstance(fields[4], int)
        or not isinstance(fields[5], str)
    ):
        raise ValueError("source audit fact has an invalid shape")
    return cast("tuple[str, str, str, str, int, str]", fields)


def _reference_key(fact: SourceOccurrence) -> tuple[str, str, str, str, int, str]:
    return (
        fact.root_code,
        fact.source_role,
        fact.filler_code,
        fact.anchor_code,
        fact.depth,
        fact.source_group_id,
    )


def validate_source_audit(raw_audit: object) -> None:
    """Require every claimed NCIt occurrence to exist in the bound source audit."""
    if (
        not isinstance(raw_audit, dict)
        or raw_audit.get("ncit_release") != _NCIT_RELEASE
    ):
        raise ValueError("source audit NCIt release does not match the registry")
    raw_facts = raw_audit.get("facts")
    if not isinstance(raw_facts, list):
        raise ValueError("source audit facts must be a list")
    audited = {_audit_fact_key(fact) for fact in raw_facts}
    referenced = {
        _reference_key(fact)
        for rule in STAGE_BUNDLE_CANDIDATES
        for member in rule.members
        for fact in member.source_occurrences
    }
    referenced.add(_reference_key(_CONTEXT_ONLY.member.source_occurrences[0]))
    if missing := referenced - audited:
        raise ValueError(f"missing semantic-bundle source fact: {min(missing)!r}")


def _not_evaluable(reason: str) -> dict[str, str]:
    return {"status": "not-evaluable", "reason": reason}


def _projected_dict(item: ProjectedConstituentEvidence) -> dict[str, object]:
    return {
        "axis": item.axis.value,
        "filler_code": item.filler_code,
        "needs_review": item.needs_review,
        "relationship_group": item.relationship_group,
        "source_role": item.source_role,
        "axis_source": item.axis_source,
        "source_fact_ids": list(item.source_fact_ids),
    }


def _claim_dict(claim: EvidenceClaim) -> dict[str, object]:
    target = claim.target
    return {
        "claim_id": claim.claim_id,
        "kind": claim.kind.value,
        "source_id": claim.source_id,
        "source_version": claim.source_version,
        "claim_identity": claim.claim_identity,
        "uri": claim.uri,
        "assertion": claim.assertion,
        "target": (
            {
                "subject_code": target.subject_code,
                "role": target.role.value,
                "filler_code": target.filler_code,
            }
            if target is not None
            else None
        ),
    }


def build_stage_bundle_report(
    engine_pairs_by_code: dict[str, tuple[ProjectedConstituentEvidence, ...]],
) -> dict[str, object]:
    """Report availability without inventing actual bundle associations."""
    rule_results: list[dict[str, object]] = []
    status_counts = {"available": 0, "deferred": 0, "incomplete": 0}
    member_counts = {"available": 0, "deferred": 0, "missing": 0}
    for candidate in STAGE_BUNDLE_CANDIDATES:
        validate_candidate_evidence(candidate, EVIDENCE_REGISTRY)
        result = evaluate_pair_availability(
            candidate,
            engine_pairs_by_code.get(candidate.subject_code, ()),
        )
        status_counts[result.status] += 1
        member_counts["available"] += len(result.available_members)
        member_counts["deferred"] += len(result.deferred_members)
        member_counts["missing"] += len(result.missing_members)
        rule_results.append(
            {
                "candidate_id": candidate.candidate_id,
                "name": candidate.name,
                "semantic_identity": candidate.semantic_identity,
                "status": result.status,
                "available_members": [
                    _member_dict(member) for member in result.available_members
                ],
                "deferred_members": [
                    _member_dict(member) for member in result.deferred_members
                ],
                "missing_members": [
                    _member_dict(member) for member in result.missing_members
                ],
                "available_engine_evidence": [
                    _projected_dict(item) for item in result.available_evidence
                ],
                "deferred_engine_evidence": [
                    _projected_dict(item) for item in result.deferred_evidence
                ],
            }
        )

    expected_bundles = len(STAGE_BUNDLE_CANDIDATES)
    expected_members = sum(len(rule.members) for rule in STAGE_BUNDLE_CANDIDATES)
    unavailable_reason = (
        "Engine evidence contains flat axis/filler pairs but no semantic bundle or "
        "within-bundle association identity; expected rules must not be projected back "
        "into actual output."
    )
    return {
        "schema_version": 2,
        "status": "FINAL-REVIEW-PENDING",
        "scope": {
            "family": BundleKind.CANCER_STAGE.value,
            "source_value_groups": len(_source_value_groups()),
            "semantic_bundle_candidates": expected_bundles,
            "excluded_context_only_subjects": ["C198031"],
        },
        "evidence_sources": [
            {
                "evidence_id": source.evidence_id,
                "title": source.title,
                "uri": source.uri,
                "access": source.access,
                "supports": source.supports,
            }
            for source in EVIDENCE_SOURCES
        ],
        "evidence_claim_registry": {
            "identity": EVIDENCE_REGISTRY.identity,
            "claims": [_claim_dict(claim) for claim in EVIDENCE_REGISTRY.claims],
        },
        "source_value_groups": _source_value_groups(),
        "semantic_bundle_candidates": [
            _rule_dict(rule) for rule in STAGE_BUNDLE_CANDIDATES
        ],
        "excluded_context_only_constructs": [_context_dict(_CONTEXT_ONLY)],
        "engine_pair_availability": {
            "interpretation": (
                "Flat engine pairs are partitioned into available, deferred, and "
                "missing "
                "members. None is evidence that the engine associated those members."
            ),
            "candidate_counts": {
                "expected": expected_bundles,
                **status_counts,
            },
            "member_occurrences": {
                "expected": expected_members,
                **member_counts,
            },
            "semantic_scores": {
                "exact_bundle": _not_evaluable(unavailable_reason),
                "contextual_member": _not_evaluable(unavailable_reason),
                "association": _not_evaluable(unavailable_reason),
            },
            "candidates": rule_results,
        },
    }


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _sha256(path: Path) -> str:
    return _sha256_bytes(path.read_bytes())


def _require_identity(path: Path, expected: str) -> None:
    actual = _sha256(path)
    if actual != expected:
        raise ValueError(
            f"{path.name} SHA-256 mismatch: expected {expected}, got {actual}"
        )


def _read_json_bytes(content: bytes) -> object:
    def reject_duplicates(pairs: list[tuple[str, object]]) -> dict[str, object]:
        result: dict[str, object] = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"duplicate JSON key: {key}")
            result[key] = value
        return result

    return json.loads(content, object_pairs_hook=reject_duplicates)


def _read_json(path: Path) -> object:
    return _read_json_bytes(path.read_bytes())


def _optional_engine_text(code: str, field: str, value: object) -> str | None:
    if value is not None and not isinstance(value, str):
        raise ValueError(f"{code} engine {field} must be text or null")
    return cast("str | None", value)


def _engine_fact_ids(code: str, value: object) -> tuple[str, ...]:
    if not isinstance(value, list) or not all(isinstance(item, str) for item in value):
        raise ValueError(f"{code} engine source_fact_ids must be a text list")
    return tuple(cast("list[str]", value))


def _projected_constituent(
    code: str, raw: object
) -> ProjectedConstituentEvidence | None:
    if not isinstance(raw, dict):
        raise ValueError(f"{code} engine constituent has an invalid shape")
    axis, filler = raw.get("axis"), raw.get("filler")
    if not isinstance(axis, str) or not isinstance(filler, str):
        raise ValueError(f"{code} engine constituent has an invalid pair")
    if axis not in {item.value for item in BundleAxis} or not filler.startswith("C"):
        return None
    needs_review = raw.get("needs_review")
    if not isinstance(needs_review, bool):
        raise ValueError(f"{code} engine needs_review must be boolean")
    relationship_group = _optional_engine_text(
        code, "relationship_group", raw.get("relationship_group")
    )
    source_role = _optional_engine_text(code, "source_role", raw.get("source_role"))
    axis_source = _optional_engine_text(code, "axis_source", raw.get("axis_source"))
    return ProjectedConstituentEvidence(
        axis=BundleAxis(axis),
        filler_code=filler,
        needs_review=needs_review,
        relationship_group=relationship_group,
        source_role=source_role,
        axis_source=cast("ProjectionAxisSource | None", axis_source),
        source_fact_ids=_engine_fact_ids(code, raw.get("source_fact_ids", [])),
    )


def _constituent_pairs(
    code: str, raw_constituents: object
) -> tuple[ProjectedConstituentEvidence, ...]:
    if not isinstance(raw_constituents, list):
        raise ValueError(f"{code} engine constituents must be a list")
    result = tuple(
        item
        for raw in raw_constituents
        if (item := _projected_constituent(code, raw)) is not None
    )
    pairs = [item.pair for item in result]
    if len(set(pairs)) != len(pairs):
        raise ValueError(f"{code} engine constituent pairs must be unique")
    return result


def _engine_pairs(
    raw_engine: object,
) -> dict[str, tuple[ProjectedConstituentEvidence, ...]]:
    if not isinstance(raw_engine, dict) or raw_engine.get("schema_version") != 1:
        raise ValueError("engine evidence schema version must be 1")
    if raw_engine.get("ncit_version") != _NCIT_RELEASE:
        raise ValueError("engine evidence NCIt release does not match the registry")
    concepts = raw_engine.get("concepts")
    if not isinstance(concepts, list):
        raise ValueError("engine evidence concepts must be a list")
    result: dict[str, tuple[ProjectedConstituentEvidence, ...]] = {}
    for concept in concepts:
        if not isinstance(concept, dict) or not isinstance(concept.get("code"), str):
            raise ValueError("engine evidence concept has an invalid shape")
        code = cast("str", concept["code"])
        if code in result:
            raise ValueError(f"duplicate engine concept: {code}")
        result[code] = _constituent_pairs(code, concept.get("constituents"))
    return result


def _engine_outcomes(raw_engine: object) -> dict[str, str]:
    if not isinstance(raw_engine, dict) or not isinstance(
        raw_engine.get("concepts"), list
    ):
        raise ValueError("engine evidence concepts must be a list")
    outcomes: dict[str, str] = {}
    for concept in raw_engine["concepts"]:
        if not isinstance(concept, dict):
            raise ValueError("engine evidence concept has an invalid shape")
        code, outcome = concept.get("code"), concept.get("outcome")
        if not isinstance(code, str) or not isinstance(outcome, str):
            raise ValueError("engine evidence concept outcome is invalid")
        if code in outcomes:
            raise ValueError(f"duplicate engine concept: {code}")
        outcomes[code] = outcome
    return outcomes


def _payload_identity(value: object) -> str:
    encoded = json.dumps(value, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def _audit_fact_dict(value: object) -> dict[str, object]:
    root, role, filler, anchor, depth, group_id = _audit_fact_key(value)
    fact_id = canonical_restriction_fact_id(anchor, group_id, role, filler)
    return {
        "root_code": root,
        "role_code": role,
        "filler_code": filler,
        "anchor_code": anchor,
        "depth": depth,
        "source_group_id": group_id,
        "fact_id": fact_id,
    }


type ContractedKey = tuple[str, str, str, str, int]


def _contracted_row_entries(row: object) -> tuple[tuple[ContractedKey, str], ...]:
    if not isinstance(row, dict):
        raise ValueError("contracted disposition row must be an object")
    root = row.get("root_code")
    role = row.get("role_code")
    filler = row.get("filler_code")
    disposition = row.get("disposition")
    anchors = row.get("anchors")
    if not all(isinstance(item, str) for item in (root, role, filler, disposition)):
        raise ValueError("contracted disposition row text is invalid")
    if not isinstance(anchors, list):
        raise ValueError("contracted disposition anchors must be a list")
    entries: list[tuple[ContractedKey, str]] = []
    for anchor in anchors:
        if (
            not isinstance(anchor, list)
            or len(anchor) != _ANCHOR_PART_COUNT
            or not isinstance(anchor[0], str)
            or not isinstance(anchor[1], int)
        ):
            raise ValueError("contracted disposition anchor is invalid")
        key = cast("ContractedKey", (root, role, filler, anchor[0], anchor[1]))
        entries.append((key, cast("str", disposition)))
    return tuple(entries)


def _contracted_dispositions(raw: object) -> dict[ContractedKey, str]:
    if not isinstance(raw, dict) or raw.get("source_identity") != _SOURCE_IDENTITY:
        raise ValueError("contracted disposition source identity is invalid")
    if raw.get("ontology_version") != _NCIT_RELEASE:
        raise ValueError("contracted disposition NCIt release is invalid")
    rows = raw.get("rows")
    if not isinstance(rows, list):
        raise ValueError("contracted disposition rows must be a list")
    result: dict[ContractedKey, str] = {}
    for row in rows:
        for key, disposition in _contracted_row_entries(row):
            if key in result:
                raise ValueError(f"duplicate contracted disposition: {key!r}")
            result[key] = disposition
    return result


def _semantic_fact_owners() -> dict[tuple[str, str], tuple[str, ...]]:
    owners: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    for candidate in STAGE_BUNDLE_CANDIDATES:
        for member in candidate.members:
            for occurrence in member.source_occurrences:
                owners[(occurrence.root_code, occurrence.fact_id)].append(
                    candidate.candidate_id
                )
    context_occurrence = _CONTEXT_ONLY.member.source_occurrences[0]
    owners[(context_occurrence.root_code, context_occurrence.fact_id)].append(
        _CONTEXT_ONLY.context_id
    )
    return {key: tuple(sorted(value)) for key, value in owners.items()}


def _fact_contracted_key(fact: dict[str, object]) -> ContractedKey:
    return cast(
        "ContractedKey",
        (
            fact["root_code"],
            fact["role_code"],
            fact["filler_code"],
            fact["anchor_code"],
            fact["depth"],
        ),
    )


def _provenance_disposition(
    fact: dict[str, object],
    semantic_owners: dict[tuple[str, str], tuple[str, ...]],
    contracted: dict[ContractedKey, str],
    concept_outcomes: dict[str, str],
) -> tuple[str, object, ContractedKey | None]:
    occurrence_key = cast("tuple[str, str]", (fact["root_code"], fact["fact_id"]))
    if owners := semantic_owners.get(occurrence_key):
        return "semantic-review-candidate", owners, None
    if fact["role_code"] == "R88":
        reason = "No stage bundle or context claim consumes this occurrence."
        return "retained-unmodeled-r88", reason, None
    r101_key = cast("tuple[str, str]", (fact["root_code"], fact["filler_code"]))
    if fact["role_code"] == "R101" and r101_key in _R101_SAME_AXIS_R82_COLLAPSES:
        retained = _R101_SAME_AXIS_R82_COLLAPSES[r101_key]
        reference = {
            "axis": "op:AssociatedRegion",
            "retained_filler": retained,
            "rule": "same-axis R82 collapse on a location axis",
        }
        return "collapsed-same-axis-r82", reference, None
    contracted_key = _fact_contracted_key(fact)
    if contracted_key in contracted:
        return "contracted-role-disposition", contracted[contracted_key], contracted_key
    root_code = cast("str", fact["root_code"])
    if (outcome := concept_outcomes.get(root_code)) != "decomposed":
        if outcome is None:
            raise ValueError(
                f"source occurrence concept has no engine outcome: {root_code}"
            )
        return "nondecomposed-concept-outcome", outcome, None
    if fact["role_code"] in {"R101", "R126"}:
        workbook = "M1-57_SME_Adjudication_Workbook_Adjudicated_v13.xlsx"
        return "constituent-workbook-review", workbook, None
    raise ValueError(f"source occurrence has no disposition: {occurrence_key!r}")


def build_provenance_ledger(
    raw_audit: object,
    raw_contracted_disposition: object,
    raw_engine: object,
) -> dict[str, object]:
    """Disposition every exact source occurrence without flattening inherited roots."""
    if not isinstance(raw_audit, dict) or not isinstance(raw_audit.get("facts"), list):
        raise ValueError("source audit facts must be a list")
    if raw_audit.get("fact_count") != _SOURCE_OCCURRENCE_COUNT:
        raise ValueError("source audit must contain exactly 304 occurrences")
    facts = [_audit_fact_dict(item) for item in raw_audit["facts"]]
    if len(facts) != _SOURCE_OCCURRENCE_COUNT:
        raise ValueError("source audit fact count does not match its occurrence list")
    occurrence_keys = [(item["root_code"], item["fact_id"]) for item in facts]
    if len(set(occurrence_keys)) != len(facts):
        raise ValueError("source audit occurrence identities must be unique per root")

    semantic_owners = _semantic_fact_owners()
    contracted = _contracted_dispositions(raw_contracted_disposition)
    concept_outcomes = _engine_outcomes(raw_engine)

    rows: list[dict[str, object]] = []
    counts: defaultdict[str, int] = defaultdict(int)
    used_contracted: set[ContractedKey] = set()
    for fact in facts:
        disposition, reference, contracted_key = _provenance_disposition(
            fact, semantic_owners, contracted, concept_outcomes
        )
        if contracted_key is not None:
            used_contracted.add(contracted_key)
        counts[disposition] += 1
        rows.append(fact | {"disposition": disposition, "reference": reference})

    if used_contracted != set(contracted):
        missing = min(set(contracted) - used_contracted)
        raise ValueError(
            f"contracted disposition has no source occurrence: {missing!r}"
        )
    return {
        "source_identity": _SOURCE_IDENTITY,
        "ncit_release": _NCIT_RELEASE,
        "occurrence_count": len(rows),
        "counts": dict(sorted(counts.items())),
        "occurrences": rows,
    }


def generate_stage_bundle_artifact(
    workbook_path: Path,
    source_audit_path: Path,
    engine_evidence_path: Path,
    contracted_disposition_path: Path,
) -> dict[str, object]:
    """Generate a hash-bound report from the v13 constituent and engine evidence."""
    workbook_bytes = workbook_path.read_bytes()
    source_audit_bytes = source_audit_path.read_bytes()
    engine_evidence_bytes = engine_evidence_path.read_bytes()
    contracted_disposition_bytes = contracted_disposition_path.read_bytes()
    for path, content, expected in (
        (workbook_path, workbook_bytes, _WORKBOOK_SHA256),
        (source_audit_path, source_audit_bytes, _SOURCE_AUDIT_SHA256),
        (engine_evidence_path, engine_evidence_bytes, _ENGINE_EVIDENCE_SHA256),
    ):
        actual = _sha256_bytes(content)
        if actual != expected:
            raise ValueError(
                f"{path.name} SHA-256 mismatch: expected {expected}, got {actual}"
            )
    raw_audit = _read_json_bytes(source_audit_bytes)
    raw_engine = _read_json_bytes(engine_evidence_bytes)
    raw_contracted_disposition = _read_json_bytes(contracted_disposition_bytes)
    validate_source_audit(raw_audit)
    report = build_stage_bundle_report(_engine_pairs(raw_engine))
    report["source_provenance"] = build_provenance_ledger(
        raw_audit,
        raw_contracted_disposition,
        raw_engine,
    )
    report["bindings"] = {
        "ncit_release": _NCIT_RELEASE,
        "constituent_workbook": {
            "file": workbook_path.name,
            "sha256": _WORKBOOK_SHA256,
            "attestation": "PENDING",
        },
        "source_audit": {
            "file": source_audit_path.name,
            "sha256": _SOURCE_AUDIT_SHA256,
        },
        "engine_evidence": {
            "file": engine_evidence_path.name,
            "sha256": _ENGINE_EVIDENCE_SHA256,
        },
        "contracted_disposition": {
            "file": contracted_disposition_path.name,
            "sha256": _sha256_bytes(contracted_disposition_bytes),
        },
    }
    report["artifact_identity"] = _payload_identity(report)
    return report


_DECISION_SHEET = "Semantic Bundle Decisions"
_DECISION_HEADER_ROW = 8
_DECISION_HEADERS = (
    "Candidate ID",
    "Semantic Identity",
    "Subject Code",
    "Candidate Name",
    "Authority",
    "Version",
    "Members",
    "Decision",
    "Rationale",
    "Reviewer",
    "Review Date",
)


def _validate_artifact_identity(artifact: dict[str, object]) -> None:
    identity = artifact.get("artifact_identity")
    payload = {
        key: value for key, value in artifact.items() if key != "artifact_identity"
    }
    if identity != _payload_identity(payload):
        raise ValueError("candidate artifact identity does not match its payload")


def _candidate_rows(artifact: dict[str, object]) -> list[dict[str, object]]:
    raw = artifact.get("semantic_bundle_candidates")
    if not isinstance(raw, list) or not all(isinstance(item, dict) for item in raw):
        raise ValueError("candidate artifact semantic bundles are invalid")
    rows = cast("list[dict[str, object]]", raw)
    candidate_ids = [item.get("candidate_id") for item in rows]
    if len(set(candidate_ids)) != len(rows):
        raise ValueError("candidate artifact IDs must be unique")
    return rows


def _member_summary(candidate: dict[str, object]) -> str:
    members = candidate.get("members")
    if not isinstance(members, list):
        raise ValueError("candidate members must be a list")
    parts: list[str] = []
    for member in members:
        if not isinstance(member, dict):
            raise ValueError("candidate member must be an object")
        parts.append(
            f"{member.get('role')}|{member.get('axis')}|{member.get('filler_code')}"
        )
    return "; ".join(parts)


def _sheet_headers(sheet: Worksheet, row: int) -> dict[str, int]:
    return {
        cast("str", sheet.cell(row, column).value): column
        for column in range(1, sheet.max_column + 1)
        if isinstance(sheet.cell(row, column).value, str)
    }


def _concept_rows(
    sheet: Worksheet, headers: dict[str, int], concept_code: str
) -> list[int]:
    code_column = headers["Concept Code"]
    return [
        row
        for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, code_column).value == concept_code
    ]


def _remove_expected_pair(
    sheet: Worksheet,
    headers: dict[str, int],
    correction: ConstituentCorrection,
) -> None:
    matches = [
        row
        for row in _concept_rows(sheet, headers, correction.concept_code)
        if sheet.cell(row, headers["SME Action"]).value in {"include", "revise"}
        and sheet.cell(row, headers["Expected Axis"]).value == correction.axis
        and sheet.cell(row, headers["Expected Filler"]).value == correction.filler_code
    ]
    if len(matches) != 1:
        raise ValueError(
            f"remove correction must match one expected pair: {correction!r}"
        )
    row = matches[0]
    sheet.cell(row, headers["SME Action"], "exclude")
    notes_column = headers["SME Notes"]
    existing = sheet.cell(row, notes_column).value
    prefix = f"{existing}\n\n" if isinstance(existing, str) and existing else ""
    sheet.cell(row, notes_column, f"{prefix}V14 CORRECTION. {correction.rationale}")


def _add_expected_pair(
    sheet: Worksheet,
    headers: dict[str, int],
    correction: ConstituentCorrection,
) -> None:
    concept_rows = _concept_rows(sheet, headers, correction.concept_code)
    if not concept_rows:
        raise ValueError(f"add correction concept is absent: {correction.concept_code}")
    duplicate = any(
        sheet.cell(row, headers["SME Action"]).value in {"include", "revise"}
        and sheet.cell(row, headers["Expected Axis"]).value == correction.axis
        and sheet.cell(row, headers["Expected Filler"]).value == correction.filler_code
        for row in concept_rows
    )
    if duplicate:
        raise ValueError(f"add correction already exists: {correction!r}")
    source_row = concept_rows[0]
    row = sheet.max_row + 1
    for field in ("Concept Order", "Concept Code", "Source Label"):
        target = cast("Cell", sheet.cell(row, headers[field]))
        target.value = sheet.cell(source_row, headers[field]).value
    values: dict[str, str | None] = {
        "Row Type": "ADD IF MISSING",
        "SME Action": "include",
        "Expected Axis": correction.axis,
        "Expected Filler": correction.filler_code,
        "Expected Group": None,
        "Expected needs_review": "FALSE",
        "Expected Provenance Status": "ncit-26.07d",
        "SME Notes": f"V14 CORRECTION. {correction.rationale}",
        "Row Complete?": "YES",
    }
    if "Expected Role Modality" in headers:
        values["Expected Role Modality"] = "asserted"
    for field, value in values.items():
        sheet.cell(row, headers[field], value)
    for column in range(headers["SME Action"], headers["Row Complete?"] + 1):
        sheet.cell(row, column).fill = PatternFill("solid", fgColor="FFF2CC")


def apply_constituent_corrections(
    workbook: Workbook,
    corrections: tuple[ConstituentCorrection, ...],
) -> None:
    """Apply an explicit reviewed correction set without changing engine evidence."""
    sheet = cast("Worksheet", workbook["Constituent Decisions"])
    headers = _sheet_headers(sheet, 4)
    required = {
        "Concept Order",
        "Concept Code",
        "Source Label",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "SME Notes",
        "Row Complete?",
    }
    if missing := required - headers.keys():
        raise ValueError(
            "constituent correction headers missing: " + ", ".join(missing)
        )
    identities = [
        (item.action, item.concept_code, item.axis, item.filler_code)
        for item in corrections
    ]
    if len(set(identities)) != len(identities):
        raise ValueError("constituent corrections must be unique")
    for correction in corrections:
        if correction.action == "remove":
            _remove_expected_pair(sheet, headers, correction)
        else:
            _add_expected_pair(sheet, headers, correction)


def _load_review_workbook_snapshot(
    base_workbook_path: Path,
    expected_base_sha256: str | None,
) -> Workbook:
    base_workbook_bytes = base_workbook_path.read_bytes()
    actual_base_sha256 = _sha256_bytes(base_workbook_bytes)
    if expected_base_sha256 is not None and actual_base_sha256 != expected_base_sha256:
        raise ValueError(
            f"{base_workbook_path.name} SHA-256 mismatch: expected "
            f"{expected_base_sha256}, got {actual_base_sha256}"
        )
    return load_workbook(BytesIO(base_workbook_bytes))


def write_review_workbook(
    base_workbook_path: Path,
    candidate_artifact: dict[str, object],
    output_path: Path,
    corrections: tuple[ConstituentCorrection, ...] = (),
    *,
    expected_base_sha256: str | None = None,
) -> None:
    """Add a reviewer-owned semantic decision sheet to a fresh workbook copy."""
    _validate_artifact_identity(candidate_artifact)
    workbook = _load_review_workbook_snapshot(
        base_workbook_path,
        expected_base_sha256,
    )
    if corrections:
        apply_constituent_corrections(workbook, corrections)
    reviewer_sheet = workbook["Reviewer & Attestation"]
    reviewer_sheet["B9"] = "PENDING"
    reviewer_sheet["C9"] = (
        "MUST REMAIN PENDING until the v14 constituent corrections and semantic "
        "bundle decisions receive final review."
    )
    workbook["Validation Summary"]["A1"] = (
        "Adjudication completeness checks - v14 pending final review"
    )
    if _DECISION_SHEET in workbook.sheetnames:
        raise ValueError(f"workbook already contains {_DECISION_SHEET}")
    sheet = workbook.create_sheet(_DECISION_SHEET)
    sheet["A1"] = "M1 #57 Semantic Bundle Final Review"
    sheet["A1"].font = Font(bold=True, size=14)
    metadata = (
        ("Candidate artifact identity", candidate_artifact["artifact_identity"]),
        ("NCIt release", _NCIT_RELEASE),
        ("Evidence registry identity", EVIDENCE_REGISTRY.identity),
        ("Attestation status", "FINAL-REVIEW-PENDING"),
        (
            "Instructions",
            "Choose ACCEPT, REJECT, or DEFER for every row and supply rationale, "
            "reviewer, and ISO review date. Set Attestation status to ATTESTED only "
            "after the complete sheet has been reviewed.",
        ),
    )
    for row, (label, value) in enumerate(metadata, start=2):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 1).font = Font(bold=True)
    for column, header in enumerate(_DECISION_HEADERS, start=1):
        cell = sheet.cell(_DECISION_HEADER_ROW, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row, candidate in enumerate(_candidate_rows(candidate_artifact), start=9):
        classification = candidate.get("classification")
        if not isinstance(classification, dict):
            raise ValueError("candidate classification must be an object")
        values = (
            candidate.get("candidate_id"),
            candidate.get("semantic_identity"),
            candidate.get("subject_code"),
            candidate.get("name"),
            classification.get("authority"),
            classification.get("version"),
            _member_summary(candidate),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        for column in range(8, 12):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor="FFF2CC")
    decision_validation = DataValidation(
        type="list",
        formula1='"ACCEPT,REJECT,DEFER"',
        allow_blank=True,
    )
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"H9:H{8 + len(_candidate_rows(candidate_artifact))}")
    sheet.freeze_panes = "A9"
    widths = (32, 68, 16, 58, 16, 20, 75, 14, 75, 28, 16)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(8, column).column_letter].width = width
    workbook.save(output_path)


def _review_text(value: object, field: str) -> str:
    if not isinstance(value, str) or not value.strip() or value.startswith("="):
        raise ValueError(f"semantic decision {field} must be reviewer-entered text")
    return value.strip()


def _validate_decision_sheet(
    sheet: Worksheet, candidate_artifact: dict[str, object]
) -> None:
    if sheet["B2"].value != candidate_artifact["artifact_identity"]:
        raise ValueError("workbook candidate artifact identity does not match")
    if sheet["B5"].value != "ATTESTED":
        raise ValueError("semantic bundle review is not ATTESTED")
    headers = tuple(
        sheet.cell(_DECISION_HEADER_ROW, column).value for column in range(1, 12)
    )
    if headers != _DECISION_HEADERS:
        raise ValueError("semantic decision headers do not match the schema")


def _decision_row_has_data(sheet: Worksheet, row: int) -> bool:
    return any(sheet.cell(row, column).value is not None for column in range(1, 12))


def _parse_decision_row(
    sheet: Worksheet,
    row: int,
    candidates: dict[str, dict[str, object]],
) -> dict[str, str]:
    candidate_id = _review_text(sheet.cell(row, 1).value, "candidate ID")
    if candidate_id not in candidates:
        raise ValueError(f"unknown semantic candidate: {candidate_id}")
    candidate = candidates[candidate_id]
    classification = cast("dict[str, object]", candidate["classification"])
    expected = (
        candidate["semantic_identity"],
        candidate["subject_code"],
        candidate["name"],
        classification["authority"],
        classification["version"],
        _member_summary(candidate),
    )
    actual = tuple(sheet.cell(row, column).value for column in range(2, 8))
    if actual != expected:
        raise ValueError(f"candidate fields changed in reviewer row {row}")
    decision = _review_text(sheet.cell(row, 8).value, "decision")
    if decision not in {"ACCEPT", "REJECT", "DEFER"}:
        raise ValueError(f"semantic decision is invalid in row {row}")
    reviewed_at = _review_text(sheet.cell(row, 11).value, "review date")
    try:
        date.fromisoformat(reviewed_at)
    except ValueError as error:
        raise ValueError(f"semantic review date is invalid in row {row}") from error
    return {
        "candidate_id": candidate_id,
        "decision": decision,
        "rationale": _review_text(sheet.cell(row, 9).value, "rationale"),
        "reviewer": _review_text(sheet.cell(row, 10).value, "reviewer"),
        "reviewed_at": reviewed_at,
    }


def import_review_decisions(
    workbook_path: Path,
    candidate_artifact: dict[str, object],
) -> dict[str, object]:
    """Generate canonical rules only from a complete, attested reviewer sheet."""
    workbook_bytes = workbook_path.read_bytes()
    return _import_review_decisions_snapshot(
        workbook_bytes,
        workbook_path.name,
        candidate_artifact,
    )


def _import_review_decisions_snapshot(
    workbook_bytes: bytes,
    workbook_name: str,
    candidate_artifact: dict[str, object],
) -> dict[str, object]:
    _validate_artifact_identity(candidate_artifact)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    if _DECISION_SHEET not in workbook.sheetnames:
        raise ValueError(f"workbook is missing {_DECISION_SHEET}")
    if workbook["Reviewer & Attestation"]["B9"].value != "ATTESTED":
        raise ValueError("reviewer attestation is not ATTESTED")
    sheet = workbook[_DECISION_SHEET]
    _validate_decision_sheet(sheet, candidate_artifact)

    candidates = {
        cast("str", item["candidate_id"]): item
        for item in _candidate_rows(candidate_artifact)
    }
    decisions: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in range(9, sheet.max_row + 1):
        if not _decision_row_has_data(sheet, row):
            continue
        decision = _parse_decision_row(sheet, row, candidates)
        candidate_id = decision["candidate_id"]
        if candidate_id in seen:
            raise ValueError(f"duplicate semantic candidate: {candidate_id}")
        seen.add(candidate_id)
        decisions.append(decision)
    if seen != set(candidates):
        raise ValueError("semantic decision sheet does not disposition every candidate")

    accepted_ids = {
        item["candidate_id"] for item in decisions if item["decision"] == "ACCEPT"
    }
    result: dict[str, object] = {
        "schema_version": 1,
        "status": "ATTESTED",
        "candidate_artifact_identity": candidate_artifact["artifact_identity"],
        "review_workbook": {
            "file": workbook_name,
            "sha256": _sha256_bytes(workbook_bytes),
        },
        "decisions": decisions,
        "semantic_bundle_rules": [
            candidate
            for candidate_id, candidate in candidates.items()
            if candidate_id in accepted_ids
        ],
    }
    result["artifact_identity"] = _payload_identity(result)
    return result


def load_constituent_corrections(path: Path) -> tuple[ConstituentCorrection, ...]:
    raw = _read_json(path)
    if not isinstance(raw, dict) or raw.get("schema_version") != 1:
        raise ValueError("constituent correction schema version must be 1")
    items = raw.get("corrections")
    if not isinstance(items, list):
        raise ValueError("constituent corrections must be a list")
    corrections: list[ConstituentCorrection] = []
    for item in items:
        if not isinstance(item, dict) or set(item) != {
            "action",
            "concept_code",
            "axis",
            "filler_code",
            "rationale",
        }:
            raise ValueError("constituent correction has an invalid shape")
        if not all(isinstance(value, str) for value in item.values()):
            raise ValueError("constituent correction values must be text")
        corrections.append(
            ConstituentCorrection(
                action=cast("Literal['add', 'remove']", item["action"]),
                concept_code=cast("str", item["concept_code"]),
                axis=cast("str", item["axis"]),
                filler_code=cast("str", item["filler_code"]),
                rationale=cast("str", item["rationale"]),
            )
        )
    return tuple(corrections)


_CANONICAL_RULE_KEYS = {
    "schema_version",
    "status",
    "candidate_artifact_identity",
    "review_workbook",
    "decisions",
    "semantic_bundle_rules",
    "artifact_identity",
}


def _validate_canonical_rules(
    raw: object,
    candidate_artifact: dict[str, object],
    workbook_name: str,
    workbook_sha256: str,
) -> None:
    if not isinstance(raw, dict) or set(raw) != _CANONICAL_RULE_KEYS:
        raise ValueError("canonical semantic rules have an invalid shape")
    if raw.get("schema_version") != 1 or raw.get("status") != "ATTESTED":
        raise ValueError("canonical semantic rules are not ATTESTED schema version 1")
    if raw.get("candidate_artifact_identity") != candidate_artifact.get(
        "artifact_identity"
    ):
        raise ValueError("canonical semantic rules bind a different candidate artifact")
    expected_workbook = {"file": workbook_name, "sha256": workbook_sha256}
    if raw.get("review_workbook") != expected_workbook:
        raise ValueError(
            "canonical semantic rules review workbook binding does not match"
        )
    identity = raw.get("artifact_identity")
    payload = {key: value for key, value in raw.items() if key != "artifact_identity"}
    if identity != _payload_identity(payload):
        raise ValueError("canonical semantic rules identity does not match its payload")


def build_verification_manifest(
    candidate_path: Path,
    review_workbook_path: Path,
    canonical_rules_path: Path | None = None,
) -> dict[str, object]:
    candidate_bytes = candidate_path.read_bytes()
    review_workbook_bytes = review_workbook_path.read_bytes()
    candidate = _read_json_bytes(candidate_bytes)
    if not isinstance(candidate, dict):
        raise ValueError("candidate artifact must be an object")
    _validate_artifact_identity(candidate)
    review_workbook = load_workbook(BytesIO(review_workbook_bytes), data_only=False)
    if _DECISION_SHEET not in review_workbook.sheetnames:
        raise ValueError(f"workbook is missing {_DECISION_SHEET}")
    if review_workbook[_DECISION_SHEET]["B2"].value != candidate["artifact_identity"]:
        raise ValueError("workbook candidate artifact identity does not match")
    review_workbook_sha256 = _sha256_bytes(review_workbook_bytes)
    files: dict[str, object] = {
        "candidate_artifact": {
            "file": candidate_path.name,
            "sha256": _sha256_bytes(candidate_bytes),
        },
        "review_workbook": {
            "file": review_workbook_path.name,
            "sha256": review_workbook_sha256,
        },
    }
    status = "FINAL-REVIEW-PENDING"
    if canonical_rules_path is not None:
        canonical_bytes = canonical_rules_path.read_bytes()
        canonical = _read_json_bytes(canonical_bytes)
        _validate_canonical_rules(
            canonical,
            candidate,
            review_workbook_path.name,
            review_workbook_sha256,
        )
        expected = _import_review_decisions_snapshot(
            review_workbook_bytes,
            review_workbook_path.name,
            candidate,
        )
        if canonical != expected:
            raise ValueError(
                "canonical semantic rules do not match the attested review workbook"
            )
        files["canonical_rules"] = {
            "file": canonical_rules_path.name,
            "sha256": _sha256_bytes(canonical_bytes),
        }
        status = "ATTESTED"
    manifest: dict[str, object] = {
        "schema_version": 1,
        "status": status,
        "candidate_artifact_identity": candidate["artifact_identity"],
        "files": files,
    }
    manifest["manifest_identity"] = _payload_identity(manifest)
    return manifest


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    commands = parser.add_subparsers(dest="command", required=True)
    prepare = commands.add_parser("prepare", help="Generate a pending review packet")
    prepare.add_argument("--workbook", required=True, type=Path)
    prepare.add_argument("--source-audit", required=True, type=Path)
    prepare.add_argument("--engine-evidence", required=True, type=Path)
    prepare.add_argument("--contracted-disposition", required=True, type=Path)
    prepare.add_argument("--output", required=True, type=Path)
    prepare.add_argument("--review-workbook-output", required=True, type=Path)
    prepare.add_argument("--corrections", type=Path)
    prepare.add_argument("--manifest-output", type=Path)
    finalize = commands.add_parser(
        "finalize",
        help="Import one reviewed workbook and bind its canonical outputs",
    )
    finalize.add_argument("--candidate", required=True, type=Path)
    finalize.add_argument("--review-workbook", required=True, type=Path)
    finalize.add_argument("--canonical-rules-output", required=True, type=Path)
    finalize.add_argument("--manifest-output", required=True, type=Path)
    return parser


def _write_json(path: Path, value: object) -> None:
    path.write_text(
        json.dumps(value, indent=2, sort_keys=True, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )


def _prepare(args: Any) -> None:
    report = generate_stage_bundle_artifact(
        args.workbook,
        args.source_audit,
        args.engine_evidence,
        args.contracted_disposition,
    )
    corrections: tuple[ConstituentCorrection, ...] = ()
    if args.corrections is not None:
        corrections = load_constituent_corrections(args.corrections)
        bindings = cast("dict[str, object]", report["bindings"])
        bindings["constituent_corrections"] = {
            "file": args.corrections.name,
            "sha256": _sha256(args.corrections),
            "count": len(corrections),
        }
        report.pop("artifact_identity")
        report["artifact_identity"] = _payload_identity(report)
    _write_json(args.output, report)
    write_review_workbook(
        args.workbook,
        report,
        args.review_workbook_output,
        corrections,
        expected_base_sha256=_WORKBOOK_SHA256,
    )
    if args.manifest_output is not None:
        manifest = build_verification_manifest(
            args.output,
            args.review_workbook_output,
        )
        _write_json(args.manifest_output, manifest)


def _finalize(args: Any) -> None:
    candidate = _read_json(args.candidate)
    if not isinstance(candidate, dict):
        raise ValueError("candidate artifact must be an object")
    canonical = import_review_decisions(args.review_workbook, candidate)
    _write_json(args.canonical_rules_output, canonical)
    manifest = build_verification_manifest(
        args.candidate,
        args.review_workbook,
        args.canonical_rules_output,
    )
    _write_json(args.manifest_output, manifest)


def main(argv: list[str] | None = None) -> None:
    args: Any = _parser().parse_args(argv)
    if args.command == "prepare":
        _prepare(args)
        return
    _finalize(args)


if __name__ == "__main__":
    main()
