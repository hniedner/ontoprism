"""Human-review packet for normalized relationship-group disagreements."""

from __future__ import annotations

import gzip
import hashlib
import json
import os
import re
import tempfile
from collections import Counter
from contextlib import suppress
from datetime import date, datetime
from pathlib import Path
from typing import Annotated, Literal, Self, cast
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from defusedxml import ElementTree as DefusedET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import Field, model_validator

from ontolib.common.boundary_models import StrictFrozenBoundaryModel
from ontolib.decomposition.evaluation import (
    PartitionComparison,
    PartitionDiagnosis,
    compare_common_pair_partition,
    compare_full_partition,
    grouping_difference_pairs,
)
from ontolib.decomposition.models import ConceptOutcome
from ontolib.decomposition.r101_conservation import R82PathEdge, R101ConservationReport

try:
    from scripts.research.current_evidence import (
        CurrentComparison,
        CurrentConceptComparison,
        CurrentConceptEvidence,
        CurrentEngineEvidence,
        CurrentMetrics,
        CurrentPartitionComparison,
        PairRelationSummary,
        validate_current_comparison,
    )
except ModuleNotFoundError:  # direct `python scripts/adjudication.py` entry point
    from research.current_evidence import (  # type: ignore[no-redef]
        CurrentComparison,
        CurrentConceptComparison,
        CurrentConceptEvidence,
        CurrentEngineEvidence,
        CurrentMetrics,
        CurrentPartitionComparison,
        PairRelationSummary,
        validate_current_comparison,
    )

_SHA256 = r"^[0-9a-f]{64}$"
_HISTORICAL_AGREEMENTS = 2
_HISTORICAL_COHORT = 20
_MIN_GROUPING_PAIRS = 2

Pair = tuple[str, str]
Partition = tuple[tuple[Pair, ...], ...]
RuleKind = Literal[
    "co-assertion-preservation",
    "routing",
    "specificity-collapse",
    "repeated-pairs",
    "reviewed-regrouping",
]


def _identity(value: object) -> str:
    return hashlib.sha256(
        json.dumps(
            value,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=True,
            default=lambda item: item.model_dump(mode="json"),
        ).encode()
    ).hexdigest()


class HistoricalAgreement(StrictFrozenBoundaryModel):
    numerator: int = Field(ge=0)
    denominator: int = Field(gt=0)
    provenance: Literal["historical-57"]


class SourceOccurrenceDocument(StrictFrozenBoundaryModel):
    occurrence_id: str = Field(pattern=_SHA256)
    root_code: str
    source_fact_id: str = Field(pattern=_SHA256)
    source_group_id: str = Field(pattern=_SHA256)
    anchor_code: str
    depth: int = Field(ge=0)
    role_code: str
    filler_code: str
    structural_path: tuple[int, ...]
    member_position: int = Field(ge=0)


class ControlConcept(StrictFrozenBoundaryModel):
    code: str = Field(pattern=r"^C[0-9]+$")
    outcome: Literal["semantic-excluded", "atomic-no-op"]
    actual_pair_count: Literal[0]
    interpretation: Literal["empty-partition-control-not-grouping-success"]


class ReviewCohort(StrictFrozenBoundaryModel):
    accepted_concept_count: int = Field(gt=0)
    outcome_counts: dict[ConceptOutcome, int]
    decomposed_codes: tuple[str, ...]
    controls: tuple[ControlConcept, ...]
    full_disagreement_codes: tuple[str, ...]
    common_pair_eligible_codes: tuple[str, ...]
    common_pair_ineligible_codes: tuple[str, ...]
    highest_fanout_code: str = Field(pattern=r"^C[0-9]+$")
    highest_fanout_occurrences: int = Field(ge=0)


class PairDeltaDiagnosis(StrictFrozenBoundaryModel):
    status: Literal[
        "no-pair-delta", "missing-pair", "extra-pair", "missing-and-extra-pairs"
    ]
    missing_pairs: tuple[Pair, ...]
    extra_pairs: tuple[Pair, ...]

    @model_validator(mode="after")
    def _status_matches_pairs(self) -> Self:
        expected = (
            "missing-and-extra-pairs"
            if self.missing_pairs and self.extra_pairs
            else "missing-pair"
            if self.missing_pairs
            else "extra-pair"
            if self.extra_pairs
            else "no-pair-delta"
        )
        if self.status != expected:
            raise ValueError("pair delta status does not match pair sets")
        return self


class GroupingDiagnosis(StrictFrozenBoundaryModel):
    kind: Literal[
        "agrees-on-common-pairs",
        "over-merge",
        "over-split",
        "misassignment",
        "ineligible-zero-shared-pairs",
        "ineligible-one-shared-pair",
    ]
    affected_pairs: tuple[Pair, ...]

    @model_validator(mode="after")
    def _affected_pair_shape(self) -> Self:
        defect = self.kind in {"over-merge", "over-split", "misassignment"}
        if defect != (len(self.affected_pairs) >= _MIN_GROUPING_PAIRS):
            raise ValueError("grouping diagnosis affected pairs are inconsistent")
        return self


class HumanReviewPendingDisposition(StrictFrozenBoundaryModel):
    status: Literal["human-review-pending"]
    proposed_diagnosis: Literal["intentionally-normalized"]
    acceptance: Literal["not-recorded"]


ReviewDisposition = HumanReviewPendingDisposition


class HistoricalSourceEvidence(StrictFrozenBoundaryModel):
    availability: Literal["unavailable-historical-oracle"]


class ActualPairEvidence(StrictFrozenBoundaryModel):
    availability: Literal["available"]
    pair: Pair
    occurrence_ids: tuple[str, ...] = Field(min_length=1)
    occurrences: tuple[SourceOccurrenceDocument, ...] = Field(min_length=1)

    @model_validator(mode="after")
    def _citations_match(self) -> Self:
        if self.occurrence_ids != tuple(
            item.occurrence_id for item in self.occurrences
        ):
            raise ValueError("pair occurrence IDs do not match cited occurrences")
        if any(item.filler_code != self.pair[1] for item in self.occurrences):
            raise ValueError("pair filler does not match cited source occurrence")
        return self


class ActualPairEvidenceUnavailable(StrictFrozenBoundaryModel):
    availability: Literal["unavailable-upstream"]
    pair: Pair
    reason: Literal["source-occurrence-unavailable-upstream"]


ActualPairEvidenceDocument = Annotated[
    ActualPairEvidence | ActualPairEvidenceUnavailable,
    Field(discriminator="availability"),
]


class ExpectedNormalizedGroup(StrictFrozenBoundaryModel):
    normalized_group_id: str = Field(pattern=_SHA256)
    pairs: tuple[Pair, ...] = Field(min_length=1)
    source_evidence: HistoricalSourceEvidence


class ActualNormalizedGroup(StrictFrozenBoundaryModel):
    normalized_group_id: str = Field(pattern=_SHA256)
    source_group_ids: tuple[str, ...]
    pairs: tuple[ActualPairEvidenceDocument, ...] = Field(min_length=1)

    @model_validator(mode="after")
    def _group_identities_are_distinct(self) -> Self:
        if self.normalized_group_id in self.source_group_ids:
            raise ValueError("normalized group identity aliases source group identity")
        cited_groups = tuple(
            sorted(
                {
                    item.source_group_id
                    for pair in self.pairs
                    if isinstance(pair, ActualPairEvidence)
                    for item in pair.occurrences
                }
            )
        )
        if self.source_group_ids != cited_groups:
            raise ValueError("source group identities do not match cited occurrences")
        return self


class TransformationRuleCatalogEntry(StrictFrozenBoundaryModel):
    kind: RuleKind
    evidence_status: Literal["derived", "derived-with-explicit-limitation"]

    @model_validator(mode="after")
    def _limitation_matches_kind(self) -> Self:
        limited = self.kind == "reviewed-regrouping"
        if limited != (self.evidence_status == "derived-with-explicit-limitation"):
            raise ValueError("transformation evidence limitation differs from kind")
        return self


class ReviewBoundary(StrictFrozenBoundaryModel):
    status: Literal["human-review-pending"]
    reason: Literal["machine-evidence-complete-human-decision-blank"]
    sme_adjudication: Literal["not-recorded"]


class RuleEvidenceRow(StrictFrozenBoundaryModel):
    row_identity: str = Field(pattern=_SHA256)
    kind: RuleKind
    concept_code: str = Field(pattern=r"^C[0-9]+$")
    source_occurrence_ids: tuple[str, ...] = Field(min_length=1)
    source_fact_ids: tuple[str, ...] = Field(min_length=1)
    source_group_ids: tuple[str, ...] = Field(min_length=1)
    source_occurrences: tuple[SourceOccurrenceDocument, ...] = Field(min_length=1)
    output_group_ids: tuple[str, ...]
    output_pairs: tuple[Pair, ...]
    r82_path: tuple[R82PathEdge, ...] = ()
    proposed_partition: Partition = ()
    affected_co_membership: tuple[Pair, ...] = ()
    machine_evidence: str
    machine_evidence_limitation: str | None = None

    @model_validator(mode="after")
    def _identity_matches(self) -> Self:
        if self.source_occurrence_ids != tuple(
            row.occurrence_id for row in self.source_occurrences
        ):
            raise ValueError("rule evidence occurrence documents differ")
        expected = _identity(self.model_dump(mode="json", exclude={"row_identity"}))
        if self.row_identity != expected:
            raise ValueError("rule evidence row identity differs")
        if self.r82_path and self.kind != "specificity-collapse":
            raise ValueError(
                "R82 path must occur only for specificity-collapse evidence"
            )
        expected_limitation = (
            "historical expected partition has no source citations"
            if self.kind == "reviewed-regrouping"
            else None
        )
        if self.machine_evidence_limitation != expected_limitation:
            raise ValueError("rule evidence limitation differs from kind")
        return self


class HistoricalGroupReviewRow(StrictFrozenBoundaryModel):
    row_identity: str = Field(pattern=_SHA256)
    concept_code: str = Field(pattern=r"^C[0-9]+$")
    review_type: Literal["pair-only", "grouping"]
    pair_delta: PairDeltaDiagnosis
    grouping_diagnosis: GroupingDiagnosis
    actual_group_ids: tuple[str, ...]
    evidence_row_ids: tuple[str, ...] = Field(min_length=1)
    machine_suggestion: str

    @model_validator(mode="after")
    def _identity_matches(self) -> Self:
        expected = _identity(self.model_dump(mode="json", exclude={"row_identity"}))
        if self.row_identity != expected:
            raise ValueError("review row identity differs")
        return self


class HistoricalGroupReviewConcept(StrictFrozenBoundaryModel):
    code: str = Field(pattern=r"^C[0-9]+$")
    expected_partition: Partition
    actual_partition: Partition
    pair_delta: PairDeltaDiagnosis
    common_pair_eligible: bool
    common_pair_agrees: bool | None
    grouping_diagnosis: GroupingDiagnosis
    expected_groups: tuple[ExpectedNormalizedGroup, ...]
    actual_groups: tuple[ActualNormalizedGroup, ...]
    disposition: ReviewDisposition

    @model_validator(mode="after")
    def _partitions_and_diagnoses_are_consistent(self) -> Self:
        expected_rows = _partition_rows(self.expected_partition)
        actual_rows = _partition_rows(self.actual_partition)
        full = compare_full_partition(expected_rows, actual_rows)
        common = compare_common_pair_partition(expected_rows, actual_rows)
        if (
            full.missing_pairs != self.pair_delta.missing_pairs
            or full.extra_pairs != self.pair_delta.extra_pairs
        ):
            raise ValueError("pair delta does not match normalized partitions")
        if (
            self.common_pair_eligible != common.eligible
            or self.common_pair_agrees != common.agrees
        ):
            raise ValueError("common-pair fields do not match normalized partitions")
        expected_kind, affected = _grouping_diagnosis(common)
        if (self.grouping_diagnosis.kind, self.grouping_diagnosis.affected_pairs) != (
            expected_kind,
            affected,
        ):
            raise ValueError("grouping diagnosis does not match normalized partitions")
        expected_pairs = tuple(
            sorted(pair for block in self.expected_partition for pair in block)
        )
        documented_expected = tuple(
            sorted(pair for group in self.expected_groups for pair in group.pairs)
        )
        actual_pairs = tuple(
            sorted(pair for block in self.actual_partition for pair in block)
        )
        documented_actual = tuple(
            sorted(pair.pair for group in self.actual_groups for pair in group.pairs)
        )
        if expected_pairs != documented_expected or actual_pairs != documented_actual:
            raise ValueError("normalized group documents do not match partitions")
        return self


class HistoricalGroupReviewPacket(StrictFrozenBoundaryModel):
    schema_version: Literal[3]
    source_identity: str = Field(pattern=_SHA256)
    ncit_version: str
    current_evidence_identity: str = Field(pattern=_SHA256)
    current_comparison_identity: str = Field(pattern=_SHA256)
    r101_report_identity: str = Field(pattern=_SHA256)
    historical_full_partition_agreement: HistoricalAgreement
    current_metrics: CurrentMetrics
    cohort: ReviewCohort
    transformation_rule_catalog: tuple[TransformationRuleCatalogEntry, ...] = Field(
        min_length=5, max_length=5
    )
    review_boundary: ReviewBoundary
    concepts: tuple[HistoricalGroupReviewConcept, ...] = Field(min_length=1)
    rule_kinds: tuple[RuleKind, ...] = Field(min_length=5, max_length=5)
    rule_evidence: tuple[RuleEvidenceRow, ...] = Field(min_length=5)
    review_rows: tuple[HistoricalGroupReviewRow, ...] = Field(min_length=1)
    packet_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_packet(self) -> Self:
        concept_codes = tuple(item.code for item in self.concepts)
        if concept_codes != self.cohort.full_disagreement_codes:
            raise ValueError("packet concepts do not match full disagreement cohort")
        if tuple(row.concept_code for row in self.review_rows) != concept_codes:
            raise ValueError("packet review rows do not match disagreement cohort")
        evidence_ids = {row.row_identity for row in self.rule_evidence}
        if any(
            not set(row.evidence_row_ids) <= evidence_ids for row in self.review_rows
        ):
            raise ValueError("review row cites absent rule evidence")
        expected = _identity(self.model_dump(mode="json", exclude={"packet_identity"}))
        if self.packet_identity != expected:
            raise ValueError("group review packet identity does not match payload")
        return self


PairRelationKind = Literal[
    "expected-emitted-review-bearing",
    "current-only-review-bearing",
    "current-only-proposed",
]


class NonScoreableEmittedPair(StrictFrozenBoundaryModel):
    relation: PairRelationKind
    pair: Pair
    source_occurrence_ids: tuple[str, ...]
    source_occurrences: tuple[SourceOccurrenceDocument, ...]
    normalized_group_membership: Literal["excluded-non-scoreable"]

    @model_validator(mode="after")
    def _occurrence_documents_match(self) -> Self:
        if self.source_occurrence_ids != tuple(
            item.occurrence_id for item in self.source_occurrences
        ):
            raise ValueError("non-scoreable pair occurrence documents differ")
        if any(item.filler_code != self.pair[1] for item in self.source_occurrences):
            raise ValueError("non-scoreable pair occurrence filler differs")
        if self.relation != "current-only-proposed" and not self.source_occurrences:
            raise ValueError(
                "release-bound non-scoreable pair lacks occurrence context"
            )
        return self


class GroupReviewRow(StrictFrozenBoundaryModel):
    row_identity: str = Field(pattern=_SHA256)
    concept_code: str = Field(pattern=r"^C[0-9]+$")
    review_type: Literal["pair-only", "grouping"]
    pair_relations: PairRelationSummary
    grouping_diagnosis: GroupingDiagnosis
    current_scoreable_group_ids: tuple[str, ...]
    evidence_row_ids: tuple[str, ...] = Field(min_length=1)
    machine_suggestion: str

    @model_validator(mode="after")
    def _identity_matches(self) -> Self:
        expected = _identity(self.model_dump(mode="json", exclude={"row_identity"}))
        if self.row_identity != expected:
            raise ValueError("review row identity differs")
        return self


class GroupReviewConcept(StrictFrozenBoundaryModel):
    code: str = Field(pattern=r"^C[0-9]+$")
    expected_partition: Partition
    actual_partition: Partition
    pair_relations: PairRelationSummary
    non_scoreable_emitted_pairs: tuple[NonScoreableEmittedPair, ...]
    common_pair_eligible: bool
    common_pair_agrees: bool | None
    grouping_diagnosis: GroupingDiagnosis
    expected_groups: tuple[ExpectedNormalizedGroup, ...]
    actual_groups: tuple[ActualNormalizedGroup, ...]
    disposition: ReviewDisposition

    @model_validator(mode="after")
    def _partitions_and_diagnoses_are_consistent(self) -> Self:
        expected_rows = _partition_rows(self.expected_partition)
        actual_rows = _partition_rows(self.actual_partition)
        common = compare_common_pair_partition(expected_rows, actual_rows)
        if (
            self.common_pair_eligible != common.eligible
            or self.common_pair_agrees != common.agrees
        ):
            raise ValueError("common-pair fields do not match normalized partitions")
        expected_kind, affected = _grouping_diagnosis(common)
        if (self.grouping_diagnosis.kind, self.grouping_diagnosis.affected_pairs) != (
            expected_kind,
            affected,
        ):
            raise ValueError("grouping diagnosis does not match normalized partitions")
        expected_pairs = tuple(
            sorted(pair for block in self.expected_partition for pair in block)
        )
        documented_expected = tuple(
            sorted(pair for group in self.expected_groups for pair in group.pairs)
        )
        actual_pairs = tuple(
            sorted(pair for block in self.actual_partition for pair in block)
        )
        documented_actual = tuple(
            sorted(pair.pair for group in self.actual_groups for pair in group.pairs)
        )
        if expected_pairs != documented_expected or actual_pairs != documented_actual:
            raise ValueError("normalized group documents do not match partitions")
        context_pairs = tuple(item.pair for item in self.non_scoreable_emitted_pairs)
        expected_context = (
            self.pair_relations.expected_emitted_review_bearing
            + self.pair_relations.current_only_review_bearing
            + self.pair_relations.current_only_proposed
        )
        if tuple(sorted(context_pairs)) != tuple(sorted(expected_context)):
            raise ValueError("non-scoreable pair context differs from pair relations")
        return self


class GroupReviewPacket(StrictFrozenBoundaryModel):
    schema_version: Literal[4]
    source_identity: str = Field(pattern=_SHA256)
    ncit_version: str
    current_evidence_identity: str = Field(pattern=_SHA256)
    current_comparison_identity: str = Field(pattern=_SHA256)
    r101_report_identity: str = Field(pattern=_SHA256)
    historical_full_partition_agreement: HistoricalAgreement
    current_metrics: CurrentMetrics
    cohort: ReviewCohort
    transformation_rule_catalog: tuple[TransformationRuleCatalogEntry, ...] = Field(
        min_length=5, max_length=5
    )
    review_boundary: ReviewBoundary
    concepts: tuple[GroupReviewConcept, ...] = Field(min_length=1)
    rule_kinds: tuple[RuleKind, ...] = Field(min_length=5, max_length=5)
    rule_evidence: tuple[RuleEvidenceRow, ...] = Field(min_length=5)
    review_rows: tuple[GroupReviewRow, ...] = Field(min_length=1)
    packet_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _validate_packet(self) -> Self:
        concept_codes = tuple(item.code for item in self.concepts)
        if concept_codes != self.cohort.full_disagreement_codes:
            raise ValueError("packet concepts do not match full disagreement cohort")
        if tuple(row.concept_code for row in self.review_rows) != concept_codes:
            raise ValueError("packet review rows do not match disagreement cohort")
        evidence_ids = {row.row_identity for row in self.rule_evidence}
        if any(
            not set(row.evidence_row_ids) <= evidence_ids for row in self.review_rows
        ):
            raise ValueError("review row cites absent rule evidence")
        expected = _identity(self.model_dump(mode="json", exclude={"packet_identity"}))
        if self.packet_identity != expected:
            raise ValueError("group review packet identity does not match payload")
        return self


def _partition_rows(partition: Partition) -> tuple[tuple[Pair, str], ...]:
    return tuple(
        (pair, f"group-{index}")
        for index, block in enumerate(partition)
        for pair in block
    )


def _grouping_diagnosis(
    comparison: PartitionComparison | CurrentPartitionComparison,
) -> tuple[
    Literal[
        "agrees-on-common-pairs",
        "over-merge",
        "over-split",
        "misassignment",
        "ineligible-zero-shared-pairs",
        "ineligible-one-shared-pair",
    ],
    tuple[Pair, ...],
]:
    eligible = comparison.eligible
    if not eligible:
        if comparison.ineligibility_reason == "zero-shared-pairs":
            return "ineligible-zero-shared-pairs", ()
        return "ineligible-one-shared-pair", ()
    if comparison.agrees:
        return "agrees-on-common-pairs", ()
    affected = grouping_difference_pairs(
        comparison.expected_partition, comparison.actual_partition
    )
    # The comparison has already induced the partitions onto shared pairs.
    diagnosis = comparison.primary_diagnosis
    if diagnosis is None:
        raise ValueError("disagreeing eligible partition lacks a diagnosis")
    if isinstance(comparison, CurrentPartitionComparison):
        typed = comparison.primary_diagnosis
        if typed is None:
            raise ValueError("disagreeing eligible partition lacks a diagnosis")
        return typed.diagnosis.value, affected
    return cast("PartitionDiagnosis", diagnosis).value, affected


def diagnose_grouping(expected: Partition, actual: Partition) -> GroupingDiagnosis:
    """Return the total common-pair diagnosis for two normalized partitions."""
    comparison = compare_common_pair_partition(
        _partition_rows(expected), _partition_rows(actual)
    )
    kind, affected = _grouping_diagnosis(comparison)
    return GroupingDiagnosis(kind=kind, affected_pairs=affected)


def _occurrence_document(value: object) -> SourceOccurrenceDocument:
    model_dump = getattr(value, "model_dump", None)
    if model_dump is None:
        raise TypeError("source occurrence must be a boundary model")
    return SourceOccurrenceDocument.model_validate(model_dump())


def _expected_groups(
    code: str, partition: Partition
) -> tuple[ExpectedNormalizedGroup, ...]:
    return tuple(
        ExpectedNormalizedGroup(
            normalized_group_id=_identity(
                {"stage": "historical-expected", "code": code, "pairs": block}
            ),
            pairs=block,
            source_evidence=HistoricalSourceEvidence(
                availability="unavailable-historical-oracle"
            ),
        )
        for block in partition
    )


def _actual_groups(
    concept: CurrentConceptEvidence, partition: Partition
) -> tuple[ActualNormalizedGroup, ...]:
    by_pair = {(item.axis, item.filler): item for item in concept.constituents}
    result: list[ActualNormalizedGroup] = []
    for block in partition:
        pairs: list[ActualPairEvidenceDocument] = []
        for pair in block:
            constituent = by_pair.get(pair)
            if constituent is None:
                raise ValueError(
                    "actual normalized partition pair is absent from evidence"
                )
            if constituent.source_occurrences:
                pairs.append(
                    ActualPairEvidence(
                        availability="available",
                        pair=pair,
                        occurrence_ids=constituent.source_occurrence_ids,
                        occurrences=tuple(
                            _occurrence_document(item)
                            for item in constituent.source_occurrences
                        ),
                    )
                )
            else:
                pairs.append(
                    ActualPairEvidenceUnavailable(
                        availability="unavailable-upstream",
                        pair=pair,
                        reason="source-occurrence-unavailable-upstream",
                    )
                )
        source_groups = tuple(
            sorted(
                {
                    item.source_group_id
                    for pair in pairs
                    if isinstance(pair, ActualPairEvidence)
                    for item in pair.occurrences
                }
            )
        )
        result.append(
            ActualNormalizedGroup(
                normalized_group_id=_identity(
                    {
                        "stage": "current-normalized",
                        "code": concept.code,
                        "pairs": block,
                    }
                ),
                source_group_ids=source_groups,
                pairs=tuple(pairs),
            )
        )
    return tuple(result)


def _validate_actual_partition(
    comparison: CurrentConceptComparison, evidence: CurrentConceptEvidence
) -> None:
    rows = tuple(
        ((item.axis, item.filler), item.relationship_group)
        for item in evidence.constituents
        if not item.needs_review and item.provenance_status == "ncit-26.07d"
    )
    actual = compare_full_partition(rows, rows).actual_partition
    if comparison.full_partition.actual_partition != actual:
        raise ValueError("actual normalized partition does not match current evidence")


def _non_scoreable_emitted_pairs(
    comparison: CurrentConceptComparison,
    evidence: CurrentConceptEvidence,
) -> tuple[NonScoreableEmittedPair, ...]:
    by_pair = {(item.axis, item.filler): item for item in evidence.constituents}
    relations: tuple[tuple[PairRelationKind, tuple[Pair, ...]], ...] = (
        (
            "expected-emitted-review-bearing",
            comparison.pair_relations.expected_emitted_review_bearing,
        ),
        (
            "current-only-review-bearing",
            comparison.pair_relations.current_only_review_bearing,
        ),
        (
            "current-only-proposed",
            comparison.pair_relations.current_only_proposed,
        ),
    )
    return tuple(
        NonScoreableEmittedPair(
            relation=relation,
            pair=pair,
            source_occurrence_ids=by_pair[pair].source_occurrence_ids,
            source_occurrences=tuple(
                _occurrence_document(item) for item in by_pair[pair].source_occurrences
            ),
            normalized_group_membership="excluded-non-scoreable",
        )
        for relation, pairs in relations
        for pair in pairs
    )


def _concept_packet(
    comparison: CurrentConceptComparison,
    evidence: CurrentConceptEvidence,
) -> GroupReviewConcept:
    _validate_actual_partition(comparison, evidence)
    full = comparison.full_partition
    common = comparison.common_pair_partition
    return GroupReviewConcept(
        code=comparison.code,
        expected_partition=full.expected_partition,
        actual_partition=full.actual_partition,
        pair_relations=comparison.pair_relations,
        non_scoreable_emitted_pairs=_non_scoreable_emitted_pairs(comparison, evidence),
        common_pair_eligible=common.eligible,
        common_pair_agrees=common.agrees,
        grouping_diagnosis=diagnose_grouping(
            common.expected_partition, common.actual_partition
        ),
        expected_groups=_expected_groups(comparison.code, full.expected_partition),
        actual_groups=_actual_groups(evidence, full.actual_partition),
        disposition=HumanReviewPendingDisposition(
            status="human-review-pending",
            proposed_diagnosis="intentionally-normalized",
            acceptance="not-recorded",
        ),
    )


def _validate_inputs(
    evidence: CurrentEngineEvidence, comparison: CurrentComparison
) -> None:
    validate_current_comparison(evidence, comparison)
    evidence_codes = tuple(item.code for item in evidence.concepts)
    comparison_codes = tuple(item.code for item in comparison.concepts)
    if len(evidence_codes) != len(set(evidence_codes)):
        raise ValueError("current evidence contains duplicate concepts")
    if len(comparison_codes) != len(set(comparison_codes)):
        raise ValueError("current comparison contains duplicate concepts")
    if set(comparison_codes) != set(evidence_codes):
        raise ValueError("current comparison concepts do not match current evidence")


def _rule_row(
    *,
    kind: RuleKind,
    concept_code: str,
    occurrences: tuple[SourceOccurrenceDocument, ...],
    output_group_ids: tuple[str, ...] = (),
    output_pairs: tuple[Pair, ...] = (),
    r82_path: tuple[R82PathEdge, ...] = (),
    proposed_partition: Partition = (),
    affected_co_membership: tuple[Pair, ...] = (),
    machine_evidence: str,
    machine_evidence_limitation: str | None = None,
) -> RuleEvidenceRow:
    canonical = tuple(sorted(occurrences, key=lambda row: row.occurrence_id))
    payload = {
        "kind": kind,
        "concept_code": concept_code,
        "source_occurrence_ids": tuple(row.occurrence_id for row in canonical),
        "source_fact_ids": tuple(sorted({row.source_fact_id for row in canonical})),
        "source_group_ids": tuple(sorted({row.source_group_id for row in canonical})),
        "source_occurrences": canonical,
        "output_group_ids": tuple(sorted(output_group_ids)),
        "output_pairs": tuple(sorted(output_pairs)),
        "r82_path": r82_path,
        "proposed_partition": proposed_partition,
        "affected_co_membership": tuple(sorted(affected_co_membership)),
        "machine_evidence": machine_evidence,
        "machine_evidence_limitation": machine_evidence_limitation,
    }
    return RuleEvidenceRow(**payload, row_identity=_identity(payload))


def _available_occurrences(
    group: ActualNormalizedGroup,
) -> tuple[SourceOccurrenceDocument, ...]:
    return tuple(
        occurrence
        for pair in group.pairs
        if isinstance(pair, ActualPairEvidence)
        for occurrence in pair.occurrences
    )


def _machine_rule_evidence(  # noqa: C901
    concepts: tuple[GroupReviewConcept, ...],
    evidence: CurrentEngineEvidence,
    report: R101ConservationReport,
) -> tuple[RuleEvidenceRow, ...]:
    result: list[RuleEvidenceRow] = []
    evidence_by_code = {row.code: row for row in evidence.concepts}
    packet_by_code = {row.code: row for row in concepts}
    for concept in concepts:
        all_occurrences = tuple(
            _occurrence_document(row)
            for row in evidence_by_code[concept.code].all_source_occurrences
        )
        for group in concept.actual_groups:
            occurrences = _available_occurrences(group)
            if occurrences:
                result.append(
                    _rule_row(
                        kind="co-assertion-preservation",
                        concept_code=concept.code,
                        occurrences=occurrences,
                        output_group_ids=(group.normalized_group_id,),
                        output_pairs=tuple(pair.pair for pair in group.pairs),
                        machine_evidence=(
                            "Exact source groups and occurrences consumed by this "
                            "normalized output group."
                        ),
                    )
                )
            for pair in group.pairs:
                if not isinstance(pair, ActualPairEvidence):
                    continue
                role_codes = ", ".join(
                    sorted({row.role_code for row in pair.occurrences})
                )
                result.append(
                    _rule_row(
                        kind="routing",
                        concept_code=concept.code,
                        occurrences=pair.occurrences,
                        output_group_ids=(group.normalized_group_id,),
                        output_pairs=(pair.pair,),
                        machine_evidence=(
                            "Source role occurrence routed to its exact normalized "
                            f"axis ({role_codes})."
                        ),
                    )
                )
                if len(pair.occurrences) > 1:
                    result.append(
                        _rule_row(
                            kind="repeated-pairs",
                            concept_code=concept.code,
                            occurrences=pair.occurrences,
                            output_group_ids=(group.normalized_group_id,),
                            output_pairs=(pair.pair,),
                            machine_evidence=(
                                "All exact source occurrences that normalize to this "
                                "repeated output pair."
                            ),
                        )
                    )
        regrouping_occurrences = (
            tuple(
                row
                for group in concept.actual_groups
                for row in _available_occurrences(group)
            )
            or all_occurrences
        )
        result.append(
            _rule_row(
                kind="reviewed-regrouping",
                concept_code=concept.code,
                occurrences=regrouping_occurrences,
                output_group_ids=tuple(
                    group.normalized_group_id for group in concept.actual_groups
                ),
                output_pairs=tuple(
                    pair for block in concept.actual_partition for pair in block
                ),
                proposed_partition=concept.expected_partition,
                affected_co_membership=(
                    concept.grouping_diagnosis.affected_pairs
                    or tuple(
                        pair for block in concept.actual_partition for pair in block
                    )
                ),
                machine_evidence=(
                    "Deterministic current-versus-historical partition proposal; "
                    "this is a review witness, not a source-stated grouping."
                ),
                machine_evidence_limitation=(
                    "historical expected partition has no source citations"
                ),
            )
        )
    occurrence_by_id = {
        row.occurrence_id: _occurrence_document(row)
        for concept in evidence.concepts
        for row in concept.all_source_occurrences
    }
    for row in report.occurrences:
        if (
            row.concept_code not in packet_by_code
            or row.retained_r82_target is None
            or row.occurrence_id not in occurrence_by_id
        ):
            continue
        concept = packet_by_code[row.concept_code]
        target = (row.retained_r82_target.axis, row.retained_r82_target.filler_code)
        groups = tuple(
            group.normalized_group_id
            for group in concept.actual_groups
            if any(pair.pair == target for pair in group.pairs)
        )
        if not groups:
            continue
        result.append(
            _rule_row(
                kind="specificity-collapse",
                concept_code=row.concept_code,
                occurrences=(occurrence_by_id[row.occurrence_id],),
                output_group_ids=groups,
                output_pairs=(target,),
                r82_path=row.r82_path,
                machine_evidence=(
                    "R101 same-axis specificity collapse with exact asserted R82 path."
                ),
            )
        )
    kinds = {row.kind for row in result}
    missing = set(cast("tuple[RuleKind, ...]", _RULE_KINDS)) - kinds
    if missing:
        raise ValueError(f"machine rule evidence is absent for: {sorted(missing)}")
    return tuple(
        sorted(result, key=lambda row: (row.concept_code, row.kind, row.row_identity))
    )


_RULE_KINDS: tuple[RuleKind, ...] = (
    "co-assertion-preservation",
    "routing",
    "specificity-collapse",
    "repeated-pairs",
    "reviewed-regrouping",
)


def _review_rows(
    concepts: tuple[GroupReviewConcept, ...], evidence: tuple[RuleEvidenceRow, ...]
) -> tuple[GroupReviewRow, ...]:
    result: list[GroupReviewRow] = []
    for concept in concepts:
        evidence_ids = tuple(
            row.row_identity for row in evidence if row.concept_code == concept.code
        )
        grouping = concept.grouping_diagnosis.kind in {
            "over-merge",
            "over-split",
            "misassignment",
        }
        payload = {
            "concept_code": concept.code,
            "review_type": "grouping" if grouping else "pair-only",
            "pair_relations": concept.pair_relations,
            "grouping_diagnosis": concept.grouping_diagnosis,
            "current_scoreable_group_ids": tuple(
                group.normalized_group_id for group in concept.actual_groups
            ),
            "evidence_row_ids": evidence_ids,
            "machine_suggestion": (
                "Review proposed co-membership against exact source occurrences."
                if grouping
                else "Review pair relations separately from scoreable grouping."
            ),
        }
        result.append(GroupReviewRow(**payload, row_identity=_identity(payload)))
    return tuple(result)


def build_group_review_packet(
    *,
    evidence: CurrentEngineEvidence,
    comparison: CurrentComparison,
    r101_report: R101ConservationReport,
) -> GroupReviewPacket:
    """Derive the current disagreement packet without making an SME decision."""
    _validate_inputs(evidence, comparison)
    if (
        r101_report.source_identity != evidence.source_identity
        or r101_report.source_release_id != evidence.ncit_version
    ):
        raise ValueError("R101 path evidence does not bind current evidence")
    by_code = {item.code: item for item in evidence.concepts}
    disagreements = tuple(
        item for item in comparison.concepts if item.full_partition.agrees is False
    )
    controls = tuple(
        ControlConcept(
            code=item.code,
            outcome=cast("Literal['semantic-excluded', 'atomic-no-op']", item.outcome),
            actual_pair_count=cast("Literal[0]", len(item.constituents)),
            interpretation="empty-partition-control-not-grouping-success",
        )
        for item in evidence.concepts
        if item.outcome in {"semantic-excluded", "atomic-no-op"}
    )
    highest = max(
        evidence.concepts,
        key=lambda item: (len(item.all_source_occurrences), item.code),
    )
    concepts = tuple(
        _concept_packet(item, by_code[item.code]) for item in disagreements
    )
    rule_evidence = _machine_rule_evidence(concepts, evidence, r101_report)
    payload = {
        "schema_version": 4,
        "source_identity": comparison.source_identity,
        "ncit_version": comparison.ncit_version,
        "current_evidence_identity": evidence.evidence_identity,
        "current_comparison_identity": comparison.comparison_identity,
        "r101_report_identity": r101_report.report_identity,
        "historical_full_partition_agreement": HistoricalAgreement(
            numerator=_HISTORICAL_AGREEMENTS,
            denominator=_HISTORICAL_COHORT,
            provenance="historical-57",
        ),
        "current_metrics": comparison.metrics,
        "cohort": ReviewCohort(
            accepted_concept_count=len(comparison.concepts),
            outcome_counts=cast(
                "dict[ConceptOutcome, int]",
                dict(
                    sorted(Counter(item.outcome for item in evidence.concepts).items())
                ),
            ),
            decomposed_codes=tuple(
                item.code for item in evidence.concepts if item.outcome == "decomposed"
            ),
            controls=controls,
            full_disagreement_codes=tuple(item.code for item in disagreements),
            common_pair_eligible_codes=tuple(
                item.code
                for item in comparison.concepts
                if item.common_pair_partition.eligible
            ),
            common_pair_ineligible_codes=tuple(
                item.code
                for item in comparison.concepts
                if not item.common_pair_partition.eligible
            ),
            highest_fanout_code=highest.code,
            highest_fanout_occurrences=len(highest.all_source_occurrences),
        ),
        "transformation_rule_catalog": tuple(
            TransformationRuleCatalogEntry(
                kind=kind,
                evidence_status=(
                    "derived-with-explicit-limitation"
                    if kind == "reviewed-regrouping"
                    else "derived"
                ),
            )
            for kind in _RULE_KINDS
        ),
        "review_boundary": ReviewBoundary(
            status="human-review-pending",
            reason="machine-evidence-complete-human-decision-blank",
            sme_adjudication="not-recorded",
        ),
        "concepts": concepts,
        "rule_kinds": _RULE_KINDS,
        "rule_evidence": rule_evidence,
        "review_rows": _review_rows(concepts, rule_evidence),
    }
    return GroupReviewPacket(**payload, packet_identity=_identity(payload))


def build_machine_group_review_packet(
    *,
    evidence: CurrentEngineEvidence,
    comparison: CurrentComparison,
    r101_report_path: Path,
) -> GroupReviewPacket:
    """Build the complete machine boundary while leaving all SME fields absent."""
    report = R101ConservationReport.model_validate_json(
        gzip.decompress(r101_report_path.read_bytes())
    )
    return build_group_review_packet(
        evidence=evidence, comparison=comparison, r101_report=report
    )


def load_group_review_packet(path: Path) -> GroupReviewPacket:
    """Load and identity-check a persisted review packet."""
    return GroupReviewPacket.model_validate_json(path.read_bytes())


def load_historical_group_review_packet(path: Path) -> HistoricalGroupReviewPacket:
    """Load the immutable schema-3 packet used only by historical rationale."""
    return HistoricalGroupReviewPacket.model_validate_json(path.read_bytes())


Decision = Literal[
    "Approve intentional normalization",
    "Require source-reproducible correction",
    "Reject proposed regrouping",
    "Abstain / escalate",
]
_DECISIONS: tuple[Decision, ...] = (
    "Approve intentional normalization",
    "Require source-reproducible correction",
    "Reject proposed regrouping",
    "Abstain / escalate",
)
_REVIEW_HEADERS = (
    "Concept",
    "Review Type",
    "Pair Relations",
    "Grouping Diagnosis",
    "Current Scoreable Group IDs",
    "Evidence Row IDs",
    "Evidence Links",
    "Machine Evidence / Suggestion",
    "Pair Decision",
    "Decision",
    "Rationale",
    "Reviewer",
    "Date",
)


class GroupDecision(StrictFrozenBoundaryModel):
    review_row_identity: str = Field(pattern=_SHA256)
    concept_code: str = Field(pattern=r"^C[0-9]+$")
    review_type: Literal["pair-only", "grouping"]
    pair_decision: Decision | None
    decision: Decision
    rationale: str = Field(min_length=1)
    reviewer: str = Field(min_length=1)
    review_date: str = Field(pattern=r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")


class GroupDecisionRegistry(StrictFrozenBoundaryModel):
    schema_version: Literal[2]
    packet_identity: str = Field(pattern=_SHA256)
    source_identity: str = Field(pattern=_SHA256)
    evidence_identity: str = Field(pattern=_SHA256)
    comparison_identity: str = Field(pattern=_SHA256)
    decisions: tuple[GroupDecision, ...] = Field(min_length=1)
    registry_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _identity_matches(self) -> Self:
        expected = _identity(
            self.model_dump(mode="json", exclude={"registry_identity"})
        )
        if self.registry_identity != expected:
            raise ValueError("group decision registry identity differs")
        return self


class GroupDecisionDryRun(StrictFrozenBoundaryModel):
    writes_performed: Literal[False]
    unresolved_count: int = Field(ge=0)
    deferred_count: int = Field(ge=0)
    affected_concepts: tuple[str, ...]
    affected_groups: tuple[str, ...]
    affected_traces: tuple[str, ...]


class BlankGroupReviewValidation(StrictFrozenBoundaryModel):
    schema_version: Literal[1]
    packet_identity: str = Field(pattern=_SHA256)
    review_workbook_identity: str = Field(pattern=_SHA256)
    correction_audit_identity: str = Field(pattern=_SHA256)
    writes_performed: Literal[False]
    human_decisions_present: Literal[False]
    authorization: Literal[False]
    readiness_eligible: Literal[False]


class GroupReviewRationaleRow(StrictFrozenBoundaryModel):
    concept_code: str = Field(pattern=r"^C[0-9]+$")
    review_row_identity: str = Field(pattern=_SHA256)
    review_type: Literal["pair-only", "grouping"]
    decision: Decision
    pair_decision: Decision | None
    rationale_sha256: str = Field(pattern=_SHA256)
    rationale_chars: int = Field(gt=0)


class GroupReviewRationaleSidecar(StrictFrozenBoundaryModel):
    schema_version: Literal[1]
    evidence_kind: Literal["group-review-rationale"]
    packet_identity: str = Field(pattern=_SHA256)
    ncit_version: str
    reviewer: str = Field(min_length=1)
    review_date: str = Field(pattern=r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")
    markdown_path: str
    markdown_sha256: str = Field(pattern=_SHA256)
    rows: tuple[GroupReviewRationaleRow, ...] = Field(min_length=18, max_length=18)
    sidecar_identity: str = Field(pattern=_SHA256)

    @model_validator(mode="after")
    def _identity_matches(self) -> Self:
        expected = _identity(self.model_dump(mode="json", exclude={"sidecar_identity"}))
        if self.sidecar_identity != expected:
            raise ValueError("group review rationale sidecar identity differs")
        return self


class LoadedGroupReviewRationaleRow(StrictFrozenBoundaryModel):
    concept_code: str
    review_row_identity: str
    review_type: Literal["pair-only", "grouping"]
    decision: Decision
    pair_decision: Decision | None
    rationale: str
    rationale_sha256: str
    rationale_chars: int


class LoadedGroupReviewRationale(StrictFrozenBoundaryModel):
    sidecar: GroupReviewRationaleSidecar
    rows: tuple[LoadedGroupReviewRationaleRow, ...]

    @property
    def reviewer(self) -> str:
        return self.sidecar.reviewer

    @property
    def review_date(self) -> str:
        return self.sidecar.review_date


def load_group_decision_registry(path: Path) -> GroupDecisionRegistry:
    """Load an identity-checked group decision artifact."""
    return GroupDecisionRegistry.model_validate_json(path.read_bytes())


_SECTION_HEADING = re.compile(r"^## ([0-9]+)\. (C[0-9]+) — .+$", re.MULTILINE)
_SELECTED_OUTCOME = re.compile(r"^- \*\*Selec[t]ed outcome:\*\* (.+)$", re.MULTILINE)
_REVIEW_TYPE = re.compile(
    r"^- \*\*Review type:\*\* (pair-only|grouping)$", re.MULTILINE
)
_DIAGNOSIS = re.compile(r"^- \*\*Diagnosis:\*\* (.+)$", re.MULTILINE)
_EXPECTED_ACTUAL = re.compile(r"^- \*\*Expected vs actual:\*\* (.+)$", re.MULTILINE)
_RATIONALE_MARKER = "Human rationale:\n"
_RATIONALE_SECTION_CODES = (
    "C181564",
    "C186620",
    "C162226",
    "C206219",
    "C115118",
    "C89995",
    "C27787",
    "C115057",
    "C101539",
    "C132677",
    "C6135",
    "C100051",
    "C4791",
    "C27262",
    "C102870",
    "C100054",
    "C198031",
    "C35756",
)


def rationale_sha256(rationale: str) -> str:
    """Digest the exact UTF-8 rationale text extracted from the Markdown boundary."""
    return hashlib.sha256(rationale.encode("utf-8")).hexdigest()


def _one_match(pattern: re.Pattern[str], section: str, field: str) -> str:
    matches = pattern.findall(section)
    if len(matches) != 1:
        raise ValueError(f"group review Markdown {field} differs")
    return matches[0]


def _parse_group_review_markdown(  # noqa: C901
    markdown: bytes, packet: HistoricalGroupReviewPacket
) -> tuple[
    str,
    str,
    tuple[tuple[HistoricalGroupReviewRow, Decision, str], ...],
]:
    try:
        text = markdown.decode("utf-8")
    except UnicodeDecodeError as error:
        raise ValueError("group review Markdown is not UTF-8") from error
    reviewer_matches = re.findall(
        r"^- Reviewer confirmed: \*\*(.+)\*\*$", text, re.MULTILINE
    )
    date_matches = re.findall(
        r"^- Review date confirmed: \*\*([0-9]{4}-[0-9]{2}-[0-9]{2})\*\*$",
        text,
        re.MULTILINE,
    )
    if len(reviewer_matches) != 1 or len(date_matches) != 1:
        raise ValueError("group review Markdown reviewer/date differs")
    _normalize_review_date(date_matches[0])
    headings = tuple(_SECTION_HEADING.finditer(text))
    if set(_RATIONALE_SECTION_CODES) != {
        row.concept_code for row in packet.review_rows
    }:
        raise ValueError("group review Markdown cohort differs from packet")
    expected_sections = tuple(enumerate(_RATIONALE_SECTION_CODES, start=1))
    observed_sections = tuple(
        (int(match.group(1)), match.group(2)) for match in headings
    )
    if observed_sections != expected_sections:
        raise ValueError("group review Markdown review sections differ")
    parsed: list[tuple[HistoricalGroupReviewRow, Decision, str]] = []
    packet_by_code = {row.concept_code: row for row in packet.review_rows}
    for position, heading in enumerate(headings):
        row = packet_by_code[heading.group(2)]
        end = (
            headings[position + 1].start()
            if position + 1 < len(headings)
            else len(text)
        )
        section = text[heading.start() : end]
        selected = _one_match(_SELECTED_OUTCOME, section, "selected outcome")
        if selected not in _DECISIONS:
            raise ValueError("group review Markdown selected outcome differs")
        review_type = _one_match(_REVIEW_TYPE, section, "review type")
        diagnosis = _one_match(_DIAGNOSIS, section, "diagnosis")
        _one_match(_EXPECTED_ACTUAL, section, "expected-vs-actual")
        if review_type != row.review_type:
            raise ValueError("group review Markdown review type differs from packet")
        if diagnosis != row.grouping_diagnosis.kind:
            raise ValueError("group review Markdown diagnosis differs from packet")
        if section.count(_RATIONALE_MARKER) != 1:
            raise ValueError("group review Markdown Human rationale marker differs")
        rationale_start = section.index(_RATIONALE_MARKER) + len(_RATIONALE_MARKER)
        remainder = section[rationale_start:]
        delimiter = remainder.find("\n---\n")
        rationale_region = remainder if delimiter == -1 else remainder[:delimiter]
        rationale = rationale_region.strip("\n")
        if not rationale.strip():
            raise ValueError("group review Markdown rationale is blank")
        parsed.append((row, cast("Decision", selected), rationale))
    parsed_by_code: dict[str, tuple[HistoricalGroupReviewRow, Decision, str]] = {
        row.concept_code: (row, decision, rationale)
        for row, decision, rationale in parsed
    }
    return (
        reviewer_matches[0],
        date_matches[0],
        tuple(parsed_by_code[row.concept_code] for row in packet.review_rows),
    )


def _canonical_markdown_path(ncit_version: str) -> str:
    return f"evidence/group-review-rationale-{ncit_version}.md"


def _build_group_review_rationale_sidecar(
    *, markdown: bytes, packet: HistoricalGroupReviewPacket
) -> GroupReviewRationaleSidecar:
    reviewer, review_date, parsed = _parse_group_review_markdown(markdown, packet)
    rows = tuple(
        GroupReviewRationaleRow(
            concept_code=row.concept_code,
            review_row_identity=row.row_identity,
            review_type=row.review_type,
            decision=decision,
            pair_decision=decision if row.review_type == "pair-only" else None,
            rationale_sha256=rationale_sha256(rationale),
            rationale_chars=len(rationale),
        )
        for row, decision, rationale in parsed
    )
    payload = {
        "schema_version": 1,
        "evidence_kind": "group-review-rationale",
        "packet_identity": packet.packet_identity,
        "ncit_version": packet.ncit_version,
        "reviewer": reviewer,
        "review_date": review_date,
        "markdown_path": _canonical_markdown_path(packet.ncit_version),
        "markdown_sha256": hashlib.sha256(markdown).hexdigest(),
        "rows": rows,
    }
    return GroupReviewRationaleSidecar(**payload, sidecar_identity=_identity(payload))


def load_group_review_rationale_evidence(  # noqa: C901
    *,
    markdown_path: Path,
    sidecar_path: Path,
    packet: HistoricalGroupReviewPacket,
) -> LoadedGroupReviewRationale:
    """Read the exact first-admission human rationale and operational binding."""
    markdown = markdown_path.read_bytes()
    sidecar = GroupReviewRationaleSidecar.model_validate_json(sidecar_path.read_bytes())
    if sidecar.packet_identity != packet.packet_identity:
        raise ValueError("group review rationale packet identity differs")
    if sidecar.ncit_version != packet.ncit_version:
        raise ValueError("group review rationale NCIt version differs")
    if sidecar.markdown_path != _canonical_markdown_path(packet.ncit_version):
        raise ValueError("group review rationale Markdown path differs")
    reviewer, review_date, parsed = _parse_group_review_markdown(markdown, packet)
    if sidecar.markdown_sha256 != hashlib.sha256(markdown).hexdigest():
        raise ValueError("group review rationale markdown digest differs")
    if (sidecar.reviewer, sidecar.review_date) != (reviewer, review_date):
        raise ValueError("group review rationale reviewer/date differs")
    loaded_rows: list[LoadedGroupReviewRationaleRow] = []
    if len(sidecar.rows) != len(parsed):
        raise ValueError("group review rationale rows differ")
    for stored, (packet_row, decision, rationale) in zip(
        sidecar.rows, parsed, strict=True
    ):
        expected_pair = decision if packet_row.review_type == "pair-only" else None
        if (
            stored.concept_code != packet_row.concept_code
            or stored.review_row_identity != packet_row.row_identity
            or stored.review_type != packet_row.review_type
            or stored.decision != decision
            or stored.pair_decision != expected_pair
        ):
            raise ValueError("group review rationale row binding differs")
        digest = rationale_sha256(rationale)
        if stored.rationale_sha256 != digest:
            raise ValueError("group review rationale digest differs")
        if stored.rationale_chars != len(rationale):
            raise ValueError("group review rationale character count differs")
        loaded_rows.append(
            LoadedGroupReviewRationaleRow(
                concept_code=stored.concept_code,
                review_row_identity=stored.review_row_identity,
                review_type=stored.review_type,
                decision=stored.decision,
                pair_decision=stored.pair_decision,
                rationale=rationale,
                rationale_sha256=digest,
                rationale_chars=len(rationale),
            )
        )
    return LoadedGroupReviewRationale(sidecar=sidecar, rows=tuple(loaded_rows))


def validate_first_evidence_inventory(
    paths: tuple[str, ...], *, ncit_version: str
) -> None:
    """Enforce only the exact reduced first evidence admission inventory."""
    expected = (
        "evidence/README.md",
        f"evidence/group-review-packet-{ncit_version}-schema3.json",
        f"evidence/group-review-rationale-{ncit_version}.json",
        f"evidence/group-review-rationale-{ncit_version}.md",
    )
    if tuple(sorted(paths)) != expected:
        raise ValueError("first evidence inventory differs")


def _evidence_links(packet: GroupReviewPacket, row: GroupReviewRow) -> str:
    group_rows = tuple(
        index
        for index, evidence_row in enumerate(_group_evidence_rows(packet), start=2)
        if evidence_row[0] == row.concept_code
    )
    rule_rows = tuple(
        index
        for index, evidence in enumerate(packet.rule_evidence, start=2)
        if evidence.row_identity in row.evidence_row_ids
    )
    return "; ".join(
        (
            "Group Evidence rows " + ", ".join(map(str, group_rows)),
            "Rule Evidence rows " + ", ".join(map(str, rule_rows)),
        )
    )


def _visible_row(packet: GroupReviewPacket, row: GroupReviewRow) -> tuple[object, ...]:
    machine = {
        "pair_relations": row.pair_relations.model_dump(mode="json"),
        "grouping_diagnosis": row.grouping_diagnosis.model_dump(mode="json"),
        "suggestion": row.machine_suggestion,
    }
    return (
        row.concept_code,
        row.review_type,
        json.dumps(row.pair_relations.model_dump(mode="json"), sort_keys=True),
        json.dumps(row.grouping_diagnosis.model_dump(mode="json"), sort_keys=True),
        json.dumps(row.current_scoreable_group_ids),
        json.dumps(row.evidence_row_ids),
        _evidence_links(packet, row),
        json.dumps(machine, sort_keys=True),
        None,
    )


_GROUP_EVIDENCE_HEADERS = (
    "Concept",
    "Side",
    "Normalized Group ID",
    "Pair",
    "Source Evidence Status",
    "Source Group IDs",
    "Source Occurrence IDs",
)
_PAIR_RELATION_EVIDENCE_HEADERS = (
    "Concept",
    "Pair Relation",
    "Pair",
    "Occurrence Context",
    "Normalized Group Membership",
    "Interpretation",
)
_SOURCE_EVIDENCE_HEADERS = (
    "Occurrence ID",
    "Root Code",
    "Source Fact ID",
    "Source Group ID",
    "Anchor Code",
    "Depth",
    "Structural Path",
    "Member Position",
    "Role Code",
    "Filler Code",
    "Label Availability",
    "Definition Availability",
)
_RULE_EVIDENCE_HEADERS = (
    "Evidence Row ID",
    "Concept",
    "Kind",
    "Source Occurrence IDs",
    "Source Fact IDs",
    "Source Group IDs",
    "Output Group IDs",
    "Output Pairs",
    "R82 Path",
    "Proposed Partition",
    "Affected Co-membership",
    "Machine Evidence",
    "Evidence Limitation",
)


def _group_evidence_rows(packet: GroupReviewPacket) -> tuple[tuple[object, ...], ...]:
    rows: list[tuple[object, ...]] = []
    for concept in packet.concepts:
        for group in concept.expected_groups:
            for pair in group.pairs:
                rows.append(
                    (
                        concept.code,
                        "expected",
                        group.normalized_group_id,
                        json.dumps(pair),
                        "EXPECTED SOURCE UNAVAILABLE - historical expected grouping "
                        "is an oracle proposal, not source-stated",
                        None,
                        None,
                    )
                )
        for group in concept.actual_groups:
            for pair in group.pairs:
                rows.append(
                    (
                        concept.code,
                        "actual",
                        group.normalized_group_id,
                        json.dumps(pair.pair),
                        pair.availability,
                        json.dumps(group.source_group_ids),
                        (
                            json.dumps(pair.occurrence_ids)
                            if isinstance(pair, ActualPairEvidence)
                            else None
                        ),
                    )
                )
    return tuple(rows)


def _pair_relation_evidence_rows(
    packet: GroupReviewPacket,
) -> tuple[tuple[object, ...], ...]:
    return tuple(
        (
            concept.code,
            context.relation,
            json.dumps(context.pair),
            json.dumps(
                tuple(
                    item.model_dump(mode="json") for item in context.source_occurrences
                ),
                sort_keys=True,
            ),
            "Excluded from scoreable grouping",
            (
                "Emitted NCIt 26.07d content carrying review status; occurrence "
                "context is shown independently from scoreable grouping."
                if context.relation != "current-only-proposed"
                else "Provisional proposed NCIt content; source occurrence context "
                "is shown when available and no normalized group is assigned."
            ),
        )
        for concept in packet.concepts
        for context in concept.non_scoreable_emitted_pairs
    )


def _source_evidence_rows(packet: GroupReviewPacket) -> tuple[tuple[object, ...], ...]:
    occurrences = {
        occurrence.occurrence_id: occurrence
        for rule in packet.rule_evidence
        for occurrence in rule.source_occurrences
    }
    return tuple(
        (
            row.occurrence_id,
            row.root_code,
            row.source_fact_id,
            row.source_group_id,
            row.anchor_code,
            row.depth,
            json.dumps(row.structural_path),
            row.member_position,
            row.role_code,
            row.filler_code,
            "Unavailable in current evidence artifact",
            "Unavailable in current evidence artifact",
        )
        for row in sorted(occurrences.values(), key=lambda item: item.occurrence_id)
    )


def _rule_evidence_rows(packet: GroupReviewPacket) -> tuple[tuple[object, ...], ...]:
    return tuple(
        (
            row.row_identity,
            row.concept_code,
            row.kind,
            json.dumps(row.source_occurrence_ids),
            json.dumps(row.source_fact_ids),
            json.dumps(row.source_group_ids),
            json.dumps(row.output_group_ids),
            json.dumps(row.output_pairs),
            json.dumps(tuple(item.model_dump(mode="json") for item in row.r82_path)),
            json.dumps(row.proposed_partition),
            json.dumps(row.affected_co_membership),
            row.machine_evidence,
            row.machine_evidence_limitation,
        )
        for row in packet.rule_evidence
    )


def _append_evidence_sheet(
    book: Workbook,
    name: str,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
) -> None:
    sheet = book.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"


def _normalize_xlsx(source: Path, destination: Path) -> None:
    with (
        ZipFile(source) as original,
        ZipFile(
            destination, "w", compression=ZIP_DEFLATED, compresslevel=9
        ) as normalized,
    ):
        for name in sorted(original.namelist()):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            content = original.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:(created|modified)[^>]*>.*?</dcterms:\1>",
                    rb'<dcterms:\1 xsi:type="dcterms:W3CDTF">'
                    rb"2000-01-01T00:00:00Z</dcterms:\1>",
                    content,
                )
            elif name == "xl/workbook.xml":
                content = re.sub(rb' calcId="[0-9]+"', b"", content)
            normalized.writestr(info, content)


def write_group_review_workbook(path: Path, packet: GroupReviewPacket) -> None:
    """Write a deterministic, formula-free workbook with blank human decisions."""
    if not path.parent.is_dir():
        raise ValueError(f"output parent does not exist: {path.parent}")
    book = Workbook()
    instructions = book.active
    if instructions is None:
        raise ValueError("workbook did not create an active sheet")
    instructions.title = "Instructions"
    instructions.append(["Manual SME group review"])
    instructions.append(
        [
            "Machine evidence and suggestions are not reviewer rationale. Complete "
            "Decision, Rationale, Reviewer, and Date for every row; Pair Decision is "
            "also required for pair-only rows and must agree with Decision."
        ]
    )
    instructions.append(
        [
            "Evidence is asymmetric: expected-side source evidence is unavailable. "
            "The expected grouping is an oracle proposal, not a source-stated group. "
            "Actual-side source occurrences and every transformation witness are "
            "displayed in the evidence sheets. Labels and definitions are explicitly "
            "marked unavailable because the bound current evidence artifact does not "
            "contain source text."
        ]
    )
    sheet = book.create_sheet("Group Review")
    sheet.append(_REVIEW_HEADERS)
    for row in packet.review_rows:
        sheet.append((*_visible_row(packet, row), None, None, None, None))
    _append_evidence_sheet(
        book, "Group Evidence", _GROUP_EVIDENCE_HEADERS, _group_evidence_rows(packet)
    )
    _append_evidence_sheet(
        book,
        "Pair Relation Evidence",
        _PAIR_RELATION_EVIDENCE_HEADERS,
        _pair_relation_evidence_rows(packet),
    )
    _append_evidence_sheet(
        book,
        "Source Evidence",
        _SOURCE_EVIDENCE_HEADERS,
        _source_evidence_rows(packet),
    )
    _append_evidence_sheet(
        book, "Rule Evidence", _RULE_EVIDENCE_HEADERS, _rule_evidence_rows(packet)
    )
    bindings = book.create_sheet("Bindings")
    bindings.append(["Name", "Value"])
    for name, value in (
        ("packet_identity", packet.packet_identity),
        ("source_identity", packet.source_identity),
        ("evidence_identity", packet.current_evidence_identity),
        ("comparison_identity", packet.current_comparison_identity),
        ("r101_report_identity", packet.r101_report_identity),
        ("schema_version", packet.schema_version),
    ):
        bindings.append([name, value])
    bindings.sheet_state = "veryHidden"
    decision_columns = tuple(
        _REVIEW_HEADERS.index(name) + 1 for name in ("Pair Decision", "Decision")
    )
    validation = DataValidation(
        type="list", formula1='"' + ",".join(_DECISIONS) + '"', allow_blank=True
    )
    sheet.add_data_validation(validation)
    for decision_column in decision_columns:
        first_decision = sheet.cell(2, decision_column).coordinate
        last_decision = sheet.cell(sheet.max_row, decision_column).coordinate
        validation.add(f"{first_decision}:{last_decision}")
    editable = {"Pair Decision", "Decision", "Rationale", "Reviewer", "Date"}
    for worksheet in book.worksheets:
        for cells in worksheet.iter_rows():
            for cell in cells:
                header = worksheet.cell(1, cast("int", cell.column)).value
                cell.protection = Protection(locked=header not in editable)
        worksheet.protection.sheet = True
    book.calculation.fullCalcOnLoad = False
    book.calculation.forceFullCalc = False
    fixed = datetime(2000, 1, 1)
    book.properties.created = fixed
    book.properties.modified = fixed
    descriptor, staging = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    try:
        book.save(staging)
        _normalize_xlsx(Path(staging), path)
    finally:
        with suppress(FileNotFoundError):
            os.unlink(staging)


_CORRECTION_AUDIT_HEADERS = (
    "Concept",
    "Pair Relation",
    "Pair",
    "Occurrence Context",
    "Current Scoreable Group IDs",
    "Scoreable Grouping Diagnosis",
    "Correction Action",
    "Correction Rationale",
    "Reviewer",
    "Date",
)
_CORRECTION_HUMAN_HEADERS = (
    "Correction Action",
    "Correction Rationale",
    "Reviewer",
    "Date",
)


def _correction_audit_rows(packet: GroupReviewPacket) -> tuple[tuple[object, ...], ...]:
    rows: list[tuple[object, ...]] = []
    for concept in packet.concepts:
        contexts = {
            item.pair: tuple(
                row.model_dump(mode="json") for row in item.source_occurrences
            )
            for item in concept.non_scoreable_emitted_pairs
        }
        group_ids = tuple(group.normalized_group_id for group in concept.actual_groups)
        for field, pairs in concept.pair_relations:
            relation = field.replace("_", "-")
            for pair in pairs:
                rows.append(
                    (
                        concept.code,
                        relation,
                        json.dumps(pair),
                        json.dumps(contexts.get(pair, ()), sort_keys=True),
                        json.dumps(group_ids),
                        concept.grouping_diagnosis.kind,
                        None,
                        None,
                        None,
                        None,
                    )
                )
        rows.append(
            (
                concept.code,
                "scoreable-grouping",
                None,
                None,
                json.dumps(group_ids),
                concept.grouping_diagnosis.kind,
                None,
                None,
                None,
                None,
            )
        )
    return tuple(rows)


def _save_deterministic_workbook(path: Path, book: Workbook) -> None:
    book.calculation.fullCalcOnLoad = False
    book.calculation.forceFullCalc = False
    fixed = datetime(2000, 1, 1)
    book.properties.created = fixed
    book.properties.modified = fixed
    descriptor, staging = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    try:
        book.save(staging)
        _normalize_xlsx(Path(staging), path)
    finally:
        with suppress(FileNotFoundError):
            os.unlink(staging)


def write_group_correction_audit(path: Path, packet: GroupReviewPacket) -> None:
    """Write machine context with every human correction field blank."""
    if not path.parent.is_dir():
        raise ValueError(f"output parent does not exist: {path.parent}")
    book = Workbook()
    instructions = book.active
    if instructions is None:
        raise ValueError("workbook did not create an active sheet")
    instructions.title = "Instructions"
    instructions.append(["Blank group-correction audit"])
    instructions.append(
        [
            "Pair relations and scoreable grouping are independent machine context. "
            "All correction fields are intentionally blank for a new human review."
        ]
    )
    sheet = book.create_sheet("Correction Audit")
    sheet.append(_CORRECTION_AUDIT_HEADERS)
    for row in _correction_audit_rows(packet):
        sheet.append(row)
    sheet.freeze_panes = "A2"
    bindings = book.create_sheet("Bindings")
    bindings.append(["Name", "Value"])
    for name, value in _expected_bindings(packet):
        bindings.append([name, value])
    bindings.sheet_state = "veryHidden"
    for worksheet in book.worksheets:
        for cells in worksheet.iter_rows():
            for cell in cells:
                header = worksheet.cell(1, cast("int", cell.column)).value
                cell.protection = Protection(
                    locked=header not in _CORRECTION_HUMAN_HEADERS
                )
        worksheet.protection.sheet = True
    _save_deterministic_workbook(path, book)


def _validate_blank_correction_audit(path: Path, packet: GroupReviewPacket) -> None:
    _reject_unsafe_workbook(path)
    book = load_workbook(path, data_only=False)
    if book.sheetnames != ["Instructions", "Correction Audit", "Bindings"]:
        raise ValueError("correction audit sheet inventory differs")
    if book["Bindings"].sheet_state != "veryHidden":
        raise ValueError("correction audit binding visibility differs")
    bindings = tuple(
        (row[0].value, row[1].value)
        for row in book["Bindings"].iter_rows(min_row=2, max_col=2)
    )
    if bindings != _expected_bindings(packet):
        raise ValueError("correction audit bindings differ")
    _validate_evidence_sheet(
        book,
        "Correction Audit",
        _CORRECTION_AUDIT_HEADERS,
        _correction_audit_rows(packet),
    )


def validate_blank_group_review_outputs(
    *,
    packet: GroupReviewPacket,
    review_workbook: Path,
    correction_audit: Path,
    output: Path,
) -> BlankGroupReviewValidation:
    """Validate blank machine boundaries without creating decisions or readiness."""
    forbidden_registry = output.with_name("must-not-exist-decision-registry.json")
    if forbidden_registry.exists():
        raise ValueError("blank validation registry output already exists")
    try:
        import_group_review_decisions(packet, review_workbook, forbidden_registry)
    except ValueError as error:
        if str(error) != "all human fields are required":
            raise
    else:
        raise ValueError("blank review workbook unexpectedly contains human decisions")
    if forbidden_registry.exists():
        raise ValueError("blank validation created a decision artifact")
    _validate_blank_correction_audit(correction_audit, packet)
    result = BlankGroupReviewValidation(
        schema_version=1,
        packet_identity=packet.packet_identity,
        review_workbook_identity=hashlib.sha256(
            review_workbook.read_bytes()
        ).hexdigest(),
        correction_audit_identity=hashlib.sha256(
            correction_audit.read_bytes()
        ).hexdigest(),
        writes_performed=False,
        human_decisions_present=False,
        authorization=False,
        readiness_eligible=False,
    )
    _write_json(output, result.model_dump(mode="json"))
    return result


def _reject_unsafe_workbook(path: Path) -> None:
    with ZipFile(path) as archive:
        names = archive.namelist()
        if any(name.casefold().endswith("vbaproject.bin") for name in names):
            raise ValueError("review workbook contains a macro")
        if any(name.startswith("xl/externalLinks/") for name in names):
            raise ValueError("review workbook contains an external link")
        formula_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f"
        for name in names:
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                root = DefusedET.fromstring(archive.read(name))
                if any(True for _ in root.iter(formula_tag)):
                    raise ValueError("review workbook contains a formula")


def _expected_bindings(packet: GroupReviewPacket) -> tuple[tuple[object, object], ...]:
    return (
        ("packet_identity", packet.packet_identity),
        ("source_identity", packet.source_identity),
        ("evidence_identity", packet.current_evidence_identity),
        ("comparison_identity", packet.current_comparison_identity),
        ("r101_report_identity", packet.r101_report_identity),
        ("schema_version", packet.schema_version),
    )


def _validate_evidence_sheet(
    book: object,
    name: str,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
) -> None:
    sheet = book[name]  # type: ignore[index]
    observed = tuple(
        tuple(cell.value for cell in row)
        for row in sheet.iter_rows(  # type: ignore[attr-defined]
            min_row=1, max_row=len(rows) + 1, max_col=len(headers)
        )
    )
    if (  # type: ignore[attr-defined]
        observed != (headers, *rows)
        or sheet.max_row != len(rows) + 1
        or sheet.max_column != len(headers)
    ):
        raise ValueError(f"{name} cells differ")


def import_group_review_decisions(  # noqa: C901, PLR0912
    packet: GroupReviewPacket, workbook: Path, output: Path
) -> GroupDecisionRegistry:
    """Import only complete human decisions after validating every machine cell."""
    _reject_unsafe_workbook(workbook)
    book = load_workbook(workbook, data_only=False)
    if book.sheetnames != [
        "Instructions",
        "Group Review",
        "Group Evidence",
        "Pair Relation Evidence",
        "Source Evidence",
        "Rule Evidence",
        "Bindings",
    ]:
        raise ValueError("review workbook sheet inventory differs")
    if book["Bindings"].sheet_state != "veryHidden":
        raise ValueError("review workbook binding visibility differs")
    observed_bindings = tuple(
        (row[0].value, row[1].value)
        for row in book["Bindings"].iter_rows(min_row=2, max_col=2)
    )
    if observed_bindings != _expected_bindings(packet):
        raise ValueError("review workbook binding cells differ")
    _validate_evidence_sheet(
        book, "Group Evidence", _GROUP_EVIDENCE_HEADERS, _group_evidence_rows(packet)
    )
    _validate_evidence_sheet(
        book,
        "Pair Relation Evidence",
        _PAIR_RELATION_EVIDENCE_HEADERS,
        _pair_relation_evidence_rows(packet),
    )
    _validate_evidence_sheet(
        book,
        "Source Evidence",
        _SOURCE_EVIDENCE_HEADERS,
        _source_evidence_rows(packet),
    )
    _validate_evidence_sheet(
        book, "Rule Evidence", _RULE_EVIDENCE_HEADERS, _rule_evidence_rows(packet)
    )
    sheet = book["Group Review"]
    if tuple(cell.value for cell in sheet[1]) != _REVIEW_HEADERS:
        raise ValueError("review workbook headers differ")
    if sheet.max_row - 1 != len(packet.review_rows):
        raise ValueError("review rows are missing or duplicated")
    headers = {str(cell.value): cast("int", cell.column) for cell in sheet[1]}
    decisions: list[GroupDecision] = []
    machine_text = {row.machine_suggestion for row in packet.review_rows} | {
        row.machine_evidence for row in packet.rule_evidence
    }
    for index, expected in enumerate(packet.review_rows, 2):
        pair_column = headers["Pair Decision"]
        observed = tuple(
            sheet.cell(index, column).value for column in range(1, pair_column + 1)
        )
        if observed[:-1] != _visible_row(packet, expected)[:-1]:
            raise ValueError("immutable review cells differ")
        pair_decision = observed[-1]
        decision = sheet.cell(index, headers["Decision"]).value
        if expected.review_type == "pair-only" and (
            pair_decision is None or str(pair_decision).strip() == ""
        ):
            raise ValueError("pair decision is required for pair-only rows")
        if pair_decision is not None and pair_decision not in _DECISIONS:
            raise ValueError("pair decision is not one of the closed values")
        if pair_decision is not None and pair_decision != decision:
            raise ValueError("pair and grouping decisions are contradictory")
        values = (
            decision,
            sheet.cell(index, headers["Rationale"]).value,
            sheet.cell(index, headers["Reviewer"]).value,
            sheet.cell(index, headers["Date"]).value,
        )
        if any(value is None or str(value).strip() == "" for value in values):
            raise ValueError("all human fields are required")
        if decision not in _DECISIONS:
            raise ValueError("decision is not one of the closed values")
        rationale = str(values[1]).strip()
        if rationale in machine_text:
            raise ValueError("machine-generated text cannot be reviewer rationale")
        review_date = _normalize_review_date(values[3])
        decisions.append(
            GroupDecision(
                review_row_identity=expected.row_identity,
                concept_code=expected.concept_code,
                review_type=expected.review_type,
                pair_decision=cast("Decision | None", pair_decision),
                decision=cast("Decision", decision),
                rationale=rationale,
                reviewer=str(values[2]).strip(),
                review_date=review_date,
            )
        )
    payload = {
        "schema_version": 2,
        "packet_identity": packet.packet_identity,
        "source_identity": packet.source_identity,
        "evidence_identity": packet.current_evidence_identity,
        "comparison_identity": packet.current_comparison_identity,
        "decisions": tuple(decisions),
    }
    registry = GroupDecisionRegistry(**payload, registry_identity=_identity(payload))
    _write_json(output, registry.model_dump(mode="json"))
    return registry


def _normalize_review_date(value: object) -> str:
    normalized = (
        value.date().isoformat()
        if isinstance(value, datetime)
        else value.isoformat()
        if isinstance(value, date)
        else str(value).strip()
    )
    try:
        if date.fromisoformat(normalized).isoformat() != normalized:
            raise ValueError
    except ValueError as error:
        raise ValueError("review date is not an ISO date") from error
    return normalized


def dry_run_group_review_decisions(
    packet: GroupReviewPacket, registry: GroupDecisionRegistry
) -> GroupDecisionDryRun:
    """Report decision impact without mutating any store or artifact."""
    if (
        registry.packet_identity != packet.packet_identity
        or registry.source_identity != packet.source_identity
        or registry.evidence_identity != packet.current_evidence_identity
        or registry.comparison_identity != packet.current_comparison_identity
    ):
        raise ValueError("decision registry binding is stale")
    by_identity = {row.row_identity: row for row in packet.review_rows}
    if tuple(row.review_row_identity for row in registry.decisions) != tuple(
        by_identity
    ):
        raise ValueError("decision registry review rows differ")
    return GroupDecisionDryRun(
        writes_performed=False,
        unresolved_count=sum(
            row.decision
            in {
                "Require source-reproducible correction",
                "Reject proposed regrouping",
            }
            for row in registry.decisions
        ),
        deferred_count=sum(
            row.decision == "Abstain / escalate" for row in registry.decisions
        ),
        affected_concepts=tuple(row.concept_code for row in registry.decisions),
        affected_groups=tuple(
            sorted(
                {
                    group
                    for decision in registry.decisions
                    for group in by_identity[
                        decision.review_row_identity
                    ].current_scoreable_group_ids
                }
            )
        ),
        affected_traces=tuple(
            sorted(
                {
                    evidence_id
                    for decision in registry.decisions
                    for evidence_id in by_identity[
                        decision.review_row_identity
                    ].evidence_row_ids
                }
            )
        ),
    )


def _write_json(path: Path, value: object) -> None:
    payload = (
        json.dumps(value, indent=2, sort_keys=True, ensure_ascii=True).encode() + b"\n"
    )
    descriptor, temporary = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.")
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    except BaseException:
        Path(temporary).unlink(missing_ok=True)
        raise


def _write_bytes(path: Path, value: bytes) -> None:
    descriptor, temporary = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.")
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(value)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    except BaseException:
        Path(temporary).unlink(missing_ok=True)
        raise


def admit_group_review_rationale_evidence(
    *,
    packet: HistoricalGroupReviewPacket,
    source_markdown: Path,
    markdown_output: Path,
    sidecar_output: Path,
) -> GroupReviewRationaleSidecar:
    """Admit a frozen Markdown copy, then bind its actual post-write bytes."""
    expected_markdown = _canonical_markdown_path(packet.ncit_version)
    if markdown_output.as_posix() != expected_markdown:
        raise ValueError(
            "group review rationale output path differs from packet release"
        )
    expected_sidecar = expected_markdown.removesuffix(".md") + ".json"
    if sidecar_output.as_posix() != expected_sidecar:
        raise ValueError(
            "group review rationale sidecar path differs from packet release"
        )
    if not source_markdown.is_file():
        raise ValueError(
            f"group review rationale source does not exist: {source_markdown}"
        )
    if not markdown_output.parent.is_dir():
        raise ValueError(
            f"evidence output parent does not exist: {markdown_output.parent}"
        )
    source = source_markdown.read_bytes()
    _parse_group_review_markdown(source, packet)
    _write_bytes(markdown_output, source)
    persisted = markdown_output.read_bytes()
    if persisted != source:
        raise ValueError("admitted group review Markdown differs from frozen source")
    sidecar = _build_group_review_rationale_sidecar(markdown=persisted, packet=packet)
    _write_json(sidecar_output, sidecar.model_dump(mode="json"))
    load_group_review_rationale_evidence(
        markdown_path=markdown_output, sidecar_path=sidecar_output, packet=packet
    )
    return sidecar


def generate_group_review_packet(
    *,
    evidence_path: Path,
    comparison_path: Path,
    r101_report_path: Path,
    output: Path,
) -> GroupReviewPacket:
    """Generate canonical JSON from the validated current evidence pair."""
    for path in (evidence_path, comparison_path, r101_report_path):
        if not path.is_file():
            raise ValueError(f"input does not exist: {path}")
    if not output.parent.is_dir():
        raise ValueError(f"output parent does not exist: {output.parent}")
    if output.resolve() in {
        evidence_path.resolve(),
        comparison_path.resolve(),
        r101_report_path.resolve(),
    }:
        raise ValueError("output must differ from inputs")
    packet = build_group_review_packet(
        evidence=CurrentEngineEvidence.model_validate_json(evidence_path.read_bytes()),
        comparison=CurrentComparison.model_validate_json(comparison_path.read_bytes()),
        r101_report=R101ConservationReport.model_validate_json(
            gzip.decompress(r101_report_path.read_bytes())
        ),
    )
    _write_json(output, packet.model_dump(mode="json"))
    if load_group_review_packet(output) != packet:
        raise ValueError("persisted group review packet did not round trip")
    return packet


def generate_group_review_boundary(
    *,
    evidence_path: Path,
    comparison_path: Path,
    r101_report_path: Path,
    output: Path,
    workbook: Path,
    correction_audit: Path,
    blank_validation: Path,
) -> GroupReviewPacket:
    """Generate the bound machine packet and blank manual-review workbook."""
    packet = generate_group_review_packet(
        evidence_path=evidence_path,
        comparison_path=comparison_path,
        r101_report_path=r101_report_path,
        output=output,
    )
    write_group_review_workbook(workbook, packet)
    write_group_correction_audit(correction_audit, packet)
    validate_blank_group_review_outputs(
        packet=packet,
        review_workbook=workbook,
        correction_audit=correction_audit,
        output=blank_validation,
    )
    return packet
