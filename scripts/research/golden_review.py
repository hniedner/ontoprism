"""Fail-closed SME adjudication import and evaluation contracts."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Annotated, Literal, NamedTuple, Self, cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import (
    AfterValidator,
    BaseModel,
    ConfigDict,
    Field,
    TypeAdapter,
    ValidationError,
    model_validator,
)

from ontolib.decomposition.axis_contracts import AXIS_CONTRACTS
from ontolib.decomposition.proposal_registry import (
    ConceptProposal,
    ProposalRegistry,
    RelationProposal,
)
from ontolib.decomposition.score import ExtractionScore, score
from ontolib.decomposition.semantic_bundles import PairProvenance  # noqa: TC001

if TYPE_CHECKING:
    from collections.abc import Iterator

    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

DecisionStatus = Literal["accepted", "rejected", "revision-needed"]
ExpectedOutcome = Literal[
    "decomposed",
    "residual",
    "semantic-excluded",
    "atomic-no-op",
]
SmeAction = Literal["include", "revise", "exclude", "not-needed"]
ConstituentRowType = Literal["ENGINE SUGGESTION", "ADD IF MISSING"]
ConstituentPair = tuple[str, str]
_ADJUDICATED_STATUS = "SME-ADJUDICATED"
_SCHEMA_VERSION = 3
_ROW_DECISION_SCHEMA_VERSION = 3
_SME_ACTIONS: tuple[SmeAction, ...] = ("include", "revise", "exclude", "not-needed")
_CONSTITUENT_ROW_TYPES: tuple[ConstituentRowType, ...] = (
    "ENGINE SUGGESTION",
    "ADD IF MISSING",
)
# The two actions that put the row's expected pair into the adjudicated oracle.
# `exclude` and `not-needed` rows are retained by the export and dropped by the
# import; three attested `exclude` rows still carry a withdrawn expectation, so the
# action -- not the presence of a pair -- decides what the SME kept.
_KEPT_SME_ACTIONS = frozenset({"include", "revise"})
_M1_REQUIRED_SEEDS = frozenset({"C4791", "C35756", "C89995"})
_M1_MIN_CONCEPTS = 20
_M1_MAX_CONCEPTS = 50
_EXPECTED_SHEETS = {
    "START HERE",
    "Reviewer & Attestation",
    "Concept Decisions",
    "Constituent Decisions",
    "Validation Summary",
    "Worked Examples",
    "Prior SME Evidence",
    "Source & Run Evidence",
}
_OPTIONAL_REVIEW_SHEETS = {"Semantic Bundle Decisions"}


class GoldenSetValidationError(ValueError):
    """The adjudication input cannot be trusted as an SME oracle."""


def _canonical_text(value: str, field: str) -> str:
    if not value or value != value.strip():
        raise ValueError(f"{field} must be non-empty without outer whitespace")
    return value


def _payload_identity(value: object) -> str:
    encoded = json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


class _StrictModel(BaseModel):
    model_config = ConfigDict(frozen=True, extra="forbid", strict=True)


class Reviewer(_StrictModel):
    name: str
    qualification_or_role: str
    reviewed_at: str

    @model_validator(mode="after")
    def _validate_reviewer(self) -> Self:
        _canonical_text(self.name, "reviewer name")
        _canonical_text(self.qualification_or_role, "reviewer qualification")
        try:
            date.fromisoformat(_canonical_text(self.reviewed_at, "reviewed_at"))
        except ValueError as error:
            raise ValueError("reviewed_at must be an ISO date") from error
        return self


class AdjudicationMetadata(_StrictModel):
    schema_version: Literal[3]
    status: Literal["SME-ADJUDICATED"]
    ncit_version: str
    source_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    sample_manifest_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    run_id: str
    run_fingerprint_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    engine_artifact_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    engine_evidence_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    corpus_evidence_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    detector_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    proposal_registry_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    workbook_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    reviewer: Reviewer

    @model_validator(mode="after")
    def _validate_text(self) -> Self:
        _canonical_text(self.ncit_version, "ncit_version")
        _canonical_text(self.run_id, "run_id")
        return self


class Constituent(_StrictModel):
    axis: str
    filler: str
    relationship_group: str | None
    needs_review: bool

    @model_validator(mode="after")
    def _validate_text(self) -> Self:
        _canonical_text(self.axis, "constituent axis")
        _canonical_text(self.filler, "constituent filler")
        if self.relationship_group is not None:
            _canonical_text(self.relationship_group, "relationship_group")
        return self

    @property
    def pair(self) -> ConstituentPair:
        return (self.axis, self.filler)


class GoldenConstituent(Constituent):
    provenance_status: PairProvenance
    proposal_id: str | None

    @model_validator(mode="after")
    def _validate_proposal_reference(self) -> Self:
        if self.provenance_status == "ncit-26.07d":
            if self.proposal_id is not None:
                raise ValueError("NCIt constituent must not carry a proposal ID")
            if re.fullmatch(r"C[0-9]+", self.filler) is None:
                raise ValueError("NCIt constituent filler must be an NCIt code")
        elif self.proposal_id is None:
            raise ValueError("augmented constituent requires a proposal ID")
        else:
            _canonical_text(self.proposal_id, "proposal ID")
        return self


class GoldenExpectation(_StrictModel):
    outcome: ExpectedOutcome
    semantic_types: tuple[str, ...]
    constituents: tuple[GoldenConstituent, ...]

    @model_validator(mode="after")
    def _validate_shape(self) -> Self:
        if not self.semantic_types:
            raise ValueError("semantic_types must not be empty")
        for semantic_type in self.semantic_types:
            _canonical_text(semantic_type, "semantic type")
        if len(self.semantic_types) != len(set(self.semantic_types)):
            raise ValueError("semantic_types must be unique")
        pairs = [item.pair for item in self.constituents]
        if len(pairs) != len(set(pairs)):
            raise ValueError("constituent axis/filler pairs must be unique")
        if self.outcome == "decomposed" and not self.constituents:
            raise ValueError(
                "decomposed expectation must contain at least one constituent"
            )
        if self.outcome != "decomposed" and self.constituents:
            raise ValueError("non-decomposed expectation must not contain constituents")
        return self


class AdjudicationDecision(_StrictModel):
    status: DecisionStatus
    rationale: str

    @model_validator(mode="after")
    def _validate_rationale(self) -> Self:
        _canonical_text(self.rationale, "adjudication rationale")
        return self


class AdjudicatedConcept(_StrictModel):
    code: str = Field(pattern=r"^C[0-9]+$")
    label: str
    adjudication: AdjudicationDecision
    expected: GoldenExpectation | None

    @model_validator(mode="after")
    def _validate_decision_shape(self) -> Self:
        _canonical_text(self.label, f"{self.code} label")
        if self.adjudication.status == "accepted" and self.expected is None:
            raise ValueError("accepted decision requires expected")
        if self.adjudication.status != "accepted" and self.expected is not None:
            raise ValueError("nonaccepted decision must not define expected")
        return self


class AdjudicationArtifact(_StrictModel):
    meta: Annotated[AdjudicationMetadata, Field(alias="_meta")]
    concepts: tuple[AdjudicatedConcept, ...]
    artifact_identity: str = Field(pattern=r"^[0-9a-f]{64}$")

    @model_validator(mode="after")
    def _validate_m1_cohort(self) -> Self:
        if not _M1_MIN_CONCEPTS <= len(self.concepts) <= _M1_MAX_CONCEPTS:
            raise ValueError("M1 golden set must contain 20 to 50 adjudicated concepts")
        codes = [concept.code for concept in self.concepts]
        if len(codes) != len(set(codes)):
            raise ValueError("concept codes must be unique")
        missing = sorted(_M1_REQUIRED_SEEDS - set(codes))
        if missing:
            raise ValueError(
                "M1 golden set is missing named seeds: " + ", ".join(missing)
            )
        if not any(
            concept.adjudication.status == "accepted" for concept in self.concepts
        ):
            raise ValueError("M1 golden set must contain at least one accepted concept")
        if self.artifact_identity != _payload_identity(
            self.model_dump(
                mode="json",
                by_alias=True,
                exclude={"artifact_identity"},
            )
        ):
            raise ValueError(
                "adjudication artifact identity does not match its payload"
            )
        return self

    @property
    def identity(self) -> str:
        return self.artifact_identity


class ExpectedTriple(NamedTuple):
    """One expectation a row records, with its members named.

    `expected_pairs()` returned bare `tuple[str, str, str]`, so an equality in a
    test read as three anonymous strings and `pair[1]` was the only way to ask for
    the axis. Naming the members does not make a transposed *construction* a type
    error -- all three are `str` -- but it does mean every read site says which is
    which, and `ExpectedTriple(code=..., axis=..., filler=...)` is available where
    a call site wants the check.
    """

    code: str
    axis: str
    filler: str


class _ConstituentRow(_StrictModel):
    """One reviewer decision row, recorded before any expectation is inferred.

    The oracle keeps only what the SME included or revised. The export keeps the
    rejected and unused rows too, because the ratio between them *is* the
    measurement of the engine: an accepted suggestion and an excluded one are
    indistinguishable once only the surviving pairs are stored.

    The three variants below differ in what they may carry, so the differences are
    enforced by the shape of each row rather than by a validator every caller then
    has to trust. `sme_action` is already a disjoint `Literal` per variant, so it
    serves as the discriminator and the wire format gains no tag field.
    """

    code: str = Field(pattern=r"^C[0-9]+$")


class KeptRow(_ConstituentRow):
    """A row whose pair entered the adjudicated expected set."""

    row_type: ConstituentRowType
    sme_action: Literal["include", "revise"]
    expected_axis: str
    expected_filler: str

    @model_validator(mode="after")
    def _validate_expectation(self) -> Self:
        _canonical_text(self.expected_axis, f"{self.code} expected axis")
        _canonical_text(self.expected_filler, f"{self.code} expected filler")
        return self

    @property
    def expected_triple(self) -> ExpectedTriple:
        return ExpectedTriple(self.code, self.expected_axis, self.expected_filler)


class ExcludedRow(_ConstituentRow):
    """A row the reviewer rejected, optionally still naming what they withdrew.

    Three rows of the attested #57 workbook carry an expectation the reviewer then
    excluded, so the pair is retained and the *action* -- never the presence of a
    pair -- decides what the SME kept.
    """

    row_type: ConstituentRowType
    sme_action: Literal["exclude"]
    expected_axis: str | None
    expected_filler: str | None

    @model_validator(mode="after")
    def _validate_expectation(self) -> Self:
        for field, value in (
            ("expected axis", self.expected_axis),
            ("expected filler", self.expected_filler),
        ):
            if value is not None:
                _canonical_text(value, f"{self.code} {field}")
        return self

    @property
    def withdrawn_triple(self) -> ExpectedTriple | None:
        """What this row withdrew, or `None` when it named nothing."""
        if self.expected_axis is None or self.expected_filler is None:
            return None
        return ExpectedTriple(self.code, self.expected_axis, self.expected_filler)


class UnusedCandidateRow(_ConstituentRow):
    """A candidate row the reviewer never had to fill in.

    `row_type` is pinned, so `ENGINE SUGGESTION` / `not-needed` has no inhabitant.
    On a suggestion the action is a non-decision: the suggestion would enter the
    acceptance denominator without the SME accepting or rejecting it, which is the
    same hole `PENDING` closes on the other side. There are no pair fields either,
    because a row nobody had to fill in has nothing to record.
    """

    row_type: Literal["ADD IF MISSING"]
    sme_action: Literal["not-needed"]


ConstituentRowDecision = Annotated[
    KeptRow | ExcludedRow | UnusedCandidateRow,
    Field(discriminator="sme_action"),
]
_ROW_DECISION_ADAPTER: TypeAdapter[ConstituentRowDecision] = TypeAdapter(
    ConstituentRowDecision
)


class RowDecisionMetadata(_StrictModel):
    """Provenance for one row-decision export.

    `source_workbook` is a **display label only**, and deliberately unconstrained
    beyond being canonical text: `export_row_decisions` passes `Path.name` and
    `export_row_decisions_bytes` accepts whatever its caller supplies, so it is
    whatever the file happened to be called when it was exported. It carries no
    authority and nothing may be matched against it. `workbook_identity` -- the
    SHA-256 of the `.xlsx` bytes -- is the identifier, and it is what
    `test_m1_baseline` compares against the oracle's.
    """

    schema_version: Literal[3]
    ncit_version: str
    source_workbook: str
    workbook_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    source_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    run_id: str
    engine_evidence_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    reviewer: Reviewer

    @model_validator(mode="after")
    def _validate_text(self) -> Self:
        _canonical_text(self.ncit_version, "ncit_version")
        _canonical_text(self.source_workbook, "source_workbook")
        _canonical_text(self.run_id, "run_id")
        return self


class EngineAcceptance(_StrictModel):
    """What the reviewer did with the suggestions of one engine run.

    There is no `not_needed` field, and its absence is the point. `not-needed`
    records a *candidate* row the SME never had to fill in; on an engine
    suggestion it would be a non-decision sitting in the acceptance denominator,
    which is the same hole `PENDING` closes on the other side. The combination has
    no inhabitant in `ConstituentRowDecision`, so this tally has no cell for it.
    """

    include: int = Field(ge=0)
    revise: int = Field(ge=0)
    exclude: int = Field(ge=0)

    @property
    def adjudicated(self) -> int:
        """Every suggestion the reviewer ruled on — the acceptance denominator."""
        return self.include + self.revise + self.exclude

    @property
    def accepted_unchanged_rate(self) -> float | None:
        """The published acceptance rate, or `None` when the run suggested nothing.

        Undefined rather than zero or an exception: a review of a run that emitted
        no constituent at all is a legitimate export with nothing to accept, and
        the same convention already governs `_residual_dict`'s empty denominator.
        """
        adjudicated = self.adjudicated
        return self.include / adjudicated if adjudicated else None


class CandidateOutcomes(_StrictModel):
    """What the reviewer did with the rows they added themselves.

    A candidate row carries no engine suggestion, so `not_needed` is a real
    outcome here and no acceptance rate is defined: nothing was proposed to
    accept.
    """

    include: int = Field(ge=0)
    revise: int = Field(ge=0)
    exclude: int = Field(ge=0)
    not_needed: int = Field(ge=0)


class RowDecisionCrossTab(_StrictModel):
    """Row type by SME action, with each row type's own set of outcomes."""

    engine_suggestion: EngineAcceptance
    add_if_missing: CandidateOutcomes


def _require_decision_rows(
    value: tuple[ConstituentRowDecision, ...],
) -> tuple[ConstituentRowDecision, ...]:
    """Reject an empty row set as a *field* failure, not a whole-model one.

    A model-level check reports the entire validated input, so the refusal that
    reaches the CLI and the CI log is `input_value={'_meta': {'engine_eviden...`
    -- the reviewer's provenance block, truncated mid-word by pydantic, and no
    part of it evidence about an empty row set. (The truncation is why the
    reviewer's name does not in fact appear; nothing in the code arranges that,
    and a shorter `_meta` would surface more of it.) Failing on the field reports
    `input_value=()`.
    """
    if not value:
        raise ValueError("row decisions must not be empty")
    return value


class RowDecisionExport(_StrictModel):
    """Every constituent decision the reviewer recorded, verbatim."""

    meta: Annotated[RowDecisionMetadata, Field(alias="_meta")]
    rows: Annotated[
        tuple[ConstituentRowDecision, ...], AfterValidator(_require_decision_rows)
    ]
    payload_identity: str = Field(pattern=r"^[0-9a-f]{64}$")

    @model_validator(mode="after")
    def _validate_rows(self) -> Self:
        kept: list[ExpectedTriple] = []
        withdrawn: list[ExpectedTriple] = []
        for row in self.rows:
            if isinstance(row, KeptRow):
                kept.append(row.expected_triple)
            elif (
                isinstance(row, ExcludedRow)
                and (triple := row.withdrawn_triple) is not None
            ):
                withdrawn.append(triple)
        both = sorted(set(kept) & set(withdrawn))
        if both:
            raise ValueError(
                "an expected pair cannot be both kept and withdrawn: "
                + ", ".join("/".join(triple) for triple in both)
            )
        identified = kept + withdrawn
        if len(identified) != len(set(identified)):
            raise ValueError(
                "rows carrying an expected pair must be unique on (code, axis, filler)"
            )
        if self.payload_identity != _payload_identity(
            self.model_dump(
                mode="json",
                by_alias=True,
                exclude={"payload_identity"},
            )
        ):
            raise ValueError("row decision payload identity does not match its payload")
        return self

    def expected_pairs(self) -> set[ExpectedTriple]:
        """The `(code, axis, filler)` triples the SME kept.

        Keyed on the variant, never on the presence of an expected pair: an
        `ExcludedRow` may still carry the expectation the reviewer withdrew.
        """
        return {row.expected_triple for row in self.rows if isinstance(row, KeptRow)}

    def cross_tab(self) -> RowDecisionCrossTab:
        """Row type by SME action, as two differently shaped tallies.

        Not a `dict[str, dict[str, int]]`: the two row types do not have the same
        columns, and a bare grid could neither say so nor compute the acceptance
        rate the grid exists to support.
        """
        counts = Counter((row.row_type, row.sme_action) for row in self.rows)
        return RowDecisionCrossTab(
            engine_suggestion=EngineAcceptance(
                include=counts[("ENGINE SUGGESTION", "include")],
                revise=counts[("ENGINE SUGGESTION", "revise")],
                exclude=counts[("ENGINE SUGGESTION", "exclude")],
            ),
            add_if_missing=CandidateOutcomes(
                include=counts[("ADD IF MISSING", "include")],
                revise=counts[("ADD IF MISSING", "revise")],
                exclude=counts[("ADD IF MISSING", "exclude")],
                not_needed=counts[("ADD IF MISSING", "not-needed")],
            ),
        )


class EngineConcept(_StrictModel):
    code: str = Field(pattern=r"^C[0-9]+$")
    outcome: ExpectedOutcome
    semantic_types: tuple[str, ...]
    constituents: tuple[Constituent, ...]

    @model_validator(mode="after")
    def _validate_shape(self) -> Self:
        if len(self.semantic_types) != len(set(self.semantic_types)):
            raise ValueError("engine semantic_types must be unique")
        pairs = [item.pair for item in self.constituents]
        if len(pairs) != len(set(pairs)):
            raise ValueError("engine constituent axis/filler pairs must be unique")
        if self.outcome == "decomposed" and not self.constituents:
            raise ValueError("engine decomposed outcome requires constituents")
        if self.outcome != "decomposed" and self.constituents:
            raise ValueError("engine non-decomposed outcome cannot carry constituents")
        return self


class EngineEvidence(_StrictModel):
    schema_version: Literal[1]
    ncit_version: str
    source_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    sample_manifest_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    run_id: str
    run_fingerprint_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    engine_artifact_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    detector_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    evidence_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    concepts: tuple[EngineConcept, ...]
    residual_precoordinated_codes: tuple[str, ...]

    @model_validator(mode="after")
    def _validate_codes(self) -> Self:
        codes = [item.code for item in self.concepts]
        if len(codes) != len(set(codes)):
            raise ValueError("engine concept codes must be unique")
        if len(self.residual_precoordinated_codes) != len(
            set(self.residual_precoordinated_codes)
        ):
            raise ValueError("residual_precoordinated_codes must be unique")
        if not set(self.residual_precoordinated_codes) <= set(codes):
            raise ValueError("engine residual codes must be a subset of concept codes")
        outcomes = {item.code: item.outcome for item in self.concepts}
        if any(
            outcomes[code] != "decomposed"
            for code in self.residual_precoordinated_codes
        ):
            raise ValueError("engine residual codes require decomposed outcomes")
        if self.evidence_identity != _payload_identity(
            self.model_dump(mode="json", exclude={"evidence_identity"})
        ):
            raise ValueError("engine evidence identity does not match its payload")
        return self


class ResidualComparisonInput(_StrictModel):
    name: str
    source_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    sample_manifest_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    run_id: str
    run_fingerprint_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    engine_artifact_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    detector_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    evidence_identity: str = Field(pattern=r"^[0-9a-f]{64}$")
    denominator_codes: tuple[str, ...]
    residual_codes: tuple[str, ...]

    @model_validator(mode="after")
    def _validate_codes(self) -> Self:
        if len(self.denominator_codes) != len(set(self.denominator_codes)):
            raise ValueError("denominator codes must be unique")
        if len(self.residual_codes) != len(set(self.residual_codes)):
            raise ValueError("residual codes must be unique")
        if not set(self.residual_codes) <= set(self.denominator_codes):
            raise ValueError("residual codes must be a subset of denominator codes")
        if self.evidence_identity != _payload_identity(
            self.model_dump(mode="json", exclude={"evidence_identity"})
        ):
            raise ValueError("residual evidence identity does not match its payload")
        return self


class ScorableGoldenSet:
    """Validated accepted expectations plus a fail-closed pair view."""

    def __init__(self, artifact: AdjudicationArtifact) -> None:
        self.ncit_version = artifact.meta.ncit_version
        self.source_identity = artifact.meta.source_identity
        self.reviewer_qualification = artifact.meta.reviewer.qualification_or_role
        self.labels = {concept.code: concept.label for concept in artifact.concepts}
        self.decisions = {
            concept.code: concept.adjudication.status for concept in artifact.concepts
        }
        self.expectations = {
            concept.code: cast("GoldenExpectation", concept.expected)
            for concept in artifact.concepts
            if concept.adjudication.status == "accepted"
        }
        self.expected = {
            code: frozenset(
                item.pair for item in expectation.constituents if not item.needs_review
            )
            for code, expectation in self.expectations.items()
        }
        self.expected_ncit_bound = {
            code: frozenset(
                item.pair
                for item in expectation.constituents
                if not item.needs_review and item.provenance_status == "ncit-26.07d"
            )
            for code, expectation in self.expectations.items()
        }
        self.expected_augmented = {
            code: frozenset(
                item.pair
                for item in expectation.constituents
                if not item.needs_review and item.provenance_status != "proposed"
            )
            for code, expectation in self.expectations.items()
        }
        self.review_exclusions = {
            code: exclusions
            for code, expectation in self.expectations.items()
            if (
                exclusions := frozenset(
                    item.pair for item in expectation.constituents if item.needs_review
                )
            )
        }
        counts = Counter(self.decisions.values())
        self.decision_counts = {
            status: counts[status]
            for status in ("accepted", "rejected", "revision-needed")
        }


def _reject_duplicate_keys(pairs: list[tuple[str, object]]) -> dict[str, object]:
    value: dict[str, object] = {}
    for key, item in pairs:
        if key in value:
            raise GoldenSetValidationError(f"duplicate JSON key: {key}")
        value[key] = item
    return value


def _model_error(error: ValidationError | ValueError) -> GoldenSetValidationError:
    return GoldenSetValidationError(str(error))


def _read_adjudication_json(path: str | Path) -> dict[str, object]:
    try:
        raw = json.loads(
            Path(path).read_text(encoding="utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
        )
    except GoldenSetValidationError:
        raise
    except (OSError, json.JSONDecodeError) as error:
        raise GoldenSetValidationError(f"cannot read golden set: {error}") from error
    if not isinstance(raw, dict):
        raise GoldenSetValidationError("golden set must be a JSON object")
    return raw


def read_json_without_duplicates(path: str | Path) -> object:
    """Read JSON while rejecting duplicate keys at every object level."""
    try:
        return json.loads(
            Path(path).read_text(encoding="utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
        )
    except GoldenSetValidationError:
        raise
    except (OSError, json.JSONDecodeError) as error:
        raise GoldenSetValidationError(f"cannot read JSON evidence: {error}") from error


def _normalize_adjudication_lists(raw: dict[str, object]) -> None:
    concepts = raw.get("concepts")
    if not isinstance(concepts, list):
        return
    for concept in concepts:
        if not isinstance(concept, dict):
            continue
        expected = concept.get("expected")
        if not isinstance(expected, dict):
            continue
        if isinstance(expected.get("semantic_types"), list):
            expected["semantic_types"] = tuple(expected["semantic_types"])
        if isinstance(expected.get("constituents"), list):
            expected["constituents"] = tuple(expected["constituents"])
    raw["concepts"] = tuple(concepts)


_PROVENANCE_PROPOSAL_STATUS = {
    "proposed": "proposed",
    "locally-approved": "locally-approved",
    "submitted": "submitted",
    "accepted-in-ncit": "accepted",
}


def _augmented_constituents(
    artifact: AdjudicationArtifact,
) -> list[GoldenConstituent]:
    return [
        item
        for concept in artifact.concepts
        if concept.expected is not None
        for item in concept.expected.constituents
        if item.provenance_status != "ncit-26.07d"
    ]


def _validate_constituent_proposal(
    constituent: GoldenConstituent,
    proposal: ConceptProposal | RelationProposal,
) -> None:
    expected_status = _PROVENANCE_PROPOSAL_STATUS[constituent.provenance_status]
    if proposal.status != expected_status:
        raise GoldenSetValidationError(
            f"proposal status does not match expected provenance: {proposal.id}"
        )
    if proposal.axis != constituent.axis:
        raise GoldenSetValidationError(
            f"proposal axis does not match augmented expectation: {proposal.id}"
        )
    if not isinstance(proposal, ConceptProposal):
        if re.fullmatch(r"C[0-9]+", constituent.filler) is None:
            raise GoldenSetValidationError(
                "relation proposal constituent filler must remain an NCIt code: "
                + proposal.id
            )
        return
    expected_filler = (
        proposal.replacement_ncit_code if proposal.status == "accepted" else proposal.id
    )
    if constituent.filler != expected_filler:
        raise GoldenSetValidationError(
            "proposal filler does not match augmented expectation: " + proposal.id
        )


def _validate_proposal_registry_binding(
    artifact: AdjudicationArtifact,
    proposal_registry: ProposalRegistry | None,
) -> None:
    augmented = _augmented_constituents(artifact)
    if proposal_registry is None:
        if augmented:
            raise GoldenSetValidationError(
                "augmented adjudication requires its proposal registry"
            )
        return
    if artifact.meta.proposal_registry_identity != proposal_registry.registry_identity:
        raise GoldenSetValidationError(
            "adjudication proposal registry identity does not match"
        )
    if artifact.meta.source_identity != proposal_registry.source_identity:
        raise GoldenSetValidationError(
            "adjudication and proposal registry source identities do not match"
        )
    if artifact.meta.ncit_version != proposal_registry.ontology_version:
        raise GoldenSetValidationError(
            "adjudication and proposal registry ontology versions do not match"
        )
    proposals = {proposal.id: proposal for proposal in proposal_registry.proposals}
    for constituent in augmented:
        proposal = proposals.get(cast("str", constituent.proposal_id))
        if proposal is None:
            raise GoldenSetValidationError(
                f"unknown proposal in augmented expectation: {constituent.proposal_id}"
            )
        _validate_constituent_proposal(constituent, proposal)


def load_adjudication(
    path: str | Path,
    proposal_registry: ProposalRegistry | None = None,
) -> AdjudicationArtifact:
    """Load one strict, complete, provenance-bearing M1 adjudication artifact."""
    raw = _read_adjudication_json(path)
    meta = raw.get("_meta")
    if not isinstance(meta, dict) or meta.get("status") != _ADJUDICATED_STATUS:
        raise GoldenSetValidationError(
            "golden set is not SME-adjudicated; automated drafts cannot be scored"
        )
    _normalize_adjudication_lists(raw)
    try:
        artifact = AdjudicationArtifact.model_validate(raw)
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error
    _validate_proposal_registry_binding(artifact, proposal_registry)
    return artifact


def load_scorable_golden(
    path: str | Path,
    proposal_registry: ProposalRegistry | None = None,
) -> ScorableGoldenSet:
    """Load accepted expectations only after complete M1 validation."""
    return ScorableGoldenSet(load_adjudication(path, proposal_registry))


def _headers(ws: Worksheet, row: int) -> dict[str, int]:
    values: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(row, column).value
        if isinstance(value, str):
            if value in values:
                raise GoldenSetValidationError(f"duplicate workbook header: {value}")
            values[value] = column
    return values


def _cell_text(ws: Worksheet, row: int, column: int, field: str) -> str:
    value = ws.cell(row, column).value
    if not isinstance(value, str):
        raise GoldenSetValidationError(f"{field} must be text")
    try:
        return _canonical_text(value, field)
    except ValueError as error:
        raise GoldenSetValidationError(str(error)) from error


def _optional_text(ws: Worksheet, row: int, column: int, field: str) -> str | None:
    value = ws.cell(row, column).value
    if value is None:
        return None
    if not isinstance(value, str):
        raise GoldenSetValidationError(f"{field} must be text or blank")
    try:
        return _canonical_text(value, field)
    except ValueError as error:
        raise GoldenSetValidationError(str(error)) from error


def _reject_reviewer_formulas(workbook: Workbook) -> None:
    for sheet_name in (
        "Reviewer & Attestation",
        "Concept Decisions",
        "Constituent Decisions",
        "Source & Run Evidence",
    ):
        for row in workbook[sheet_name].iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise GoldenSetValidationError(
                        f"formula cells are not permitted in reviewer input: "
                        f"{sheet_name}!{cell.coordinate}"
                    )


def hidden_column_indexes(sheet: Worksheet, *, limit: int) -> set[int]:
    """Every 1-based column index hidden on *sheet*, including grouped ranges.

    ``column_dimensions`` is keyed by the letter of each range's ``min`` only, so a
    grouped hide of ``C:E`` is stored as a single entry under ``"C"`` with
    ``min=3, max=5``. Probing ``column_dimensions["D"]`` auto-creates a default with
    ``hidden=False``, which would let a reviewer conceal data by hiding a range whose
    first column is blank. Expand the stored ranges instead of probing per letter.

    *limit* caps the expansion and is required, because an uncapped expansion is a
    hang. Excel writes ``<col min="8" max="16384" hidden="1"/>`` when a reviewer
    hides the trailing columns, so without the cap this returns ~16k indexes and a
    caller that then reads cells materialises the whole grid. A hidden column that
    actually *contains* data always has a parsed cell, hence an index within
    ``max_column``, so capping there loses no tamper signal.
    """
    indexes: set[int] = set()
    for dimension in sheet.column_dimensions.values():
        if not dimension.hidden:
            continue
        # `min`/`max` are populated by the reader, by `DimensionHolder.group()`,
        # and by `reindex()` at save time; a bare
        # `column_dimensions["D"].hidden = True` leaves them None. Fall back to the
        # key rather than skipping, so the gate cannot degrade to "nothing is
        # hidden".
        start = (
            dimension.min
            if dimension.min is not None
            else column_index_from_string(dimension.index)
        )
        end = min(dimension.max if dimension.max is not None else start, limit)
        indexes.update(range(start, end + 1))
    return indexes


def _reject_hidden_reviewer_columns(workbook: Workbook) -> None:
    for sheet_name in (
        "Reviewer & Attestation",
        "Concept Decisions",
        "Constituent Decisions",
        "Source & Run Evidence",
    ):
        sheet = workbook[sheet_name]
        # Read both extents once: `max_row`/`max_column` are O(cells), and reading
        # them inside the scan would make it quadratic.
        max_row, max_column = sheet.max_row, sheet.max_column
        hidden = [
            get_column_letter(column)
            for column in sorted(hidden_column_indexes(sheet, limit=max_column))
            if any(
                sheet.cell(row, column).value is not None
                for row in range(1, max_row + 1)
            )
        ]
        if hidden:
            raise GoldenSetValidationError(
                f"hidden reviewer columns are not permitted in {sheet_name}: "
                + ", ".join(hidden)
            )


def _evidence_values(workbook: Workbook) -> dict[str, str]:
    ws = workbook["Source & Run Evidence"]
    result: dict[str, str] = {}
    for row in range(5, ws.max_row + 1):
        if ws.row_dimensions[row].hidden and _row_has_data(ws, row):
            raise GoldenSetValidationError(
                f"hidden source evidence rows are not permitted: {row}"
            )
        key = ws.cell(row, 1).value
        value = ws.cell(row, 2).value
        if isinstance(key, str) and isinstance(value, str):
            if key in result:
                raise GoldenSetValidationError(
                    f"duplicate Source & Run Evidence key: {key}"
                )
            result[key] = value
    return result


def _row_has_data(ws: Worksheet, row: int) -> bool:
    return any(
        ws.cell(row, column).value is not None for column in range(1, ws.max_column + 1)
    )


def _parse_semantic_types(value: str, field: str) -> tuple[str, ...]:
    items = tuple(value.split("; "))
    for item in items:
        try:
            _canonical_text(item, field)
        except ValueError as error:
            raise GoldenSetValidationError(str(error)) from error
    if len(items) != len(set(items)):
        raise GoldenSetValidationError(f"{field} must be unique")
    return items


def _parse_review_bool(value: object, field: str) -> bool:
    if value == "TRUE":
        return True
    if value == "FALSE":
        return False
    raise GoldenSetValidationError(f"{field} must be TRUE or FALSE")


def _constituent_row_identity(
    ws: Worksheet, row: int, headers: dict[str, int]
) -> tuple[str, str, str]:
    code = ws.cell(row, headers["Concept Code"]).value
    if not isinstance(code, str):
        raise GoldenSetValidationError("constituent concept code must be text")
    if ws.row_dimensions[row].hidden:
        raise GoldenSetValidationError(
            f"hidden constituent rows are not permitted: {row}"
        )
    row_type = _cell_text(ws, row, headers["Row Type"], f"{code} row type")
    if row_type not in _CONSTITUENT_ROW_TYPES:
        raise GoldenSetValidationError(f"{code} has invalid row type: {row_type}")
    action = _cell_text(ws, row, headers["SME Action"], f"{code} SME action")
    return code, row_type, action


def _constituent_decision_sheet(
    workbook: Workbook,
) -> tuple[Worksheet, dict[str, int]]:
    ws = workbook["Constituent Decisions"]
    headers = _headers(ws, 4)
    required = {
        "Concept Code",
        "Row Type",
        "SME Action",
        "Expected Axis",
        "Expected Filler",
        "Expected Group",
        "Expected needs_review",
        "Expected Provenance Status",
        "Expected Proposal ID",
        "Row Complete?",
    }
    if missing := required - headers.keys():
        raise GoldenSetValidationError(
            "Constituent Decisions is missing headers: " + ", ".join(sorted(missing))
        )
    return ws, headers


def _constituent_row_decision(
    ws: Worksheet, row: int, headers: dict[str, int]
) -> ConstituentRowDecision:
    """Validate one reviewer decision row without inferring an expectation.

    `Expected Axis` and `Expected Filler` are read on *every* row, not only on the
    kept ones: required on `include`/`revise`, optional but canonically validated
    on `exclude`/`not-needed`. That is a deliberate tightening over the reader this
    replaced, which skipped non-kept rows before reaching those columns and so
    accepted a padded or non-textual value there. The export writes those cells out
    verbatim, so a defect in them is a workbook defect worth failing on rather than
    something to record and propagate. It is fail-closed and cannot alter the
    oracle: a row it rejects would previously have been dropped, never kept.

    On a `not-needed` row the pair must be *blank*: `UnusedCandidateRow` has no
    field to hold it, and a populated pair contradicts the action that says the
    reviewer never had to fill the row in.

    `Row Complete?` is waived only for a `not-needed` row, and the two checks that
    precede the waiver -- `PENDING`, then the `ENGINE SUGGESTION` / `not-needed`
    combination -- ensure it cannot reach a suggestion. Those checks live here, at
    the workbook boundary, so the reviewer is told which cell to fix; the invariant
    itself is carried by `UnusedCandidateRow`, which every other entry point sees.
    """
    code, row_type, action = _constituent_row_identity(ws, row, headers)
    complete = _cell_text(
        ws, row, headers["Row Complete?"], f"{code} constituent completeness"
    )
    if row_type == "ENGINE SUGGESTION" and action == "PENDING":
        raise GoldenSetValidationError(f"{code} has pending constituent action")
    if action not in _SME_ACTIONS:
        raise GoldenSetValidationError(f"{code} has invalid SME action: {action}")
    if row_type == "ENGINE SUGGESTION" and action == "not-needed":
        raise GoldenSetValidationError(
            f"{code} engine suggestion cannot be left not-needed"
        )
    if action != "not-needed" and complete != "YES":
        raise GoldenSetValidationError(f"{code} has incomplete constituent row")
    axis_column, filler_column = headers["Expected Axis"], headers["Expected Filler"]
    axis_field, filler_field = f"{code} expected axis", f"{code} expected filler"
    read = _cell_text if action in _KEPT_SME_ACTIONS else _optional_text
    axis = read(ws, row, axis_column, axis_field)
    filler = read(ws, row, filler_column, filler_field)
    payload: dict[str, object] = {
        "code": code,
        "row_type": row_type,
        "sme_action": action,
    }
    if action == "not-needed":
        if axis is not None or filler is not None:
            raise GoldenSetValidationError(
                f"{code} not-needed candidate row must not record an expected pair"
            )
    else:
        payload["expected_axis"] = axis
        payload["expected_filler"] = filler
    try:
        return _ROW_DECISION_ADAPTER.validate_python(payload)
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error


def _constituent_decisions(
    ws: Worksheet, headers: dict[str, int]
) -> Iterator[tuple[int, ConstituentRowDecision]]:
    """Yield every populated reviewer decision row with its sheet row number."""
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, headers["Concept Code"]).value is None:
            if _row_has_data(ws, row):
                raise GoldenSetValidationError(
                    f"populated constituent row has blank concept code: {row}"
                )
            continue
        yield row, _constituent_row_decision(ws, row, headers)


def _kept_constituent(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    decision: KeptRow,
) -> GoldenConstituent:
    """Build the expectation a kept row contributes to the oracle."""
    code = decision.code
    return GoldenConstituent(
        axis=decision.expected_axis,
        filler=decision.expected_filler,
        relationship_group=_optional_text(
            ws, row, headers["Expected Group"], f"{code} expected group"
        ),
        needs_review=_parse_review_bool(
            ws.cell(row, headers["Expected needs_review"]).value,
            f"{code} expected needs_review",
        ),
        provenance_status=cast(
            "PairProvenance",
            _cell_text(
                ws,
                row,
                headers["Expected Provenance Status"],
                f"{code} expected provenance status",
            ),
        ),
        proposal_id=_optional_text(
            ws,
            row,
            headers["Expected Proposal ID"],
            f"{code} expected proposal ID",
        ),
    )


def _declared_concept_codes(ws: Worksheet, headers: dict[str, int]) -> set[str]:
    """Every textual code the reviewer declared on `Concept Decisions`."""
    return {
        code
        for row in range(5, ws.max_row + 1)
        if isinstance((code := ws.cell(row, headers["Concept Code"]).value), str)
    }


def _reject_orphan_constituents(
    constituent_codes: set[str], declared_codes: set[str]
) -> None:
    """Fail on a constituent row whose concept was never adjudicated.

    Nothing licenses such a row: the SME recorded no decision for that concept, so
    its constituent decisions are unattested. Both the oracle import and the
    row-decision export apply this, because an orphan row silently enlarges the
    acceptance denominator.
    """
    orphaned = sorted(constituent_codes - declared_codes)
    if orphaned:
        raise GoldenSetValidationError(
            "constituent rows reference unknown concepts: " + ", ".join(orphaned)
        )


def _workbook_constituents(
    workbook: Workbook,
) -> tuple[dict[str, list[GoldenConstituent]], set[str]]:
    ws, headers = _constituent_decision_sheet(workbook)
    result: dict[str, list[GoldenConstituent]] = {}
    row_codes: set[str] = set()
    for row, decision in _constituent_decisions(ws, headers):
        row_codes.add(decision.code)
        if isinstance(decision, KeptRow):
            result.setdefault(decision.code, []).append(
                _kept_constituent(ws, row, headers, decision)
            )
    return result, row_codes


def _load_review_workbook(content: bytes) -> Workbook:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    except (BadZipFile, OSError, ValueError) as error:
        raise GoldenSetValidationError(
            f"cannot read adjudication workbook: {error}"
        ) from error
    sheet_names = set(workbook.sheetnames)
    unexpected_sheets = sheet_names - _EXPECTED_SHEETS
    if not sheet_names >= _EXPECTED_SHEETS or not unexpected_sheets <= (
        _OPTIONAL_REVIEW_SHEETS
    ):
        raise GoldenSetValidationError(
            "workbook sheets do not match the review contract"
        )
    for sheet_name in _EXPECTED_SHEETS:
        if workbook[sheet_name].sheet_state != "visible":
            raise GoldenSetValidationError(
                f"reviewer input sheet must be visible: {sheet_name}"
            )
    _reject_hidden_reviewer_columns(workbook)
    _reject_reviewer_formulas(workbook)
    return workbook


def _reviewer_from_workbook(workbook: Workbook) -> Reviewer:
    reviewer_ws = workbook["Reviewer & Attestation"]
    hidden = [
        row
        for row in range(1, reviewer_ws.max_row + 1)
        if reviewer_ws.row_dimensions[row].hidden and _row_has_data(reviewer_ws, row)
    ]
    if hidden:
        raise GoldenSetValidationError(
            "hidden reviewer rows are not permitted: "
            + ", ".join(str(row) for row in hidden)
        )
    reviewed_at_value = reviewer_ws["B7"].value
    if isinstance(reviewed_at_value, datetime):
        reviewed_at = reviewed_at_value.date().isoformat()
    elif isinstance(reviewed_at_value, date):
        reviewed_at = reviewed_at_value.isoformat()
    else:
        reviewed_at = reviewed_at_value
    if reviewer_ws["B9"].value != "ATTESTED":
        raise GoldenSetValidationError("reviewer attestation is pending")
    return Reviewer(
        name=_cell_text(reviewer_ws, 5, 2, "reviewer name"),
        qualification_or_role=_cell_text(reviewer_ws, 6, 2, "reviewer qualification"),
        reviewed_at=cast("str", reviewed_at),
    )


def _required_evidence(workbook: Workbook) -> dict[str, str]:
    evidence = _evidence_values(workbook)
    required_evidence = {
        "NCIt release",
        "Source identity",
        "Sample identity",
        "Engine run",
        "Run fingerprint identity",
        "Artifact SHA-256",
        "Engine evidence identity",
        "Corpus evidence identity",
        "Detector identity",
        "Proposal registry identity",
    }
    if missing := required_evidence - evidence.keys():
        raise GoldenSetValidationError(
            "Source & Run Evidence is missing: " + ", ".join(sorted(missing))
        )
    return evidence


def _concept_headers(ws: Worksheet) -> dict[str, int]:
    headers = _headers(ws, 4)
    required_headers = {
        "Concept Code",
        "Source Label",
        "Source Semantic Types",
        "Expected Semantic Types",
        "SME Decision Status",
        "Expected Outcome",
        "Rationale / Required Follow-up",
        "Source Reviewed?",
        "Concept Complete?",
    }
    if missing := required_headers - headers.keys():
        raise GoldenSetValidationError(
            "Concept Decisions is missing headers: " + ", ".join(sorted(missing))
        )
    return headers


def _concept_decision_sheet(workbook: Workbook) -> tuple[Worksheet, dict[str, int]]:
    """Return `Concept Decisions` only once its declarations can be trusted.

    Everything downstream asks this sheet which concepts the reviewer declared, and
    two ways of answering that question were wrong before the question was asked: a
    hidden row declares a concept no human reading the workbook can see, and a
    populated row with an empty code cell declares nothing while looking adjudicated.

    Both guards used to live inside `_workbook_concepts`, so only the oracle import
    ran them; `export_row_decisions_bytes` called `_concept_headers` and
    `_declared_concept_codes` directly and reached the orphan check with a code set
    the import would have refused to build. A hidden row declaring `C999999` plus a
    matching constituent row therefore exported cleanly and failed to import. The
    preconditions live here so neither path can reach the declared codes without
    them.
    """
    ws = workbook["Concept Decisions"]
    headers = _concept_headers(ws)
    hidden = [
        row
        for row in range(5, ws.max_row + 1)
        if ws.row_dimensions[row].hidden
        and ws.cell(row, headers["Concept Code"]).value is not None
    ]
    if hidden:
        raise GoldenSetValidationError(
            "hidden concept rows are not permitted: "
            + ", ".join(str(row) for row in hidden)
        )
    blank = [
        row
        for row in range(5, ws.max_row + 1)
        if ws.cell(row, headers["Concept Code"]).value is None
        and _row_has_data(ws, row)
    ]
    if blank:
        raise GoldenSetValidationError(
            "populated concept row has blank concept code: "
            + ", ".join(str(row) for row in blank)
        )
    return ws, headers


def _decision_status(
    ws: Worksheet, row: int, headers: dict[str, int], code: str
) -> DecisionStatus:
    status = _cell_text(
        ws, row, headers["SME Decision Status"], f"{code} adjudication status"
    )
    if status == "PENDING":
        raise GoldenSetValidationError(f"{code} has pending adjudication")
    if status not in {"accepted", "rejected", "revision-needed"}:
        raise GoldenSetValidationError(f"{code} has invalid adjudication status")
    if ws.cell(row, headers["Source Reviewed?"]).value != "YES":
        raise GoldenSetValidationError(f"{code} source review is incomplete")
    if ws.cell(row, headers["Concept Complete?"]).value != "YES":
        raise GoldenSetValidationError(f"{code} adjudication is incomplete")
    return cast("DecisionStatus", status)


def _expectation_from_row(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    code: str,
    status: DecisionStatus,
    expected_constituents: dict[str, list[GoldenConstituent]],
) -> GoldenExpectation | None:
    if status != "accepted":
        if expected_constituents.get(code):
            raise GoldenSetValidationError(
                f"{code} nonaccepted decision must not define expected constituents"
            )
        return None
    outcome = _cell_text(
        ws, row, headers["Expected Outcome"], f"{code} expected outcome"
    )
    semantic_types = _parse_semantic_types(
        _cell_text(
            ws,
            row,
            headers["Expected Semantic Types"],
            f"{code} semantic types",
        ),
        f"{code} semantic types",
    )
    return GoldenExpectation(
        outcome=cast("ExpectedOutcome", outcome),
        semantic_types=semantic_types,
        constituents=tuple(expected_constituents.get(code, [])),
    )


def _concept_from_row(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    expected_constituents: dict[str, list[GoldenConstituent]],
) -> AdjudicatedConcept | None:
    code_value = ws.cell(row, headers["Concept Code"]).value
    if code_value is None:
        if _row_has_data(ws, row):
            raise GoldenSetValidationError(
                f"populated concept row has blank concept code: {row}"
            )
        return None
    if not isinstance(code_value, str):
        raise GoldenSetValidationError("concept code must be text")
    status = _decision_status(ws, row, headers, code_value)
    return AdjudicatedConcept(
        code=code_value,
        label=_cell_text(ws, row, headers["Source Label"], f"{code_value} label"),
        adjudication=AdjudicationDecision(
            status=status,
            rationale=_cell_text(
                ws,
                row,
                headers["Rationale / Required Follow-up"],
                f"{code_value} rationale",
            ),
        ),
        expected=_expectation_from_row(
            ws,
            row,
            headers,
            code_value,
            status,
            expected_constituents,
        ),
    )


def _workbook_concepts(workbook: Workbook) -> tuple[AdjudicatedConcept, ...]:
    ws, headers = _concept_decision_sheet(workbook)
    expected_constituents, constituent_codes = _workbook_constituents(workbook)
    _reject_orphan_constituents(constituent_codes, _declared_concept_codes(ws, headers))
    concepts = tuple(
        concept
        for row in range(5, ws.max_row + 1)
        if (concept := _concept_from_row(ws, row, headers, expected_constituents))
    )
    return concepts


def import_adjudication_workbook_bytes(
    workbook_bytes: bytes,
    proposal_registry: ProposalRegistry | None = None,
) -> AdjudicationArtifact:
    """Import and validate one immutable issue #57 workbook snapshot."""
    workbook = _load_review_workbook(workbook_bytes)
    reviewer = _reviewer_from_workbook(workbook)
    evidence = _required_evidence(workbook)
    try:
        payload = {
            "_meta": AdjudicationMetadata(
                schema_version=_SCHEMA_VERSION,
                status=_ADJUDICATED_STATUS,
                ncit_version=evidence["NCIt release"],
                source_identity=evidence["Source identity"],
                sample_manifest_identity=evidence["Sample identity"],
                run_id=evidence["Engine run"],
                run_fingerprint_identity=evidence["Run fingerprint identity"],
                engine_artifact_identity=evidence["Artifact SHA-256"],
                engine_evidence_identity=evidence["Engine evidence identity"],
                corpus_evidence_identity=evidence["Corpus evidence identity"],
                detector_identity=evidence["Detector identity"],
                proposal_registry_identity=evidence["Proposal registry identity"],
                workbook_identity=hashlib.sha256(workbook_bytes).hexdigest(),
                reviewer=reviewer,
            ),
            "concepts": _workbook_concepts(workbook),
        }
        artifact = AdjudicationArtifact.model_validate(
            {
                **payload,
                "artifact_identity": _payload_identity(
                    {
                        "_meta": payload["_meta"].model_dump(mode="json"),
                        "concepts": [
                            concept.model_dump(mode="json")
                            for concept in payload["concepts"]
                        ],
                    }
                ),
            }
        )
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error
    _validate_proposal_registry_binding(artifact, proposal_registry)
    return artifact


def import_adjudication_workbook(
    path: str | Path,
    proposal_registry: ProposalRegistry | None = None,
) -> AdjudicationArtifact:
    """Import the issue #57 workbook without inferring any reviewer decision."""
    workbook_path = Path(path)
    try:
        workbook_bytes = workbook_path.read_bytes()
    except OSError as error:
        raise GoldenSetValidationError(
            f"cannot read adjudication workbook: {error}"
        ) from error
    return import_adjudication_workbook_bytes(workbook_bytes, proposal_registry)


def export_row_decisions_bytes(
    workbook_bytes: bytes, source_workbook: str
) -> RowDecisionExport:
    """Export every constituent decision row of one workbook snapshot.

    The oracle import discards `exclude` and `not-needed` rows, which is what makes
    the engine's acceptance rate unrecoverable from the artifact alone. This keeps
    the rows instead.

    It runs the workbook-level tamper gates and the shared row reader, and no more.
    Specifically it runs:

    - `_load_review_workbook`: the sheet-name contract, every reviewer sheet
      visible, no hidden reviewer column carrying data, no formula cell in reviewer
      input.
    - `_reviewer_from_workbook`: no hidden populated reviewer row, `ATTESTED`
      attestation, canonical reviewer name and qualification, ISO `reviewed_at`.
    - `_required_evidence`: no hidden populated `Source & Run Evidence` row, no
      duplicate evidence key, all ten required evidence keys present.
    - `_constituent_decision_sheet`: no duplicate header, all ten required
      `Constituent Decisions` headers present.
    - `_constituent_decisions` / `_constituent_row_decision`: no populated row with
      a blank concept code, no hidden constituent row, textual concept code
      matching `^C[0-9]+$`, a known row type and SME action, no `PENDING` engine
      suggestion, no `not-needed` engine suggestion, `Row Complete?` `YES` on every
      row but a `not-needed` one, `Expected Axis`/`Expected Filler` canonical on
      every row and required on a kept one, and blank on a `not-needed` one.
    - `_concept_decision_sheet` / `_reject_orphan_constituents`: all nine required
      `Concept Decisions` headers present, no hidden row declaring a concept, no
      populated row with a blank concept code, and no constituent row referencing a
      concept the reviewer never declared. These are the `Concept Decisions` gates
      the export shares with the import, because an orphan row enlarges the
      acceptance denominator with a concept nobody adjudicated — and a hidden or
      blank-coded declaration decides which rows count as orphans.
    - `RowDecisionExport`: a nonempty row set in which every row carrying an
      expected pair is unique on `(code, axis, filler)` and no such triple is both
      kept and withdrawn, signed by a `payload_identity` over the rows themselves.
      `workbook_identity` hashes the `.xlsx`; `payload_identity` hashes what was
      read out of it, so an edit to the tracked JSON is a load failure. A row
      carrying *no* expected pair has no identity here — the export does not record
      the engine's suggested axis and filler — so such rows are unconstrained by
      the uniqueness rule.

    It does **not** run `_kept_constituent`, so the `Expected Provenance Status`,
    `Expected needs_review`, `Expected Group` and `Expected Proposal ID` gates —
    and the whole of `GoldenConstituent` — belong to the import alone. Nor does it
    read any *decision* on `Concept Decisions` or apply the M1 cohort,
    artifact-identity and proposal-registry gates. A workbook the export accepts
    can therefore still be rejected by `import-workbook`; the export is evidence of
    what the reviewer recorded, never a substitute for validating the oracle.
    """
    workbook = _load_review_workbook(workbook_bytes)
    reviewer = _reviewer_from_workbook(workbook)
    evidence = _required_evidence(workbook)
    ws, headers = _constituent_decision_sheet(workbook)
    rows = tuple(decision for _, decision in _constituent_decisions(ws, headers))
    concept_ws, concept_headers = _concept_decision_sheet(workbook)
    _reject_orphan_constituents(
        {decision.code for decision in rows},
        _declared_concept_codes(concept_ws, concept_headers),
    )
    try:
        meta = RowDecisionMetadata(
            schema_version=_ROW_DECISION_SCHEMA_VERSION,
            ncit_version=evidence["NCIt release"],
            source_workbook=source_workbook,
            workbook_identity=hashlib.sha256(workbook_bytes).hexdigest(),
            source_identity=evidence["Source identity"],
            run_id=evidence["Engine run"],
            engine_evidence_identity=evidence["Engine evidence identity"],
            reviewer=reviewer,
        )
        # Bind the identity *after* the payload exists (D61). Nothing here is
        # pre-declared: the hash is taken over the rows that were actually read.
        payload: dict[str, object] = {
            "_meta": meta.model_dump(mode="json"),
            "rows": tuple(row.model_dump(mode="json") for row in rows),
        }
        return RowDecisionExport.model_validate(
            {**payload, "payload_identity": _payload_identity(payload)}
        )
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error


def export_row_decisions(path: str | Path) -> RowDecisionExport:
    """Export the row-level SME decisions from an attested review workbook."""
    workbook_path = Path(path)
    try:
        workbook_bytes = workbook_path.read_bytes()
    except OSError as error:
        raise GoldenSetValidationError(
            f"cannot read adjudication workbook: {error}"
        ) from error
    return export_row_decisions_bytes(workbook_bytes, workbook_path.name)


def load_row_decisions(path: str | Path) -> RowDecisionExport:
    """Load the tracked row-decision export, rejecting a hand edit that broke it."""
    raw = read_json_without_duplicates(path)
    if not isinstance(raw, dict):
        raise GoldenSetValidationError("row decisions must be a JSON object")
    if isinstance(raw.get("rows"), list):
        raw["rows"] = tuple(raw["rows"])
    try:
        return RowDecisionExport.model_validate(raw)
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error


def _normalize_engine(value: dict[str, object]) -> dict[str, object]:
    normalized = dict(value)
    concepts = normalized.get("concepts")
    if isinstance(concepts, list):
        normalized["concepts"] = tuple(
            {
                **concept,
                "semantic_types": tuple(concept.get("semantic_types", ())),
                "constituents": tuple(concept.get("constituents", ())),
            }
            if isinstance(concept, dict)
            else concept
            for concept in concepts
        )
    residual = normalized.get("residual_precoordinated_codes")
    if isinstance(residual, list):
        normalized["residual_precoordinated_codes"] = tuple(residual)
    return normalized


def _normalize_residual(value: dict[str, object]) -> dict[str, object]:
    normalized = dict(value)
    for field in ("denominator_codes", "residual_codes"):
        field_value = normalized.get(field)
        if isinstance(field_value, list):
            normalized[field] = tuple(field_value)
    return normalized


def _model_validate(model: type[_StrictModel], value: object) -> object:
    if model is EngineEvidence and isinstance(value, dict):
        value = _normalize_engine(value)
    elif model is ResidualComparisonInput and isinstance(value, dict):
        value = _normalize_residual(value)
    try:
        return model.model_validate(value)
    except (ValidationError, ValueError) as error:
        raise _model_error(error) from error


def _require_engine_identity(
    artifact: AdjudicationArtifact, engine: EngineEvidence
) -> None:
    checks = (
        ("NCIt version", artifact.meta.ncit_version, engine.ncit_version),
        ("source identity", artifact.meta.source_identity, engine.source_identity),
        (
            "sample manifest identity",
            artifact.meta.sample_manifest_identity,
            engine.sample_manifest_identity,
        ),
        ("run id", artifact.meta.run_id, engine.run_id),
        (
            "run fingerprint identity",
            artifact.meta.run_fingerprint_identity,
            engine.run_fingerprint_identity,
        ),
        (
            "engine artifact identity",
            artifact.meta.engine_artifact_identity,
            engine.engine_artifact_identity,
        ),
        (
            "engine evidence identity",
            artifact.meta.engine_evidence_identity,
            engine.evidence_identity,
        ),
        (
            "detector identity",
            artifact.meta.detector_identity,
            engine.detector_identity,
        ),
    )
    for name, expected, actual in checks:
        if expected != actual:
            raise GoldenSetValidationError(f"{name} does not match adjudication")
    artifact_codes = [concept.code for concept in artifact.concepts]
    engine_codes = [concept.code for concept in engine.concepts]
    if artifact_codes != engine_codes:
        raise GoldenSetValidationError(
            "engine worklist does not match adjudication order"
        )


def _group_partition(
    constituents: tuple[Constituent, ...], excluded: set[ConstituentPair]
) -> dict[str, object]:
    grouped: dict[str, set[ConstituentPair]] = {}
    ungrouped: set[ConstituentPair] = set()
    for item in constituents:
        if item.pair in excluded:
            continue
        if item.relationship_group is None:
            ungrouped.add(item.pair)
        else:
            grouped.setdefault(item.relationship_group, set()).add(item.pair)
    groups = sorted(
        (sorted([list(pair) for pair in members]) for members in grouped.values()),
        key=lambda value: json.dumps(value, separators=(",", ":")),
    )
    return {
        "groups": groups,
        "ungrouped": sorted([list(pair) for pair in ungrouped]),
    }


def _score_dict(result: ExtractionScore) -> dict[str, object]:
    precision = result.precision if result.actual else None
    recall = result.recall if result.expected else None
    f1 = result.f1 if precision is not None and recall is not None else None
    return {
        "expected": result.expected,
        "actual": result.actual,
        "true_positive": result.true_positive,
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "missing": sorted([list(pair) for pair in result.missing]),
        "extra": sorted([list(pair) for pair in result.extra]),
        "deferred": sorted([list(pair) for pair in result.deferred]),
    }


def _micro_score(expected: int, actual: int, true_positive: int) -> dict[str, object]:
    return {
        "expected": expected,
        "actual": actual,
        "true_positive": true_positive,
        "precision": true_positive / actual if actual else None,
        "recall": true_positive / expected if expected else None,
    }


def _axis_scores(
    counts: dict[str, Counter[str]],
) -> dict[str, dict[str, object]]:
    return {
        axis: _micro_score(
            values["expected"], values["actual"], values["true_positive"]
        )
        for axis, values in sorted(counts.items())
    }


def _proposal_counts(
    expected: GoldenExpectation, actual: EngineConcept
) -> tuple[Counter[str], Counter[str], int, tuple[str, ...]]:
    provenance = Counter(item.provenance_status for item in expected.constituents)
    proposal_statuses = Counter(
        item.provenance_status
        for item in expected.constituents
        if item.provenance_status != "ncit-26.07d"
    )
    augmented_expected = sum(
        not item.needs_review
        and item.provenance_status
        in {"locally-approved", "submitted", "accepted-in-ncit"}
        for item in expected.constituents
    )
    engine_ids = tuple(
        item.filler for item in actual.constituents if item.filler.startswith("MINT-")
    )
    return provenance, proposal_statuses, augmented_expected, engine_ids


def _update_axis_counts(
    counts: dict[str, Counter[str]],
    expected: set[ConstituentPair],
    actual: set[ConstituentPair],
) -> None:
    for pair in expected | actual:
        values = counts.setdefault(pair[0], Counter())
        values["expected"] += pair in expected
        values["actual"] += pair in actual
        values["true_positive"] += pair in expected & actual


def _pair_modality(pair: ConstituentPair) -> str:
    contract = AXIS_CONTRACTS.get(pair[0])
    if contract is not None:
        return contract.modality
    if re.fullmatch(r"R[0-9]+", pair[0]) is not None:
        return "asserted"
    raise GoldenSetValidationError(f"unknown normalized axis: {pair[0]}")


def _is_non_defining(pair: ConstituentPair) -> bool:
    return _pair_modality(pair) == "non-defining"


def _score_concept_views(
    expected: GoldenExpectation,
    actual: EngineConcept,
    aggregates: dict[str, dict[str, int]],
    axis_aggregates: dict[str, dict[str, Counter[str]]],
    deferrals: dict[str, dict[str, int]],
) -> tuple[
    dict[str, tuple[GoldenConstituent, ...]],
    dict[str, ExtractionScore],
    set[ConstituentPair],
    set[ConstituentPair],
]:
    actual_pairs = {item.pair for item in actual.constituents}
    engine_review_flags = {
        item.pair for item in actual.constituents if item.needs_review
    }
    view_items = {
        "ncit_bound": tuple(
            item
            for item in expected.constituents
            if item.provenance_status == "ncit-26.07d"
        ),
        "augmented": tuple(
            item
            for item in expected.constituents
            if item.provenance_status != "proposed"
        ),
        "defining_only": tuple(
            item
            for item in expected.constituents
            if item.provenance_status == "ncit-26.07d"
            and not _is_non_defining(item.pair)
        ),
        "non_defining": tuple(
            item
            for item in expected.constituents
            if item.provenance_status == "ncit-26.07d" and _is_non_defining(item.pair)
        ),
    }
    results: dict[str, ExtractionScore] = {}
    augmented_exclusions: set[ConstituentPair] = set()
    for view, items in view_items.items():
        view_actual_pairs = (
            {pair for pair in actual_pairs if not _is_non_defining(pair)}
            if view == "defining_only"
            else {pair for pair in actual_pairs if _is_non_defining(pair)}
            if view == "non_defining"
            else actual_pairs
        )
        expected_pairs = {item.pair for item in items}
        expected_exclusions = {item.pair for item in items if item.needs_review}
        result = score(
            expected_pairs,
            view_actual_pairs,
            expected_needs_review=expected_exclusions,
        )
        results[view] = result
        for field in ("expected", "actual", "true_positive"):
            aggregates[view][field] += getattr(result, field)
        deferrals[view]["expected"] += len(items)
        deferrals[view]["deferred"] += len(expected_exclusions)
        deferrals[view]["engine_matches"] += len(
            expected_exclusions & view_actual_pairs
        )
        exclusions = expected_exclusions
        _update_axis_counts(
            axis_aggregates[view],
            expected_pairs - exclusions,
            view_actual_pairs - exclusions,
        )
        if view == "augmented":
            augmented_exclusions = expected_exclusions
    return view_items, results, engine_review_flags, augmented_exclusions


def _concept_report(
    concept: AdjudicatedConcept,
    expected: GoldenExpectation,
    actual: EngineConcept,
    view_items: dict[str, tuple[GoldenConstituent, ...]],
    results: dict[str, ExtractionScore],
    engine_review_flags: set[ConstituentPair],
    expected_exclusions: set[ConstituentPair],
) -> dict[str, object]:
    exclusions = expected_exclusions
    expected_groups = _group_partition(view_items["augmented"], exclusions)
    actual_groups = _group_partition(actual.constituents, exclusions)
    return {
        "code": concept.code,
        "expected_outcome": expected.outcome,
        "actual_outcome": actual.outcome,
        "outcome_match": expected.outcome == actual.outcome,
        "expected_semantic_types": sorted(expected.semantic_types),
        "actual_semantic_types": sorted(actual.semantic_types),
        "semantic_types_match": set(expected.semantic_types)
        == set(actual.semantic_types),
        "expected_review_exclusions": sorted(
            [list(pair) for pair in expected_exclusions]
        ),
        "engine_review_flags": sorted([list(pair) for pair in engine_review_flags]),
        "expected_pair_modality": [
            {
                "axis": item.axis,
                "filler": item.filler,
                "modality": _pair_modality(item.pair),
            }
            for item in sorted(view_items["augmented"], key=lambda value: value.pair)
        ],
        "pair_score": {view: _score_dict(result) for view, result in results.items()},
        "expected_group_partition": expected_groups,
        "actual_group_partition": actual_groups,
        "group_match": expected_groups == actual_groups,
    }


def _residual_dict(value: ResidualComparisonInput) -> dict[str, object]:
    denominator = len(value.denominator_codes)
    count = len(value.residual_codes)
    return {
        "name": value.name,
        "source_identity": value.source_identity,
        "sample_manifest_identity": value.sample_manifest_identity,
        "run_id": value.run_id,
        "run_fingerprint_identity": value.run_fingerprint_identity,
        "engine_artifact_identity": value.engine_artifact_identity,
        "detector_identity": value.detector_identity,
        "evidence_identity": value.evidence_identity,
        "denominator_codes": list(value.denominator_codes),
        "residual_codes": list(value.residual_codes),
        "count": count,
        "denominator": denominator,
        "rate": count / denominator if denominator else None,
    }


def evaluate_adjudication(
    artifact: AdjudicationArtifact,
    raw_engine: object,
    raw_corpus_comparison: object,
) -> dict[str, object]:
    """Evaluate accepted expectations without laundering identity or review failures."""
    engine = cast("EngineEvidence", _model_validate(EngineEvidence, raw_engine))
    _require_engine_identity(artifact, engine)
    corpus = cast(
        "ResidualComparisonInput",
        _model_validate(ResidualComparisonInput, raw_corpus_comparison),
    )
    if corpus.source_identity != artifact.meta.source_identity:
        raise GoldenSetValidationError(
            "corpus source identity does not match adjudication"
        )
    if corpus.detector_identity != artifact.meta.detector_identity:
        raise GoldenSetValidationError(
            "corpus detector identity does not match adjudication"
        )
    # The corpus evidence_identity is recorded, never pre-declared. It hashes a payload
    # containing run_id = f"{branch}-{uuid4()}" (`_new_run_id`), so a value
    # written into the workbook before the corpus run exists can never be matched by
    # that run. _residual_dict(corpus) still carries the identity into the report, which
    # keeps the comparison auditable; asserting equality here only made the gate
    # unsatisfiable by construction.
    engine_by_code = {concept.code: concept for concept in engine.concepts}
    concept_reports: list[dict[str, object]] = []
    aggregates = {
        "ncit_bound": {"expected": 0, "actual": 0, "true_positive": 0},
        "augmented": {"expected": 0, "actual": 0, "true_positive": 0},
        "defining_only": {"expected": 0, "actual": 0, "true_positive": 0},
        "non_defining": {"expected": 0, "actual": 0, "true_positive": 0},
    }
    axis_aggregates: dict[str, dict[str, Counter[str]]] = {
        "ncit_bound": {},
        "augmented": {},
        "defining_only": {},
        "non_defining": {},
    }
    deferrals = {
        "ncit_bound": {"expected": 0, "deferred": 0, "engine_matches": 0},
        "augmented": {"expected": 0, "deferred": 0, "engine_matches": 0},
        "defining_only": {"expected": 0, "deferred": 0, "engine_matches": 0},
        "non_defining": {"expected": 0, "deferred": 0, "engine_matches": 0},
    }
    provenance_counts: Counter[str] = Counter()
    proposal_status_counts: Counter[str] = Counter()
    augmented_proposal_expected = 0
    engine_proposal_emissions = 0
    engine_proposal_ids: set[str] = set()
    actual_decomposed: list[str] = []
    group_matches: list[bool] = []
    for concept in artifact.concepts:
        if concept.adjudication.status != "accepted":
            continue
        expected = cast("GoldenExpectation", concept.expected)
        actual = engine_by_code[concept.code]
        provenance, statuses, accepted, engine_ids = _proposal_counts(expected, actual)
        provenance_counts.update(provenance)
        proposal_status_counts.update(statuses)
        augmented_proposal_expected += accepted
        engine_proposal_emissions += len(engine_ids)
        engine_proposal_ids.update(engine_ids)
        view_items, results, engine_review_flags, expected_exclusions = (
            _score_concept_views(
                expected, actual, aggregates, axis_aggregates, deferrals
            )
        )
        if actual.outcome == "decomposed":
            actual_decomposed.append(concept.code)
        concept_report = _concept_report(
            concept,
            expected,
            actual,
            view_items,
            results,
            engine_review_flags,
            expected_exclusions,
        )
        group_matches.append(cast("bool", concept_report["group_match"]))
        concept_reports.append(concept_report)
    residual_payload = {
        "name": "accepted-adjudication",
        "source_identity": artifact.meta.source_identity,
        "sample_manifest_identity": artifact.meta.sample_manifest_identity,
        "run_id": artifact.meta.run_id,
        "run_fingerprint_identity": artifact.meta.run_fingerprint_identity,
        "engine_artifact_identity": artifact.meta.engine_artifact_identity,
        "detector_identity": artifact.meta.detector_identity,
        "denominator_codes": tuple(actual_decomposed),
        "residual_codes": tuple(
            code
            for code in actual_decomposed
            if code in set(engine.residual_precoordinated_codes)
        ),
    }
    adjudication_residual = ResidualComparisonInput(
        **residual_payload,
        evidence_identity=_payload_identity(residual_payload),
    )
    adjudication_residual_dict = _residual_dict(adjudication_residual)
    corpus_dict = _residual_dict(corpus)
    return {
        "schema_version": 2,
        "adjudication_identity": artifact.identity,
        "accepted_concepts": len(concept_reports),
        "decision_counts": dict(
            Counter(c.adjudication.status for c in artifact.concepts)
        ),
        "pair_micro": {
            view: _micro_score(
                counts["expected"], counts["actual"], counts["true_positive"]
            )
            for view, counts in aggregates.items()
        },
        "pair_by_axis": {
            view: _axis_scores(counts) for view, counts in axis_aggregates.items()
        },
        "expected_pair_provenance": dict(sorted(provenance_counts.items())),
        "expected_pair_deferrals": {
            view: dict(values) for view, values in sorted(deferrals.items())
        },
        "proposal_governance": {
            "engine_emissions": engine_proposal_emissions,
            "distinct_engine_proposals": len(engine_proposal_ids),
            "augmented_expected": augmented_proposal_expected,
            "expected_by_status": dict(sorted(proposal_status_counts.items())),
        },
        "group_partition_agreement": {
            "concepts_agree": sum(group_matches),
            "concepts_disagree": len(group_matches) - sum(group_matches),
        },
        "concepts": concept_reports,
        "residual_comparison": {
            "metric": "D37 detector-relative residual_precoordination",
            "adjudication": adjudication_residual_dict,
            "corpus_sample": corpus_dict,
            "absolute_rate_delta": (
                abs(
                    cast("float", adjudication_residual_dict["rate"])
                    - cast("float", corpus_dict["rate"])
                )
                if adjudication_residual_dict["rate"] is not None
                and corpus_dict["rate"] is not None
                else None
            ),
            "rates_averaged": False,
        },
    }


def write_evaluation_report(report: dict[str, object], path: str | Path) -> None:
    """Write canonical JSON so identical evidence produces byte-identical reports."""
    rendered = json.dumps(
        report,
        indent=2,
        sort_keys=True,
        ensure_ascii=True,
    )
    Path(path).write_text(rendered + "\n", encoding="utf-8")
